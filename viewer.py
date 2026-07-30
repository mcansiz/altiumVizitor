"""
@file viewer.py
@brief Altium projelerini interaktif HTML görüntüleyiciye ve veri dosyalarına dönüştürür.

@details
Altium şematik (.SchDoc) ve PCB (.PcbDoc) dosyalarını okuyup tek-dosya, çevrimdışı
çalışan HTML görüntüleyiciler üretir: interaktif şematik, Altium benzeri PCB katman
görüntüleyici ve gerçek STEP modelleri + bakır/silkscreen dokusu içeren 3B board.
Ayrıca BOM / Pick&Place (CSV), IC bağlantı haritası / MCU pin listesi (Excel) ve
AI/LLM analizi için kompakt JSON dışa aktarımı sağlar. Üretilen HTML tamamen
bağımsızdır (Three.js gömülü, harici bağımlılık yok).

@author Mikail Cansız
@date 2026
@version 2.9.27
"""
# Schematic Viz Generator — Altium projelerini interaktif HTML görüntüleyiciye dönüştürür.
# Copyright (C) 2026  Mikail Cansız <cansizmikail@gmail.com>
#
# SPDX-License-Identifier: AGPL-3.0-or-later
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
from pathlib import Path
import json
import re
import datetime
import xml.etree.ElementTree as ET
from typing import Callable

from altium_monkey.altium_prjpcb import AltiumPrjPcb
from altium_monkey.altium_schdoc import AltiumSchDoc

# Uygulama sürümü — tek kaynak burası; gui.py buradan import eder.
# HTML çıktılarında sağ üst köşedeki rozette görünür (build saati yerine).
APP_VERSION = "2.9.41"

# Dikey pin adlarının doğru render edildiği minimum altium_monkey sürümü.
# Bu sürümden öncesinde STM32 gibi IC'lerde dikey pinler yatay çiziliyordu.
MIN_RECOMMENDED_AM = "2026.6.21"


def _check_altium_monkey_version(log):
    """@brief Kurulu altium_monkey sürümünü kontrol et, eskiyse dikey pin uyarısı ver.
    
    @param log Log mesajı callback'i (str alır)
    """
    try:
        from importlib import metadata
        ver = metadata.version("altium-monkey")
    except Exception:
        return
    # CalVer karşılaştırması (YYYY.M.D) — string parça parça sayısal
    def parse(v):
        """@brief parse()
        
        @param v
        @return Üretilen sonuç.
        """
        try:
            return tuple(int(x) for x in v.split("."))
        except Exception:
            return (0,)
    if parse(ver) < parse(MIN_RECOMMENDED_AM):
        log(f"  ! Not: altium_monkey {ver} kullanılıyor. Dikey pin adları "
            f"(STM32 vb.) {MIN_RECOMMENDED_AM} öncesinde yatay render edilir. "
            f"Güncelleme önerilir: pip install --upgrade altium-monkey")


THUMB_W = 700
THUMB_H = 500
GRID_COLS = 4
GAP = 80
SVG_NS = "http://www.w3.org/2000/svg"


def safe_id(name: str) -> str:
    """@brief safe_id()
    
    @param name Ad
    @return Üretilen sonuç.
    """
    s = re.sub(r"[^a-zA-Z0-9_-]", "_", name)
    return re.sub(r"_+", "_", s).strip("_") or "sheet"


def get_obj_text(obj):
    """@brief get_obj_text()
    
    @param obj
    @return Üretilen sonuç.
    """
    for a in ("text", "name", "net", "label", "signal", "net_name"):
        v = getattr(obj, a, None)
        if v:
            # Altium overbar notation'ını kaldır: 'ADC1_C\S\' -> 'ADC1_CS'
            # (SVG'de overbar görsel olarak üstçizgiyle çizilir, text'in içinde
            # backslash yoktur. Eşleşme için normalize ediyoruz.)
            return str(v).replace("\\", "")
    return None


def get_comp_field(comp, names, default=""):
    """@brief get_comp_field()
    
    @param comp
    @param names
    @param default Varsayılan değer
    @return Üretilen sonuç.
    """
    for n in names:
        v = getattr(comp, n, None)
        if v:
            return str(v)
    return default


def get_component_parameters(c) -> dict:
    """@brief Bir komponentin tüm ek parametrelerini topla (Manufacturer, Supplier, Stock vs).
    
    altium_monkey API'sinde parametreler farklı yapılarda olabilir, çoklu fallback denenir.
    
    @param c
    @return Üretilen sonuç.
    """
    result = {}
    # Yöntem 1: c.parameters (dict veya iterable)
    try:
        p = getattr(c, "parameters", None)
        if p:
            if hasattr(p, "items"):  # dict-like
                for k, v in p.items():
                    if v:
                        result[str(k)] = str(v)
            else:  # iterable of parameter objects
                for param in p:
                    name = getattr(param, "name", None)
                    val = getattr(param, "value", None) or getattr(param, "text", None)
                    if name and val:
                        result[str(name)] = str(val)
    except Exception:
        pass
    # Yöntem 2: get_parameter_dict()
    if not result:
        try:
            getter = getattr(c, "get_parameter_dict", None)
            if getter:
                p = getter()
                if p:
                    result = {str(k): str(v) for k, v in p.items() if v}
        except Exception:
            pass
    return result


def strip_aspect_ratio(svg_str: str) -> str:
    """@brief strip_aspect_ratio()
    
    @param svg_str SVG metni
    @return Üretilen sonuç.
    """
    if "preserveAspectRatio" in svg_str:
        return re.sub(
            r'preserveAspectRatio="[^"]*"',
            'preserveAspectRatio="none"',
            svg_str,
            count=1,
        )
    return svg_str.replace("<svg", '<svg preserveAspectRatio="none"', 1)


def collect_net_names_from_sheet(schdoc) -> set:
    """@brief Bir sayfadaki tüm net adlarını topla — sheet entries / block pinleri /
    
    cross-sheet & harness connector'lar dahil.
    
    @param schdoc Şema (.SchDoc) nesnesi
    @return Üretilen sonuç.
    """
    names = set()
    # Doğrudan net objeleri
    for getter_name in (
        "get_net_labels",
        "get_power_ports",
        "get_ports",
        "get_cross_sheet_connectors",
        "get_harness_connectors",
    ):
        try:
            getter = getattr(schdoc, getter_name, None)
            if getter:
                for obj in getter():
                    n = get_obj_text(obj)
                    if n:
                        names.add(n)
        except Exception:
            pass

    # Top-level sheet entries (eğer böyle bir property varsa)
    for attr in ("sheet_entries",):
        try:
            container = getattr(schdoc, attr, None)
            if container:
                for entry in container:
                    n = get_obj_text(entry)
                    if n:
                        names.add(n)
        except Exception:
            pass

    # Sheet symbol → child entries (block pinleri)
    try:
        for ss in schdoc.get_sheet_symbols():
            for child_attr in ("sheet_entries", "entries", "child_labels"):
                try:
                    children = getattr(ss, child_attr, None)
                    if children:
                        for c in children:
                            n = get_obj_text(c)
                            if n:
                                names.add(n)
                except Exception:
                    pass
    except Exception:
        pass

    return names


def extract_label_positions(svg_str: str, target_names: set) -> dict:
    """@brief extract_label_positions()
    
    @param svg_str SVG metni
    @param target_names Hedef ad listesi
    @return Üretilen sonuç.
    """
    if not target_names:
        return {}
    try:
        root = ET.fromstring(svg_str)
    except ET.ParseError:
        return {}

    vb = root.get("viewBox")
    if vb:
        try:
            vb_x, vb_y, vb_w, vb_h = (float(x) for x in vb.split())
        except ValueError:
            return {}
    else:
        try:
            w_raw = re.sub(r"[^\d.]", "", root.get("width", "1000")) or "1000"
            h_raw = re.sub(r"[^\d.]", "", root.get("height", "800")) or "800"
            vb_w, vb_h = float(w_raw), float(h_raw)
            vb_x, vb_y = 0.0, 0.0
        except ValueError:
            return {}
    if vb_w <= 0 or vb_h <= 0:
        return {}

    ns = "{%s}" % SVG_NS
    positions = {}
    for text_el in root.iter(ns + "text"):
        content = "".join(text_el.itertext()).strip()
        if not content or content not in target_names:
            continue
        x_str = text_el.get("x")
        y_str = text_el.get("y")
        if x_str is None or y_str is None:
            continue
        try:
            x = float(x_str.split()[0])
            y = float(y_str.split()[0])
        except ValueError:
            continue
        rx = (x - vb_x) / vb_w
        ry = (y - vb_y) / vb_h
        if 0 <= rx <= 1 and 0 <= ry <= 1:
            positions.setdefault(content, []).append((rx, ry))
    return positions


def _resolve_schdoc_paths(project, project_path, log):
    """@brief SchDoc dosyalarını bul (cross-platform).
    
    Sıra:
    1. project.get_reachable_schdoc_paths() — projenin kendi çözümü
    2. PrjPcb içindeki `DocumentPath=...SchDoc` satırlarını oku, ters slash'i
       normalize et (Windows'ta kaydedilen proje Linux'ta açılınca SCH ters-slash
       dosya, SCH ileri-slash dosya olur). Dosya gerçekten varsa kullan.
    3. Son çare: PrjPcb klasöründe rglob ile *.SchDoc tara.

    Bu, Windows'ta yapılıp Linux'ta açılan (veya tersi) projelerin sorunsuz
    çalışmasını sağlar — path ayracı farkı ve taşınma durumlarına dayanıklı.
    
    @param project Altium proje nesnesi / yolu
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    try:
        paths = list(project.get_reachable_schdoc_paths() or [])
    except Exception:
        paths = []
    if paths:
        return paths

    base = Path(project_path).parent

    # 2) PrjPcb metnini parse et
    try:
        text = Path(project_path).read_text(encoding="utf-8", errors="ignore")
        refs = re.findall(r"DocumentPath\s*=\s*(.+\.SchDoc)", text, re.IGNORECASE)
        resolved = []
        for ref in refs:
            # Ters slash → ileri slash, baştaki/sondaki boşlukları at
            rel = ref.strip().replace("\\", "/")
            candidate = (base / rel)
            if candidate.exists():
                resolved.append(candidate)
            else:
                # case-insensitive arama (Linux dosya sistemi büyük/küçük duyarlı)
                name = Path(rel).name.lower()
                match = next((p for p in base.rglob("*.SchDoc")
                             if p.name.lower() == name), None)
                if match:
                    resolved.append(match)
        # Tekrarsız, sıralı
        seen = set()
        unique = []
        for p in resolved:
            key = str(p).lower()
            if key not in seen:
                seen.add(key)
                unique.append(p)
        if unique:
            log(f"  · PrjPcb referanslarından {len(unique)} SchDoc çözüldü "
                f"(path normalize edildi).")
            return unique
    except Exception as e:
        log(f"  · PrjPcb parse edilemedi: {e}")

    # 3) Son çare: klasör taraması
    log("  · Klasörden taranıyor (*.SchDoc)...")
    found = sorted(base.rglob("*.SchDoc"))
    if found:
        log(f"  · {len(found)} SchDoc dosyadan bulundu.")
    return found


def _resolve_pcbdoc_paths(project_path, log):
    """@brief PcbDoc dosyalarını bul (cross-platform, kardeş klasör dahil).

    SchDoc çözümünün PCB karşılığı. Kör `rglob` proje klasörü altını tarar; oysa
    PcbDoc çoğu projede proje dosyasının KARDEŞ klasöründedir (ör. PrjPcb
    `PCB PROJECT/` içinde, PcbDoc `PCB/` içinde, referans `..\\PCB\\x.PcbDoc`).
    Bu durumda proje klasörü altını tarayan rglob PCB'yi bulamaz.

    Sıra:
    1. PrjPcb içindeki `DocumentPath=...PcbDoc` satırlarını oku, ters slash'i
       normalize et, `..\\` ile yukarı çıkışları çöz. Dosya gerçekten varsa kullan
       (case-insensitive eşleştirme de yapılır — Linux dosya sistemi için).
    2. Son çare: proje klasörü VE bir üst dizini (kardeş klasörler dahil) rglob
       ile *.PcbDoc tara.

    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param log Log mesajı callback'i (str alır)
    @return PcbDoc yollarının listesi (Path), tekrarsız.
    """
    base = Path(project_path).parent

    # 1) PrjPcb metnini parse et
    try:
        text = Path(project_path).read_text(encoding="utf-8", errors="ignore")
        refs = re.findall(r"DocumentPath\s*=\s*(.+\.PcbDoc)", text, re.IGNORECASE)
        resolved = []
        for ref in refs:
            rel = ref.strip().replace("\\", "/")
            candidate = (base / rel)
            if candidate.exists():
                resolved.append(candidate.resolve())
            else:
                # case-insensitive arama (base + üst dizin)
                name = Path(rel).name.lower()
                for root in (base, base.parent):
                    match = next((p for p in root.rglob("*.PcbDoc")
                                 if p.name.lower() == name), None)
                    if match:
                        resolved.append(match)
                        break
        seen = set()
        unique = []
        for p in resolved:
            key = str(p).lower()
            if key not in seen:
                seen.add(key)
                unique.append(p)
        if unique:
            log(f"  · PrjPcb referanslarından {len(unique)} PcbDoc çözüldü "
                f"(path normalize edildi).")
            return unique
    except Exception as e:
        log(f"  · PrjPcb PcbDoc için parse edilemedi: {e}")

    # 2) Son çare: klasör + üst dizin taraması (kardeş klasörler dahil)
    found = sorted(base.rglob("*.PcbDoc"))
    if not found and base.parent != base:
        found = sorted(base.parent.rglob("*.PcbDoc"))
    return found


def _collect_data(project_path: str, log, with_pcb=False, progress=None):
    """@brief Projeyi yükle, tüm sayfa/net/komponent verilerini topla.
    
    Hem generate_viewer hem generate_json bunu kullanır.
    with_pcb=True ise PCB cross-probe verisi de toplanır (maliyetli, sadece
    HTML viewer için).
    progress(frac, label): 0..1 arası şematik-faz ilerlemesi (opsiyonel).
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param log Log mesajı callback'i (str alır)
    @param with_pcb PCB dahil mi (bool)
    @param progress İlerleme callback'i (yüzde:int, etiket:str)
    @return Üretilen sonuç.
    """
    prog = progress or (lambda frac, label: None)
    log(f"Proje: {project_path}")
    _check_altium_monkey_version(log)
    project = AltiumPrjPcb(project_path)
    sch_paths = _resolve_schdoc_paths(project, project_path, log)
    log(f"{len(sch_paths)} şema bulundu.\n")

    sheets_raw = []
    components = []
    all_net_names = set()
    loaded_schdocs = []  # netlist derlemesi için SchDoc objeleri

    # Sheet adı → id eşleştirmesi (block hedeflerini çözmek için)
    sheet_name_to_id = {p.stem: safe_id(p.stem) for p in sch_paths}

    log("Pass 1: SchDoc yükleme + SVG render...")
    n_sheets = max(1, len(sch_paths))
    for idx, sch_path in enumerate(sch_paths):
        sheet_name = sch_path.stem
        prog(0.45 * idx / n_sheets, f"Sayfa render: {sheet_name}")
        sheet_id = safe_id(sheet_name)
        col = idx % GRID_COLS
        row = idx // GRID_COLS
        cx = col * (THUMB_W + GAP)
        cy = row * (THUMB_H + GAP)

        try:
            schdoc = AltiumSchDoc(sch_path)
            loaded_schdocs.append(schdoc)
            try:
                svg = schdoc.to_svg()
            except TypeError:
                svg = schdoc.to_svg(project_parameters=project.parameters)

            # Komponent kutusu (highlight) için sheet'in viewBox yüksekliği:
            # mils → SVG viewBox dönüşümü (1 viewBox birimi = 10 mils, y ters).
            _vbm = re.search(r'viewBox="[-\d.]+ [-\d.]+ [-\d.]+ ([-\d.]+)"', svg)
            vb_h = float(_vbm.group(1)) if _vbm else None

            sheet_nets = collect_net_names_from_sheet(schdoc)
            all_net_names.update(sheet_nets)

            try:
                for c in schdoc.get_components():
                    # Multi-part bilgisi (örn STM32 IC2A/IC2B/IC2C)
                    rec = getattr(c, "record", None)
                    part_id = getattr(rec, "current_part_id", None) if rec else None
                    part_count = getattr(rec, "part_count", 1) if rec else 1
                    # Tam komponent sınırı (SVG viewBox koordinatında [x,y,w,h])
                    sch_box = None
                    if vb_h is not None:
                        try:
                            fb = c.full_bounds_mils()
                            sch_box = [round(fb.x1_mils / 10, 2),
                                       round(vb_h - fb.y2_mils / 10, 2),
                                       round((fb.x2_mils - fb.x1_mils) / 10, 2),
                                       round((fb.y2_mils - fb.y1_mils) / 10, 2)]
                        except Exception:
                            sch_box = None
                    components.append({
                        "designator": get_comp_field(c, ["designator", "logical_designator", "ref"], "?"),
                        "value": get_comp_field(c, ["comment", "value"], ""),
                        "description": get_comp_field(c, ["description"], ""),
                        "footprint": get_comp_field(c, ["footprint", "current_footprint"], ""),
                        "library_reference": get_comp_field(c, ["library_reference", "lib_ref"], ""),
                        "library_name": get_comp_field(c, ["library_name"], ""),
                        "parameters": get_component_parameters(c),
                        "unique_id": getattr(c, "unique_id", "") or "",
                        "part_id": part_id,
                        "part_count": part_count or 1,
                        "sheet_id": sheet_id,
                        "sheet_name": sheet_name,
                        "sch_box": sch_box,
                    })
            except Exception as e:
                log(f"  ! Component okuma hatası ({sheet_name}): {e}")

            # Sheet symbol (block) bilgisi — block navigation için
            blocks_for_sheet = []
            try:
                for ss in schdoc.get_sheet_symbols():
                    filename = (getattr(ss, "file_name", None) or
                                getattr(ss, "child_filename", None) or "")
                    designator = (getattr(ss, "designator", None) or
                                  getattr(ss, "sheet_name_text", None) or "")
                    if not filename:
                        continue
                    # Filename'i sheet adına çevir (uzantı ve path at)
                    target_name = str(filename).rsplit(".", 1)[0]
                    target_name = target_name.replace("\\", "/").split("/")[-1]
                    target_id = sheet_name_to_id.get(target_name)
                    if target_id:
                        blocks_for_sheet.append({
                            "designator": str(designator) if designator else "",
                            "filename": str(filename),
                            "target_name": target_name,
                            "target_id": target_id,
                        })
            except Exception as e:
                log(f"  ! Block okuma hatası ({sheet_name}): {e}")

            sheets_raw.append({
                "id": sheet_id, "name": sheet_name,
                "x": cx, "y": cy, "svg": svg, "sheet_nets": sheet_nets,
                "blocks": blocks_for_sheet,
            })
            log(f"  OK {sheet_name}  ({len(sheet_nets)} net, {len(blocks_for_sheet)} block)")

        except Exception as e:
            log(f"  ERR {sheet_name}: {e}")

    log(f"\nToplam {len(all_net_names)} farklı net adı toplandı (tüm sayfalar).")

    log("\nPass 2: Pozisyon çıkarımı (block pinleri dahil)...")
    sheets = []
    net_index = {}

    for j, raw in enumerate(sheets_raw):
        prog(0.45 + 0.45 * j / max(1, len(sheets_raw)),
             f"Pozisyon çıkarımı: {raw['name']}")
        positions = extract_label_positions(raw["svg"], all_net_names)
        total_pos = sum(len(v) for v in positions.values())

        found_names = set()
        for net_name, locs in positions.items():
            found_names.add(net_name)
            for rx, ry in locs:
                net_index.setdefault(net_name, []).append({
                    "sheet_id": raw["id"], "sheet_name": raw["name"],
                    "rx": rx, "ry": ry,
                })

        for net_name in raw["sheet_nets"]:
            if net_name not in found_names:
                net_index.setdefault(net_name, []).append({
                    "sheet_id": raw["id"], "sheet_name": raw["name"],
                    "rx": 0.5, "ry": 0.5,
                })

        sheets.append({
            "id": raw["id"], "name": raw["name"],
            "x": raw["x"], "y": raw["y"], "w": THUMB_W, "h": THUMB_H,
            "svg": strip_aspect_ratio(raw["svg"]),
            "blocks": raw["blocks"],
        })
        log(f"  + {raw['name']}: {total_pos} kesin pozisyon")

    net_list = [
        {"name": n, "occurrences": occs, "count": len(occs)}
        for n, occs in net_index.items()
    ]
    net_list.sort(key=lambda x: (-x["count"], x["name"].lower()))

    # === Multi-part komponentleri birleştir ===
    # STM32 gibi IC'ler şemada IC2A/IC2B/IC2C diye part'lara bölünür ama tek
    # mantıksal komponenttir. Aynı designator'ı tek girişte topla, her part'ın
    # bulunduğu sayfayı ve part_id'sini koru (tıklama eşleşmesi için).
    merged = {}
    for c in components:
        # Birleştirme anahtarı: designator (aynı designator = aynı komponent).
        # unique_id de aynı olur ama designator daha sağlam (parça parça okunur).
        key = c["designator"]
        if key not in merged:
            merged[key] = {
                "designator": c["designator"],
                "value": c["value"],
                "description": c["description"],
                "footprint": c["footprint"],
                "library_reference": c["library_reference"],
                "library_name": c["library_name"],
                "parameters": c["parameters"],
                "part_count": c.get("part_count", 1) or 1,
                # Hangi sayfa(lar)da, hangi part olarak görünüyor
                "placements": [],
                # Geriye dönük uyumluluk: ilk görüldüğü sayfa
                "sheet_id": c["sheet_id"],
                "sheet_name": c["sheet_name"],
            }
        # Boş alanları doldur (bir part'ta value boş, diğerinde dolu olabilir)
        for fld in ("value", "description", "footprint", "library_reference", "library_name"):
            if not merged[key][fld] and c.get(fld):
                merged[key][fld] = c[fld]
        if c.get("parameters") and not merged[key]["parameters"]:
            merged[key]["parameters"] = c["parameters"]
        merged[key]["placements"].append({
            "sheet_id": c["sheet_id"],
            "sheet_name": c["sheet_name"],
            "part_id": c.get("part_id"),
            "sch_box": c.get("sch_box"),
        })

    components = list(merged.values())
    # NOT: altium_monkey'in part_count'u güvenilmez (tek dirençte bile 2 dönebiliyor),
    # bu yüzden gerçek multi-part göstergesi = birden fazla yerde çizilmiş olması.
    multipart_count = sum(1 for c in components if len(c["placements"]) > 1)
    components.sort(key=lambda c: (c["sheet_name"], c["designator"]))

    if multipart_count:
        log(f"\n{len(net_list)} net · {len(components)} komponent "
            f"({multipart_count} multi-part birleştirildi)")
    else:
        log(f"\n{len(net_list)} net · {len(components)} komponent")

    # === Netlist derlemesi (gerçek pin→net elektriksel bağlantı) ===
    prog(0.92, "Netlist derleniyor")
    netlist = compile_project_netlist(loaded_schdocs, project, log)
    if netlist:
        try:
            _merge_netlist_with_pcb(netlist, project_path, log)
        except Exception as e:
            log(f"  ! PCB netlist doğrulaması atlandı: {e}")

    # === BOM / Pick&Place / Varyant (AltiumDesign API) ===
    prog(0.96, "Tasarım verileri")
    design_extras = collect_design_extras(project_path, log)

    # === PCB cross-probe (opsiyonel, maliyetli) ===
    if with_pcb:
        prog(0.98, "PCB konumları (cross-probe)")
    pcb_data = collect_pcb_placement(project_path, log) if with_pcb else {"available": False}
    prog(1.0, "Şematik verisi hazır")

    return {
        "sheets": sheets,
        "net_list": net_list,
        "components": components,
        "netlist": netlist,
        "design_extras": design_extras,
        "pcb": pcb_data,
        "project_name": Path(project_path).stem,
    }


def collect_pcb_placement(project_path, log):
    """@brief PCB'den komponent konumlarını ve board outline SVG'sini topla
    
    (şematik-PCB cross-probe için). PrjPcb klasöründe .PcbDoc arar.

    Döner: {
        "available": bool,
        "board_svg": str,            # board outline SVG (TOP görünüm)
        "board_w_mm", "board_h_mm",  # board boyutu (mm)
        "board_x0_mm", "board_y0_mm",# board sol-alt köşe ofseti (mm)
        "components": {designator: {"x_mm", "y_mm", "layer", "rotation", "footprint"}}
    }
    x_mm/y_mm board sol-alt köşeye göre normalize edilmemiş, board SVG viewBox
    ile aynı koordinat sisteminde (üst-sol orijin, mm).
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    result = {"available": False, "board_svg": "", "components": {}}
    try:
        from altium_monkey.altium_pcbdoc import AltiumPcbDoc
    except Exception:
        log("\n! AltiumPcbDoc API yok — PCB cross-probe atlanıyor.")
        return result

    # PcbDoc bul (kardeş klasör dahil, Panel olmayanı tercih et)
    pcb_paths = _resolve_pcbdoc_paths(project_path, log)
    if not pcb_paths:
        log("\n· PCB dosyası bulunamadı — cross-probe atlanıyor.")
        return result
    # "Panel" içermeyen ilk PCB'yi seç (panel = üretim paneli, asıl board değil)
    pcb_path = next((p for p in pcb_paths if "panel" not in p.stem.lower()), pcb_paths[0])

    log(f"\nPCB cross-probe: {pcb_path.name}")
    try:
        pcb = AltiumPcbDoc.from_file(pcb_path)
    except Exception as e:
        log(f"! PCB parse hatası: {e}")
        return result

    if not pcb.components:
        log("! PCB'de komponent bulunamadı (parse boş).")
        return result

    MM = 0.0254  # 1 mil = 0.0254 mm

    # Board sınırlarını gerçek outline bounding_box'tan al (mil).
    # NOT: to_board_outline_svg() bazen panel/mekanik katmanı çiziyor ve
    # komponent koordinatlarıyla hizalı olmuyor. bounding_box güvenilir.
    bb = None
    try:
        bb = pcb.board.outline.bounding_box  # (left, bottom, right, top) mil
    except Exception:
        bb = None

    if bb and len(bb) == 4:
        bx_min, by_min, bx_max, by_max = bb
    else:
        # Fallback: komponentlerden sınır çıkar (sonra hesaplanır)
        bx_min = by_min = bx_max = by_max = None

    # Komponent konumları (mil → mm), board sol-alt köşeye göre relatif.
    comps = {}
    all_x = []
    all_y = []
    raw = {}
    for i, c in enumerate(pcb.components):
        try:
            x_mil, y_mil = pcb.get_component_pnp_position_mils(i, origin_relative=False)
        except Exception:
            continue
        desig = getattr(c, "designator", "") or ""
        if not desig:
            continue
        rot = getattr(c, "rotation", 0)
        try:
            rot = float(str(rot).strip())
        except Exception:
            rot = 0.0
        raw[desig] = {
            "x_mil": x_mil, "y_mil": y_mil,
            "layer": getattr(c, "layer", "TOP") or "TOP",
            "rotation": round(rot, 1),
            "footprint": getattr(c, "footprint", "") or "",
        }
        all_x.append(x_mil)
        all_y.append(y_mil)

    if not raw:
        log("! PCB komponent konumu çıkarılamadı.")
        return result

    # Sınır yoksa komponentlerden türet (biraz pay bırak)
    if bx_min is None:
        pad = 50
        bx_min, bx_max = min(all_x) - pad, max(all_x) + pad
        by_min, by_max = min(all_y) - pad, max(all_y) + pad

    board_w_mil = bx_max - bx_min
    board_h_mil = by_max - by_min

    # Her komponenti board sol-üst köşeye göre mm cinsinden normalize et.
    # SVG koordinatı: üst-sol orijin, Y aşağı artar. Altium: Y yukarı artar.
    # Bu yüzden Y'yi ters çevir (by_max - y).
    for desig, r in raw.items():
        x_rel_mm = round((r["x_mil"] - bx_min) * MM, 3)
        y_rel_mm = round((by_max - r["y_mil"]) * MM, 3)  # Y flip
        comps[desig] = {
            "x_mm": x_rel_mm,
            "y_mm": y_rel_mm,
            "layer": r["layer"],
            "rotation": r["rotation"],
            "footprint": r["footprint"],
            # Mutlak konum da sakla (popup'ta "gerçek" koordinat göstermek için)
            "abs_x_mm": round(r["x_mil"] * MM, 2),
            "abs_y_mm": round(r["y_mil"] * MM, 2),
        }

    bw = round(board_w_mil * MM, 2)
    bh = round(board_h_mil * MM, 2)

    log(f"  ✓ {len(comps)} komponent konumu · board {bw:.0f}×{bh:.0f}mm")
    result.update({
        "available": True,
        "board_w_mm": bw, "board_h_mm": bh,
        "components": comps,
        "pcb_name": pcb_path.name,
    })
    return result


def compile_project_netlist(schdocs, project, log):
    """@brief Tüm projeyi multi-sheet netlist olarak derle.
    
    Cross-sheet bağlantıları (port/net label eşleşmesi) altium_monkey'in
    compile_netlist'i çözer. Dönen değer her net için pin terminallerini
    içeren bir dict (pin→net bağlantısı = gerçek elektriksel netlist).
    Hata olursa None döner (JSON yine de net özetiyle üretilir).
    
    @param schdocs Şema nesneleri listesi
    @param project Altium proje nesnesi / yolu
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    if not schdocs:
        return None
    try:
        from altium_monkey.altium_netlist_compilation import compile_netlist
    except Exception as e:
        log(f"\n! Netlist modülü import edilemedi: {e}")
        return None

    log("\nNetlist derleniyor (pin→net bağlantısı)...")
    # KRİTİK: Projenin KENDİ netlist ayarlarıyla derle (scope, kanal formatı).
    # options=None varsayılanı GLOBAL scope kullanır; hiyerarşik projelerde
    # (port ↔ sheet entry bağlantılı) sayfalar arası netler birleşmez — MCU
    # pinleri sahte 'NC' görünür ve kanal designator'ları PCB ile uyuşmaz
    # (U2A vs U2_1). from_prjpcb Altium'un derlemesini birebir izler.
    options = None
    if project is not None:
        try:
            from altium_monkey.altium_netlist_options import NetlistOptions
            options = NetlistOptions.from_prjpcb(project)
        except Exception as e:
            log(f"  ! Proje netlist ayarları okunamadı, varsayılan: {e}")
    try:
        nl = compile_netlist(schdocs, project, options)
        raw = nl.to_json()
    except Exception as e:
        log(f"! Netlist derleme hatası: {e}")
        return None

    # Port adı → yön (io_type) haritası: (sayfa_dosyaadı_lower, port_adı).
    # Net endpoints'inde port adı/sayfası var ama io_type YOK — şematik
    # dokümanlarından tamamlanır. Yön, GPIO_IN/GPIO_OUT gösterimi için.
    port_io = {}
    for doc in schdocs:
        try:
            fname = Path(str(getattr(doc, "filepath", ""))).name.lower()
            for prt in (doc.get_ports() or []):
                nm = getattr(prt, "name", "") or ""
                io = getattr(getattr(prt, "io_type", None), "name", "") or ""
                if nm:
                    port_io.setdefault((fname, nm), io)
        except Exception:
            continue

    # Tüm komponent pinleri kataloğu: {designator: {pin_no: pin_adı}}.
    # Netlist yalnız BAĞLI pinleri içerir; bağlı olmayan (NC) pinlerin
    # varlığı/adları buradan gelir (Excel'de "NC" satırları için).
    # Multipart soneki atılır (IC2A → IC2, netlist'in birleşik designator'ı).
    all_pins = {}
    for doc in schdocs:
        try:
            for comp in doc.objects:
                if type(comp).__name__ != "AltiumSchComponent":
                    continue
                desig, pins = "", {}
                for ch in (getattr(comp, "children", None) or []):
                    tn = type(ch).__name__
                    if tn == "AltiumSchDesignator":
                        desig = getattr(ch, "text", "") or ""
                    elif tn == "AltiumSchPin":
                        pno = str(getattr(ch, "designator", "") or "")
                        if pno:
                            pins[pno] = getattr(ch, "name", "") or ""
                if desig and pins:
                    desig = re.sub(r"(?<=\d)[A-Z]$", "", desig)
                    all_pins.setdefault(desig, {}).update(pins)
        except Exception:
            continue

    nets_out = []
    total_terminals = 0
    for n in raw.get("nets", []):
        terminals = []
        for t in n.get("terminals", []):
            terminals.append({
                "designator": t.get("designator", ""),
                "pin": t.get("pin", ""),
                "pin_name": t.get("pin_name", ""),
                "pin_type": t.get("pin_type", ""),
            })
        total_terminals += len(terminals)
        # Net'e bağlı portlar (endpoints role=='port') + yönleri
        ports, seen_p = [], set()
        for ep in n.get("endpoints", []) or []:
            if ep.get("role") != "port":
                continue
            nm = ep.get("name", "") or ""
            ss = ep.get("source_sheet", "") or ""
            k = (nm, ss)
            if not nm or k in seen_p:
                continue
            seen_p.add(k)
            ports.append({"name": nm, "source_sheet": ss,
                          "io_type": port_io.get((Path(ss).name.lower(), nm), "")})
        nets_out.append({
            "name": n.get("name", ""),
            "auto_named": n.get("auto_named", False),
            "source_sheets": n.get("source_sheets", []),
            "terminals": terminals,
            "ports": ports,
        })

    log(f"  ✓ {len(nets_out)} net, {total_terminals} pin bağlantısı çıkarıldı")
    return {
        "schema": raw.get("schema", ""),
        "nets": nets_out,
        "all_pins": all_pins,
    }


def _merge_netlist_with_pcb(netlist, project_path, log):
    """@brief Netlist'i PCB'nin (rotalanmış board) bağlantısından yeniden kur.

    Şematik derlemesi iki tür hataya açık: (1) altium_monkey'in hiyerarşik
    derlemesi bazı port↔sheet-entry köprülerini kaçırıyor (örn. BRK-209'da
    MCU sayfasının SCL/SDA/PEX_INT/PEX_RST portları → MCU pini sahte 'NC');
    (2) kanal-tekrarlı sayfalarda (Repeat) kanal indeksi sırası Altium'un
    board annotasyonundan FARKLI çıkabiliyor (şematikte Q2_2 olan parça
    fiziksel board'da Q2_5). PCB dosyası Altium'un KENDİ derlemesi + gerçek
    designator'lardır = kesin doğru. Bu yüzden PCB varsa netler PCB pad
    listesinden YENİDEN kurulur; pin adları şematik netlist'inden tamamlanır
    (önce tam designator, yoksa kanal soneki atılmış taban designator — aynı
    sembolün pin adları kanaldan bağımsız aynıdır). Yalnızca şematikte olup
    PCB'de hiç pad'i olmayan netler (örn. yerleştirilmemiş parçalar) şematik
    halleriyle korunur. netlist sözlüğü YERİNDE güncellenir.

    Güvenlik: PCB komponent designator'larının en az yarısı şematikte yoksa
    (yanlış PcbDoc seçilmiş olabilir) hiçbir şey değiştirilmez.
    PCB yoksa/yüklenemezse netlist DEĞİŞMEZ (no-op, log'a not düşer).

    @param netlist compile_project_netlist çıktısı ({"nets": [...]})
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param log Log mesajı callback'i (str alır)
    """
    nets = netlist.get("nets") or []
    if not nets:
        return
    try:
        paths = _resolve_pcbdoc_paths(project_path, lambda m: None)
    except Exception:
        paths = []
    if not paths:
        log("  · PCB bulunamadı — netlist PCB ile doğrulanamadı (şematik esas).")
        return
    # Panel olmayan PcbDoc tercih (collect_pcb_placement ile aynı kural)
    pcb_path = next((p for p in paths if "panel" not in p.name.lower()), paths[0])

    log(f"  · Netlist PCB'den doğrulanıyor: {pcb_path.name} (büyük board'da sürebilir)...")
    try:
        from altium_monkey.altium_pcbdoc import AltiumPcbDoc
        pcb = (AltiumPcbDoc.from_file(str(pcb_path))
               if hasattr(AltiumPcbDoc, "from_file") else AltiumPcbDoc(str(pcb_path)))
        comps = {i: str(getattr(c, "designator", "") or "")
                 for i, c in enumerate(pcb.components)}
        pcb_net_names = {i: getattr(n, "name", "") or "" for i, n in enumerate(pcb.nets)}
        pcb_groups = {}  # net_index -> [(desig, pad_no)]
        for pad in pcb.pads:
            ni = getattr(pad, "net_index", None)
            if ni is None or ni < 0:
                continue
            d = comps.get(getattr(pad, "component_index", -1), "")
            p = str(getattr(pad, "designator", "") or "")
            if d and p:
                pcb_groups.setdefault(ni, []).append((d, p))
    except Exception as e:
        log(f"  · PCB netlist doğrulaması atlandı (PCB okunamadı): {e}")
        return

    if not pcb_groups:
        log("  · PCB'de net'e bağlı pad yok — şematik netlist'i korunuyor.")
        return

    # Güvenlik: yanlış PcbDoc'a karşı designator örtüşme kontrolü
    sch_desigs = {t["designator"] for n in nets for t in n["terminals"]}

    def base_desig(d):
        """@brief Kanal sonekini at: 'Q2_5' → 'Q2' (pin adı eşlemesi için).

        @param d Designator
        @return Taban designator
        """
        return re.sub(r"_\d+$", "", d)

    sch_bases = {base_desig(d) for d in sch_desigs}
    pcb_desigs = {d for pins in pcb_groups.values() for d, _ in pins}
    hit = sum(1 for d in pcb_desigs
              if d in sch_desigs or base_desig(d) in sch_bases)
    if hit < len(pcb_desigs) * 0.5:
        log(f"  ! PCB komponentleri şematikle örtüşmüyor "
            f"({hit}/{len(pcb_desigs)}) — yanlış PcbDoc olabilir, "
            f"şematik netlist'i korunuyor.")
        return

    # Pin adı/tipi sözlükleri: tam designator → taban designator fallback'i.
    # Ayrıca pin → şematik net kaydı (ad + portlar): PCB net'i otomatik adlıysa
    # (NetU5_20) şematikteki etiket/PORT adıyla (PWR_ARM_MCU) adlandırılır.
    pin_name_exact, pin_name_base, pin_sch_net = {}, {}, {}
    for n in nets:
        for t in n["terminals"]:
            k = (t["designator"], t["pin"])
            pin_name_exact[k] = (t.get("pin_name", ""), t.get("pin_type", ""))
            pin_name_base.setdefault((base_desig(t["designator"]), t["pin"]),
                                     (t.get("pin_name", ""), t.get("pin_type", "")))
            pin_sch_net[k] = n

    new_nets = []
    named_pins = 0
    total_pins = 0
    renamed = 0
    # Yeniden adlandırma çakışma koruması: aday ad (etiket/port) başka bir PCB
    # net'inin adıysa DOKUNMA — yoksa iki ayrı net aynı adı alır ve ada göre
    # kurulan sözlüklerde biri diğerini ezer (örn. R arkasındaki NetR162_1 de
    # 'SPI_MISO' olunca gerçek SPI_MISO'nun U5.22 ucu kaybolmuştu).
    taken_names = {v for v in pcb_net_names.values() if v}
    for ni in sorted(pcb_groups):
        name = pcb_net_names.get(ni, "") or f"PCBNET_{ni}"
        members = sorted(set(pcb_groups[ni]))
        terminals = []
        sch_labels, ports, seen_p = set(), [], set()
        for d, p in members:
            info = pin_name_exact.get((d, p))
            if info is None:
                info = pin_name_base.get((base_desig(d), p), ("", ""))
            else:
                named_pins += 1
            total_pins += 1
            terminals.append({"designator": d, "pin": p,
                              "pin_name": info[0] or "", "pin_type": info[1] or ""})
            rec = pin_sch_net.get((d, p))
            if rec:
                if not rec.get("auto_named"):
                    sch_labels.add(rec["name"])
                for pt in rec.get("ports") or []:
                    pk = (pt.get("name"), pt.get("source_sheet"))
                    if pk not in seen_p:
                        seen_p.add(pk)
                        ports.append(pt)
        auto = name.startswith(("Net", "PCBNET_"))
        if auto:
            # Şematikteki tek etiket adı > en yaygın port adı > PCB adı
            cand = None
            if len(sch_labels) == 1:
                cand = next(iter(sch_labels))
            elif ports:
                from collections import Counter
                cand = Counter(pt["name"] for pt in ports).most_common(1)[0][0]
            if cand and cand != name and cand not in taken_names:
                taken_names.add(cand)
                name = cand
                renamed += 1
                auto = name.startswith(("Net", "PCBNET_"))
        new_nets.append({"name": name, "auto_named": auto,
                         "source_sheets": [], "terminals": terminals,
                         "ports": ports})

    # PCB'de hiç pad'i olmayan şematik netleri koru (yerleştirilmemiş parçalar)
    pcb_pins = {dp for pins in pcb_groups.values() for dp in pins}
    kept_sch = 0
    for n in nets:
        if not any((t["designator"], t["pin"]) in pcb_pins for t in n["terminals"]):
            new_nets.append(n)
            kept_sch += 1

    netlist["nets"] = new_nets
    log(f"  ✓ Netlist PCB'den kuruldu: {len(pcb_groups)} PCB neti "
        f"({total_pins} pad; {named_pins} pin adı şematikten eşleşti; "
        f"{renamed} otomatik ad şematik etiket/port adıyla değiştirildi) "
        f"+ {kept_sch} şematik-yalnız net korundu.")


def collect_design_extras(project_path, log):
    """@brief AltiumDesign API'si üzerinden BOM, Pick&Place ve varyant bilgisini topla.
    
    Bu API güncel altium_monkey sürümlerinde mevcut (2026.6.x+). Eski sürümde
    veya hata durumunda her alan için boş/None döner — uygulama yine çalışır.

    Döner: {"variants": [...], "bom": [...], "pnp": [...], "has_pcb": bool}
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    result = {"variants": [], "bom": [], "pnp": [], "has_pcb": False,
              "available": False}
    try:
        from altium_monkey.altium_design import AltiumDesign
    except Exception:
        log("\n! AltiumDesign API yok (eski altium_monkey sürümü) — "
            "BOM/PnP atlanıyor.")
        return result

    log("\nTasarım verileri (BOM / Pick&Place / Varyant)...")
    try:
        design = AltiumDesign.from_prjpcb(project_path)
    except Exception as e:
        log(f"! AltiumDesign yüklenemedi: {e}")
        return result

    result["available"] = True

    # Varyantlar
    try:
        result["variants"] = list(design.get_variants() or [])
        if result["variants"]:
            log(f"  ✓ {len(result['variants'])} varyant: "
                f"{', '.join(result['variants'])}")
        else:
            log("  · Varyant tanımlı değil")
    except Exception as e:
        log(f"  ! Varyant okuma hatası: {e}")

    # BOM (tüm komponentler, varyant=None → hepsi)
    try:
        bom = design.to_bom(variant=None)
        result["bom"] = bom or []
        log(f"  ✓ BOM: {len(result['bom'])} komponent")
    except Exception as e:
        log(f"  ! BOM hatası: {e}")

    # Pick & Place (PCB gerekir — yoksa graceful)
    try:
        pnp_entries = design.to_pnp(variant=None, units="mm")
        pnp_list = []
        for e in (pnp_entries or []):
            # PnpEntry objesi → dict (to_json varsa onu kullan)
            if hasattr(e, "to_json"):
                pnp_list.append(e.to_json())
            else:
                pnp_list.append({
                    "designator": getattr(e, "designator", ""),
                    "comment": getattr(e, "comment", ""),
                    "layer": getattr(e, "layer", ""),
                    "footprint": getattr(e, "footprint", ""),
                    "center_x": getattr(e, "center_x", 0.0),
                    "center_y": getattr(e, "center_y", 0.0),
                    "rotation": getattr(e, "rotation", 0.0),
                })
        result["pnp"] = pnp_list
        result["has_pcb"] = bool(pnp_list)
        log(f"  ✓ Pick&Place: {len(pnp_list)} yerleşim (mm)")
    except Exception as e:
        # PCB yoksa ValueError beklenir — sessizce geç
        log(f"  · Pick&Place atlandı (PCB yok veya hata): {e}")

    return result


def _svg_path_bbox(d):
    """@brief SVG path `d` attribute'undan bounding box (minx,miny,maxx,maxy).

    Yalnızca UÇ-NOKTA koordinatlarını kullanır: yay (A) yarıçap/flag'leri ve
    eğri (C/S/Q) kontrol noktaları bbox'ı kirletmez. Hem mutlak hem göreli
    komutları takip eder (Altium çoğunlukla mutlak üretir). Naive "tüm sayıları
    koordinat say" yaklaşımı yuvarlatılmış/yaylı outline'larda yanlış merkez
    veriyordu — bu parser onu düzeltir.

    @param d Path `d` attribute içeriği
    @return (minx, miny, maxx, maxy) veya None
    """
    toks = re.findall(r"[MmLlHhVvCcSsQqTtAaZz]|-?\d*\.?\d+(?:[eE][-+]?\d+)?", d)
    xs, ys = [], []
    cx = cy = sx = sy = 0.0
    cmd = None
    i, n = 0, len(toks)

    def num():
        nonlocal i
        v = float(toks[i]); i += 1; return v

    while i < n:
        if toks[i].isalpha():
            cmd = toks[i]; i += 1
            if cmd in ("Z", "z"):
                cx, cy = sx, sy
                continue
        if cmd is None or i >= n:
            i += 1; continue
        rel = cmd.islower(); C = cmd.upper()
        try:
            if C == "M":
                x = num(); y = num()
                if rel: x += cx; y += cy
                cx, cy = x, y; sx, sy = x, y
                cmd = "l" if rel else "L"   # sonraki çiftler lineto
            elif C in ("L", "T"):
                x = num(); y = num()
                if rel: x += cx; y += cy
                cx, cy = x, y
            elif C == "H":
                x = num()
                cx = cx + x if rel else x
            elif C == "V":
                y = num()
                cy = cy + y if rel else y
            elif C == "C":
                num(); num(); num(); num(); x = num(); y = num()
                if rel: x += cx; y += cy
                cx, cy = x, y
            elif C in ("S", "Q"):
                num(); num(); x = num(); y = num()
                if rel: x += cx; y += cy
                cx, cy = x, y
            elif C == "A":
                num(); num(); num(); num(); num(); x = num(); y = num()
                if rel: x += cx; y += cy
                cx, cy = x, y
            else:
                i += 1; continue
        except (IndexError, ValueError):
            break
        xs.append(cx); ys.append(cy)
    if not xs or not ys:
        return None
    return (min(xs), min(ys), max(xs), max(ys))


def _build_board_surface(all_layers, view_w, view_h, log=print, board_wh_mm=None):
    """@brief Board yüzeyi için bakır + pad + silkscreen dokusu üret (3D'de board

    üzerine bindirilir). Katman SVG'leri yeniden renklendirilip birleştirilir,
    gzip+base64 ile gömülür (tarayıcıda DecompressionStream ile açılır).
    Hizalama: board outline'ının SVG'deki merkezi bulunur → 3D düzlemine eşlenir.
    Outline RENGE göre değil, gerçek BOYUTA göre (board_wh_mm) eşleştirilir ki
    farklı renk/katman/yaylı outline'larda da doğru çalışsın; viewBox origin'i
    sıfır olmayan projelerde de doğru hizalar.

    @param all_layers Katman adı → SVG string sözlüğü
    @param view_w Görüntü genişliği (mm) — fallback
    @param view_h Görüntü yüksekliği (mm) — fallback
    @param log Log mesajı callback'i (str alır)
    @param board_wh_mm (board_genişlik_mm, board_yükseklik_mm) — outline boyut eşleştirme için
    @return Üretilen sonuç.
    """
    import re, gzip, base64
    def inner(s):
        """@brief inner()
        
        @param s
        @return Üretilen sonuç.
        """
        if not s:
            return ""
        i = s.find(">", s.find("<svg")) + 1
        j = s.rfind("</svg>")
        return s[i:j] if (i > 0 and j > i) else ""
    def recolor(s, color):
        """@brief recolor()
        
        @param s
        @param color Renk (hex, ör. #RRGGBB)
        @return Üretilen sonuç.
        """
        t = inner(s)
        t = re.sub(r'(fill|stroke)="#[0-9A-Fa-f]{6}"',
                   lambda m: '%s="%s"' % (m.group(1), color), t)
        t = re.sub(r'(fill|stroke):#[0-9A-Fa-f]{6}',
                   lambda m: '%s:%s' % (m.group(0).split(":")[0], color), t)
        return t
    # Kaynak viewBox'ı (origin dahil) al — wrapper SVG ve hizalama math'i bunu kullanır.
    top_raw = all_layers.get("TOP", "")
    vbx, vby, vbw, vbh = 0.0, 0.0, view_w, view_h
    try:
        vm = re.search(r'viewBox="\s*([\d.\-eE]+)\s+([\d.\-eE]+)\s+'
                       r'([\d.\-eE]+)\s+([\d.\-eE]+)\s*"', top_raw)
        if vm:
            vbx, vby, vbw, vbh = (float(vm.group(1)), float(vm.group(2)),
                                  float(vm.group(3)), float(vm.group(4)))
    except Exception:
        pass

    # board outline'ın SVG koordinatındaki merkezi (bcx,bcy).
    # Strateji: tüm path'lerin uç-nokta bbox'ını çıkar; board GERÇEK BOYUTUNA
    # (board_wh_mm) uyan path'i seç (renkten/katmandan bağımsız). #C0A000 ipucu
    # eşit boyutlu adaylar arasında öncelik kazanır.
    bcx, bcy = vbx + vbw / 2.0, vby + vbh / 2.0
    # Kırpılmış viewBox (varsayılan = tam viewBox; board tespit edilirse daraltılır)
    cvbx, cvby, cvbw, cvbh = vbx, vby, vbw, vbh
    aligned = False   # board outline güvenle bulundu mu (delik alfa-delme kapısı)
    board_w_mm = board_h_mm = None
    if board_wh_mm and board_wh_mm[0] and board_wh_mm[1]:
        board_w_mm, board_h_mm = float(board_wh_mm[0]), float(board_wh_mm[1])
    try:
        tol = (max(1.0, 0.03 * max(board_w_mm, board_h_mm))
               if board_w_mm else None)
        candidates = []   # (öncelik, cx, cy, bbox)  öncelik 0=hint+boyut, 1=boyut
        hint_only = None  # (cx, cy) — #C0A000 path, boyut filtresiz son çare
        for m in re.finditer(r"<path\b[^>]*>", top_raw):
            el = m.group(0)
            dm = re.search(r'\bd="([^"]+)"', el)
            if not dm:
                continue
            bbox = _svg_path_bbox(dm.group(1))
            if not bbox:
                continue
            pminx, pminy, pmaxx, pmaxy = bbox
            pw, ph = pmaxx - pminx, pmaxy - pminy
            pcx, pcy = (pminx + pmaxx) / 2.0, (pminy + pmaxy) / 2.0
            is_hint = "C0A000" in el
            if is_hint and hint_only is None:
                hint_only = (pcx, pcy)
            if board_w_mm and tol:
                if ((abs(pw - board_w_mm) <= tol and abs(ph - board_h_mm) <= tol) or
                        (abs(pw - board_h_mm) <= tol and abs(ph - board_w_mm) <= tol)):
                    candidates.append((0 if is_hint else 1, pcx, pcy, bbox))
        if candidates:
            candidates.sort(key=lambda c: c[0])
            bcx, bcy = candidates[0][1], candidates[0][2]
            aligned = True
            win_bbox = candidates[0][3]
            log("  · board outline merkezi: %s eşleşme (bcx=%.2f, bcy=%.2f, %d aday)"
                % ("boyut+#C0A000" if candidates[0][0] == 0 else "boyut",
                   bcx, bcy, len(candidates)))
            # Board, viewBox'tan çok küçükse (büyük fab/mekanik içerik board dışında)
            # doku kanvası tüm viewBox'a yayılıp board yazılarını bulanıklaştırır.
            # viewBox'ı board bbox'ı + margin'e KIRP (kaynak viewBox'a clamp) → 2048px
            # kanvas board'a yoğunlaşır, silk net olur. Hizalama formülü kırpılmış
            # viewBox ile yeniden türetilir (board merkezi yine dünya origin'ine düşer).
            mxn, myn, mxx, myx = win_bbox
            mrg = max(2.0, 0.06 * max(mxx - mxn, myx - myn))
            cx0 = max(vbx, mxn - mrg);          cy0 = max(vby, myn - mrg)
            cx1 = min(vbx + vbw, mxx + mrg);     cy1 = min(vby + vbh, myx + mrg)
            if (cx1 - cx0) > 1 and (cy1 - cy0) > 1 and \
               (cx1 - cx0) < vbw * 0.92:   # anlamlı kırpma varsa uygula
                cvbx, cvby = cx0, cy0
                cvbw, cvbh = cx1 - cx0, cy1 - cy0
                log("  · doku viewBox board'a kırpıldı: %.0f×%.0f → %.0f×%.0fmm "
                    "(silk çözünürlüğü artar)" % (vbw, vbh, cvbw, cvbh))
        elif hint_only is not None:
            bcx, bcy = hint_only
            aligned = True
            log("  · board outline boyut eşleşmedi, #C0A000 ipucu kullanıldı.")
        else:
            log("  · board outline bulunamadı — doku merkezlendi (hizalama yaklaşık).")
    except Exception as e:
        log(f"  · board merkezi tespiti başarısız ({e}) — doku merkezlendi.")
    COPPER = "#c07a35"; GOLD = "#e0b030"; SILK = "#f0f0f0"; DRILL = "#141414"
    def _recolor(s, color):
        s = re.sub(r'(fill|stroke)="#[0-9A-Fa-f]{6}"',
                   lambda m: '%s="%s"' % (m.group(1), color), s)
        s = re.sub(r'(fill|stroke):#[0-9A-Fa-f]{6}',
                   lambda m: '%s:%s' % (m.group(0).split(":")[0], color), s)
        return s
    def grp(key, color, opacity=1.0, white_none=False):
        """@brief grp()

        @param key
        @param color Renk (hex, ör. #RRGGBB)
        @param opacity Saydamlık (0-1)
        @param white_none Beyaz knockout'ları şeffaf yap (bool)
        @return Üretilen sonuç.
        """
        t = inner(all_layers.get(key, ""))
        if not t:
            return ""
        if white_none:   # beyaz knockout/clearance → şeffaf (yeşil board görünsün)
            t = re.sub(r'fill="#[Ff]{6}"', 'fill="none"', t)
        return '<g opacity="%s">%s</g>' % (opacity, _recolor(t, color))
    def prim(key, prim_vals, color, opacity=1.0):
        """@brief Bir katmandan yalnız belirli data-primitive elemanlarını al, renklendir.

        Pad'leri (through-hole + SMD) altın yapmak ve DRILLS delik circle'larını koyu
        çizmek için. Board outline gibi diğer path'ler hariç tutulur.
        @param key Katman anahtarı (TOP/BOTTOM/DRILLS …)
        @param prim_vals data-primitive değeri veya değer listesi (ör. "pad",
                         ("pad-hole","via-hole"))
        @param color Renk (hex)
        @param opacity Saydamlık (0-1)
        @return <g> string (eşleşme yoksa boş).
        """
        t = inner(all_layers.get(key, ""))
        if not t:
            return ""
        if isinstance(prim_vals, str):
            prim_vals = (prim_vals,)
        alt = "|".join(re.escape(v) for v in prim_vals)
        els = re.findall(
            r'<(?:rect|circle|ellipse|path|polygon)\b[^>]*?data-primitive="(?:%s)"[^>]*?/?>'
            % alt, t)
        if not els:
            return ""
        return '<g opacity="%s">%s</g>' % (opacity, _recolor("".join(els), color))
    def comp(copper_key, paste_key, silk_key):
        # bakır yarı-saydam (yeşil maske altında görünür) + altın pad + delik + beyaz silk
        """@brief comp()

        @param copper_key Bakır katman anahtarı
        @param paste_key Lehim pastası (pad) katman anahtarı
        @param silk_key Silkscreen katman anahtarı
        @return Üretilen sonuç.
        """
        # Katman sırası (alttan üste): izler(sönük bakır) → pad'ler(altın, opak) →
        # SMD pasta → multi-layer pad → silk → DELİKLER(koyu, en üstte). Pad'ler
        # copper katmanında ama %50 bakırla sönük kaldığından ayrıca altın overlay
        # ediliyor; through-hole delikleri Altium'daki gibi koyu görünsün diye DRILLS
        # pad-hole circle'ları en üste çiziliyor (yoksa pad'ler dolu altın disk kalır).
        parts = (grp(copper_key, COPPER, 0.5, white_none=True)
                 + prim(copper_key, "pad", GOLD, 0.95)
                 + grp(paste_key, GOLD, 0.95)
                 + grp("MULTILAYER", GOLD, 0.95)
                 + grp(silk_key, SILK, 1.0)
                 + prim("DRILLS", ("pad-hole", "via-hole"), DRILL, 1.0))
        # Wrapper viewBox = kırpılmış board bölgesi (board bulunamazsa tam viewBox).
        # Origin (cvbx,cvby) dahil; SVG içeriği aynı kalır, sadece görünüm penceresi
        # daralır → kanvas board'a yoğunlaşır, hizalama dönüş formülünde telafi edilir.
        svg = ('<svg xmlns="http://www.w3.org/2000/svg" viewBox="%s %s %s %s">%s</svg>'
               % (cvbx, cvby, cvbw, cvbh, parts))
        return base64.b64encode(gzip.compress(svg.encode("utf-8"), 8)).decode()
    try:
        top = comp("TOP", "TOPPASTE", "TOPOVERLAY")
        bot = comp("BOTTOM", "BOTTOMPASTE", "BOTTOMOVERLAY")
    except Exception as e:
        log(f"  · yüzey dokusu üretilemedi: {e}")
        return None
    log(f"  ✓ 3D yüzey dokusu: top {len(top)//1024}KB + bot {len(bot)//1024}KB (gzip)")
    # Düzlem boyutu = KIRPILMIŞ viewBox boyutu; ofset kırpılmış origin'i hesaba katar.
    # Board merkezi (bcx,bcy) yine dünya origin'ine düşer (board mesh de orada).
    # "ok": board outline güvenle bulundu → doku dünya-gerçeğine demirli; delik
    # alfa-delme yalnız bu durumda yapılır. Fallback (merkezleme sezgiseli)
    # dokuyu kaydırabilir — delme dünya-koordinatlı olduğundan koyu delik boyası
    # hilal olarak açığa çıkardı; delmeyi kapatmak eski (boyalı) görünümü korur.
    return {"top": top, "bot": bot, "gz": 1, "ok": 1 if aligned else 0,
            "cx": round(cvbx + cvbw / 2.0 - bcx, 3),
            "cy": round(bcy - cvby - cvbh / 2.0, 3),
            "w": round(cvbw, 3), "h": round(cvbh, 3)}


def _check_step_deps():
    """@brief 3D STEP modeli çıkarımı için gereken bağımlılıkları denetle.

    `cascadio` (STEP→GLB tessellation) ve `trimesh` (mesh ayrıştırma) gömülü STEP
    modellerini gerçek geometriyle çizmek için zorunludur. Eksik olanların adını
    döndürür; ikisi de varsa boş liste.

    @return Eksik modül adlarının listesi (boşsa hepsi kurulu).
    """
    missing = []
    for mod in ("cascadio", "trimesh"):
        try:
            __import__(mod)
        except Exception:
            missing.append(mod)
    return missing


def _extract_step_models(pcb, log=print):
    """@brief Gömülü STEP modellerini cascadio ile mesh'e çevir (opsiyonel bağımlılık).
    
    Dönen: { model_id: {"name", "parts":[{"v":[...mm], "f":[...idx], "c":"#rgb"}]} }.
    cascadio/trimesh yoksa boş döner → çağıran extrude gövdelere düşer.
    
    @param pcb AltiumPcbDoc PCB nesnesi
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    try:
        import cascadio, trimesh
    except Exception:
        log("  · cascadio/trimesh yok — STEP yerine extrude gövdeler kullanılacak.")
        return {}
    import zlib, tempfile, os
    import numpy as np
    try:
        ents = pcb.get_embedded_model_entries()
    except Exception as e:
        log(f"  · gömülü model girişleri okunamadı: {e}")
        return {}
    td = tempfile.mkdtemp(prefix="amstep_")
    out = {}
    for mdl, raw in ents:
        mid = getattr(mdl, "id", None)
        name = getattr(mdl, "name", "?")
        if not mid:
            continue
        try:
            data = zlib.decompress(raw)
            sp = os.path.join(td, "m.step"); open(sp, "wb").write(data)
            gp = os.path.join(td, "m.glb")
            cascadio.step_to_glb(sp, gp, tol_linear=0.06, tol_angular=0.5)
            sc = trimesh.load(gp, force="scene")
            parts = []
            for g in sc.dump(concatenate=False):     # parça transform'ları bake'li
                v = np.asarray(g.vertices, float) * 1000.0      # glb metre → mm
                col = [170, 170, 176]
                try:
                    bc = getattr(g.visual.material, "baseColorFactor", None)
                    if bc is not None:
                        col = [max(0, min(255, int(c*255) if c <= 1 else int(c)))
                               for c in bc[:3]]
                except Exception:
                    pass
                parts.append({
                    "v": [round(float(x), 3) for x in v.ravel()],
                    "f": [int(x) for x in np.asarray(g.faces).ravel()],
                    "c": "#%02x%02x%02x" % (col[0], col[1], col[2]),
                })
            if parts:
                out[mid] = {"name": name, "parts": parts}
        except Exception as e:
            log(f"  · STEP '{name}' atlandı: {repr(e)[:70]}")
    if out:
        nt = sum(len(p["f"])//3 for m in out.values() for p in m["parts"])
        log(f"  ✓ 3D STEP: {len(out)} model tessellate edildi (~{nt} üçgen)")
    return out


def _extract_3d(pcb, log=print):
    """@brief PCB'den 3D görünüm verisi çıkar: board dış hattı + kalınlık + komponent
    
    gövdeleri (extrude için dış hat poligonu + yükseklik + katman + renk).
    Koordinatlar mm, board merkezi orijine alınır.
    
    @param pcb AltiumPcbDoc PCB nesnesi
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    MIL2MM = 0.0254
    try:
        ol = pcb.board.outline
        pts = [(float(x) * MIL2MM, float(y) * MIL2MM) for (x, y) in ol.points_mils]
        if len(pts) < 3:
            return {"available": False}
        xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
        minx, maxx, miny, maxy = min(xs), max(xs), min(ys), max(ys)
        cx, cy = (minx + maxx) / 2, (miny + maxy) / 2
        outline = [[round(x - cx, 3), round(y - cy, 3)] for (x, y) in pts]
        # Board kalınlığı (katman yığını): bakır + dielektrik toplamı
        th = 0.0
        try:
            for L in pcb.board.layer_stackup:
                for a in ("copper_thickness", "diel_height"):
                    try:
                        th += float(getattr(L, a, 0) or 0)
                    except Exception:
                        pass
        except Exception:
            pass
        th_mm = th * MIL2MM if th > 5 else th
        if not (0.3 < th_mm < 6):
            th_mm = 1.6
        # Designator önekine göre nazik varsayılan renk (gövde rengi hep gri ise)
        PREFIX_COL = {"U": "#1d1d22", "Q": "#17171a", "R": "#2c2c30", "C": "#6e5634",
                      "L": "#3a3a42", "D": "#5a2222", "Y": "#9aa0a6", "X": "#9aa0a6",
                      "J": "#26262a", "P": "#26262a", "S": "#2a2a2e"}
        # Gömülü STEP modelleri (varsa) — model olan gövde gerçek mesh ile,
        # olmayan extrude prizma ile çizilir.
        models = _extract_step_models(pcb, log)
        comps = pcb.components
        bodies = []        # modelsiz gövdeler (extrude fallback)
        placements = []    # STEP modelli gövdeler
        U2MM = 1e-4 * MIL2MM    # model_2d_x/y ve model_3d_dz birimi: 0.1 µmil → mm
        for b in pcb.component_bodies:
            vo = getattr(b, "outline", None)
            if not vo:
                continue
            poly = []
            for v in vo:
                try:
                    x, y = float(v.x_mils), float(v.y_mils)
                except Exception:
                    continue
                poly.append([round(x * MIL2MM - cx, 3), round(y * MIL2MM - cy, 3)])
            if len(poly) < 3:
                continue
            pcx = sum(p[0] for p in poly) / len(poly)
            pcy = sum(p[1] for p in poly) / len(poly)
            ci = b.component_index
            desig = None; layer = "top"
            if isinstance(ci, int) and 0 <= ci < len(comps):
                desig = comps[ci].designator
                layer = "bottom" if str(comps[ci].layer).upper().startswith("B") else "top"
            h = (b.overall_height_mils or 0) * MIL2MM
            if h <= 0.001:
                h = 0.4
            so = (b.standoff_height_mils or 0) * MIL2MM
            mid = getattr(b, "model_id", None)
            if mid and mid in models:
                # KESİN Altium yerleşimi (v2.9.31): R = Rz(rotz)·Ry(roty)·Rx(rotx),
                # konum = model_2d anchor'ı (model orijini oraya düşer), dikey =
                # model_3d_dz (orijin board yüzeyinin dz üstünde). Alt katmanda parça
                # anchor'dan geçen X ekseni etrafında 180° DÖNDÜRÜLÜR (proper
                # rotation — eski scale.z=-1 aynası D-sub gibi yönlü parçaların
                # yüzünü ters gösteriyordu). Sıra/eksen/anchor/dz, iki board'da tüm
                # STEP parçaların bacak↔pad hizalamasıyla ampirik doğrulandı
                # (BRK-213 274 gövde + Smart_MCU 75 gövde; ort. hata ~0.3mm).
                # Eski outline-fit/flip/pins_up sezgiselleri bu kesin veriyle
                # gereksizleşti ve kaldırıldı (bkz. Çözülen Sorunlar v2.9.31).
                rxv = float(getattr(b, "model_3d_rotx", 0) or 0)
                ryv = float(getattr(b, "model_3d_roty", 0) or 0)
                rzv = float(getattr(b, "model_3d_rotz", 0) or 0)
                dzv = float(getattr(b, "model_3d_dz", 0) or 0) * U2MM
                m2x = float(getattr(b, "model_2d_x", 0) or 0)
                m2y = float(getattr(b, "model_2d_y", 0) or 0)
                if m2x == 0 and m2y == 0:
                    # model_2d verisi yok (eski dosya) → outline centroid'e düş.
                    # NOT: anchor-centroid mesafesi kötü-veri sinyali DEĞİL —
                    # origin'i kenarda modellenmiş parçalarda 30mm+ meşru fark var
                    # (Smart_MCU P1/U1), mesafe eşiği onları bozuyordu.
                    ax, ay = pcx, pcy
                else:
                    ax = m2x * U2MM - cx
                    ay = m2y * U2MM - cy
                placements.append({
                    "m": mid, "d": desig, "layer": layer,
                    "cx": round(ax, 3), "cy": round(ay, 3), "dz": round(dzv, 3),
                    "rx": rxv, "ry": ryv, "rz": rzv,
                })
            else:
                col = int(b.body_color_3d or 0x808080)
                if col in (0x808080, 0x7F7F7F, 0):
                    pref = (desig or "?")[0].upper()
                    color = PREFIX_COL.get(pref, "#3a3a3e")
                else:
                    r = col & 0xFF; g = (col >> 8) & 0xFF; bl = (col >> 16) & 0xFF
                    color = "#%02x%02x%02x" % (r, g, bl)
                bodies.append({"d": desig, "layer": layer, "h": round(h, 3),
                               "z0": round(so, 3), "color": color, "poly": poly})
        # GERÇEK delikler (v2.9.32): ≥0.6mm çaplı yuvarlak delikler (THT pad +
        # büyük via) JS'te board geometrisinden shape.holes ile kesilir, doku
        # alfası delinir → içinden arka plan görünür (eski görünüm: koyu boyalı
        # disk). 4. eleman plated bayrağı: 1 → altın barrel (delik duvarı) çizilir,
        # 0 (NPTH montaj deliği) → çıplak FR4 duvar. Küçük via'lar (<0.6mm)
        # kesilmez — dokudaki koyu nokta (tented via görünümü) yeterli ve yüzlerce
        # mini delik earcut üçgenlemesini şişirir.
        drills = []
        try:
            # earcut ASLA throw etmez — bozuk girdi (outline dışı/kenarı kesen
            # delik, çakışan pad+via) sessizce bozuk üçgenleme üretir; JS'teki
            # try/catch bu modu YAKALAYAMAZ. Emniyet Python'da: delik çemberi
            # board bbox'ının tamamen içinde olmalı, (x,y) dedupe edilir.
            hw = (maxx - minx) / 2 - 0.05
            hh = (maxy - miny) / 2 - 0.05
            seen_xy = set()
            def _add(x, y, r, plated):
                if abs(x) + r > hw or abs(y) + r > hh:
                    return
                key = (round(x, 1), round(y, 1))
                if key in seen_xy:
                    return
                seen_xy.add(key)
                drills.append([round(x, 2), round(y, 2), round(r, 3), plated])
            for pd in pcb.pads:
                try:
                    hs = float(getattr(pd, "hole_size_mils", 0) or 0) * MIL2MM
                    if hs < 0.6 or int(getattr(pd, "hole_shape", 0) or 0) != 0:
                        continue      # yuvarlak olmayan (slot/kare) delik kesilmez
                    _add(float(pd.x_mils) * MIL2MM - cx,
                         float(pd.y_mils) * MIL2MM - cy, hs / 2,
                         1 if getattr(pd, "is_plated", True) else 0)
                except Exception:
                    continue
            for vv in (getattr(pcb, "vias", None) or []):
                try:
                    hs = float(getattr(vv, "hole_size_mils", 0) or 0) * MIL2MM
                    if hs < 0.6:
                        continue
                    _add(float(vv.x_mils) * MIL2MM - cx,
                         float(vv.y_mils) * MIL2MM - cy, hs / 2, 1)
                except Exception:
                    continue
            if len(drills) > 1200:    # earcut emniyeti: en büyükler öncelikli
                drills.sort(key=lambda d: -d[2])
                drills = drills[:1200]
        except Exception:
            drills = []
        log(f"  ✓ 3D: board {maxx-minx:.0f}×{maxy-miny:.0f}mm · {th_mm:.2f}mm ·"
            f" {len(placements)} STEP + {len(bodies)} extrude gövde ·"
            f" {len(drills)} gerçek delik")
        return {"available": True, "thickness": round(th_mm, 3),
                "outline": outline, "bodies": bodies,
                "models": models, "placements": placements, "drills": drills,
                "w": round(maxx - minx, 2), "h": round(maxy - miny, 2)}
    except Exception as e:
        log(f"  · 3D çıkarılamadı: {e}")
        return {"available": False}


def _gen_distinct_color(i):
    """@brief i. ayırt edilebilir rengi üret (golden-angle HSL ile geniş ton dağılımı).

    Katman sayısı kadar farklı renk gerektiğinde (mech/other çakışmaları, palet taşması)
    kullanılır. Altın açı (0.618) ile ardışık renkler hue'da olabildiğince uzak olur.
    @param i Sıra indeksi (0,1,2,…)
    @return "#RRGGBB"
    """
    import colorsys
    h = (0.137 + i * 0.6180339887) % 1.0
    s = 0.50 + 0.18 * (i % 3) / 2.0
    l = 0.50 + 0.10 * (i % 2)
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return "#%02x%02x%02x" % (int(r * 255 + 0.5), int(g * 255 + 0.5), int(b * 255 + 0.5))


def _recolor_pcb_layer(svg, color, role):
    """@brief 2D PCB katmanını swatch rengine boyar (render ↔ swatch renk uyumu).

    altium_monkey her katmanı sabit ham renkle çiziyor (TOP kırmızı, BOTTOM mavi,
    silk sarı …) — bunlar sidebar swatch renkleriyle uyuşmuyordu. Burada tüm
    primitive fill/stroke renkleri katmanın swatch rengine (`color`) boyanır →
    render swatch ile birebir eşleşir, Altium gibi çok-renkli katman görünümü olur.
    Beyaz knockout (#FFFFFF) şeffaf yapılır (drill rolünde delik = color); board
    outline (#C0A000) nötr gri kenara çekilir. Bakır pour'u (shapebased-region,
    ham #000000) de boyandığından dolu bakır alan görünür (önceden siyah=görünmez).

    PLANE rolünde altium_monkey pour'u (dolu güç düzlemi) HİÇ vermiyor — yalnız split
    çizgileri var. Bu yüzden board outline path'inden yarı-saydam dolu bir pour
    SENTEZLENİR ki Altium'daki gibi dolu güç düzlemi alanı görünsün (anti-pad/clearance
    verisi olmadığından yaklaşık; görselleştirme amaçlı).

    @param svg Ham katman SVG'si
    @param color Katman (swatch) rengi (hex)
    @param role Katman rolü (copper/silk/plane/drill/mask/…)
    @return Boyanmış SVG.
    """
    EDGE = "#5a5a5a"
    # PLANE pour sentezi (recolor'dan ÖNCE): board outline'ını yarı-saydam dolu kopyala
    if role == "plane":
        md = re.search(r'<path\b[^>]*?\bd="([^"]+)"[^>]*?C0A000', svg)
        if md:
            fill = ('<path d="%s" fill="%s" fill-opacity="0.5" stroke="none" '
                    'data-feature="plane-fill"/>' % (md.group(1), color))
            svg = re.sub(r'(<svg\b[^>]*>)', lambda m: m.group(1) + fill, svg, count=1)

    def repl(m):
        attr, hx = m.group(1), m.group(2).upper()
        if hx == "C0A000":
            return '%s="%s"' % (attr, EDGE)
        if hx == "FFFFFF":
            return ('%s="%s"' % (attr, color)) if role == "drill" else ('%s="none"' % attr)
        return '%s="%s"' % (attr, color)
    svg = re.sub(r'(fill|stroke)="(#[0-9A-Fa-f]{6})"', repl, svg)
    svg = re.sub(r'(fill|stroke):#[0-9A-Fa-f]{6}',
                 lambda m: '%s:%s' % (m.group(0).split(":")[0], color), svg)
    return svg


def collect_pcb_layers(project_path, log, max_layer_mb=8):
    """@brief PCB'nin tüm katmanlarını SVG olarak render et (tam görüntüleyici için).
    
    Her katman ayrı SVG, data-component metadata'sı korunur (cross-probe).

    max_layer_mb: tek bir katman bu boyutu (MB) aşarsa atlanır (örn devasa
    mekanik katmanlar HTML'i şişirmesin).

    Döner: {"available": bool, "layers": [{"name", "display", "role", "svg",
            "default_on", "color"}], "components": {desig: {...}},
            "view_w", "view_h", "pcb_name"}
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param log Log mesajı callback'i (str alır)
    @param max_layer_mb Katman başına maksimum boyut (MB)
    @return Üretilen sonuç.
    """
    result = {"available": False, "layers": [], "components": {}}
    try:
        from altium_monkey.altium_pcbdoc import AltiumPcbDoc
    except Exception:
        log("\n! AltiumPcbDoc API yok — PCB görüntüleyici atlanıyor.")
        return result

    pcb_paths = _resolve_pcbdoc_paths(project_path, log)
    if not pcb_paths:
        log("\n· PCB dosyası bulunamadı.")
        return result
    # Komponenti olan PCB'yi seç (panel genelde boş)
    pcb_path = None
    pcb = None
    for p in sorted(pcb_paths, key=lambda x: "panel" in x.stem.lower()):
        try:
            cand = AltiumPcbDoc.from_file(p)
            if cand.components:
                pcb_path, pcb = p, cand
                break
        except Exception:
            continue
    if pcb is None:
        log("\n! PCB parse edilemedi veya komponent yok.")
        return result

    log(f"\nPCB görüntüleyici: {pcb_path.name} ({len(pcb.components)} komponent)")

    # Katman render sırası ve renkleri (Altium'a yakın)
    # role: bakır / silkscreen / pasta / lehim / mekanik / drill
    LAYER_STYLE = {
        "TOP": ("Top Copper", "copper", "#ff0000", True),
        "BOTTOM": ("Bottom Copper", "copper", "#0000ff", False),
        "TOPOVERLAY": ("Top Silkscreen", "silk", "#e8e8e8", True),
        "BOTTOMOVERLAY": ("Bottom Silkscreen", "silk", "#999999", False),
        "TOPSOLDER": ("Top Solder Mask", "mask", "#1a6b3a", False),
        "BOTTOMSOLDER": ("Bottom Solder Mask", "mask", "#0d4d28", False),
        "TOPPASTE": ("Top Paste", "paste", "#b0b0b0", False),
        "BOTTOMPASTE": ("Bottom Paste", "paste", "#808080", False),
        "MULTILAYER": ("Multi-Layer (Pads)", "copper", "#d4a020", True),
        "DRILLS": ("Drills", "drill", "#222222", True),
        "DRILLDRAWING": ("Drill Drawing", "drill", "#555555", False),
    }
    # İç katman/plane renk paletleri (her numara için ayrı, ayırt edilebilir renk)
    INNER_PALETTE = ["#d08a3e", "#7fa6d9", "#cf5b6a", "#5fb487", "#b07cc6",
                     "#d4b94a", "#5bb0b0", "#9aa84e"]
    PLANE_PALETTE = ["#6b8e23", "#23788e", "#8e6b23", "#8e2370", "#4a8e23",
                     "#236b8e"]

    try:
        all_layers = pcb.to_layer_svgs()
    except Exception as e:
        log(f"! Katman render hatası: {e}")
        return result

    view_w = view_h = 0
    layers_out = []
    for name, svg in all_layers.items():
        size_mb = len(svg) / 1024 / 1024
        if size_mb > max_layer_mb:
            log(f"  · {name} atlandı ({size_mb:.0f}MB > {max_layer_mb}MB limit)")
            continue
        # viewBox'ı ilk katmandan al
        if not view_w:
            m = re.search(r'viewBox="[\d.\-]+\s+[\d.\-]+\s+([\d.\-]+)\s+([\d.\-]+)"', svg)
            if m:
                view_w, view_h = float(m.group(1)), float(m.group(2))
        disp, role, color, default_on = LAYER_STYLE.get(
            name, (name.title(), "other", "#888888", False))
        # Mekanik katmanlar: jenerik
        if name.startswith("MECHANICAL"):
            disp, role, color, default_on = (name.replace("MECHANICAL", "Mech "),
                                             "mech", "#7a6f4a", False)
        # İç sinyal/güç katmanları (MID1, MID2, ...) — her biri AYRI renk (paletten,
        # katman numarasına göre). Önceden hepsi tek renkti, ayırt edilemiyordu.
        elif name.startswith("MID"):
            num = name[3:]
            nm = re.search(r"(\d+)$", name)
            idx = (int(nm.group(1)) - 1) if nm else 0
            disp, role, color, default_on = (
                f"Inner {num}", "inner", INNER_PALETTE[idx % len(INNER_PALETTE)], False)
        elif name.startswith("INTERNALPLANE") or name.startswith("PLANE"):
            nm = re.search(r"(\d+)$", name)
            idx = (int(nm.group(1)) - 1) if nm else 0
            disp, role, color, default_on = (
                name.title(), "plane", PLANE_PALETTE[idx % len(PLANE_PALETTE)], False)
        # Ham SVG + seçilen rengi sakla; recolor (yıkıcı) benzersizlik geçişinden
        # SONRA, FINAL renkle yapılır (aşağıda).
        layers_out.append({
            "name": name, "display": disp, "role": role,
            "_raw": svg, "default_on": default_on, "color": color,
        })

    # TÜM katmanlara BENZERSİZ renk garanti et: kaç katman varsa hepsi farklı görünsün.
    # Curated/palet renkler (Top turuncu, Bottom mavi, silk beyaz, inner/plane paletleri)
    # ilk geldiğinde KORUNUR; çakışanlar (mech hep #7a6f4a, other hep #888888, palet
    # taşması) golden-angle HSL ile üretilen ayırt edilebilir renklerle değiştirilir.
    # Recolor her zaman HAM svg'den FINAL renge yapılır (recolor yıkıcı, üst üste binmez).
    used = set()
    gi = 0
    for lyr in layers_out:
        c = lyr["color"]
        if c.lower() in used:
            while True:
                c = _gen_distinct_color(gi); gi += 1
                if c.lower() not in used:
                    break
        used.add(c.lower())
        lyr["color"] = c
        lyr["svg"] = _recolor_pcb_layer(lyr.pop("_raw"), c, lyr["role"])

    # Komponent konumları (cross-probe popup için, mil→mm board-relative)
    MM = 0.0254
    bb = None
    try:
        bb = pcb.board.outline.bounding_box
    except Exception:
        pass
    comps = {}
    for i, c in enumerate(pcb.components):
        desig = getattr(c, "designator", "") or ""
        if not desig:
            continue
        try:
            x_mil, y_mil = pcb.get_component_pnp_position_mils(i, origin_relative=False)
        except Exception:
            x_mil = y_mil = 0
        rot = getattr(c, "rotation", 0)
        try:
            rot = float(str(rot).strip())
        except Exception:
            rot = 0.0
        comps[desig] = {
            "abs_x_mm": round(x_mil * MM, 2),
            "abs_y_mm": round(y_mil * MM, 2),
            "layer": getattr(c, "layer", "TOP") or "TOP",
            "rotation": round(rot, 1),
            "footprint": getattr(c, "footprint", "") or "",
        }

    # Render sırasına göre sırala (alttan üste: solder→copper→silk)
    role_order = {"mask": 0, "paste": 1, "copper": 2, "inner": 2, "plane": 2,
                  "drill": 3, "silk": 4, "mech": 5, "other": 6}
    layers_out.sort(key=lambda l: role_order.get(l["role"], 9))

    log(f"  ✓ {len(layers_out)} katman render edildi · "
        f"{len(comps)} komponent · görüntü {view_w:.0f}×{view_h:.0f}mm")
    board3d = _extract_3d(pcb, log)
    if board3d.get("available"):
        try:
            # board gerçek boyutu (mm) — yüzey dokusu outline'ı boyuta göre eşler
            board_wh_mm = None
            if bb and len(bb) == 4:
                board_wh_mm = (abs(bb[2] - bb[0]) * MM, abs(bb[3] - bb[1]) * MM)
            board3d["surf"] = _build_board_surface(
                all_layers, view_w, view_h, log, board_wh_mm)
        except Exception as e:
            log(f"  · yüzey dokusu atlandı: {e}")
            board3d["surf"] = None
    result.update({
        "available": True,
        "layers": layers_out,
        "components": comps,
        "view_w": view_w, "view_h": view_h,
        "pcb_name": pcb_path.name,
        "board3d": board3d,
    })
    return result


def generate_viewer(
    project_path: str,
    output_path: str,
    inter_sheet_color: str = "#4ec9b0",
    intra_sheet_color: str = "#ff9800",
    log: Callable[[str], None] = print,
    progress: Callable[[int, str], None] = None,
):
    """@brief Tam interaktif HTML viewer üret (gömülü SVG'lerle, büyük dosya).
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param output_path Çıktı dosyası yolu
    @param inter_sheet_color Sayfalar arası bağlantı rengi (hex)
    @param intra_sheet_color Sayfa içi bağlantı rengi (hex)
    @param log Log mesajı callback'i (str alır)
    @param progress İlerleme callback'i (yüzde:int, etiket:str)
    """
    prog = progress or (lambda percent, label: None)
    prog(2, "Şematik verisi toplanıyor")
    data = _collect_data(
        project_path, log, with_pcb=True,
        progress=lambda frac, label: prog(2 + int(frac * 88), label),
    )
    prog(92, "HTML oluşturuluyor")
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    html = build_html(
        data["sheets"], data["net_list"], data["components"], timestamp,
        inter_sheet_color, intra_sheet_color, pcb=data.get("pcb"),
        project_name=data["project_name"],
    )
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    log(f"\n✓ HTML üretildi: {out}")
    log(f"  build: {timestamp}  ({out.stat().st_size / 1024 / 1024:.1f} MB)")
    prog(100, "Tamamlandı")


def _extract_svg_inner(svg_str):
    """@brief Bir SVG string'inin dış svg sarmalını soyup iç içeriğini döndür.
    
    @param svg_str SVG metni
    @return Üretilen sonuç.
    """
    start = svg_str.find(">", svg_str.find("<svg"))
    end = svg_str.rfind("</svg>")
    if start == -1 or end == -1:
        return svg_str
    return svg_str[start + 1:end]


def generate_pcb_viewer(project_path, output_path, log=print, progress=None):
    """@brief Tam ekran PCB görüntüleyici HTML üret — Altium benzeri.
    
    Tüm katmanlar (seçilebilir), gerçek komponent yerleşimi, pan/zoom,
    komponente tıklayınca detay (cross-probe).
    progress(percent, label): GUI ilerleme çubuğu (opsiyonel; <0 = marquee).
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param output_path Çıktı dosyası yolu
    @param log Log mesajı callback'i (str alır)
    @param progress İlerleme callback'i (yüzde:int, etiket:str)
    @return Üretilen sonuç.
    """
    prog = progress or (lambda percent, label: None)
    # to_layer_svgs() süresi kestirilemeyen ağır adım → marquee.
    prog(-1, "PCB katmanları render ediliyor (uzun sürebilir)…")
    pcb = collect_pcb_layers(project_path, log)
    if not pcb.get("available"):
        log("! PCB görüntüleyici üretilemedi (PCB yok veya parse edilemedi).")
        return False

    comp_info = {}
    try:
        prog(70, "Şematik komponent bilgisi")
        log("\nŞematik komponent bilgisi (cross-probe zenginleştirme)...")
        data = _collect_data(project_path, lambda m: None, with_pcb=False)
        for c in data.get("components", []):
            comp_info[c["designator"]] = {
                "value": c.get("value", ""),
                "description": c.get("description", ""),
                "sheet": c.get("sheet_name", ""),
            }
    except Exception as e:
        log(f"  · Şematik bilgisi alınamadı (yine de devam): {e}")

    prog(90, "PCB HTML oluşturuluyor")
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    html = build_pcb_html(pcb, comp_info, timestamp, Path(project_path).stem)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    log(f"\n✓ PCB görüntüleyici üretildi: {out}")
    log(f"  build: {timestamp}  ({out.stat().st_size / 1024 / 1024:.1f} MB)")
    prog(100, "Tamamlandı")
    return True


def generate_combined_viewer(
    project_path: str,
    output_path: str,
    inter_sheet_color: str = "#4ec9b0",
    intra_sheet_color: str = "#ff9800",
    log: Callable[[str], None] = print,
    progress: Callable[[int, str], None] = None,
):
    """@brief Şematik + PCB tek HTML'de yan yana, çift yönlü cross-probe.
    
    Sol panel şematik, sağ panel PCB, ortada sürüklenebilir ayraç.
    Birinde komponente tıklayınca diğeri o komponenti gösterir (postMessage).

    İki viewer iframe içinde izole çalışır (JS çakışması olmaz). Her viewer'ın
    HTML'i kabuk sayfaya gömülür, runtime'da iframe'e srcdoc ile yüklenir.

    progress(percent, label): GUI ilerleme çubuğu için (opsiyonel). percent<0
    = belirsiz/marquee (örn. süresi kestirilemeyen PCB katman render adımı).
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param output_path Çıktı dosyası yolu
    @param inter_sheet_color Sayfalar arası bağlantı rengi (hex)
    @param intra_sheet_color Sayfa içi bağlantı rengi (hex)
    @param log Log mesajı callback'i (str alır)
    @param progress İlerleme callback'i (yüzde:int, etiket:str)
    @return Üretilen sonuç.
    """
    prog = progress or (lambda percent, label: None)

    # 3D STEP modelleri birleşik görünümün ayrılmaz parçası (3D sekmesi). Bağımlılık
    # eksikse sessizce extrude kutulara düşmek yerine üretimi açıkça durdur (kullanıcı
    # eksikliği fark etsin — bkz. CLAUDE.md "STEP gösterme özelliği çalışmıyor").
    missing = _check_step_deps()
    if missing:
        raise RuntimeError(
            "3D STEP modelleri için gerekli bağımlılık(lar) eksik: "
            + ", ".join(missing)
            + ".\nKur:  py -3.12 -m pip install " + " ".join(missing)
            + "\n(Bu paketler olmadan birleşik görünümün 3D sekmesi gerçek STEP "
            "geometrisi yerine basit extrude kutular gösterir.)"
        )

    log("Birleşik görünüm: şematik + PCB toplanıyor...")
    prog(2, "Şematik verisi toplanıyor")

    # Tek sefer veri topla (şematik + PCB birlikte). _collect_data 0..1
    # ilerlemesini genel barda %2..%50 bandına eşle.
    data = _collect_data(
        project_path, log, with_pcb=True,
        progress=lambda frac, label: prog(2 + int(frac * 48), label),
    )
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")

    # 1) Şematik HTML
    prog(52, "Şematik HTML oluşturuluyor")
    sch_html = build_html(
        data["sheets"], data["net_list"], data["components"], timestamp,
        inter_sheet_color, intra_sheet_color, pcb=data.get("pcb"),
        project_name=data["project_name"],
    )

    # 2) PCB HTML (tam katmanlı görüntüleyici)
    # to_layer_svgs() tek bloklu, süresi kestirilemeyen ağır çağrı → marquee.
    prog(-1, "PCB katmanları render ediliyor (uzun sürebilir)…")
    pcb_layers = collect_pcb_layers(project_path, log)
    have_pcb = pcb_layers.get("available", False)
    prog(86, "PCB HTML oluşturuluyor")
    if have_pcb:
        comp_info = {
            c["designator"]: {
                "value": c.get("value", ""),
                "description": c.get("description", ""),
                "sheet": c.get("sheet_name", ""),
            }
            for c in data.get("components", [])
        }
        pcb_html = build_pcb_html(pcb_layers, comp_info, timestamp,
                                  Path(project_path).stem)
    else:
        log("  · PCB yok — birleşik görünümde sağ panel boş olacak.")
        pcb_html = ("<!DOCTYPE html><html><body style='background:#0a0a0a;"
                    "color:#888;font-family:sans-serif;display:flex;"
                    "align-items:center;justify-content:center;height:100vh;'>"
                    "<div>Bu projede okunabilir PCB dosyası bulunamadı.</div>"
                    "</body></html>")

    # 3) 3D görünüm HTML'i (board + komponent gövdeleri)
    board3d = pcb_layers.get("board3d", {}) if have_pcb else {}
    have_3d = bool(board3d and board3d.get("available"))
    td_html = build_3d_html(board3d, timestamp, Path(project_path).stem) if have_3d else ""

    # 4) Kabuk: üç iframe + splitter + köprü
    prog(94, "Birleştiriliyor ve yazılıyor")
    shell = build_combined_shell(sch_html, pcb_html, timestamp,
                                 Path(project_path).stem, have_pcb,
                                 td_html=td_html, have_3d=have_3d)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(shell, encoding="utf-8")
    log(f"\n✓ Birleşik görünüm üretildi: {out}")
    log(f"  build: {timestamp}  ({out.stat().st_size / 1024 / 1024:.1f} MB)")
    prog(100, "Tamamlandı")
    return True


def build_3d_html(d3d, timestamp, project_name):
    """@brief Three.js ile 3D PCB: board levhası + extrude komponent gövdeleri,
    
    orbit kontrol (döndür/zoom/pan), tıkla→cross-probe. Self-contained (Three.js
    gömülü). f-string DEĞİL: büyük JS bloğunda brace kaçışı olmasın diye
    placeholder + .replace kullanılır (Python .replace tamamen literaldir).
    
    @param d3d 3B veri sözlüğü (board3d)
    @param timestamp Zaman damgası metni
    @param project_name Proje adı
    @return Üretilen sonuç.
    """
    import json
    tpl = r'''<!DOCTYPE html><html><head><meta charset="utf-8"><style>
  html,body{margin:0;height:100%;background:#9499a0;overflow:hidden;
            font-family:"Segoe UI",system-ui,sans-serif;}
  #c3d{display:block;width:100%;height:100%;}
  #tb3d{position:absolute;top:8px;right:8px;display:flex;gap:4px;z-index:5;}
  .b3d{background:#1a1a1a;border:1px solid #333;color:#bbb;font-size:11px;
       padding:4px 11px;border-radius:4px;cursor:pointer;}
  .b3d:hover{border-color:#4ec9b0;color:#4ec9b0;}
  .b3d.on{background:#4ec9b0;color:#0a0a0a;border-color:#4ec9b0;font-weight:bold;}
  #lbl3d{position:absolute;left:0;top:0;background:rgba(20,20,20,.9);color:#4ec9b0;
         font-size:12px;padding:2px 7px;border-radius:3px;pointer-events:none;
         display:none;transform:translate(10px,-130%);z-index:6;white-space:nowrap;}
  #info3d{position:absolute;left:8px;bottom:8px;color:#666;font-size:11px;z-index:5;
          pointer-events:none;}
  #err3d{position:absolute;inset:0;display:flex;align-items:center;
         justify-content:center;color:#888;font-size:13px;text-align:center;padding:20px;}
</style></head><body>
<canvas id="c3d"></canvas>
<div id="tb3d">
  <button class="b3d" id="v-iso">3B</button>
  <button class="b3d" id="v-top">Üst</button>
  <button class="b3d" id="v-bot">Alt</button>
  <button class="b3d" id="v-rot">Döndür</button>
  <button class="b3d on" id="v-comp" title="Komponentleri gizle/göster — çıplak board'u incele">Parçalar</button>
  <button class="b3d on" id="v-lod" title="LOD: döndürme/zoom sırasında çözünürlük düşürülür (akıcılık), durunca netleşir">LOD</button>
</div>
<div id="lbl3d"></div>
<div id="info3d">Sürükle: döndür · Tekerlek: zoom · Sağ-sürükle: kaydır · Tıkla: komponent</div>
<script>__THREE__</script>
<script>
const D = __DATA__;
const canvas = document.getElementById('c3d');
function fail(msg){ const e=document.createElement('div'); e.id='err3d'; e.textContent=msg;
  document.body.appendChild(e); }
let renderer, scene, camera;
const meshByDesig = {};   // desig -> [mesh,...]
const pickList = [];
let autoRot = false, selectedDesig = null;
const orbit = { r:120, az:-0.7, el:0.6, tx:0, ty:0, tz:0 };
const DEG = Math.PI/180;
const dimMats = [];   // seçimde karartılacak tüm materyaller
function dimReg(mat, desig){
  if(!mat) return;
  mat.userData.dDesig = desig || null;
  mat.userData.col0 = mat.color ? mat.color.clone() : null;
  mat.userData.emis0 = mat.emissive ? mat.emissive.getHex() : 0;
  mat.userData.op0 = (mat.opacity !== undefined) ? mat.opacity : 1;
  dimMats.push(mat);
}
function registerMesh(mesh, desig){
  mesh.userData.desig = desig || null;
  pickList.push(mesh);
  if(desig){ (meshByDesig[desig] = meshByDesig[desig] || []).push(mesh); }
  dimReg(mesh.material, desig);
}

function shapeFrom(poly){
  const s = new THREE.Shape();
  s.moveTo(poly[0][0], poly[0][1]);
  for(let i=1;i<poly.length;i++) s.lineTo(poly[i][0], poly[i][1]);
  s.closePath();
  return s;
}
function buildScene(){
  scene = new THREE.Scene();
  scene.background = new THREE.Color(0x9499a0);
  scene.add(new THREE.HemisphereLight(0xffffff, 0x35404e, 1.0));
  scene.add(new THREE.AmbientLight(0xffffff, 0.22));
  const d1 = new THREE.DirectionalLight(0xffffff, 0.8); d1.position.set(60,90,120); scene.add(d1);
  const d2 = new THREE.DirectionalLight(0xbfd0ff, 0.28); d2.position.set(-90,-40,60); scene.add(d2);
  const d3 = new THREE.DirectionalLight(0xfff0e0, 0.5); d3.position.set(20,40,-120); scene.add(d3);

  const th = D.thickness || 1.6;
  // Board levhası (yeşil FR4). Delikler (≥0.6mm: THT pad + büyük via) v2.9.32'den
  // itibaren shape.holes ile GERÇEKTEN kesilir → içinden arka plan görünür
  // (doku alfası da addSurface'te deliniyor). ExtrudeGeometry hole winding'ini
  // kendisi normalize eder; kesim patlarsa deliksiz geometriye düşülür.
  const bshape = shapeFrom(D.outline);
  (D.drills||[]).forEach(dr=>{
    const p = new THREE.Path();
    // aClockwise=false (CCW) ŞART: r128 ExtrudeGeometry delik winding'ini yalnız
    // dış kontur CCW gelip ters çevrildiğinde normalize eder; bu board'ların
    // outline'ı CW geldiğinden CW delik duvar normalleri ters kalıyordu (NPTH
    // duvarı backface-cull ile görünmezdi). CCW delik iki dalda da kanonik.
    p.absarc(dr[0], dr[1], dr[2], 0, Math.PI*2, false);
    bshape.holes.push(p);
  });
  let bgeo;
  try { bgeo = new THREE.ExtrudeGeometry(bshape, {depth:th, bevelEnabled:false, curveSegments:16}); }
  catch(e){ bgeo = new THREE.ExtrudeGeometry(shapeFrom(D.outline), {depth:th, bevelEnabled:false}); }
  const board = new THREE.Mesh(bgeo, new THREE.MeshStandardMaterial(
    {color:0x0c6b32, metalness:0.12, roughness:0.72}));
  board.position.z = -th/2; scene.add(board); dimReg(board.material, null);
  // Kenar çizgisi DELİKSİZ geometriden: yüzlerce delik çemberi çizgi kalabalığı
  // yapmasın (delik kenarını zaten barrel/duvar gösteriyor).
  const edge = new THREE.LineSegments(new THREE.EdgesGeometry(
    new THREE.ExtrudeGeometry(shapeFrom(D.outline), {depth:th, bevelEnabled:false})),
    new THREE.LineBasicMaterial({color:0x0a431f}));
  edge.position.z = -th/2; scene.add(edge); dimReg(edge.material, null);
  // Kaplamalı (plated) delik duvarları: altın barrel — tek InstancedMesh (tek draw
  // call). Yarıçap deliğinkinden %4 küçük: board'un yeşil delik duvarıyla z-fight
  // olmasın (altın önde kalır). NPTH montaj delikleri barrel almaz (çıplak FR4).
  const plated = (D.drills||[]).filter(d=>d[3]);
  if(plated.length && THREE.InstancedMesh){
    const cyl = new THREE.CylinderGeometry(1, 1, 1, 16, 1, true);
    cyl.rotateX(Math.PI/2);   // silindir ekseni Y→Z
    const bmat = new THREE.MeshStandardMaterial({color:0xb99043, metalness:0.72,
      roughness:0.40, side:THREE.DoubleSide});
    const inst = new THREE.InstancedMesh(cyl, bmat, plated.length);
    const m4 = new THREE.Matrix4();
    plated.forEach((d,i)=>{
      // boşluk: %4 ama en az 50µm — küçük deliklerde (0.6mm) 12µm oransal boşluk
      // kameranın derinlik hassasiyetinin altında kalıp z-fight yapabiliyordu
      const rr = Math.min(d[2]*0.96, d[2]-0.05);
      m4.makeScale(rr, rr, th+0.04);
      m4.setPosition(d[0], d[1], 0);
      inst.setMatrixAt(i, m4);
    });
    inst.instanceMatrix.needsUpdate = true;
    // r128 InstancedMesh'i TABAN geometrinin (birim silindir, origin) bounding
    // sphere'iyle frustum-cull eder → köşeye zoom'da tüm barrel'lar kaybolurdu
    inst.frustumCulled = false;
    scene.add(inst); dimReg(bmat, null);
  }
  addSurface();   // bakır + pad + silkscreen dokusu

  // Komponentler (extrude gövdeler)
  (D.bodies||[]).forEach(b=>{
    if(!b.poly || b.poly.length<3) return;
    const h = Math.max(b.h||0.4, 0.1);
    let geo;
    try { geo = new THREE.ExtrudeGeometry(shapeFrom(b.poly), {depth:h, bevelEnabled:false}); }
    catch(e){ return; }
    const mat = new THREE.MeshStandardMaterial(
      {color:new THREE.Color(b.color||'#3a3a3e'), metalness:0.30, roughness:0.5});
    const mesh = new THREE.Mesh(geo, mat);
    if(b.layer==='bottom') mesh.position.z = -(th/2 + (b.z0||0) + h);
    else                   mesh.position.z =  (th/2 + (b.z0||0));
    scene.add(mesh);
    registerMesh(mesh, b.d);
  });

  // STEP modelleri (gerçek mesh, gömülü) — placement listesi
  buildModels();

  // başlangıç kamera mesafesi board boyutuna göre
  orbit.r = Math.max(D.w||60, D.h||60) * 1.7;
}

async function decompB64(b64, gz){
  const bin = Uint8Array.from(atob(b64), c=>c.charCodeAt(0));
  if(gz && typeof DecompressionStream !== 'undefined'){
    const st = new Blob([bin]).stream().pipeThrough(new DecompressionStream('gzip'));
    const buf = await new Response(st).arrayBuffer();
    return new TextDecoder().decode(buf);
  }
  return new TextDecoder().decode(bin);   // gzip yoksa düz metin (fallback)
}
function addSurface(){
  const s = D.surf; if(!s || !s.top) return;
  if(s.gz && typeof DecompressionStream === 'undefined') return;  // eski tarayıcı → düz yeşil
  const th = D.thickness || 1.6;
  function mk(b64, top){
    decompB64(b64, s.gz).then(svg=>{
      const img = new Image();
      img.onload = ()=>{
        const MAX = 2048, ar = s.w/s.h;
        const cw = ar>=1 ? MAX : Math.round(MAX*ar);
        const chh = ar>=1 ? Math.round(MAX/ar) : MAX;
        const cv = document.createElement('canvas'); cv.width=cw; cv.height=chh;
        const ctx = cv.getContext('2d');
        ctx.drawImage(img, 0, 0, cw, chh);
        // Delik alfa-delme (v2.9.32): geometriden kesilen deliklerin içi dokuda da
        // ŞEFFAF yapılır → gerçek delikten arka plan görünür (altın annular ring
        // korunur; +0.8px pay koyu delik boyasının AA kalıntısını temizler).
        // Dünya→kanvas eşlemesi düzlem yerleşiminden türetilir: düzlem dünyada
        // [s.cx−w/2, s.cx+w/2]×[s.cy−h/2, s.cy+h/2] aralığını kaplar, kanvas v=0
        // üst kenar = +Y (dokular üst/alt AYNI XY eşlemesini kullanır, bkz. v2.9.4).
        if(s.ok && (D.drills||[]).length){
          ctx.save();
          ctx.globalCompositeOperation = 'destination-out';
          const k = cw / s.w;
          (D.drills||[]).forEach(d=>{
            const u = (d[0] - (s.cx - s.w/2)) * k;
            const v = ((s.cy + s.h/2) - d[1]) * (chh / s.h);
            const r = d[2] * k + 0.8;
            if(u < -r || v < -r || u > cw+r || v > chh+r) return;
            ctx.beginPath(); ctx.arc(u, v, r, 0, Math.PI*2); ctx.fill();
          });
          ctx.restore();
        }
        const tex = new THREE.CanvasTexture(cv);
        try { tex.anisotropy = renderer.capabilities.getMaxAnisotropy(); } catch(e){}
        const mat = new THREE.MeshStandardMaterial({map:tex, transparent:true,
          metalness:0.0, roughness:0.82, polygonOffset:true, polygonOffsetFactor:-2,
          depthWrite:false, side:THREE.DoubleSide});
        const pl = new THREE.Mesh(new THREE.PlaneGeometry(s.w, s.h), mat);
        // Üst ve alt doku AYNI XY eşlemesini kullanır (board feature (x,y)→dünya (x,y)).
        // Alt yüz İÇİN ROTASYON YOK: alt komponentler (scale.z=-1) XY'lerini koruduğu
        // için doku da korumalı; through-hole pad değişmezi bunu gerektirir. Eski
        // rotation.x=π Y-flip'i dokuyu komponentlerden kaydırıyordu (bakır board dışına
        // taşıyor, hizasız). DoubleSide ile aşağıdan (Alt kamera) görünür; SVG'deki
        // aynalı alt-silk + Alt kameranın X-aynası birbirini götürür → metin düz okunur.
        if(top){ pl.position.set(s.cx, s.cy, th/2 + 0.02); }
        else   { pl.position.set(s.cx, s.cy, -(th/2 + 0.02)); }
        scene.add(pl); dimReg(mat, null);
      };
      img.onerror = ()=>{};
      img.src = 'data:image/svg+xml;charset=utf-8,' + encodeURIComponent(svg);
    }).catch(()=>{});
  }
  mk(s.top, true);
  if(s.bot) mk(s.bot, false);
}

function buildModels(){
  const th = D.thickness || 1.6;
  const models = D.models || {};
  (D.placements || []).forEach(pl=>{
    const md = models[pl.m]; if(!md || !md.parts) return;
    const inner = new THREE.Group();
    md.parts.forEach(pt=>{
      if(!pt.v || !pt.f || pt.v.length < 9) return;
      const geo = new THREE.BufferGeometry();
      geo.setAttribute('position', new THREE.BufferAttribute(new Float32Array(pt.v), 3));
      geo.setIndex(pt.f);
      geo.computeVertexNormals();
      const mat = new THREE.MeshStandardMaterial(
        {color:new THREE.Color(pt.c||'#b0b0b6'), metalness:0.5, roughness:0.42,
         side:THREE.DoubleSide});
      inner.add(new THREE.Mesh(geo, mat));
    });
    if(!inner.children.length) return;
    // KESİN Altium yerleşimi (v2.9.31): dünya = R·v + (cx, cy, ±(th/2+dz));
    // R = Rz(rotz)·Ry(roty)·Rx(rotx). Sıra, anchor (model_2d) ve dz, iki board'da
    // tüm STEP parçaların bacak↔pad hizalamasıyla ampirik doğrulandı (~0.3mm).
    // Alt katman: anchor'dan geçen X ekseni etrafında 180° PROPER rotation —
    // eski scale.z=-1 aynası parçanın ayna görüntüsünü çiziyordu (D-sub gibi
    // yönlü parçalarda yüz 180° ters, D-şekli ters elli görünüyordu).
    const AX = new THREE.Vector3(1,0,0), AY = new THREE.Vector3(0,1,0),
          AZ = new THREE.Vector3(0,0,1);
    const q = new THREE.Quaternion().setFromAxisAngle(AZ,(pl.rz||0)*DEG)
      .multiply(new THREE.Quaternion().setFromAxisAngle(AY,(pl.ry||0)*DEG))
      .multiply(new THREE.Quaternion().setFromAxisAngle(AX,(pl.rx||0)*DEG));
    if(pl.layer==='bottom'){
      // Rx(180)·(R·v + dz·ẑ) + (cx,cy,−th/2)  →  konum z = −th/2 − dz
      q.premultiply(new THREE.Quaternion().setFromAxisAngle(AX, Math.PI));
      inner.position.set(pl.cx, pl.cy, -(th/2 + (pl.dz||0)));
    } else {
      inner.position.set(pl.cx, pl.cy, th/2 + (pl.dz||0));
    }
    inner.quaternion.copy(q);
    scene.add(inner);
    inner.children.forEach(m=>registerMesh(m, pl.d));
  });
}

// === Kamera / orbit (up=+Y; üst/alt görünümler el=0'da kararlı, kutuplar kenar-bakışta) ===
function applyCam(){
  const el = Math.max(-1.48, Math.min(1.48, orbit.el));
  const ce = Math.cos(el), se = Math.sin(el);
  camera.position.set(
    orbit.tx + orbit.r*ce*Math.sin(orbit.az),
    orbit.ty + orbit.r*se,
    orbit.tz + orbit.r*ce*Math.cos(orbit.az));
  camera.lookAt(orbit.tx, orbit.ty, orbit.tz);
}
// === LOD: etkileşim sırasında dinamik çözünürlük (2D LOD'un WebGL karşılığı):
// döndürme/pan/zoom BOYUNCA pixelRatio düşürülür (~6x az piksel), hareket
// durunca 220ms'de tam çözünürlüğe dönülür. Döndür (autoRot) etkilenmez.
let lodOn = true; try { lodOn = localStorage.getItem('schviz-3dlod') !== '0'; } catch(e){}
let lodLow = false, lodT = null;
function lodSetLow(low){
  low = low && lodOn;
  if (low === lodLow || !renderer) return;
  lodLow = low;
  const basePR = Math.min(devicePixelRatio||1, 1.5);
  renderer.setPixelRatio(low ? Math.min(basePR, 0.6) : basePR);
  resize();   // drawing buffer'ı yeni pixelRatio ile yeniden boyutlandırır
}
function lodTouch(){ lodSetLow(true); clearTimeout(lodT);
  lodT = setTimeout(()=>lodSetLow(false), 220); }
const lodBtn3d = document.getElementById('v-lod');
if (!lodOn) lodBtn3d.classList.remove('on');
lodBtn3d.onclick = ()=>{ lodOn = !lodOn; lodBtn3d.classList.toggle('on', lodOn);
  try { localStorage.setItem('schviz-3dlod', lodOn?'1':'0'); } catch(e){}
  if (!lodOn) lodSetLow(false); };

let drag=null, lastX=0, lastY=0, moved=false;
canvas.addEventListener('mousedown', e=>{ drag = (e.button===2?'pan':'rot');
  lastX=e.clientX; lastY=e.clientY; moved=false; e.preventDefault(); });
window.addEventListener('mousemove', e=>{
  if(!drag) return;
  const dx=e.clientX-lastX, dy=e.clientY-lastY; lastX=e.clientX; lastY=e.clientY;
  if(Math.abs(dx)>2||Math.abs(dy)>2) moved=true;
  if(drag==='rot'){ orbit.az -= dx*0.0095; orbit.el += dy*0.0095; }
  else {
    const k = orbit.r*0.0016;
    const right = new THREE.Vector3().setFromMatrixColumn(camera.matrix,0);
    const up    = new THREE.Vector3().setFromMatrixColumn(camera.matrix,1);
    orbit.tx += (-right.x*dx + up.x*dy)*k;
    orbit.ty += (-right.y*dx + up.y*dy)*k;
    orbit.tz += (-right.z*dx + up.z*dy)*k;
  }
  lodTouch();
  applyCam();
});
window.addEventListener('mouseup', ()=>{ drag=null; });
canvas.addEventListener('contextmenu', e=>e.preventDefault());
canvas.addEventListener('wheel', e=>{ e.preventDefault();
  lodTouch();
  orbit.r *= (e.deltaY<0 ? 0.9 : 1.1);
  orbit.r = Math.max(4, Math.min(3000, orbit.r)); applyCam(); }, {passive:false});

// === Seçim + cross-probe ===
const ray = new THREE.Raycaster(), ndc = new THREE.Vector2();
function pick(e){
  const r=canvas.getBoundingClientRect();
  ndc.x = ((e.clientX-r.left)/r.width)*2-1;
  ndc.y = -((e.clientY-r.top)/r.height)*2+1;
  ray.setFromCamera(ndc, camera);
  // Gizli komponentler tıklama/hover hedefi olmasın (raycaster visible'a bakmaz)
  const hits = ray.intersectObjects(pickList.filter(m=>m.visible), false);
  return hits.length ? hits[0].object : null;
}
function setSel(desig){
  const has = !!(desig && meshByDesig[desig] && meshByDesig[desig].length);
  dimMats.forEach(m=>{
    // önce eski haline döndür
    if(m.userData.col0) m.color.copy(m.userData.col0);
    if(m.emissive) m.emissive.setHex(m.userData.emis0 || 0);
    m.opacity = m.userData.op0;
    if(!has) return;
    if(m.userData.dDesig === desig){
      // seçili komponent: tam renk + yeşil parıltı
      if(m.emissive) m.emissive.setHex(0x37b257);
    } else {
      // diğer her şey kararsın (board + komponentler + doku)
      if(m.userData.col0) m.color.copy(m.userData.col0).multiplyScalar(0.20);
      if(m.transparent) m.opacity = (m.userData.op0 || 1) * 0.42;
    }
  });
  selectedDesig = has ? desig : null;
}
canvas.addEventListener('click', e=>{
  if(moved) return;
  const o = pick(e);
  if(o && o.userData.desig){ setSel(o.userData.desig); crossOut(o.userData.desig); }
  else { setSel(null); }
});
const lbl = document.getElementById('lbl3d');
// Hover raycast PAHALI (yüzlerce mesh) — her mousemove'da değil, frame başına 1 kez
// işle (requestAnimationFrame throttle). Hover gecikmesi göze çarpmaz, FPS korunur.
let hoverEv=null, hoverQueued=false;
canvas.addEventListener('mousemove', e=>{
  hoverEv=e;
  if(hoverQueued) return;
  hoverQueued=true;
  requestAnimationFrame(()=>{
    hoverQueued=false;
    const ev=hoverEv; if(!ev) return;
    if(drag){ lbl.style.display='none'; return; }
    const o = pick(ev);
    if(o && o.userData.desig){
      const r=canvas.getBoundingClientRect();
      lbl.textContent=o.userData.desig;
      lbl.style.left=(ev.clientX-r.left)+'px'; lbl.style.top=(ev.clientY-r.top)+'px';
      lbl.style.display='block'; canvas.style.cursor='pointer';
    } else { lbl.style.display='none'; canvas.style.cursor='default'; }
  });
});
const IN_FRAME = window.parent && window.parent!==window;
function crossOut(d){ if(IN_FRAME) window.parent.postMessage(
  {type:'xprobe', source:'3d', designator:d}, '*'); }
window.addEventListener('message', ev=>{
  const m=ev.data; if(!m||m.type!=='xprobe'||m.source==='3d') return;
  // Diğer panelden seçim geldi — parçalar gizliyse geri aç ki seçim görünsün
  if(!compsVisible) compBtn.onclick();
  let d = m.designator;
  if(d && !meshByDesig[d]){
    // Şematikten MANTIKSAL ad gelmiş olabilir (hiyerarşik kanal: R103) →
    // kanal-sonekli ilk fiziksel kopyaya çözümle (R103_diffI2C_1)
    const q = (d+'_').toUpperCase();
    const c = Object.keys(meshByDesig).filter(k=>k.toUpperCase().startsWith(q)).sort();
    if(c.length) d = c[0];
  }
  setSel(d);
});

// === Butonlar ===
document.getElementById('v-iso').onclick=()=>{ orbit.el=0.6; orbit.az=-0.7; applyCam(); };
document.getElementById('v-top').onclick=()=>{ orbit.el=0; orbit.az=0; applyCam(); };
document.getElementById('v-bot').onclick=()=>{ orbit.el=0; orbit.az=Math.PI; applyCam(); };
const rotBtn=document.getElementById('v-rot');
rotBtn.onclick=()=>{ autoRot=!autoRot; rotBtn.classList.toggle('on',autoRot); };

// === Komponentleri gizle/göster: çıplak board'u (bakır/silkscreen dokusu)
//     incelemek için. Board levhası + yüzey dokusu kalır; desig'li tüm
//     mesh'ler (extrude gövdeler + STEP modelleri) gizlenir. ===
const compBtn=document.getElementById('v-comp');
let compsVisible=true;
compBtn.onclick=()=>{
  compsVisible=!compsVisible;
  compBtn.classList.toggle('on', compsVisible);
  pickList.forEach(m=>{ if(m.userData.desig) m.visible=compsVisible; });
  if(!compsVisible) setSel(null);   // gizli parçada seçim asılı kalmasın
};

function resize(){
  if(!renderer) return;
  const W=canvas.clientWidth||800, H=canvas.clientHeight||600;
  camera.aspect=W/H; camera.updateProjectionMatrix(); renderer.setSize(W,H,false);
}
function loop(){
  requestAnimationFrame(loop);
  if(autoRot){ orbit.az -= 0.0042; applyCam(); }
  renderer.render(scene, camera);
}
function init(){
  const W=canvas.clientWidth||800, H=canvas.clientHeight||600;
  camera = new THREE.PerspectiveCamera(42, W/H, 0.1, 8000);
  camera.up.set(0,1,0);
  try { renderer = new THREE.WebGLRenderer({canvas, antialias:true}); }
  catch(e){ fail('WebGL bu tarayıcıda kullanılamıyor.'); return; }
  renderer.setPixelRatio(Math.min(devicePixelRatio||1, 1.5));  // yüksek-DPI'de piksel sınırı → FPS
  renderer.setSize(W,H,false);
  buildScene();
  applyCam();
  new ResizeObserver(resize).observe(canvas);
  window.addEventListener('resize', resize);
  loop();
}
if(D && D.available && D.outline && D.outline.length>=3) init();
else fail('Bu projede 3D verisi bulunamadı.');
</script>
</body></html>'''
    return (tpl
            .replace("__THREE__", _three_js_source())
            .replace("__DATA__", json.dumps(d3d)))


def build_combined_shell(sch_html, pcb_html, timestamp, project_name, have_pcb,
                         td_html="", have_3d=False):
    """@brief Şematik + PCB (+ opsiyonel 3D) iframe'lerini bir kabukta toplar.
    
    HTML'ler JSON string olarak gömülür, srcdoc ile iframe'e yüklenir. 3D ağır
    olduğundan tembel yüklenir (ilk 3D moduna geçişte). Cross-probe üç yönlüdür.
    
    @param sch_html Şematik iç HTML'i
    @param pcb_html PCB iç HTML'i
    @param timestamp Zaman damgası metni
    @param project_name Proje adı
    @param have_pcb PCB mevcut mu (bool)
    @param td_html 3B görünümün iç HTML'i
    @param have_3d 3B görünüm mevcut mu (bool)
    @return Üretilen sonuç.
    """
    # Performans/dosya boyutu: iç HTML'ler gzip+base64 gömülür (ham JSON string yerine),
    # runtime'da DecompressionStream ile açılır. base64 olduğundan `</script>` kaçışına
    # da gerek kalmaz. ~3-6× daha küçük gömme (HTML iyi sıkışır).
    import gzip as _gzip, base64 as _b64
    def _gz(s):
        return _b64.b64encode(_gzip.compress((s or "").encode("utf-8"), 6)).decode()
    sch_b64 = _gz(sch_html)
    pcb_b64 = _gz(pcb_html)
    td_b64 = _gz(td_html) if have_3d else ""
    btn_3d = ('<button class="vm-btn" id="vm-3d" title="3D görünüm ( 4 )">3D</button>'
              if have_3d else '')
    pane_3d = ('<div class="pane" id="pane-3d" style="display:none;width:100%">'
               '<div class="pane-loading" id="load-3d">3D hazırlanıyor…</div>'
               '<iframe id="frame-3d"></iframe></div>'
               if have_3d else '')
    have_3d_js = 'true' if have_3d else 'false'
    return f"""<!DOCTYPE html>
<html lang="tr"><head>
<meta charset="utf-8">
<title>Şematik + PCB · {project_name} · {timestamp}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  html,body {{ height:100%; overflow:hidden; background:#0a0a0a;
               font-family:'Segoe UI',sans-serif; }}
  #topbar {{ height:34px; background:#161616; border-bottom:1px solid #333;
             display:flex; align-items:center; padding:0 12px; gap:14px;
             color:#ccc; font-size:12px; }}
  #topbar .title {{ color:#4ec9b0; font-weight:bold; }}
  #topbar .hint {{ color:#666; }}
  #topbar .badge {{ margin-left:auto; color:#555; font-size:11px; }}
  #view-modes {{ display:flex; gap:2px; margin-left:18px; }}
  .vm-btn {{ background:#222; border:1px solid #3a3a3a; color:#bbb;
             font-size:11px; padding:3px 12px; cursor:pointer; }}
  .vm-btn:first-child {{ border-radius:4px 0 0 4px; }}
  .vm-btn:last-child {{ border-radius:0 4px 4px 0; }}
  .vm-btn:hover {{ color:#4ec9b0; border-color:#4ec9b0; }}
  .vm-btn.active {{ background:#4ec9b0; color:#0a0a0a; border-color:#4ec9b0;
                    font-weight:bold; }}
  #split {{ display:flex; height:calc(100vh - 34px); width:100%; }}
  .pane {{ height:100%; overflow:hidden; position:relative; }}
  #pane-sch {{ width:50%; }}
  #pane-pcb {{ flex:1; }}
  .pane iframe {{ width:100%; height:100%; border:none; display:block; }}
  #divider {{ width:6px; background:#333; cursor:col-resize; flex-shrink:0;
              position:relative; }}
  #divider:hover {{ background:#4ec9b0; }}
  #divider::after {{ content:'⋮'; position:absolute; top:50%; left:50%;
                     transform:translate(-50%,-50%); color:#666; font-size:14px; }}
  /* Ağır panel (PCB/3D) yüklenirken dönen gösterge */
  .pane-loading {{ position:absolute; inset:0; display:flex; align-items:center;
                   justify-content:center; color:#4ec9b0; font-size:13px;
                   background:#0f0f0f; z-index:5; }}
  .pane-loading::before {{ content:''; width:16px; height:16px;
                           border:2px solid #333; border-top-color:#4ec9b0;
                           border-radius:50%; margin-right:10px;
                           animation:spin 0.9s linear infinite; }}
  @keyframes spin {{ to {{ transform:rotate(360deg); }} }}
</style>
</head>
<body>
<div id="topbar">
  <span class="title">Şematik + PCB</span>
  <span class="hint">Bir tarafta komponente tıkla → diğerlerinde otomatik gösterilir</span>
  <div id="view-modes">
    <button class="vm-btn active" id="vm-sch" title="Sadece şematik ( 1 )">Şematik</button>
    <button class="vm-btn" id="vm-both" title="Yan yana ( 2 )">Böl</button>
    <button class="vm-btn" id="vm-pcb" title="Sadece PCB ( 3 )">PCB</button>
    {btn_3d}
  </div>
  <span class="badge">{project_name} · v{APP_VERSION}</span>
</div>
<div id="split">
  <div class="pane" id="pane-sch">
    <iframe id="frame-sch"></iframe>
  </div>
  <div id="divider"></div>
  <div class="pane" id="pane-pcb">
    <div class="pane-loading" id="load-pcb">PCB hazırlanıyor…</div>
    <iframe id="frame-pcb"></iframe>
  </div>
  {pane_3d}
</div>
<script>
// İç HTML'ler gzip+base64 gömülü — runtime'da DecompressionStream ile açılır.
const SCH_GZ = "{sch_b64}";
const PCB_GZ = "{pcb_b64}";
const TD_GZ  = "{td_b64}";
const HAVE_3D  = {have_3d_js};

const frameSch = document.getElementById('frame-sch');
const framePcb = document.getElementById('frame-pcb');
const frame3d  = HAVE_3D ? document.getElementById('frame-3d') : null;
let td_loaded = false, pcb_loaded = false, lastSel = null;

// gzip+base64 → metin (DecompressionStream). Yoksa null (eski tarayıcı).
async function gunzipB64(b64) {{
  const bin = Uint8Array.from(atob(b64), c => c.charCodeAt(0));
  if (typeof DecompressionStream === 'undefined') return null;
  const st = new Blob([bin]).stream().pipeThrough(new DecompressionStream('gzip'));
  return await new Response(st).text();
}}
// İç HTML'i aç + srcdoc'a yaz (blob file:// altında engelli; srcdoc üst origin'i
// miras alır, postMessage cross-probe çalışır). Eski tarayıcıda uyarı gösterir.
async function loadFrame(id, b64) {{
  const fr = document.getElementById(id);
  if (!fr) return;
  const ld = document.getElementById(id.replace('frame-', 'load-'));
  if (ld) fr.addEventListener('load', () => {{ ld.style.display = 'none'; }}, {{ once: true }});
  const html = await gunzipB64(b64);
  fr.srcdoc = (html === null)
    ? '<body style="font:14px sans-serif;color:#999;background:#111;padding:24px">'
      + 'Bu tarayıcı sıkıştırılmış görünümü açamıyor (DecompressionStream gerekli). '
      + 'Lütfen tarayıcıyı güncelleyin.</body>'
    : html;
}}
// Performans: açılış SADECE şematik — PCB ve 3D tembel yüklenir (ilk o moda
// geçişte). Açılışta ekstra yük yok; şematik hemen etkileşime hazır.
let curMode = 'sch';
loadFrame('frame-sch', SCH_GZ);
function repostSel(fr) {{ if (lastSel) setTimeout(() =>
  postTo(fr, {{type:'xprobe', source:'sch', designator:lastSel}}), 400); }}
function ensurePcbLoaded() {{
  if (!pcb_loaded) {{ pcb_loaded = true; loadFrame('frame-pcb', PCB_GZ).then(() => repostSel(framePcb)); }}
}}
function ensure3dLoaded() {{
  if (HAVE_3D && !td_loaded) {{ td_loaded = true; loadFrame('frame-3d', TD_GZ).then(() => repostSel(frame3d)); }}
}}

// === Cross-probe yönlendirme (üç yönlü) ===
function postTo(fr, d) {{ try {{ if (fr && fr.contentWindow) fr.contentWindow.postMessage(d, '*'); }} catch(e) {{}} }}
window.addEventListener('message', ev => {{
  const d = ev.data;
  if (!d || d.type !== 'xprobe') return;
  if (d.designator) lastSel = d.designator;
  // PCB'yi yalnızca görünür bir modda yükle — sadece-şematik modunda arka planda
  // ağır PCB yüklenmesin (lastSel saklanır, moda geçince repostSel ile iletilir).
  if (d.source !== 'pcb' && (curMode === 'both' || curMode === 'pcb')) ensurePcbLoaded();
  if (d.source === 'sch') {{ postTo(framePcb, d); postTo(frame3d, d); }}
  else if (d.source === 'pcb') {{ postTo(frameSch, d); postTo(frame3d, d); }}
  else if (d.source === '3d') {{ postTo(frameSch, d); postTo(framePcb, d); }}
}});

// === Sürüklenebilir ayraç ===
const divider = document.getElementById('divider');
const paneSch = document.getElementById('pane-sch');
const split = document.getElementById('split');
let resizing = false, lastSplitPct = 50;
divider.addEventListener('mousedown', () => {{
  resizing = true; document.body.style.userSelect = 'none';
  frameSch.style.pointerEvents = 'none'; framePcb.style.pointerEvents = 'none';
}});
window.addEventListener('mousemove', e => {{
  if (!resizing) return;
  const rect = split.getBoundingClientRect();
  let pct = ((e.clientX - rect.left) / rect.width) * 100;
  pct = Math.max(15, Math.min(85, pct));
  paneSch.style.width = pct + '%'; lastSplitPct = pct;
}});
window.addEventListener('mouseup', () => {{
  resizing = false; document.body.style.userSelect = '';
  frameSch.style.pointerEvents = ''; framePcb.style.pointerEvents = '';
}});

// === Görünüm modu: şematik / böl / PCB / 3D ===
const paneSchEl = document.getElementById('pane-sch');
const panePcbEl = document.getElementById('pane-pcb');
const pane3dEl  = HAVE_3D ? document.getElementById('pane-3d') : null;
const dividerEl = document.getElementById('divider');
const vmButtons = {{ sch:document.getElementById('vm-sch'),
                     both:document.getElementById('vm-both'),
                     pcb:document.getElementById('vm-pcb') }};
if (HAVE_3D) vmButtons.td = document.getElementById('vm-3d');
function setViewMode(mode) {{
  curMode = mode;
  const show3d = (mode === 'td');
  if (pane3dEl) pane3dEl.style.display = show3d ? '' : 'none';
  if (show3d) {{
    paneSchEl.style.display='none'; panePcbEl.style.display='none';
    dividerEl.style.display='none';
    ensure3dLoaded();
    // son seçili komponenti 3D'ye ilet (yüklendikten sonra)
    if (lastSel) setTimeout(() => postTo(frame3d, {{type:'xprobe',source:'sch',designator:lastSel}}), 350);
  }} else if (mode === 'sch') {{
    paneSchEl.style.display=''; paneSchEl.style.width='100%';
    panePcbEl.style.display='none'; dividerEl.style.display='none';
  }} else if (mode === 'pcb') {{
    ensurePcbLoaded();
    paneSchEl.style.display='none'; dividerEl.style.display='none';
    panePcbEl.style.display='';
  }} else {{ // both
    ensurePcbLoaded();
    paneSchEl.style.display=''; paneSchEl.style.width=lastSplitPct + '%';
    panePcbEl.style.display=''; dividerEl.style.display='';
  }}
  Object.entries(vmButtons).forEach(([k, b]) => b && b.classList.toggle('active', k === mode));
}}
// Klavye: 1=Şematik 2=Böl 3=PCB 4=3D (odak kabuktayken)
document.addEventListener('keydown', e => {{
  if (e.key === '1') setViewMode('sch');
  else if (e.key === '2') setViewMode('both');
  else if (e.key === '3') setViewMode('pcb');
  else if (e.key === '4' && HAVE_3D) setViewMode('td');
}});
vmButtons.sch.onclick = () => setViewMode('sch');
vmButtons.pcb.onclick = () => setViewMode('pcb');
vmButtons.both.onclick = () => setViewMode('both');
if (HAVE_3D) vmButtons.td.onclick = () => setViewMode('td');
// Açılış modu: SADECE şematik (Böl/PCB/3D moda geçince ilgili panel tembel yüklenir).
setViewMode('sch');
// Komponent seçilince görünüm modu DEĞİŞMEZ; cross-probe arka planda çalışır.
</script>
</body></html>"""


def build_pcb_html(pcb, comp_info, timestamp, project_name):
    """@brief PCB görüntüleyici HTML'i oluştur.
    
    @param pcb AltiumPcbDoc PCB nesnesi
    @param comp_info Komponent bilgi sözlüğü
    @param timestamp Zaman damgası metni
    @param project_name Proje adı
    @return Üretilen sonuç.
    """
    view_w = pcb["view_w"] or 200
    view_h = pcb["view_h"] or 200

    # Performans: VARSAYILAN KAPALI katmanlar açılışta DOM'a konmaz (boş <g>), içerikleri
    # LAZY_SVG'de tutulur ve ilk gösterimde enjekte edilir → başlangıç DOM/parse maliyeti
    # düşer. Net-highlight / pad-label gibi çapraz-katman işlemler tüm gizli katmanları da
    # taradığından, onlar tetiklendiğinde loadAllLazyLayers() ile hepsi bir kerede yüklenir.
    layer_divs = []
    layer_meta = []
    lazy_svgs = {}
    for i, lyr in enumerate(pcb["layers"]):
        inner = _extract_svg_inner(lyr["svg"])
        if lyr["default_on"]:
            layer_divs.append(
                f'<g class="pcb-layer" id="layer-{i}" '
                f'data-layer="{lyr["name"]}" style="display:block">{inner}</g>'
            )
        else:
            # boş placeholder; içerik LAZY_SVG[i]'de, ilk gösterimde enjekte
            layer_divs.append(
                f'<g class="pcb-layer" id="layer-{i}" data-lazy="1" '
                f'data-layer="{lyr["name"]}" style="display:none"></g>'
            )
            lazy_svgs[i] = inner
        layer_meta.append({
            "id": i, "name": lyr["name"], "display": lyr["display"],
            "role": lyr["role"], "color": lyr["color"],
            "default_on": lyr["default_on"],
        })
    layers_svg = "\n".join(layer_divs)

    return f"""<!DOCTYPE html>
<html lang="tr"><head>
<meta charset="utf-8">
<title>PCB Görüntüleyici · {project_name} · {timestamp}</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ background:#0a0a0a; color:#ddd; font-family:'Segoe UI',sans-serif;
          overflow:hidden; height:100vh; }}
  #app {{ display:flex; height:100vh; }}
  #sidebar {{ width:240px; background:#161616; border-right:1px solid #333;
              display:flex; flex-direction:column; flex-shrink:0;
              position:relative; overflow:hidden;
              transition:width .18s ease; }}
  /* Sol panel tamamen katlanabilir — dar bir şerit + ▸ butonu kalır */
  #sidebar.collapsed {{ width:26px; }}
  #sidebar.collapsed > *:not(#sb-toggle) {{ display:none !important; }}
  #sb-toggle {{ position:absolute; top:9px; right:8px; width:20px; height:20px;
                padding:0; background:#1a1a1a; border:1px solid #333;
                color:#888; font-size:11px; line-height:18px; text-align:center;
                cursor:pointer; border-radius:3px; z-index:5; }}
  #sb-toggle:hover {{ color:#4ec9b0; border-color:#4ec9b0; }}
  #sidebar.collapsed #sb-toggle {{ position:static; margin:6px auto 0;
                                   display:block; }}
  #sidebar h2 {{ font-size:13px; padding:12px 32px 12px 12px;
                 border-bottom:1px solid #333;
                 color:#4ec9b0; text-transform:uppercase; letter-spacing:1px; }}
  .build-info {{ font-size:10px; color:#555; padding:4px 12px; }}
  #layer-list {{ flex:1; overflow-y:auto; padding:6px; }}
  .layer-item {{ display:flex; align-items:center; padding:6px 8px;
                 cursor:pointer; border-radius:4px; font-size:12px;
                 user-select:none; }}
  .layer-item:hover {{ background:#222; }}
  .layer-swatch {{ width:14px; height:14px; border-radius:3px; margin-right:8px;
                   border:1px solid #444; flex-shrink:0; }}
  .layer-name {{ flex:1; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
  .layer-top {{ background:transparent; border:1px solid #444; color:#999;
                cursor:pointer; border-radius:4px; font-size:11px; line-height:1;
                padding:2px 5px; margin:0 6px 0 4px; flex-shrink:0; }}
  .layer-top:hover {{ border-color:#4ec9b0; color:#4ec9b0; }}
  .layer-top.on {{ background:#2a4a6a; border-color:#4ec9b0; color:#4ec9b0; }}
  .layer-check {{ color:#4ec9b0; font-size:13px; flex-shrink:0; }}
  .layer-off {{ opacity:0.4; }}
  .layer-off .layer-check {{ visibility:hidden; }}
  #search-box {{ padding:8px 12px; border-bottom:1px solid #333; }}
  #pcb-search-toggle {{ width:100%; text-align:left; background:#1a1a1a;
                        border:1px solid #333; color:#888; padding:4px 8px;
                        font-size:11px; cursor:pointer; border-radius:2px;
                        font-family:inherit; }}
  #pcb-search-toggle:hover {{ color:#4ec9b0; border-color:#4ec9b0; }}
  #search-box.collapsed input {{ display:none; }}
  #search-box input {{ width:100%; background:#0d0d0d; border:1px solid #333;
                       color:#ddd; padding:6px 8px; border-radius:4px;
                       font-size:12px; margin-top:6px; }}
  #canvas-wrap {{ flex:1; position:relative; overflow:hidden; cursor:crosshair;
                  background:#0a0a0a;
                  background-image:radial-gradient(circle, #1a1a1a 1px, transparent 1px);
                  background-size:20px 20px; }}
  #canvas-wrap.grabbing {{ cursor:crosshair; }}
  #pcb-svg {{ position:absolute; transform-origin:0 0; will-change:transform; }}
  /* Pan sırasında hit-testing kapalı (Chromium mousemove başına binlerce elemanı
     hit-test ediyordu). Sınıf gerçek harekette eklenir; hareketsiz tıklama hedefi
     değişmez, pan sonrası tıklama zaten `moved` bayrağıyla yutulur. */
  #canvas-wrap.panning #pcb-svg {{ pointer-events:none; }}
  #pcb-svg .pcb-layer * {{ vector-effect:non-scaling-stroke; }}
  /* LOD: etkileşim sırasında katmanların yerine tek board bitmap'i (svg'nin
     ALTINDA durur → overlay'ler — net-hl, hl-marker, pad etiketi — üstte canlı
     kalır). Katmanlar visibility:hidden (display değil — inline display katman
     aç/kapa durumunu taşıyor, ona dokunmuyoruz). lod-fade: bitmap→SVG dönüşünde
     bitmap 160ms daha görünür kalır, SVG karoları arkada rasterize olur. */
  #lod-canvas {{ position:absolute; left:0; top:0; transform-origin:0 0;
                 display:none; will-change:transform; }}
  #canvas-wrap.lod #lod-canvas,
  #canvas-wrap.lod-fade #lod-canvas {{ display:block; }}
  #canvas-wrap.lod .pcb-layer {{ visibility:hidden; }}
  #toolbar {{ position:absolute; top:10px; right:10px; display:flex; gap:6px;
              z-index:50; }}
  .tool-btn {{ background:rgba(30,30,30,0.92); border:1px solid #444;
               color:#ddd; padding:6px 12px; border-radius:5px; cursor:pointer;
               font-size:12px; }}
  .tool-btn:hover {{ border-color:#4ec9b0; color:#4ec9b0; }}
  .tool-btn.active {{ background:#2a4a6a; border-color:#4ec9b0; color:#4ec9b0; }}
  /* Komponent detayı: sağda yüzen popup yerine sol sidebar'a dock edilmiş,
     katlanabilir (ok) + boyutlandırılabilir (üst tutamaç) panel. */
  #comp-popup {{ display:none; flex-direction:column; flex-shrink:0;
                 background:#161616; border-top:1px solid #333;
                 height:280px; min-height:32px; max-height:65%; overflow:hidden;
                 font-size:12px; }}
  #comp-popup.open {{ display:flex; }}
  #comp-popup.collapsed {{ height:auto !important; }}
  #comp-popup.collapsed .popup-body,
  #comp-popup.collapsed #popup-resize {{ display:none; }}
  #popup-resize {{ height:7px; cursor:ns-resize; background:#222;
                   flex-shrink:0; border-bottom:1px solid #333; }}
  #popup-resize:hover {{ background:#4ec9b0; }}
  .popup-hdr {{ padding:8px 10px; background:#2a4a6a; display:flex;
                align-items:center; gap:6px; flex-shrink:0; }}
  .popup-hdr .desig {{ font-weight:bold; color:#4ec9b0; font-size:14px;
                       font-family:Consolas,monospace; flex:1;
                       overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }}
  .popup-collapse {{ background:none; border:none; color:#bcd; cursor:pointer;
                     font-size:13px; padding:0 2px; line-height:1; }}
  .popup-collapse:hover {{ color:#fff; }}
  .popup-x {{ background:none; border:none; color:#aaa; cursor:pointer;
              font-size:18px; line-height:1; }}
  .popup-x:hover {{ color:#fff; }}
  .popup-body {{ padding:10px 12px; overflow-y:auto; flex:1; }}
  .prow {{ display:flex; padding:3px 0; border-bottom:1px solid #282828; }}
  .prow:last-child {{ border-bottom:none; }}
  .pk {{ color:#888; min-width:95px; }}
  .pv {{ color:#eee; flex:1; word-break:break-word; }}
  #info-bar {{ position:absolute; bottom:10px; left:10px;
               background:rgba(20,20,20,0.85); padding:5px 12px;
               border-radius:4px; font-size:11px; color:#888; z-index:40; }}
  .comp-hover-label {{ position:fixed; background:rgba(0,0,0,0.85);
                       color:#4ec9b0; padding:2px 6px; border-radius:3px;
                       font-size:11px; font-family:Consolas,monospace;
                       pointer-events:none; z-index:200; display:none; }}
  /* Seçili komponent işaretçisi: tek temiz kutu + etiket. Çizgi kalınlığı/
     köşe/yazı JS'te 1/scale ile ölçeklenir (ekranda sabit px) — zoom CSS
     transform ile yapıldığından non-scaling-stroke çalışmaz. */
  #hl-marker {{ pointer-events:none; }}
  #hl-marker .hl-box {{ fill:none; stroke:#00e5ff; stroke-linejoin:round;
    animation:hlpulse 1.1s ease-in-out infinite; }}
  #hl-marker .hl-text {{ fill:#00e5ff; font-family:Consolas,monospace;
    font-weight:bold; paint-order:stroke; stroke:#04222a;
    stroke-linejoin:round; }}
  @keyframes hlpulse {{ 0%,100%{{stroke-opacity:0.5}} 50%{{stroke-opacity:1}} }}
</style>
</head>
<body>
<div id="app">
  <div id="sidebar">
    <button id="sb-toggle" title="Paneli gizle ( B )">◂</button>
    <h2>PCB · {pcb["pcb_name"]}</h2>
    <div class="build-info">{view_w:.0f}×{view_h:.0f}mm</div>
    <div id="search-box" class="collapsed">
      <button id="pcb-search-toggle"><span id="pcb-search-caret">▸</span> Ara</button>
      <input type="text" id="comp-search" placeholder="Komponent ara + Enter... ( / )">
    </div>
    <div id="layer-list"></div>
    <div id="comp-popup">
      <div id="popup-resize" title="Sürükle: yeniden boyutlandır"></div>
      <div class="popup-hdr">
        <button class="popup-collapse" id="pp-collapse" title="Küçült / Büyüt">▾</button>
        <span class="desig" id="pp-desig"></span>
        <button class="popup-x" id="pp-close" title="Kapat">×</button>
      </div>
      <div class="popup-body" id="pp-body"></div>
    </div>
  </div>
  <div id="canvas-wrap">
    <canvas id="lod-canvas"></canvas>
    <svg id="pcb-svg" xmlns="http://www.w3.org/2000/svg"
         width="{view_w}" height="{view_h}"
         viewBox="0 0 {view_w} {view_h}">
      {layers_svg}
    </svg>
    <div id="toolbar">
      <button class="tool-btn" id="zoom-in-btn" title="Yaklaş">+</button>
      <button class="tool-btn" id="zoom-out-btn" title="Uzaklaş">−</button>
      <button class="tool-btn" id="fit-btn">Sığdır</button>
      <button class="tool-btn active" id="lod-toggle"
              title="LOD: gezinirken board bitmap çizilir (Chromium'da akıcılık). Kapatınca her zaman canlı SVG.">LOD</button>
      <button class="tool-btn" id="flip-btn">Üst/Alt</button>
      <button class="tool-btn" id="pin-btn" title="Pad pin no + net adı">Pin</button>
      <button class="tool-btn" id="bg-btn" title="Arka plan rengini değiştir">Zemin</button>
      <button class="tool-btn" id="all-on">Hepsi</button>
      <button class="tool-btn" id="all-off">Temizle</button>
    </div>
    <div class="comp-hover-label" id="hover-label"></div>
    <div id="info-bar">Sürükle: kaydır · Tekerlek: zoom · Komponente tıkla: detay</div>
  </div>
</div>
<script>
const LAYERS = {json.dumps(layer_meta)};
const COMPONENTS = {json.dumps(pcb["components"])};
const COMP_INFO = {json.dumps(comp_info)};
const VIEW_W = {view_w}, VIEW_H = {view_h};
// Tembel katman içerikleri (varsayılan kapalı katmanlar) — ilk gösterimde enjekte.
let LAZY_SVG = {json.dumps(lazy_svgs)};

const wrap = document.getElementById('canvas-wrap');
const svg = document.getElementById('pcb-svg');
const lodCanvas = document.getElementById('lod-canvas');
// Bir katmanın içeriğini (tembelse) DOM'a enjekte et.
function ensureLayerLoaded(id) {{
  const g=document.getElementById('layer-'+id);
  if(g && g.dataset.lazy) {{
    if(LAZY_SVG[id]!==undefined) {{ g.innerHTML=LAZY_SVG[id]; delete LAZY_SVG[id]; }}
    delete g.dataset.lazy;
  }}
}}
// Çapraz-katman işlemler (net highlight, pad etiketleri) gizli katmanları da tarar →
// tetiklenince TÜM tembel katmanları yükle (tek seferlik, doğal etkileşim anında).
function loadAllLazyLayers() {{
  for(const k in LAZY_SVG) ensureLayerLoaded(+k);
}}
let scale = 1, tx = 0, ty = 0;
function applyTransform() {{
  svg.style.transform = `translate(${{tx}}px, ${{ty}}px) scale(${{scale}})`;
  lodCanvas.style.transform = svg.style.transform;   // bitmap svg ile senkron
  updateMarkerMetrics();
  updatePadLabelVis();
}}
let autoFit=true;   // kullanıcı pan/zoom/seçim yapana kadar her yeniden boyutta sığdır
function fitView() {{
  const r = wrap.getBoundingClientRect();
  if(!r.width || !r.height) return;
  const pad = 40;
  scale = Math.min((r.width-pad*2)/VIEW_W, (r.height-pad*2)/VIEW_H);
  tx = (r.width - VIEW_W*scale)/2;
  ty = (r.height - VIEW_H*scale)/2;
  applyTransform();
}}
let dragging=false, lastX=0, lastY=0, moved=false;
wrap.addEventListener('mousedown', e => {{
  dragging=true; moved=false; lastX=e.clientX; lastY=e.clientY;
  wrap.classList.add('grabbing');
  svg.style.transition='none';
}});
window.addEventListener('mousemove', e => {{
  if (!dragging) return;
  const dx=e.clientX-lastX, dy=e.clientY-lastY;
  if (Math.abs(dx)>2||Math.abs(dy)>2) {{
    // Gerçek pan başladı: SVG hit-testing'i kapat (bkz. #canvas-wrap.panning CSS'i)
    // + etkileşim boyunca bitmap moduna geç (akıcı sürükleme).
    if (!moved) {{ wrap.classList.add('panning'); panInteract=true; pcbLodUpdate(); }}
    moved=true; autoFit=false;
  }}
  tx+=dx; ty+=dy; lastX=e.clientX; lastY=e.clientY; applyTransform();
}});
window.addEventListener('mouseup', () => {{
  dragging=false; wrap.classList.remove('grabbing', 'panning');
  if (panInteract) {{ panInteract=false; pcbLodUpdate(); }}
}});
// Tekerlek zoom'u rAF ile birleştirilir: çarpanlar wheelF'te birikir, kare başına
// TEK transform uygulanır (Chromium her scale değişiminde tüm görünür karoları
// yeniden rasterize ettiğinden event başına uygulamak takılma yaratıyordu).
let wheelF=1, wheelPend=false, wheelMx=0, wheelMy=0;
wrap.addEventListener('wheel', e => {{
  e.preventDefault();
  pcbLodTouchWheel();   // zoom serisi boyunca bitmap modu (akıcı tekerlek)
  autoFit=false;
  svg.style.transition='none';
  const r=wrap.getBoundingClientRect();
  wheelMx=e.clientX-r.left; wheelMy=e.clientY-r.top;
  wheelF*=e.deltaY<0?1.15:1/1.15;
  if (wheelPend) return;
  wheelPend=true;
  requestAnimationFrame(() => {{
    wheelPend=false;
    const ns=Math.max(0.05,Math.min(80,scale*wheelF)); wheelF=1;
    tx=wheelMx-(wheelMx-tx)*(ns/scale); ty=wheelMy-(wheelMy-ty)*(ns/scale);
    scale=ns; applyTransform();
  }});
}}, {{ passive:false }});

// === LOD: etkileşim sırasında board bitmap'i ==============================
// Şematikteki desenin PCB uyarlaması: pan sürüklemesi / tekerlek zoom serisi
// BOYUNCA görünür katmanların tek bitmap'i gösterilir (compositor'da ucuz),
// hareket durunca canlı SVG'ye dönülür → tıklama/highlight/pad etiketi aynen.
// Şematikten farklar: (1) dinlenmede HEP canlı SVG (PCB'de uzak zoom'da da
// komponent/net tıklaması yaygın); (2) bitmap TEK parça (board tek SVG);
// (3) katman aç/kapa/sıralama bitmap'i eskitir (lodGen) → sakin anda yeniden
// üretilir, hazır olana dek canlı SVG. Net highlight karartması bitmap'e
// yansımaz (harekette karartmasız görünür, durunca geri gelir) — bilinen kısıt.
let pcbLodEnabled = lsGet().pcbLod !== false;
let lodReady=false, lodBuilding=false, lodK=0, lodActive=false, lodGen=0, lodBuiltGen=-1;
let panInteract=false, wheelInteract=false, wheelIdleT=null, lodFadeT=null;
function pcbLodUpdate() {{
  // Bitmap, kendi çözünürlüğünün ~4 katına dek kabul edilebilir (harekette);
  // daha yakın zoom'da canlı SVG (görünür alan küçük → raster zaten ucuz).
  const maxS = lodK ? (lodK * 4) / (window.devicePixelRatio || 1) : 0;
  const want = pcbLodEnabled && lodReady && lodBuiltGen === lodGen &&
               (panInteract || wheelInteract) && scale <= maxS;
  if (want === lodActive) return;
  lodActive = want;
  if (want) {{
    clearTimeout(lodFadeT); wrap.classList.remove('lod-fade');
    wrap.classList.add('lod');
  }} else {{
    // Bitmap'i hemen söndürme: 160ms daha görünür kalsın → SVG karoları
    // arkada rasterize edilir, boş karo/flash görünmez (bkz. .lod-fade CSS).
    wrap.classList.remove('lod');
    wrap.classList.add('lod-fade');
    clearTimeout(lodFadeT);
    lodFadeT = setTimeout(() => wrap.classList.remove('lod-fade'), 160);
  }}
}}
function pcbLodTouchWheel() {{
  wheelInteract = true; pcbLodUpdate();
  clearTimeout(wheelIdleT);
  wheelIdleT = setTimeout(() => {{ wheelInteract = false; pcbLodUpdate(); }}, 180);
}}
// Görünür katmanlardan bitmap üret (async; katman değişince yeniden).
function pcbLodBuild() {{
  if (lodBuilding || !pcbLodEnabled) return;
  lodBuilding = true;
  const gen = lodGen;
  const parts = [];
  for (const ch of svg.children) {{
    if (ch.classList && ch.classList.contains('pcb-layer') && ch.style.display !== 'none')
      parts.push(new XMLSerializer().serializeToString(ch));
  }}
  const k = 2600 / Math.max(VIEW_W, VIEW_H);
  const W = Math.max(1, Math.round(VIEW_W * k)), H = Math.max(1, Math.round(VIEW_H * k));
  const src = '<svg xmlns="http://www.w3.org/2000/svg" width="' + W + '" height="' + H
            + '" viewBox="0 0 ' + VIEW_W + ' ' + VIEW_H + '">' + parts.join('') + '</svg>';
  const img = new Image();
  let url = '', tried = false;
  const done = ok => {{
    if (url) {{ URL.revokeObjectURL(url); url = ''; }}
    if (ok) {{
      try {{
        lodCanvas.width = W; lodCanvas.height = H;
        lodCanvas.getContext('2d').drawImage(img, 0, 0, W, H);
        lodCanvas.style.width = VIEW_W + 'px'; lodCanvas.style.height = VIEW_H + 'px';
        lodK = k; lodReady = true; lodBuiltGen = gen;
      }} catch (e) {{ /* bitmap üretilemedi → canlı SVG'de kal */ }}
    }}
    lodBuilding = false;
    if (lodBuiltGen !== lodGen) setTimeout(pcbLodBuild, 100);  // üretim sırasında eskidi
    pcbLodUpdate();
  }};
  img.onload = () => done(true);
  img.onerror = () => {{
    // blob: bazı ortamlarda engellenebilir → data: URI ile bir kez daha dene
    if (tried) {{ done(false); return; }}
    tried = true;
    if (url) {{ URL.revokeObjectURL(url); url = ''; }}
    img.src = 'data:image/svg+xml;charset=utf-8,' + encodeURIComponent(src);
  }};
  url = URL.createObjectURL(new Blob([src], {{ type: 'image/svg+xml;charset=utf-8' }}));
  img.src = url;
}}
// Katman görünürlüğü/sırası değişti → bitmap'i eskit, sakin anda yeniden üret.
function pcbLodInvalidate() {{
  lodGen++; pcbLodUpdate();
  clearTimeout(pcbLodInvalidate._t);
  pcbLodInvalidate._t = setTimeout(pcbLodBuild, 500);
}}
// Toolbar LOD toggle'ı (tercih localStorage'da).
const pcbLodBtn = document.getElementById('lod-toggle');
function setPcbLod(on) {{
  pcbLodEnabled = on;
  pcbLodBtn.classList.toggle('active', on);
  if (on && lodBuiltGen !== lodGen) pcbLodBuild();
  pcbLodUpdate();
  lsSet({{ pcbLod: on }});
}}
pcbLodBtn.addEventListener('click', () => setPcbLod(!pcbLodEnabled));
if (!pcbLodEnabled) pcbLodBtn.classList.remove('active');
(window.requestIdleCallback || (f => setTimeout(f, 400)))(pcbLodBuild);

const layerList = document.getElementById('layer-list');
let raisedId=null;   // en üste getirilen katman (null = orijinal sıra)
// Katmanların orijinal (id) sırasına dön — overlay'lerin (pad etiketi/highlight) ALTINA,
// diğer katmanların yerli yerine yerleştirir.
function restoreLayerOrder() {{
  let anchor=null;
  for(const ch of svg.children) {{ if(!ch.classList.contains('pcb-layer')) {{ anchor=ch; break; }} }}
  LAYERS.forEach(l => {{
    const gg=document.getElementById('layer-'+l.id);
    if(gg) {{ if(anchor) svg.insertBefore(gg, anchor); else svg.appendChild(gg); }}
  }});
}}
// Seçilen katmanı diğer KATMANLARIN üstüne taşı (overlay'lerin altında kalır).
// Aynı katmana tekrar basınca orijinal sıraya döner (toggle).
function bringLayerToTop(id) {{
  pcbLodInvalidate();                           // katman sırası değişiyor
  if(raisedId===id) {{ raisedId=null; restoreLayerOrder(); renderLayerList(); return; }}
  raisedId=id;
  ensureLayerLoaded(id);                      // tembelse içeriği yükle
  restoreLayerOrder();                       // önce normalleştir
  const gg=document.getElementById('layer-'+id);
  let anchor=null;
  for(const ch of svg.children) {{ if(!ch.classList.contains('pcb-layer')) {{ anchor=ch; break; }} }}
  if(anchor) svg.insertBefore(gg, anchor); else svg.appendChild(gg);
  gg.style.display='block';                  // üste getirilen katman görünür olsun
  renderLayerList();
}}
function renderLayerList() {{
  layerList.innerHTML='';
  LAYERS.forEach(l => {{
    const g=document.getElementById('layer-'+l.id);
    const on=g && g.style.display!=='none';
    const item=document.createElement('div');
    item.className='layer-item'+(on?'':' layer-off');
    item.innerHTML=`<span class="layer-swatch" style="background:${{l.color}}"></span>`
      +`<span class="layer-name">${{l.display}}</span>`
      +`<button class="layer-top${{raisedId===l.id?' on':''}}" `
      +`title="En üste getir (tekrar bas: orijinal sıra)">↑</button>`
      +`<span class="layer-check">✓</span>`;
    item.onclick=() => {{
      const gg=document.getElementById('layer-'+l.id);
      const vis=gg.style.display!=='none';
      if(!vis) ensureLayerLoaded(l.id);        // açarken tembel içeriği yükle
      gg.style.display=vis?'none':'block';
      item.classList.toggle('layer-off',vis);
      pcbLodInvalidate();
    }};
    item.querySelector('.layer-top').onclick=(ev) => {{
      ev.stopPropagation(); bringLayerToTop(l.id);
    }};
    layerList.appendChild(item);
  }});
}}
document.getElementById('all-on').onclick=() => {{
  loadAllLazyLayers();
  LAYERS.forEach(l => document.getElementById('layer-'+l.id).style.display='block');
  renderLayerList();
  pcbLodInvalidate();
}};
document.getElementById('all-off').onclick=() => {{
  LAYERS.forEach(l => document.getElementById('layer-'+l.id).style.display='none');
  renderLayerList();
  pcbLodInvalidate();
}};
document.getElementById('fit-btn').onclick=fitView;
// Görünüm merkezinde yakınlaş/uzaklaş (toolbar +/− butonları)
function zoomBy(f) {{
  autoFit=false; svg.style.transition='none';
  const r=wrap.getBoundingClientRect();
  const mx=r.width/2, my=r.height/2;
  const ns=Math.max(0.05,Math.min(80,scale*f));
  tx=mx-(mx-tx)*(ns/scale); ty=my-(my-ty)*(ns/scale); scale=ns; applyTransform();
}}
document.getElementById('zoom-in-btn').onclick=() => zoomBy(1.35);
document.getElementById('zoom-out-btn').onclick=() => zoomBy(1/1.35);

// === UI tercihlerini hatırla (file:// altında kısıtlıysa sessizce atla) ===
const LS_KEY='schviz-ui';
function lsGet() {{ try {{ return JSON.parse(localStorage.getItem(LS_KEY)||'{{}}'); }} catch(e) {{ return {{}}; }} }}
function lsSet(patch) {{ try {{
  localStorage.setItem(LS_KEY, JSON.stringify(Object.assign(lsGet(), patch)));
}} catch(e) {{}} }}

// === Katlanabilir sol panel (şematik tarafıyla aynı davranış) ===
const sbEl=document.getElementById('sidebar');
const sbToggle=document.getElementById('sb-toggle');
function setSbOpen(open) {{
  sbEl.classList.toggle('collapsed', !open);
  sbToggle.textContent = open ? '◂' : '▸';
  sbToggle.title = (open ? 'Paneli gizle' : 'Paneli göster') + ' ( B )';
  lsSet({{ pcbSidebar: open }});
}}
sbToggle.addEventListener('click', () => setSbOpen(sbEl.classList.contains('collapsed')));

// === Katlanabilir arama (varsayılan kapalı; / açar, Esc kapatır) ===
const pcbSearchBox=document.getElementById('search-box');
const pcbSearchInput=document.getElementById('comp-search');
function setPcbSearchOpen(open) {{
  if (open) setSbOpen(true);
  pcbSearchBox.classList.toggle('collapsed', !open);
  document.getElementById('pcb-search-caret').textContent = open ? '▾' : '▸';
  if (open) pcbSearchInput.focus(); else pcbSearchInput.blur();
}}
document.getElementById('pcb-search-toggle').addEventListener('click', () =>
  setPcbSearchOpen(pcbSearchBox.classList.contains('collapsed')));

(function restorePcbUi() {{
  if (lsGet().pcbSidebar === false) setSbOpen(false);
}})();
// === Arka plan rengi: üst toolbar'dan döngüyle değiştirilir (siyah↔gri↔açık).
//     Nokta ızgarası rengi zemine göre kontrastlı seçilir. ===
const BG_PRESETS = [
  {{name:'Siyah',  bg:'#0a0a0a', dot:'#1a1a1a'}},
  {{name:'Koyu gri', bg:'#3a3d42', dot:'#4a4e54'}},
  {{name:'Gri',    bg:'#808890', dot:'#6f767e'}},
  {{name:'Açık',   bg:'#c8ccd2', dot:'#b3b7be'}},
];
let bgIndex=0;
function applyBg() {{
  const p = BG_PRESETS[bgIndex];
  wrap.style.background = p.bg;
  wrap.style.backgroundImage =
    'radial-gradient(circle, '+p.dot+' 1px, transparent 1px)';
  const btn = document.getElementById('bg-btn');
  if(btn) btn.title = 'Arka plan: '+p.name+' (tıkla: değiştir)';
}}
document.getElementById('bg-btn').onclick=() => {{
  bgIndex = (bgIndex+1) % BG_PRESETS.length; applyBg();
}};
applyBg();
// === Pad etiketleri: pin no + net adı (Altium gibi). Dünya koordinatında çizilir
//     (zoom ile ölçeklenir); yeterince yakınlaşınca otomatik görünür, uzakta gizlenir.
//     Varsayılan AÇIK. ===
let padLabelsGroup=null, padLabelsOn=true;
const PAD_LABEL_MIN_SCALE = 14;   // bu zoom'un altında etiketler gizlenir (okunmaz)
function buildPadLabels() {{
  loadAllLazyLayers();             // pad'ler tüm katmanlarda → tembel olanları da yükle
  const NS='http://www.w3.org/2000/svg';
  const g=document.createElementNS(NS,'g');
  g.setAttribute('id','pad-labels'); g.style.pointerEvents='none';
  const seen=new Set();
  const vr=wrap.getBoundingClientRect();
  svg.querySelectorAll('[data-primitive="pad"]').forEach(el => {{
    const comp=el.getAttribute('data-component')||'';
    const num=el.getAttribute('data-pad-number')||el.getAttribute('data-pad-designator')||'';
    if(!num) return;
    const key=comp+'/'+num;
    if(seen.has(key)) return;
    // getCTM yerine ekran kutusunu kök uzaya çevir (Firefox getCTM farkına bağışık).
    // Gizli katmandaki eleman 0-boyut döner → seen'e EKLEMEDEN atla ki pad'in
    // görünür bir elemanı sonra etiketlensin.
    const er=el.getBoundingClientRect();
    if(!er.width || !er.height) return;
    seen.add(key);
    const cx=(er.left+er.width/2 - vr.left - tx)/scale;   // kök koordinat merkezi
    const cy=(er.top +er.height/2 - vr.top  - ty)/scale;
    const padW=er.width/scale, padH=er.height/scale;
    const small=Math.min(padW,padH);
    if(small<=0) return;
    const net=el.getAttribute('data-net')||'';
    // Pad'e sığacak, kompakt yazı (zoom-threshold zaten okunurluğu sağlar)
    const fs=Math.min(small*0.40, Math.max(padW,padH)*0.24);
    const mk=(txt,dy,size,bold)=>{{
      const t=document.createElementNS(NS,'text');
      t.setAttribute('x',cx); t.setAttribute('y',cy+dy);
      t.setAttribute('text-anchor','middle');
      t.setAttribute('dominant-baseline','central');
      t.setAttribute('font-size',size);
      t.setAttribute('font-family','Consolas,monospace');
      if(bold) t.setAttribute('font-weight','bold');
      t.setAttribute('fill','#fff');
      t.setAttribute('stroke','#000'); t.setAttribute('stroke-width',size*0.16);
      t.setAttribute('paint-order','stroke'); t.setAttribute('stroke-linejoin','round');
      t.textContent=txt; return t;
    }};
    if(net) {{
      g.appendChild(mk(num, -fs*0.56, fs*0.92, true));
      g.appendChild(mk(net,  fs*0.58, fs*0.78, false));
    }} else {{
      g.appendChild(mk(num, 0, fs, true));
    }}
  }});
  svg.appendChild(g);
  return g;
}}
// Zoom'a göre görünürlük: açık VE yeterince yakınsa kur+göster, değilse gizle
function updatePadLabelVis() {{
  if(!padLabelsOn || scale < PAD_LABEL_MIN_SCALE) {{
    if(padLabelsGroup) padLabelsGroup.style.display='none';
    return;
  }}
  if(!padLabelsGroup) {{
    padLabelsGroup=buildPadLabels();
  }}
  padLabelsGroup.style.display='';
  // Net highlight aktifse iz/klonlar pad etiketlerinin ÜSTÜNDE kalsın
  if(netHlGroup) svg.appendChild(netHlGroup);
}}
function setPadLabels(on) {{
  padLabelsOn=on;
  updatePadLabelVis();
  document.getElementById('pin-btn').classList.toggle('active', on);
}}
document.getElementById('pin-btn').onclick=() => setPadLabels(!padLabelsOn);
let showingTop=true;
document.getElementById('flip-btn').onclick=() => {{
  showingTop=!showingTop;
  LAYERS.forEach(l => {{
    const g=document.getElementById('layer-'+l.id);
    const n=l.name.toUpperCase();
    if(n.startsWith('TOP')) g.style.display=showingTop?'block':'none';
    else if(n.startsWith('BOTTOM')) g.style.display=showingTop?'none':'block';
  }});
  renderLayerList();
  pcbLodInvalidate();
}};

const popup=document.getElementById('comp-popup');
function showComp(desig) {{
  setSbOpen(true);  // popup sidebar'a dock'lu — panel kapalıysa aç
  const pc=COMPONENTS[desig]||{{}};
  const info=COMP_INFO[desig]||{{}};
  document.getElementById('pp-desig').textContent=desig;
  const row=(k,v)=>v?`<div class="prow"><span class="pk">${{k}}</span><span class="pv">${{v}}</span></div>`:'';
  let html='';
  html+=row('Değer',info.value);
  html+=row('Açıklama',info.description);
  html+=row('Şema Sayfası',info.sheet);
  html+=row('Footprint',pc.footprint);
  html+=row('Katman',pc.layer);
  html+=row('Konum (mm)',(pc.abs_x_mm!==undefined)?`X=${{pc.abs_x_mm}} Y=${{pc.abs_y_mm}}`:'');
  if(pc.rotation) html+=row('Dönüş',pc.rotation+'°');
  if(!html) html='<div style="color:#888">Bilgi yok</div>';
  document.getElementById('pp-body').innerHTML=html;
  if (typeof clearNetHighlight === 'function') clearNetHighlight();
  popup.classList.add('open');
  popup.classList.remove('collapsed');
  document.getElementById('pp-collapse').textContent='▾';
  highlightComp(desig);
}}
document.getElementById('pp-close').onclick=() => {{
  popup.classList.remove('open'); clearHighlight();
}};
// Katla / aç (sol üstteki ok) — simge durumuna küçült
document.getElementById('pp-collapse').onclick=() => {{
  const c=popup.classList.toggle('collapsed');
  document.getElementById('pp-collapse').textContent=c?'▸':'▾';
}};
// Üst tutamaçtan dikey boyutlandırma
(function(){{
  const handle=document.getElementById('popup-resize');
  let rz=false, sy=0, sh=0;
  handle.addEventListener('mousedown', e => {{
    rz=true; sy=e.clientY; sh=popup.offsetHeight;
    document.body.style.userSelect='none'; e.preventDefault();
  }});
  window.addEventListener('mousemove', e => {{
    if(!rz) return;
    let h=sh+(sy-e.clientY);                 // yukarı sürükle → büyüt
    h=Math.max(70, Math.min(window.innerHeight*0.8, h));
    popup.style.height=h+'px';
  }});
  window.addEventListener('mouseup', () => {{ rz=false; document.body.style.userSelect=''; }});
}})();
const SVGNS='http://www.w3.org/2000/svg';
let highlightMarker=null;
let pendingComp=null;   // PCB paneli gizliyken seçilen komponent (görünür olunca uygulanır)
function clearHighlight() {{
  if(highlightMarker) {{ highlightMarker.remove(); highlightMarker=null; }}
}}
// Komponentin tüm primitive'lerinin birleşik sınır kutusunu KÖK user-space'te
// (mm) hesapla → tek temiz kutu + etiket çiz → komponente odaklan.
// Not: getBBox() öğenin kendi yerel uzayını verir; getCTM ile kök uzaya çevrilir.
function highlightComp(desig) {{
  loadAllLazyLayers();             // komponent alt katmanda olabilir → tembelleri yükle
  clearHighlight();
  // PCB paneli gizli/0-boyut ise (ör. "Şematik" tek-panel modu) getBBox/getCTM
  // çalışmaz → marker oluşmaz, board kayardı. Komponenti beklet; panel görünür
  // olunca ResizeObserver uygular.
  const vr=wrap.getBoundingClientRect();
  if(!vr.width || !vr.height) {{ pendingComp=desig; return; }}
  const els=[...document.querySelectorAll(`[data-component="${{desig}}"]`)];
  if(!els.length) return;
  let x0=1e9,y0=1e9,x1=-1e9,y1=-1e9;
  els.forEach(el => {{
    // getCTM yerine ekran kutusunu kendi tx/ty/scale'imizle kök uzaya çevir
    // (Firefox getCTM farkına bağışık). transform-origin:0 0 olduğundan:
    //   ekran_x_rel = tx + uzay_x*scale  →  uzay_x = (ekran_x_rel - tx)/scale
    const er=el.getBoundingClientRect();
    if(!er.width && !er.height) return;   // gizli katman → atla
    const ax0=(er.left -vr.left -tx)/scale, ay0=(er.top    -vr.top -ty)/scale;
    const ax1=(er.right-vr.left -tx)/scale, ay1=(er.bottom -vr.top -ty)/scale;
    x0=Math.min(x0,ax0); y0=Math.min(y0,ay0); x1=Math.max(x1,ax1); y1=Math.max(y1,ay1);
  }});
  if(x0>x1) return;
  // Sıkı kutu: küçük oransal + küçük sabit pay
  const pad=Math.max(x1-x0,y1-y0)*0.06+0.12;
  const bx=x0-pad, by=y0-pad, bw=(x1-x0)+pad*2, bh=(y1-y0)+pad*2;

  const g=document.createElementNS(SVGNS,'g');
  g.setAttribute('id','hl-marker');
  const r=document.createElementNS(SVGNS,'rect');
  r.setAttribute('x',bx); r.setAttribute('y',by);
  r.setAttribute('width',bw); r.setAttribute('height',bh);
  r.setAttribute('class','hl-box');
  g.appendChild(r);
  const t=document.createElementNS(SVGNS,'text');
  t.setAttribute('class','hl-text'); t.textContent=desig;
  g.appendChild(t);
  g.dataset.bx=bx; g.dataset.by=by;
  svg.appendChild(g);
  highlightMarker=g;
  updateMarkerMetrics();        // stroke/rx/font ekran-sabit
  focusBox(bx,by,bw,bh);
}}
// Zoom CSS transform ile yapıldığından non-scaling-stroke çalışmaz; stroke/rx/
// font'u 1/scale ile ölçekleyerek ekran-pikselinde sabit tutuyoruz.
function updateMarkerMetrics() {{
  if(!highlightMarker) return;
  const k=1/scale;
  const box=highlightMarker.querySelector('.hl-box');
  if(box) {{ box.setAttribute('stroke-width',1.6*k); box.setAttribute('rx',2.5*k); }}
  const t=highlightMarker.querySelector('.hl-text');
  if(t) {{
    const fs=12*k;
    t.setAttribute('font-size',fs);
    t.setAttribute('stroke-width',0.18*fs);
    t.setAttribute('x',+highlightMarker.dataset.bx);
    t.setAttribute('y',(+highlightMarker.dataset.by)-3*k);
  }}
}}
// Komponenti görüş alanının ortasına getir ve rahat görünecek kadar yakınlaş
// (Altium viewer gibi). Sadece gerekiyorsa yakınlaşır, kullanıcı daha
// yakındaysa o zoom'u korur.
function focusBox(bx,by,bw,bh) {{
  const r=wrap.getBoundingClientRect();
  if(!r.width||!r.height) return;
  autoFit=false;
  const cx=bx+bw/2, cy=by+bh/2;
  // Komponent şu an ekranda ne kadar görünüyor?
  const onScreen=Math.max(bw,bh)*scale;
  const minSide=Math.min(r.width,r.height);
  let ns=scale;
  // Yalnızca çok küçük görünüyorsa hafifçe yakınlaş (bağlamı koru, aşırıya kaçma)
  if(onScreen < minSide*0.05) {{
    ns=(minSide*0.14)/Math.max(bw,bh,0.1);
    ns=Math.max(scale, Math.min(ns, 12));
  }}
  scale=ns;
  tx=r.width/2 - cx*scale;
  ty=r.height/2 - cy*scale;
  svg.style.transition='transform 0.35s ease';
  applyTransform();
  clearTimeout(focusBox._t);
  focusBox._t=setTimeout(()=>{{ svg.style.transition='none'; }}, 400);
}}
svg.addEventListener('click', e => {{
  if(moved) return;
  let el=e.target, desig=null;
  while(el && el!==svg) {{
    if(el.getAttribute && el.getAttribute('data-component')) {{
      desig=el.getAttribute('data-component'); break;
    }}
    el=el.parentNode;
  }}
  if(desig) {{ showComp(desig); crossProbeOut(desig); }}
  // Boş alana tıklama (komponent yok, bakır/net yok) = net highlight'ı iptal
  else if(!netAt(e.target)) clearNetHighlight();
}});
// SVG'nin dışındaki zemine tıklama da highlight'ı temizler
wrap.addEventListener('click', e => {{
  if(!moved && e.target===wrap) clearNetHighlight();
}});

// === Bakır yol / net highlight: çift tıkla → net'i TÜM katmanlarda göster,
//     gerisini karart. data-net render edilmiş bakır elemanlarda mevcut. ===
const INFO_DEFAULT = 'Sürükle: kaydır · Tekerlek: zoom · Tıkla: komponent · Çift tıkla: bakır yol';
const infoBar = document.getElementById('info-bar');
infoBar.textContent = INFO_DEFAULT;
function netAt(target) {{
  let el = target;
  while (el && el !== svg) {{
    if (el.getAttribute && el.getAttribute('data-net')) return el.getAttribute('data-net');
    el = el.parentNode;
  }}
  return null;
}}
let netHlGroup = null, netDimmed = false;
function clearNetHighlight() {{
  if (netHlGroup) {{ netHlGroup.remove(); netHlGroup = null; infoBar.textContent = INFO_DEFAULT; }}
  if (netDimmed) {{
    document.querySelectorAll('#pcb-svg [id^="layer-"]').forEach(g => {{ g.style.filter = ''; }});
    netDimmed = false;
  }}
}}
function highlightNet(netName) {{
  loadAllLazyLayers();              // net tüm katmanlarda → gizli/tembel olanları da yükle
  clearNetHighlight();
  clearHighlight();                 // komponent kutusu varsa temizle
  autoFit=false;
  popup.classList.remove('open');
  const els = [...svg.querySelectorAll('[data-net]')]
                .filter(e => e.getAttribute('data-net') === netName);
  if (!els.length) return;
  const NS = 'http://www.w3.org/2000/svg';
  // 1) Karartma: tüm katmanları grileştir + karart (Altium gibi) → net renkleri
  //    (kırmızı/mavi/...) öne çıksın. Net klonları filtresiz, tam renkte üstte.
  document.querySelectorAll('#pcb-svg [id^="layer-"]').forEach(g => {{
    g.style.filter = 'grayscale(1) brightness(0.58)';
  }});
  netDimmed = true;
  // 2) Net elemanlarını KENDİ katman renkleriyle klonla (Top kırmızı, Bottom mavi,
  //    plane'ler yeşil...) → en üstte tam parlaklıkta göster.
  const g = document.createElementNS(NS, 'g');
  g.setAttribute('id', 'net-hl'); g.style.pointerEvents = 'none';
  els.forEach(el => {{
    // Katman gruplarında transform yok; eleman kendi transform'unu (pad rotate vb.)
    // taşıyor → cloneNode bunu korur, #net-hl kök altında olduğundan konum aynı kalır.
    // getCTM KULLANMA: Firefox getCTM'e SVG'nin CSS zoom'unu da katıp çift-transform
    // yapıyordu (klonlar ekran dışına düşüyordu).
    const c = el.cloneNode(true);
    const cs = getComputedStyle(el);
    if (cs.fill && cs.fill !== 'none')   c.setAttribute('fill', boostColor(cs.fill));
    if (cs.stroke && cs.stroke !== 'none') c.setAttribute('stroke', boostColor(cs.stroke));
    c.style.vectorEffect = 'non-scaling-stroke';
    c.removeAttribute('data-net');
    g.appendChild(c);
  }});
  svg.appendChild(g);
  netHlGroup = g;
  infoBar.textContent = `Net: ${{netName}} · ${{els.length}} bakır eleman · Esc temizler`;
}}
// Katman rengini biraz parlatıp doygunlaştır ki gri zeminde net görünsün
function boostColor(col) {{
  const m = col.match(/rgba?\\(([^)]+)\\)/);
  if (!m) return col;
  let [r, g, b] = m[1].split(',').map(s => parseFloat(s));
  const lift = v => Math.round(v + (255 - v) * 0.42);
  return `rgb(${{lift(r)}},${{lift(g)}},${{lift(b)}})`;
}}
svg.addEventListener('dblclick', e => {{
  const net = netAt(e.target);
  if (net) {{ e.preventDefault(); highlightNet(net); }}
}});
const hoverLabel=document.getElementById('hover-label');
svg.addEventListener('mousemove', e => {{
  let el=e.target, desig=null, net=null;
  while(el && el!==svg) {{
    if(el.getAttribute) {{
      if(!desig && el.getAttribute('data-component')) desig=el.getAttribute('data-component');
      if(!net && el.getAttribute('data-net')) net=el.getAttribute('data-net');
    }}
    el=el.parentNode;
  }}
  const label = desig || (net ? ('⚡ ' + net) : null);
  if(label) {{
    hoverLabel.textContent=label;
    hoverLabel.style.display='block';
    hoverLabel.style.left=(e.clientX+12)+'px';
    hoverLabel.style.top=(e.clientY+12)+'px';
  }} else hoverLabel.style.display='none';
}});
// Kanal-farkındalıklı çözümleme (v2.9.33): hiyerarşik Repeat projelerinde
// şematikteki MANTIKSAL designator (R103) board'da kanal-sonekli FİZİKSEL
// kopyalara açılır (R103_diffI2C_1..3). Tam eşleşme yoksa 'AD_' önekli
// kopyalar bulunur (arama + cross-probe bunu kullanır).
function channelCopies(base) {{
  const q = String(base||'').toUpperCase() + '_';
  return Object.keys(COMPONENTS).filter(d => d.toUpperCase().startsWith(q)).sort();
}}
let searchCycleKey=null, searchCycleIdx=0;
document.getElementById('comp-search').addEventListener('keydown', e => {{
  if(e.key!=='Enter') return;
  const q=e.target.value.trim().toUpperCase();
  if(!q) return;
  const desig=Object.keys(COMPONENTS).find(d => d.toUpperCase()===q);
  if(desig) {{ showComp(desig); crossProbeOut(desig); return; }}
  // Mantıksal ad → kanal kopyaları; Enter'a her basışta sıradaki kopyaya geçilir
  const copies = channelCopies(q);
  if(copies.length) {{
    if(searchCycleKey !== q) {{ searchCycleKey = q; searchCycleIdx = 0; }}
    const d = copies[searchCycleIdx % copies.length];
    searchCycleIdx++;
    showComp(d); crossProbeOut(d);
    return;
  }}
  alert(q+' bulunamadı');
}});
// Esc: net / komponent highlight'ı temizle, detay panelini kapat
window.addEventListener('keydown', e => {{
  if (document.activeElement === pcbSearchInput) {{
    if (e.key === 'Escape') setPcbSearchOpen(false);
    return;
  }}
  if (e.key === 'Escape') {{
    clearNetHighlight();
    clearHighlight();
    popup.classList.remove('open');
  }}
  else if (e.key === 'b' || e.key === 'B')
    setSbOpen(sbEl.classList.contains('collapsed'));
  else if (e.key === '/') {{ e.preventDefault(); setPcbSearchOpen(true); }}
}});

// === Cross-probe köprüsü (birleşik görünüm için) ===
const IN_FRAME = window.parent && window.parent !== window;
function crossProbeOut(designator) {{
  if (IN_FRAME) {{
    window.parent.postMessage({{type:'xprobe', source:'pcb', designator:designator}}, '*');
  }}
}}
window.addEventListener('message', ev => {{
  const d = ev.data;
  if (!d || d.type !== 'xprobe' || d.source === 'pcb') return;
  if (COMPONENTS[d.designator]) {{ showComp(d.designator); return; }}
  // Şematikten MANTIKSAL ad gelmiş olabilir (hiyerarşik kanal: R103) →
  // ilk fiziksel kopyayı göster (R103_diffI2C_1)
  const copies = channelCopies(d.designator);
  if (copies.length) showComp(copies[0]);
}});

renderLayerList();
document.getElementById('pin-btn').classList.add('active');  // pad etiketleri varsayılan açık
// İlk sığdırma: iframe/flex layout boyutu tam oturduktan sonra çalışsın.
// (srcdoc iframe ilk render'da getBoundingClientRect 0 verebilir → board
//  ekran dışına düşerdi.) rAF + küçük gecikme ile garanti altına al.
function fitWhenReady(tries) {{
  const r = wrap.getBoundingClientRect();
  if (r.width > 0 && r.height > 0) {{ fitView(); return; }}
  if (tries > 0) requestAnimationFrame(() => fitWhenReady(tries - 1));
}}
requestAnimationFrame(() => fitWhenReady(30));
// Panel gizliyken ("Şematik" modu) cross-probe geldiyse marker oluşamıyordu;
// panel görünür olunca (Böl/PCB'ye geçince) ilk sığdırmayı yap ve bekleyen
// komponenti vurgula.
new ResizeObserver(() => {{
  const r = wrap.getBoundingClientRect();
  if (r.width <= 0 || r.height <= 0) return;
  if (autoFit) fitView();
  if (pendingComp) {{ const d = pendingComp; pendingComp = null; highlightComp(d); }}
}}).observe(wrap);
// Pencere/pane yeniden boyutlandığında otomatik sığdırma yok (kullanıcı
// zoom'unu bozmasın diye), ama ilk yükte garanti.
</script>
</body></html>"""


def generate_json(
    project_path: str,
    output_path: str,
    log: Callable[[str], None] = print,
):
    """@brief AI'ya/LLM'e verilebilir kompakt JSON çıktı üret.
    
    SVG verisi DAHİL DEĞİL — sadece yapısal bilgi:
    sayfa adları, her sayfadaki komponentler, her netin sayfa-bazlı dağılımı.
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param output_path Çıktı dosyası yolu
    @param log Log mesajı callback'i (str alır)
    """
    data = _collect_data(project_path, log)

    def classify_net(name):
        """@brief classify_net()
        
        @param name Ad
        @return Üretilen sonuç.
        """
        n = name.upper()
        if re.match(r"^A?GND", n) or n in ("VSS", "AGND", "DGND") or n.endswith("GND"):
            return "ground"
        if re.match(r"^[+\-]?\d*\.?\d+V", n) or re.match(r"^V(CC|DD|IN|OUT|BAT|REF)", n) or n in ("VCC", "VDD"):
            return "power"
        return "signal"

    # Netlist'ten komponent.pin → net adı eşlemesi kur
    # (AI "U5 pin 11 hangi net'e bağlı?" diye sorabilsin)
    netlist_early = data.get("netlist") or {}
    pin_to_net = {}   # "U5.11" -> "ADC1_CS"
    comp_pins = {}    # "U5" -> {"11": {"net": "ADC1_CS", "pin_name": "CS"}}
    for nn in netlist_early.get("nets", []):
        net_name = nn["name"]
        for t in nn.get("terminals", []):
            desig = t.get("designator", "")
            pin = t.get("pin", "")
            if not desig or not pin:
                continue
            key = f"{desig}.{pin}"
            pin_to_net[key] = net_name
            comp_pins.setdefault(desig, {})[pin] = {
                "net": net_name,
                "pin_name": t.get("pin_name", ""),
            }

    # Komponentleri sayfa bazında grupla
    comps_by_sheet = {}
    for c in data["components"]:
        comp_entry = {
            "designator": c["designator"],
            "value": c["value"],
        }
        # Ek alanlar (boş değilse)
        for k in ("description", "footprint", "library_reference"):
            if c.get(k):
                comp_entry[k] = c[k]
        if c.get("parameters"):
            comp_entry["parameters"] = c["parameters"]
        # Pin → net bağlantıları (netlist'ten)
        pins = comp_pins.get(c["designator"])
        if pins:
            # pin numarasına göre sırala (sayısalsa sayısal sırada)
            def _pin_key(p):
                """@brief _pin_key()
                
                @param p
                @return Üretilen sonuç.
                """
                return (0, int(p)) if p.isdigit() else (1, p)
            comp_entry["pins"] = {
                p: pins[p] for p in sorted(pins.keys(), key=_pin_key)
            }
        comps_by_sheet.setdefault(c["sheet_id"], []).append(comp_entry)

    sheets_json = [
        {
            "id": s["id"],
            "name": s["name"],
            "components": comps_by_sheet.get(s["id"], []),
        }
        for s in data["sheets"]
    ]

    # Netlist'ten net adı → pin terminalleri eşlemesi
    # (gerçek elektriksel bağlantı: hangi pin hangi net'e bağlı)
    netlist = data.get("netlist") or {}
    terminals_by_net = {}
    for nn in netlist.get("nets", []):
        terminals_by_net[nn["name"]] = nn

    nets_json = []
    matched_terminals = 0
    for net in data["net_list"]:
        by_sheet = {}
        for occ in net["occurrences"]:
            by_sheet[occ["sheet_name"]] = by_sheet.get(occ["sheet_name"], 0) + 1
        entry = {
            "name": net["name"],
            "type": classify_net(net["name"]),
            "count": net["count"],
            "sheets": by_sheet,
        }
        # Gerçek pin bağlantıları (netlist'ten)
        nl_net = terminals_by_net.get(net["name"])
        if nl_net and nl_net.get("terminals"):
            conns = []
            for t in nl_net["terminals"]:
                pin_ref = f"{t['designator']}.{t['pin']}"
                if t.get("pin_name"):
                    pin_ref += f" ({t['pin_name']})"
                conns.append(pin_ref)
            entry["connections"] = conns
            matched_terminals += len(conns)
        nets_json.append(entry)

    # Netlist'te olup net_list'te olmayan net'leri de ekle
    # (auto-named net'ler, tek-pin net'ler vs. — SVG'de label'ı olmayanlar)
    known_names = {net["name"] for net in data["net_list"]}
    for nn in netlist.get("nets", []):
        if nn["name"] in known_names:
            continue
        if not nn.get("terminals"):
            continue
        conns = []
        for t in nn["terminals"]:
            pin_ref = f"{t['designator']}.{t['pin']}"
            if t.get("pin_name"):
                pin_ref += f" ({t['pin_name']})"
            conns.append(pin_ref)
        nets_json.append({
            "name": nn["name"],
            "type": classify_net(nn["name"]),
            "count": len(conns),
            "sheets": {},
            "connections": conns,
            "auto_named": nn.get("auto_named", False),
        })
        matched_terminals += len(conns)

    extras = data.get("design_extras") or {}
    bom = extras.get("bom") or []
    pnp = extras.get("pnp") or []
    variants = extras.get("variants") or []

    json_obj = {
        "project": {
            "name": data["project_name"],
            "path": str(project_path),
        },
        "summary": {
            "sheet_count": len(data["sheets"]),
            "net_count": len(nets_json),
            "component_count": len(data["components"]),
            "pin_connection_count": matched_terminals,
            "has_netlist": bool(netlist.get("nets")),
            "has_bom": bool(bom),
            "has_pnp": bool(pnp),
            "variant_count": len(variants),
        },
        "variants": variants,
        "sheets": sheets_json,
        "nets": nets_json,
    }
    # BOM ve Pick&Place yalnızca veri varsa ekle (sade tut)
    if bom:
        json_obj["bom"] = bom
    if pnp:
        json_obj["pnp"] = pnp

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        json.dumps(json_obj, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    log(f"\n✓ JSON üretildi: {out}")
    log(f"  Boyut: {out.stat().st_size / 1024:.1f} KB")


def generate_bom_csv(project_path, output_path, variant=None,
                     log: Callable[[str], None] = print):
    """@brief BOM'u CSV dosyası olarak üret. variant verilirse o varyanta filtrele.
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param output_path Çıktı dosyası yolu
    @param variant Varyant adı
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    import csv
    extras = collect_design_extras(project_path, log)
    bom = extras.get("bom") or []
    if not bom:
        log("! BOM verisi yok — CSV üretilemedi.")
        return False

    # Tüm parametre anahtarlarını topla (sütun başlıkları için)
    base_cols = ["designator", "value", "footprint", "library_ref",
                 "description", "dnp"]
    param_keys = set()
    for row in bom:
        for k in (row.get("parameters") or {}).keys():
            param_keys.add(k)
    param_cols = sorted(param_keys)

    out = Path(output_path).with_suffix(".csv")
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(base_cols + param_cols)
        for row in bom:
            params = row.get("parameters") or {}
            line = [row.get(c, "") for c in base_cols]
            line += [params.get(k, "") for k in param_cols]
            writer.writerow(line)

    log(f"\n✓ BOM CSV üretildi: {out}")
    log(f"  {len(bom)} komponent · {len(param_cols)} parametre sütunu")
    return True


def generate_pnp_csv(project_path, output_path, variant=None, units="mm",
                     log: Callable[[str], None] = print):
    """@brief Pick&Place'i CSV dosyası olarak üret (PCB gerektirir).
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param output_path Çıktı dosyası yolu
    @param variant Varyant adı
    @param units Birim
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    import csv
    extras = collect_design_extras(project_path, log)
    pnp = extras.get("pnp") or []
    if not pnp:
        log("! Pick&Place verisi yok (PCB dosyası gerekli) — CSV üretilemedi.")
        return False

    cols = ["designator", "comment", "layer", "footprint",
            "center_x", "center_y", "rotation"]
    out = Path(output_path).with_suffix(".csv")
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Designator", "Comment", "Layer", "Footprint",
                         f"X ({units})", f"Y ({units})", "Rotation"])
        for e in pnp:
            writer.writerow([e.get(c, "") for c in cols])

    log(f"\n✓ Pick&Place CSV üretildi: {out}")
    log(f"  {len(pnp)} yerleşim ({units})")
    return True


def _infer_pin_function(pin_name, net, net_type):
    """@brief Pin adı ve net adından fonksiyonel blok kategorisini çıkar.
    
    NOT: Bu otomatik sınıflandırmadır, tasarımcının özel açıklaması değil.
    
    @param pin_name Pin adı
    @param net Net adı
    @param net_type Net tipi
    @return Üretilen sonuç.
    """
    pn = (pin_name or "").upper()
    nu = (net or "").upper()
    if (net_type == "ground" or
            pn in ("VSS", "GND", "VSSA", "AGND", "DGND", "EPAD") or
            pn.startswith(("VSS", "AVSS"))):
        return "Güç — Toprak (GND)"
    if (net_type == "power" or
            re.search(r"^V(DD|CC|IN|OUT|BAT|REF|SYS|DDQ|DDA)", pn) or
            pn.startswith(("VDD", "AVDD", "VBAT"))):
        return "Güç — Besleme"
    # Özel bloklar (TX/RX'ten önce kontrol — Ethernet TXRX yanlış eşleşmesin)
    if re.search(r"ENET|RGMII|MDIO|MDC|GBE|\bETH|TXRX|TRX\d", pn) or re.search(r"\bETH|ENET|RGMII", nu):
        return "Ethernet"
    if re.search(r"USB", pn) or re.search(r"USB", nu):
        return "USB"
    if re.search(r"PCIE|PCI_E", pn) or re.search(r"PCIE", nu):
        return "PCIe"
    if re.search(r"MIPI|DSI|CSI", pn) or re.search(r"MIPI|DSI|CSI", nu):
        return "MIPI (Kamera/Ekran)"
    if re.search(r"DRAM|DDR|\bDQ\d|DQS|LPDDR", pn) or re.search(r"DRAM|LPDDR", nu):
        return "DRAM"
    if re.search(r"SD\d|SDIO|MMC|EMMC|NAND", pn) or re.search(r"EMMC|SD\d|NAND", nu):
        return "SD/MMC/Flash"
    if re.search(r"SDA", pn):
        return "I2C — Data"
    if re.search(r"SCL", pn):
        return "I2C — Clock"
    if re.search(r"MISO", pn):
        return "SPI — MISO"
    if re.search(r"MOSI", pn):
        return "SPI — MOSI"
    if re.search(r"SCK|SCLK", pn):
        return "SPI — Clock"
    if re.search(r"\bCS\b|CS_|_CS|NSS|\bSS\b", pn):
        return "SPI — Chip Select"
    if re.search(r"RESET|RST|NRST", pn):
        return "Reset"
    if re.search(r"\bINT\b|IRQ|_INT|_WAKE", pn):
        return "Kesme / Wake"
    if re.search(r"UART.*TX|\bTXD\b|\bTX\b", pn):
        return "UART — TX"
    if re.search(r"UART.*RX|\bRXD\b|\bRX\b", pn):
        return "UART — RX"
    if re.search(r"ADC|AIN|\bVIN\b", pn):
        return "Analog Giriş"
    if re.search(r"CLK|OSC|XTAL", pn):
        return "Saat / Osilatör"
    if re.search(r"GPIO|IO\d", pn):
        return "GPIO"
    return "Sinyal"


def _detect_interface(pin_names):
    """@brief Pin adlarından IC'nin haberleşme arayüzünü tahmin et (I2C / SPI).
    
    @param pin_names Pin adları listesi
    @return Üretilen sonuç.
    """
    names = " ".join((n or "").upper() for n in pin_names)
    has_i2c = bool(re.search(r"\bSDA\b|\bSCL\b", names))
    has_spi = bool(re.search(r"\bSCK\b|\bMISO\b|\bMOSI\b|\bSCLK\b", names))
    if has_i2c and has_spi:
        return "I2C / SPI"
    if has_i2c:
        return "I2C"
    if has_spi:
        return "SPI"
    return ""


def _natural_pin_key(p):
    """@brief BGA/numeric pin doğal sıralama: harf öneki + sayı + kalan.
    
    @param p
    @return Üretilen sonuç.
    """
    m = re.match(r"^([A-Za-z]*)(\d*)(.*)$", p or "")
    if m:
        return (m.group(1), int(m.group(2)) if m.group(2) else 0, m.group(3))
    return (p, 0, "")


# Seri pasif tespiti: R12, C4, L3, FB2, F1 gibi 2-pinli parçalar netlist
# izlemesinde "atlanabilir" sayılır (RN/CN/LED gibi öneklere \d şartı takılmaz).
_JUMPABLE_RE = re.compile(r"^(?:FB|R|C|L|F)\d", re.I)


def _trace_net_endpoints(start_net, net_terminals, comp_pins, is_power,
                         exclude=frozenset(), max_hops=4):
    """@brief Net'i seri pasifler ÜZERİNDEN izleyip gerçek uç noktaları bulur.

    Bir MCU pini çoğu zaman IC'ye doğrudan değil seri direnç/ferrit üzerinden
    gider; ham netlist bu durumda karşı-uç olarak sadece direnci gösterir.
    Bu fonksiyon 2-pinli pasiflerin (R/C/L/FB/F) öbür bacağındaki net'e
    atlayarak (BFS, max_hops pasif derinliği) sinyalin ULAŞTIĞI IC/konnektör
    pinlerini döndürür. Pasif üzerinden bir güç/toprak net'ine varılırsa bu
    ayrıca pull-up/pull-down bilgisi olarak raporlanır.

    @param start_net Başlangıç net adı
    @param net_terminals {net: [(desig, pin, pin_name), ...]}
    @param comp_pins {desig: {pin: {"net", "pin_name"}}}
    @param is_power Net adını alıp güç/toprak ise True dönen callable
    @param exclude İzlemede yok sayılacak designator kümesi (örn. MCU'nun kendisi)
    @param max_hops Atlanacak maksimum seri pasif sayısı
    @return (endpoints, pulls): endpoints=[{"desig","pin","pin_name","via"}],
            pulls=[(via_tuple, güç_net_adı)]
    """
    endpoints = {}
    pulls = []
    seen_nets = {start_net}
    queue = [(start_net, ())]
    while queue:
        net, via = queue.pop(0)
        for d, p, pn in net_terminals.get(net, []):
            if not d or not p or d in exclude:
                continue
            jumpable = bool(_JUMPABLE_RE.match(d)) and len(comp_pins.get(d, {})) == 2
            if jumpable:
                if len(via) >= max_hops:
                    continue
                for p2, info2 in comp_pins[d].items():
                    net2 = info2["net"]
                    if net2 in seen_nets:
                        continue
                    seen_nets.add(net2)
                    if is_power(net2):
                        pulls.append((via + (d,), net2))
                    else:
                        queue.append((net2, via + (d,)))
            else:
                key = (d, p)
                if key not in endpoints or len(via) < len(endpoints[key]["via"]):
                    endpoints[key] = {"desig": d, "pin": p, "pin_name": pn, "via": via}
    return list(endpoints.values()), pulls


def _endpoint_sort_key(e):
    """@brief Uç noktaları önem sırasına dizer: IC'ler önce, testpoint'ler sonda.

    @param e _trace_net_endpoints endpoint sözlüğü
    @return Sıralama anahtarı (öncelik, pasif-derinliği, designator)
    """
    d = e["desig"].upper()
    if re.match(r"^(U\d|IC|Q\d)", d):
        pri = 0
    elif re.match(r"^(J\d|CN|P\d|X\d|K\d|S\w?\d)", d):
        pri = 1
    elif re.match(r"^(TP|MP|H\d)", d):
        pri = 3
    else:
        pri = 2
    return (pri, len(e["via"]), _natural_pin_key(e["desig"]))


def _fmt_endpoint(e):
    """@brief Uç noktayı 'U6.5 (SDA) [R12 üzerinden]' biçiminde yazar.

    @param e _trace_net_endpoints endpoint sözlüğü
    @return Biçimlenmiş metin
    """
    s = f"{e['desig']}.{e['pin']}"
    if e["pin_name"]:
        s += f" ({e['pin_name']})"
    if e["via"]:
        s += f" [{'+'.join(e['via'])} üzerinden]"
    return s


def _fmt_pull(via, pnet):
    """@brief Pasif→güç bağlantısını 'R5→+3V3 (pull-up)' biçiminde yazar.

    Son pasif kondansatörse DC pull olamaz → 'filtre C' diye etiketlenir.

    @param via Atlanan pasif designator zinciri (tuple)
    @param pnet Ulaşılan güç/toprak net adı
    @return Biçimlenmiş metin
    """
    u = pnet.upper()
    gndish = u.startswith(("GND", "VSS", "AGND", "DGND")) or u.endswith("GND")
    if via[-1][:1].upper() == "C":
        kind = "filtre C"
    else:
        kind = "pull-down" if gndish else "pull-up"
    return f"{'+'.join(via)}→{pnet} ({kind})"


def _net_port_dir(ports, main_sheet):
    """@brief Net'in şematik PORT'undan sinyal yönünü çıkar: 'IN'/'OUT'/'I/O'/''.

    Aynı isimli port birden çok sayfada olabilir (hiyerarşide karşılıklı yönler).
    main_sheet (küçük harf, örn. 'mcu') verilirse dosya adı onu içeren sayfadaki
    port tercih edilir — yön MCU perspektifinden okunur (şematikteki ok yönü).

    @param ports Net'in ports listesi ([{"name","source_sheet","io_type"}])
    @param main_sheet Ana işlemcinin sayfa adı (küçük harf) veya ''
    @return 'IN' | 'OUT' | 'I/O' | '' (port yok / yön belirsiz)
    """
    if not ports:
        return ""
    pick = None
    if main_sheet:
        for pt in ports:
            if main_sheet in (pt.get("source_sheet") or "").lower():
                pick = pt
                break
    if pick is None:
        pick = ports[0]
    return {"INPUT": "IN", "OUTPUT": "OUT",
            "BIDIRECTIONAL": "I/O"}.get((pick.get("io_type") or "").upper(), "")


def generate_ic_map_xlsx(project_path, output_path, min_pins=4,
                         main_designators=None, exclude_prefixes=None,
                         log: Callable[[str], None] = print):
    """@brief IC Bağlantı Haritası Excel üret (mcu.xlsx örneği düzeni).

    Tek tabloda IC grupları: No | Kontrol Entegresi | Sinyal Adı (Net) |
    Kontrol Arayüzü | I2C Adres | Entegre Portu | MCU Portu | Fonksiyonel Blok.
    Yalnız SİNYAL pinleri listelenir (güç/toprak satırları yok). 'MCU Portu'
    ham netlist karşı-ucu değil, seri pasifler (R/L/C/FB) üzerinden
    _trace_net_endpoints ile izlenen GERÇEK ana işlemci pinidir
    (örn. 'PC4 (R12 üzerinden)'). I2C adresi şematikten türetilemez, '-' konur.

    Parametreler:
      min_pins: bu sayıdan az pinli komponentler atlanır (varsayılan 4)
      main_designators: ana işlemci(ler)in designator'ı. Şunlar olabilir:
        - None / boş → otomatik tespit (en çok pinli aktif U* komponenti)
        - "U2" (tek string) → tek ana işlemci
        - ["U2", "U7"] (liste) veya "U2,U7" (virgüllü) → çoklu ana işlemci
      Ana işlemci grubu tablonun BAŞINDA yer alır; bağlı olmayan pinleri de
      "NC" satırı olarak listelenir.
      exclude_prefixes: designator HARF önekiyle komponent hariç tutma
        ("J,P,TP" veya liste) — çok pinli ama gereksiz parçaları (konnektör,
        header) listeden çıkarmak için. Ana işlemci asla hariç tutulmaz.

    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param output_path Çıktı dosyası yolu
    @param min_pins Minimum pin sayısı eşiği
    @param main_designators Ana işlemci designator listesi
    @param exclude_prefixes Hariç tutulacak designator harf önekleri (str/list)
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        log(f"! openpyxl yok, Excel üretilemedi: {e}  (pip install openpyxl)")
        return False

    data = _collect_data(project_path, log)
    netlist = data.get("netlist") or {}
    nets = netlist.get("nets") or []
    if not nets:
        log("! Netlist yok — IC haritası üretilemedi.")
        return False

    log("\nIC Bağlantı Haritası hazırlanıyor...")

    def strip_ob(s):
        """@brief strip_ob()
        
        @param s
        @return Üretilen sonuç.
        """
        return (s or "").replace("\\", "")

    # net -> terminaller [(desig, pin, pin_name)] + net -> portlar (yön için)
    net_terminals = {}
    net_ports = {}
    for n in nets:
        nm = strip_ob(n.get("name", ""))
        net_terminals[nm] = [
            (t.get("designator", ""), t.get("pin", ""), strip_ob(t.get("pin_name", "")))
            for t in n.get("terminals", [])
        ]
        net_ports[nm] = n.get("ports") or []

    # komponent -> {pin: {net, pin_name}}  ve pin sayıları
    comp_pins = {}
    for nm, terms in net_terminals.items():
        for dd, pp, pnn in terms:
            if dd and pp:
                comp_pins.setdefault(dd, {})[pp] = {"net": nm, "pin_name": pnn}

    if not comp_pins:
        log("! Pin bağlantısı bulunamadı.")
        return False

    # === Ana işlemci(ler) belirle ===
    # Girdiyi normalize et: None/str/list → temizlenmiş designator listesi
    if main_designators is None:
        main_list = []
    elif isinstance(main_designators, str):
        main_list = [d.strip() for d in main_designators.split(",") if d.strip()]
    else:
        main_list = [str(d).strip() for d in main_designators if str(d).strip()]

    # Verilenleri doğrula (gerçekten projede var mı, case-insensitive)
    valid_main = []
    avail_upper = {d.upper(): d for d in comp_pins}
    for d in main_list:
        real = avail_upper.get(d.upper())
        if real:
            valid_main.append(real)
        else:
            log(f"  ! Uyarı: '{d}' ana işlemci olarak girildi ama projede bulunamadı, atlanıyor.")

    if not valid_main:
        # Otomatik tespit: en çok pinli U* (yoksa en çok pinli herhangi)
        by_count = sorted(comp_pins.items(), key=lambda kv: -len(kv[1]))
        auto = next((d for d, _ in by_count if d.upper().startswith("U")), by_count[0][0])
        valid_main = [auto]
        log(f"  · Ana işlemci otomatik seçildi: {auto} "
            f"({len(comp_pins[auto])} pin). Belirli bir IC istiyorsan "
            f"main_designators parametresiyle ver.")
    else:
        log(f"  · Ana işlemci(ler): {', '.join(valid_main)}")

    main_set = set(valid_main)

    # Ana işlemcinin bağlı olmayan (NC) pinlerini şematik pin kataloğundan
    # ekle (net='' işareti, görüntüde "NC" — mcu.xlsx örneğindeki NC satırları)
    all_pins_cat = netlist.get("all_pins") or {}
    for d in valid_main:
        cat = (all_pins_cat.get(d) or
               all_pins_cat.get(re.sub(r"_\d+$", "", d)) or {})
        for p, pn in cat.items():
            if p not in comp_pins.get(d, {}):
                comp_pins.setdefault(d, {})[p] = {"net": "",
                                                  "pin_name": strip_ob(pn)}

    # komponent meta (value/footprint/sheet) — data["components"]'ten
    meta = {}
    for c in data.get("components", []):
        meta[c["designator"]] = c

    def classify_net(name):
        """@brief classify_net()
        
        @param name Ad
        @return Üretilen sonuç.
        """
        u = name.upper()
        if (re.match(r"^A?GND", u) or u.startswith(("VSS", "AVSS")) or
                u in ("AGND", "DGND", "GND") or u.endswith("GND")):
            return "ground"
        if (re.match(r"^[+\-]?\d*\.?\d+V", u) or
                re.match(r"^A?V(CC|DD|IN|OUT|BAT|REF|SYS)", u) or u in ("VCC", "VDD")):
            return "power"
        return "signal"

    def is_power(name):
        """@brief Net güç/toprak mı? (_trace_net_endpoints durdurma koşulu)

        @param name Net adı
        @return bool
        """
        return classify_net(name) != "signal"

    def mcu_port_for(net, self_desig):
        """@brief Net'i seri pasifler üzerinden izleyip ana işlemci portunu bul.

        Ham netlist'te karşı-uç bir dirençse bu fonksiyon direncin öbür
        bacağından devam edip MCU pinine ulaşır (örn. 'PC4 (R12 üzerinden)').

        @param net Net adı
        @param self_desig Komponentin kendi designator'ı
        @return Biçimlenmiş MCU port metni ('' = MCU'ya ulaşmıyor)
        """
        eps, _ = _trace_net_endpoints(net, net_terminals, comp_pins, is_power,
                                      exclude={self_desig})
        mains = [e for e in eps if e["desig"] in main_set]
        if not mains:
            return ""
        mains.sort(key=lambda e: (len(e["via"]), _natural_pin_key(e["pin"])))
        outs = []
        for e in mains[:3]:
            t = e["pin_name"] or f"pin {e['pin']}"
            if len(main_set) > 1:
                t = f"{e['desig']}.{t}"
            if e["via"]:
                t += f" ({'+'.join(e['via'])} üzerinden)"
            outs.append(t)
        if len(mains) > 3:
            outs.append(f"… +{len(mains) - 3}")
        return " ; ".join(outs)

    # Hariç tutulacak designator HARF önekleri (örn. "J,P,TP" → konnektör/
    # header'lar listelenmez). Ana işlemci asla hariç tutulmaz.
    if exclude_prefixes is None:
        excl = set()
    elif isinstance(exclude_prefixes, str):
        excl = {x.strip().upper() for x in exclude_prefixes.split(",") if x.strip()}
    else:
        excl = {str(x).strip().upper() for x in exclude_prefixes if str(x).strip()}

    def desig_prefix(d):
        """@brief Designator'ın harf önekini döndür ('U2_1' → 'U').

        @param d Designator
        @return Harf öneki (büyük harf)
        """
        m = re.match(r"^([A-Za-z]+)", d)
        return m.group(1).upper() if m else ""

    # ≥min_pins komponentleri: ANA İŞLEMCİ(LER) TABLONUN BAŞINDA,
    # çevre IC'ler designator sırasıyla arkasından
    selected = [(d, p) for d, p in comp_pins.items()
                if len(p) >= min_pins and
                (d in main_set or desig_prefix(d) not in excl)]
    if excl:
        n_excl = sum(1 for d in comp_pins
                     if len(comp_pins[d]) >= min_pins and d not in main_set and
                     desig_prefix(d) in excl)
        log(f"  · Hariç tutulan önekler: {', '.join(sorted(excl))} "
            f"({n_excl} komponent atlandı)")

    def desig_key(item):
        """@brief desig_key()

        @param item
        @return Üretilen sonuç.
        """
        m = re.match(r"^([A-Za-z]+)(\d+)", item[0])
        return (m.group(1), int(m.group(2))) if m else (item[0], 0)
    selected.sort(key=lambda it: (it[0] not in main_set, desig_key(it)))

    # === Excel kur (mcu.xlsx düzeni: tek tablo, IC grupları birleşik hücre) ===
    wb = Workbook()
    ws = wb.active
    ws.title = "IC Baglanti Haritasi"
    F = "Calibri"
    title_font = Font(name=F, bold=True, size=14, color="1F3864")
    meta_font = Font(name=F, size=9, color="555555")
    colhdr_font = Font(name=F, bold=True, color="FFFFFF", size=10)
    colhdr_fill = PatternFill("solid", fgColor="1F3864")
    ic_font = Font(name=F, bold=True, size=9, color="1F3864")
    cell_font = Font(name=F, size=9)
    mono_font = Font(name="Consolas", size=9)
    no_font = Font(name=F, size=9, color="888888")
    dash_font = Font(name=F, size=9, color="AAAAAA")
    alt_fill = PatternFill("solid", fgColor="F2F6FB")
    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="medium", color="1F3864")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap_top = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

    main_label = ", ".join(valid_main)
    COLS = ["No", "Kontrol Entegresi", "Desig", "Sinyal Adı (Net)",
            "Kontrol\nArayüzü", "I2C\nAdres", "Entegre Portu", "Pin\nSay.",
            f"MCU Portu ({main_label})", "Fonksiyonel Blok"]
    widths = [5, 32, 9, 26, 11, 9, 20, 7, 26, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    NCOL = len(COLS)

    def merge_row(r, text, font, height=None):
        """@brief Tek satırı tüm sütunlar boyunca birleştirip yazar.

        @param r Satır numarası
        @param text Metin
        @param font Font
        @param height Satır yüksekliği
        """
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        c = ws.cell(row=r, column=1)
        c.value = text
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if height:
            ws.row_dimensions[r].height = height

    main_vals = ", ".join(f"{d} ({meta.get(d, {}).get('value', '')})" for d in valid_main)
    merge_row(1, f"{Path(project_path).stem} — IC Bağlantı Haritası", title_font, height=24)
    merge_row(2, f"Ana işlemci: {main_vals}  ·  "
              f"'MCU Portu' = seri pasifler (R/L/FB) üzerinden izlenen ana işlemci pini  ·  "
              f"güç/toprak pinleri listelenmez  ·  I2C adresleri şematikten "
              f"çıkarılamaz, elle doldurun", meta_font)

    hdr_row = 4
    for c, name in enumerate(COLS, 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.value = name
        cell.font = colhdr_font
        cell.fill = colhdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[hdr_row].height = 26
    ws.freeze_panes = f"A{hdr_row + 1}"

    # Ana işlemcinin sayfası (port yönü MCU perspektifinden okunsun diye)
    mcu_sheet = (meta.get(valid_main[0], {}).get("sheet_name") or "").lower()

    def row_iface(func, net, pname=""):
        """@brief Satır bazında Kontrol Arayüzü: I2C/SPI/USART/SWD/USB veya yönlü GPIO.

        Arayüz önce pin fonksiyonu + net/pin ADINDAN tespit edilir (SPI_SCK →
        SPI, SWDIO → SWD, MCU_TX → USART; mcu.xlsx örneğindeki gibi — CS/IRQ
        gibi tekil kontrol hatları GPIO sayılır). Arayüz değilse yön, net'e
        bağlı şematik PORT'un ok yönünden gelir (GPIO_IN/GPIO_OUT).

        @param func _infer_pin_function çıktısı
        @param net Net adı
        @param pname Pin adı
        @return Arayüz metni
        """
        s = f"{net} {pname}".upper()
        if "USB" in s or "USB" in func:
            return "USB"
        if re.search(r"SWDIO|SWCLK|(^|[_ /])SWO?([_ /]|$)|JTAG|JTCK|JTMS|JMS", s):
            return "SWD"
        if "I2C" in func or re.search(r"I2C|SCL(?!K)|SDA", s):
            return "I2C"
        if "SPI" in func or re.search(r"SPI|SCK|SCLK|MISO|MOSI", s):
            return "SPI"
        if ("UART" in func or
                re.search(r"UART|USART|RS485|(^|[_ /])(TXD?|RXD?|RE|DE)([_ /]|$)", s)):
            return "USART"
        d = _net_port_dir(net_ports.get(net) or [], mcu_sheet)
        return f"GPIO_{d}" if d else "GPIO"

    row = hdr_row + 1
    no = 1
    group_alt = False
    written_ics = 0
    for desig, pins in selected:
        # Sinyal pinleri: güç/toprak NET'leri VE güç görevli pinler (VDD/VBAT
        # gibi — kullanıcı isteği: besleme pinleri listelenmez) atlanır
        rows_data = []
        for p in sorted(pins, key=_natural_pin_key):
            net = pins[p]["net"]
            pname = pins[p]["pin_name"]
            if classify_net(net) != "signal":
                continue
            func = _infer_pin_function(pname, net, "signal")
            if func.startswith("Güç"):
                continue
            rows_data.append((p, net, pname, "NC" if not net else func))
        if not rows_data:
            continue
        # Kanal-sonekli designator (U2_1) meta'da yoksa taban adla (U2) dene
        m = meta.get(desig) or meta.get(re.sub(r"_\d+$", "", desig)) or {}
        is_main_ic = desig in main_set
        fill = alt_fill if group_alt else None
        start = row

        for p, net, pname, func in rows_data:
            if is_main_ic:
                dev_port = f"Pin {p}"
                mcu_port = pname or f"pin {p}"
            else:
                dev_port = f"{pname} (Pin {p})" if pname else f"Pin {p}"
                mcu_port = mcu_port_for(net, desig) or "—"
            vals = [no, "", "", net or "NC", row_iface(func, net, pname), "",
                    dev_port, "", mcu_port, func]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c)
                cell.value = v
                cell.border = border
                if fill:
                    cell.fill = fill
                if c == 1:
                    cell.font = no_font
                    cell.alignment = center
                elif c == 4 or c == 7:
                    cell.font = mono_font
                elif c == 5:
                    cell.font = cell_font
                    cell.alignment = center
                elif c == 9:
                    cell.font = dash_font if v == "—" else mono_font
                else:
                    cell.font = cell_font
            row += 1
            no += 1
        end = row - 1

        # Grup hücreleri: Kontrol Entegresi / Desig / I2C Adres / Pin Sayısı
        # dikey birleşik (Arayüz SATIR bazında: GPIO_IN/GPIO_OUT yönleriyle)
        ic_lines = [m.get("value", "") or desig]
        if m.get("description"):
            ic_lines.append(m["description"][:60])
        base_d = re.sub(r"_\d+$", "", desig)
        total_pin_count = (len(all_pins_cat.get(desig) or
                               all_pins_cat.get(base_d) or {}) or len(pins))
        desig_txt = desig + ("\n★ ANA" if is_main_ic else "")
        for col, val, al in ((2, "\n".join(ic_lines), wrap_top),
                             (3, desig_txt, Alignment(horizontal="center",
                                                      vertical="center",
                                                      wrap_text=True)),
                             (6, "-", center),
                             (8, total_pin_count, center)):
            if end > start:
                ws.merge_cells(start_row=start, start_column=col,
                               end_row=end, end_column=col)
            cell = ws.cell(row=start, column=col)
            cell.value = val
            cell.font = ic_font if col in (2, 3) else cell_font
            cell.alignment = al
        # Grup üst sınırı kalın — IC blokları ayrışsın
        for c in range(1, NCOL + 1):
            cur = ws.cell(row=start, column=c).border
            ws.cell(row=start, column=c).border = Border(
                left=cur.left, right=cur.right, top=thick, bottom=cur.bottom)
        group_alt = not group_alt
        written_ics += 1

    out = Path(output_path).with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log(f"\n✓ IC Bağlantı Haritası üretildi: {out}")
    log(f"  {written_ics} IC · {no - 1} sinyal satırı · ana işlemci: {', '.join(valid_main)}")
    return True


def generate_mcu_pinout_xlsx(project_path, output_path, mcu_designator,
                             include_power=True,
                             log: Callable[[str], None] = print):
    """@brief MCU merkezli pin listesi Excel üret.

    Verilen MCU'nun HER pini bir satır: pin no, pin adı, net, ve sinyalin
    ULAŞTIĞI gerçek IC portları. Hedef sütununda seri pasifler (R/L/C/FB)
    _trace_net_endpoints ile atlanır: direnç yerine 'IC6.9 (P0_0)
    [R12 üzerinden]' yazılır; pasif üzerinden güç net'ine giden bağlantılar
    'R5→+3V3 (pull-up)' olarak raporlanır.

    Parametreler:
      mcu_designator: zorunlu — listesi çıkarılacak MCU'nun designator'ı (örn "U2")
      include_power: False ise GND/VDD/VCC/VSS gibi güç-toprak pinleri atlanır
    
    @param project_path Altium proje dosyası (.PrjPcb) yolu
    @param output_path Çıktı dosyası yolu
    @param mcu_designator MCU designator'ı
    @param include_power Güç netlerini dahil et (bool)
    @param log Log mesajı callback'i (str alır)
    @return Üretilen sonuç.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        log(f"! openpyxl yok, Excel üretilemedi: {e}  (pip install openpyxl)")
        return False

    if not mcu_designator or not str(mcu_designator).strip():
        log("! MCU designator boş — hangi entegrenin pin listesi çıkarılacak "
            "belirtilmeli (örn 'U2').")
        return False
    mcu_designator = str(mcu_designator).strip()

    data = _collect_data(project_path, log)
    netlist = data.get("netlist") or {}
    nets = netlist.get("nets") or []
    if not nets:
        log("! Netlist yok — MCU pin listesi üretilemedi.")
        return False

    def strip_ob(s):
        """@brief strip_ob()
        
        @param s
        @return Üretilen sonuç.
        """
        return (s or "").replace("\\", "")

    # net -> terminaller [(desig, pin, pin_name)] + net -> portlar (yön için)
    net_terminals = {}
    net_ports = {}
    for n in nets:
        nm = strip_ob(n.get("name", ""))
        net_terminals[nm] = [
            (t.get("designator", ""), t.get("pin", ""), strip_ob(t.get("pin_name", "")))
            for t in n.get("terminals", [])
        ]
        net_ports[nm] = n.get("ports") or []

    # komponent -> {pin: {net, pin_name}} — pasif-izleme için tüm parçalar
    comp_pins = {}
    for nm, terms in net_terminals.items():
        for dd, pp, pnn in terms:
            if dd and pp:
                comp_pins.setdefault(dd, {})[pp] = {"net": nm, "pin_name": pnn}

    # MCU designator'ını doğrula (case-insensitive)
    all_desigs = {d.upper() for terms in net_terminals.values() for d, _, _ in terms}
    if mcu_designator.upper() not in all_desigs:
        log(f"! '{mcu_designator}' projede bulunamadı. "
            f"Mevcut entegrelerden birini gir.")
        return False
    # Gerçek designator yazımını bul
    real = next((d for terms in net_terminals.values()
                 for d, _, _ in terms if d.upper() == mcu_designator.upper()),
                mcu_designator)
    mcu_designator = real

    # MCU pinleri: pin -> {net, pin_name}
    mcu_pins = {}
    for nm, terms in net_terminals.items():
        for d, p, pn in terms:
            if d == mcu_designator and p:
                mcu_pins[p] = {"net": nm, "pin_name": pn}

    if not mcu_pins:
        log(f"! {mcu_designator} için pin bulunamadı.")
        return False

    # Bağlı olmayan (NC) pinleri şematik pin kataloğundan ekle (net='' işareti,
    # görüntüde "NC"). Kanal-sonekli designator için taban ad fallback'i.
    all_pins_cat = netlist.get("all_pins") or {}
    cat = (all_pins_cat.get(mcu_designator) or
           all_pins_cat.get(re.sub(r"_\d+$", "", mcu_designator)) or {})
    nc_added = 0
    for p, pn in cat.items():
        if p not in mcu_pins:
            mcu_pins[p] = {"net": "", "pin_name": strip_ob(pn)}
            nc_added += 1
    if nc_added:
        log(f"  · {nc_added} bağlı olmayan (NC) pin eklendi.")

    log(f"\nMCU pin listesi hazırlanıyor: {mcu_designator} ({len(mcu_pins)} pin)")

    # meta (MCU değeri başlıkta göstermek için)
    meta = {c["designator"]: c for c in data.get("components", [])}
    mcu_val = meta.get(mcu_designator, {}).get("value", "")
    # MCU'nun sayfası: port yönü (GPIO_IN/OUT) MCU perspektifinden okunur
    mcu_sheet = (meta.get(mcu_designator, {}).get("sheet_name") or "").lower()

    def func_with_dir(pname, net, nt):
        """@brief Fonksiyon çıkarımı + şematik port yönü (GPIO_IN/GPIO_OUT).

        @param pname Pin adı
        @param net Net adı
        @param nt Net tipi
        @return Fonksiyon metni
        """
        if not net:  # bağlı olmayan pin
            return "NC"
        func = _infer_pin_function(pname, net, nt)
        if func in ("Sinyal", "GPIO"):
            # Pin adı (PB8 vs) arayüzü ele vermez — NET adından tespit et
            nu = (net or "").upper()
            if re.search(r"SCL(?!K)", nu):
                return "I2C — Clock"
            if re.search(r"SDA", nu):
                return "I2C — Data"
            if re.search(r"SPI|SCK|SCLK|MISO|MOSI", nu):
                return "SPI"
            if re.search(r"SWDIO|SWCLK|(^|[_ /])SWO?([_ /]|$)|JTAG|JTCK|JMS", nu):
                return "SWD"
            if re.search(r"UART|USART|RS485|(^|[_ /])(TXD?|RXD?|RE|DE)([_ /]|$)", nu):
                return "USART"
            d = _net_port_dir(net_ports.get(net) or [], mcu_sheet)
            if d:
                return f"GPIO_{d}"
        return func

    def is_power_net(name):
        """@brief is_power_net()
        
        @param name Ad
        @return Üretilen sonuç.
        """
        u = name.upper()
        return (u.startswith(("GND", "VSS", "VDD", "VCC", "AGND", "DGND",
                              "AVSS", "AVDD")) or
                u.endswith("GND") or bool(re.match(r"^[+\-]?\d*\.?\d+V", u)))

    def natural_pin_key(p):
        """@brief natural_pin_key()
        
        @param p
        @return Üretilen sonuç.
        """
        m = re.match(r"^([A-Za-z]*)(\d*)(.*)$", p or "")
        if m:
            return (m.group(1), int(m.group(2)) if m.group(2) else 0, m.group(3))
        return (p, 0, "")

    # === Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = f"{mcu_designator} Pinout"[:31]
    F = "Calibri"
    title_font = Font(name=F, bold=True, size=14, color="1F3864")
    meta_font = Font(name=F, size=9, color="555555")
    hdr_font = Font(name=F, bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    cell_font = Font(name=F, size=10)
    mono_font = Font(name="Consolas", size=10)
    pwr_fill = PatternFill("solid", fgColor="FFF2CC")
    gnd_fill = PatternFill("solid", fgColor="E2EFDA")
    nc_font = Font(name=F, size=10, color="AAAAAA", italic=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    COLS = ["MCU Pin", "Pin Adı (MCU)", "Fonksiyon / Arayüz",
            "Net (Sinyal)", "Hedef IC Portu (seri pasifler atlanır)"]
    widths = [10, 26, 22, 28, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    NCOL = len(COLS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    ws.cell(row=1, column=1).value = f"{mcu_designator} ({mcu_val}) — Pin Bağlantı Listesi"
    ws.cell(row=1, column=1).font = title_font
    ws.cell(row=1, column=1).alignment = left
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
    ws.cell(row=2, column=1).value = (
        f"{len(mcu_pins)} pin  ·  {Path(project_path).stem}  ·  "
        f"Hedef sütunu seri pasifleri (R/L/FB) atlayıp ulaşılan GERÇEK IC "
        f"portunu gösterir; pull-up/down dirençleri ayrıca belirtilir. "
        f"Çoklu hedef ';' ile ayrılır.")
    ws.cell(row=2, column=1).font = meta_font

    # Fonksiyon dağılımı özeti (satır 3)
    from collections import Counter
    func_counter = Counter()
    for pp in mcu_pins.values():
        nn = pp["net"]
        uu2 = nn.upper()
        nt2 = ("ground" if (uu2.startswith(("GND", "VSS", "AGND", "DGND")) or uu2.endswith("GND"))
               else "power" if (bool(re.match(r"^[+\-]?\d*\.?\d+V", uu2)) or uu2.startswith(("VDD", "VCC")))
               else "signal")
        func_counter[func_with_dir(pp["pin_name"], nn, nt2)] += 1
    summary = "  ·  ".join(f"{k}: {v}" for k, v in func_counter.most_common())
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NCOL)
    ws.cell(row=3, column=1).value = "Fonksiyon dağılımı →  " + summary
    ws.cell(row=3, column=1).font = Font(name=F, size=9, color="1F6F54", italic=True)

    hdr_row = 5
    for c, name in enumerate(COLS, 1):
        cell = ws.cell(row=hdr_row, column=c)
        cell.value = name
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = border
    ws.freeze_panes = f"A{hdr_row + 1}"

    row = hdr_row + 1
    written = 0
    func_fill = PatternFill("solid", fgColor="EAF1FB")  # fonksiyon sütunu hafif mavi
    for p in sorted(mcu_pins.keys(), key=natural_pin_key):
        net = mcu_pins[p]["net"]
        pname = mcu_pins[p]["pin_name"]
        powernet = is_power_net(net)
        if powernet and not include_power:
            continue

        others = [(d, pp, pn) for d, pp, pn in net_terminals.get(net, [])
                  if d != mcu_designator]

        if powernet:
            target_str = f"({len(others)} bağlantı — güç/toprak ağı)" if others else "—"
        elif not others:
            target_str = "(bağlı değil / NC)"
        else:
            # Seri pasifler (R/L/C/FB) üzerinden atlayıp GERÇEK IC portunu bul:
            # 'R12' yerine 'IC6.9 (P0_0) [R12 üzerinden]'. Pasif bir güç
            # net'ine gidiyorsa pull-up/pull-down olarak raporla.
            eps, pulls = _trace_net_endpoints(
                net, net_terminals, comp_pins, is_power_net,
                exclude={mcu_designator})
            eps.sort(key=_endpoint_sort_key)
            parts = [_fmt_endpoint(e) for e in eps]
            seen_pull = set()
            for via, pnet in pulls:
                s = _fmt_pull(via, pnet)
                if s not in seen_pull:
                    seen_pull.add(s)
                    parts.append(s)
            if not parts:  # yalnız pasife gidip kör noktada bitiyorsa ham liste
                parts = [f"{d}.{pp} ({pn})" if pn else f"{d}.{pp}"
                         for d, pp, pn in others]
            target_str = " ; ".join(parts)

        # Fonksiyon/arayüz çıkarımı (pin adı + net adından)
        uu = net.upper()
        nt = ("ground" if (uu.startswith(("GND", "VSS", "AGND", "DGND")) or uu.endswith("GND"))
              else "power" if (bool(re.match(r"^[+\-]?\d*\.?\d+V", uu)) or uu.startswith(("VDD", "VCC")))
              else "signal")
        func = func_with_dir(pname, net, nt)

        fill = pwr_fill if nt == "power" else (gnd_fill if nt == "ground" else None)

        vals = [p, pname, func, net or "NC", target_str]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c)
            cell.value = v
            cell.border = border
            if c == 1:
                cell.font = mono_font
                cell.alignment = center
            elif c == 2:
                cell.font = cell_font
            elif c == 3:  # fonksiyon
                cell.font = cell_font
                if not fill:
                    cell.fill = func_fill
            elif c == 4:  # net
                cell.font = mono_font
            else:  # hedef
                cell.font = nc_font if target_str.startswith(("(bağlı", "—")) else cell_font
            if fill and c in (1, 3, 4):
                cell.fill = fill
        row += 1
        written += 1

    out = Path(output_path).with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log(f"\n✓ MCU pin listesi üretildi: {out}")
    log(f"  {mcu_designator} — {written} pin yazıldı")
    return True


def build_html(sheets, net_list, components, timestamp,
               inter_color, intra_color, pcb=None, project_name=""):
    """@brief Şematik viewer'ın tam HTML belgesini kurar (gömülü SVG + JS).
    
    @param sheets
    @param net_list Net listesi
    @param components Komponent listesi
    @param timestamp Zaman damgası metni
    @param inter_color Sayfalar arası bağlantı rengi (hex)
    @param intra_color Sayfa içi bağlantı rengi (hex)
    @param pcb AltiumPcbDoc PCB nesnesi
    @param project_name Proje adı (not/kutu localStorage anahtarı + kayıt dosya adı)
    @return Üretilen sonuç.
    """
    pcb = pcb or {"available": False}
    max_x = max((s["x"] + s["w"] for s in sheets), default=2000) + 100
    max_y = max((s["y"] + s["h"] for s in sheets), default=1200) + 100

    sheet_divs = "\n".join(
        f'<div class="sheet-card" id="sheet-{s["id"]}" data-sheet-id="{s["id"]}" '
        f'style="left:{s["x"]}px;top:{s["y"]}px;width:{s["w"]}px;height:{s["h"]}px;">'
        f'<div class="sheet-title">{s["name"]}</div>'
        f'<div class="sheet-body">{s["svg"]}</div>'
        f'</div>'
        for s in sheets
    )

    sheet_positions = {
        s["id"]: {"x": s["x"], "y": s["y"], "w": s["w"], "h": s["h"],
                  "name": s["name"], "blocks": s.get("blocks", [])}
        for s in sheets
    }
    # Komponent highlight kutuları: {sheet_id: {designator: [x,y,w,h]}}
    # (SVG viewBox koordinatında tam komponent sınırı — full_bounds_mils'ten)
    sch_boxes = {}
    for c in components:
        for pl in c.get("placements", []):
            box = pl.get("sch_box")
            if box and pl.get("sheet_id"):
                sch_boxes.setdefault(pl["sheet_id"], {})[c["designator"]] = box

    return f"""<!DOCTYPE html>
<html lang="tr"><head>
<meta charset="utf-8">
<meta http-equiv="cache-control" content="no-cache">
<title>Schematic Viz · {timestamp}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ margin:0; background:#1a1a1a; color:#ddd;
          font-family: 'Consolas','Courier New', monospace;
          height:100vh; overflow:hidden; display:flex; user-select:none; }}
  #sidebar {{ width:320px; min-width:320px; background:#202020; padding:10px;
              overflow:hidden; display:flex; flex-direction:column;
              border-right:1px solid #333; position:relative;
              transition:width .18s ease, min-width .18s ease, padding .18s ease; }}
  /* Sol panel tamamen katlanabilir — dar bir şerit + ▸ butonu kalır */
  #sidebar.collapsed {{ width:26px; min-width:26px; padding:6px 2px; }}
  #sidebar.collapsed > *:not(#sidebar-toggle) {{ display:none !important; }}
  /* Küçük ok butonu — panelin sağ üst köşesinde */
  #sidebar-toggle {{ position:absolute; top:8px; right:8px; width:20px; height:20px;
                     padding:0; background:#1a1a1a; border:1px solid #333;
                     color:#888; font-family:inherit; font-size:11px;
                     line-height:18px; text-align:center; cursor:pointer;
                     border-radius:3px; z-index:5; }}
  #sidebar-toggle:hover {{ color:{inter_color}; border-color:{inter_color}; }}
  #sidebar.collapsed #sidebar-toggle {{ position:static; margin:0 auto;
                                        display:block; }}
  .stat {{ padding-right:26px; }}
  .stat {{ color:#666; font-size:11px; margin-bottom:8px; }}
  .tabs {{ display:flex; gap:4px; margin-bottom:6px; }}
  .tab {{ flex:1; background:#1a1a1a; border:1px solid #333; color:#888;
          padding:5px 8px; font-family:inherit; font-size:11px; cursor:pointer;
          border-radius:2px; text-transform:uppercase; letter-spacing:1px; }}
  .tab.active {{ background:#2a4a6a; color:{inter_color}; border-color:{inter_color}; }}
  /* Katlanabilir arama bölümü — varsayılan KAPALI ( / kısayolu açar) */
  #search-wrap {{ margin-bottom:6px; }}
  #search-toggle {{ width:100%; text-align:left; background:#1a1a1a;
                    border:1px solid #333; color:#888; padding:4px 8px;
                    font-family:inherit; font-size:11px; cursor:pointer;
                    border-radius:2px; }}
  #search-toggle:hover {{ color:{inter_color}; border-color:{inter_color}; }}
  #search-wrap.collapsed #search {{ display:none; }}
  #search {{ width:100%; padding:6px 8px; background:#111; border:1px solid #333;
             color:#fff; border-radius:2px; margin-top:4px; font-size:12px;
             font-family:inherit; outline:none; }}
  #search:focus {{ border-color:{inter_color}; }}
  .list-container {{ flex:1; overflow-y:auto; }}
  .list-container.hidden {{ display:none; }}
  .list-container::-webkit-scrollbar {{ width:8px; }}
  .list-container::-webkit-scrollbar-thumb {{ background:#444; border-radius:4px; }}
  .net-item, .comp-item {{ padding:3px 6px; cursor:pointer; font-size:11px;
                            border-radius:2px; white-space:nowrap;
                            overflow:hidden; text-overflow:ellipsis; }}
  .net-item {{ color:#bbb; }}
  .net-item.power {{ color:#ff8a65; }}
  .net-item.ground {{ color:#81c784; }}
  .net-item:hover, .comp-item:hover {{ background:#2a3a4a; color:#fff; }}
  .net-item.active {{ background:#2a4a6a; color:{inter_color}; font-weight:bold; }}
  .net-item.active.color-1 {{ background:#5a3015; color:#ff9800; }}
  .net-item.active.color-2 {{ background:#5a1535; color:#e91e63; }}
  .net-item.active.color-3 {{ background:#4a4a15; color:#ffeb3b; }}
  .badge {{ float:right; color:#666; font-size:10px; margin-left:4px; }}
  .comp-item .sheet-tag {{ float:right; color:#666; font-size:10px; }}
  #viewport {{ flex:1; position:relative; overflow:hidden; cursor:grab;
               background: radial-gradient(circle at 50% 50%, #252525, #151515); }}
  #viewport.grabbing {{ cursor:grabbing; }}
  /* Pan sırasında SVG hit-testing kapalı: Chromium her mousemove'da binlerce
     SVG elemanını hit-test edip :hover stil değişimleriyle repaint tetikliyor.
     Sınıf mousedown'da değil GERÇEK harekette eklenir (panMoved eşiği) —
     hareketsiz tıklamanın hedef elemanı değişmez. */
  #viewport.panning .sheet-body svg {{ pointer-events:none; }}
  #canvas {{ position:absolute; transform-origin:0 0; left:0; top:0;
             width:{max_x}px; height:{max_y}px; will-change:transform; }}
  /* contain: hover/highlight repaint'ini tek karta sınırlar (tüm kanvası boyatmaz) */
  .sheet-card {{ position:absolute; background:#fff; overflow:hidden;
                 contain:layout paint;
                 box-shadow: 0 0 0 1px #444, 0 6px 24px rgba(0,0,0,0.6);
                 display:flex; flex-direction:column; }}
  .sheet-title {{ height:30px; background:#252525; color:#ccc; padding:6px 10px;
                   font-size:14px; border-bottom:1px solid #3a3a3a; text-align:center;
                   flex-shrink:0; line-height:18px; }}
  .sheet-body {{ flex:1; overflow:hidden; background:#fff; position:relative; }}
  .sheet-body svg {{ width:100%; height:100%; display:block; }}
  /* LOD: uzak zoom'da sayfa yerine bir kez üretilmiş bitmap gösterilir (bkz.
     buildLods JS'i). SVG display:none DEĞİL visibility:hidden ile gizlenir —
     highlight/arama getBoundingClientRect ölçümleri çalışmaya devam eder. */
  .lod-bitmap {{ position:absolute; left:0; top:0; width:100%; height:100%;
                 display:none; }}
  #canvas.lod .sheet-body.lod-ready svg {{ visibility:hidden; }}
  /* lod-fade: bitmap→SVG dönüşünde bitmap ~160ms daha ÜSTTE kalır (SVG sonra
     eklendiği için değil — bitmap DOM'da svg'den sonra) → Chromium SVG
     karolarını bitmap'in arkasında rasterize eder, beyaz parlama görünmez. */
  #canvas.lod .sheet-body.lod-ready .lod-bitmap,
  #canvas.lod-fade .sheet-body.lod-ready .lod-bitmap {{ display:block; }}
  /* Şematik metinleri PDF'teki gibi seçilebilir/kopyalanabilir (body user-select:none
     bunu global kapatıyor; text elemanlarında geri açılır). Tıklanabilir sınıflar
     (net/block/designator) pointer imlecini korur — hem seçilir hem tıklanır. */
  .sheet-body svg text {{ user-select:text; -webkit-user-select:text; cursor:text; }}
  .sheet-body svg text::selection {{ background:#4ec9b0; color:#000; }}
  .sheet-body svg text.clickable-net, .sheet-body svg text.block-link,
  .sheet-body svg text.comp-designator {{ cursor:pointer; }}
  .sheet-card.hit {{ box-shadow: 0 0 0 3px {inter_color}, 0 0 40px rgba(0,0,0,0.5); }}
  .sheet-card.hit-1 {{ box-shadow: 0 0 0 3px #ff9800, 0 0 40px rgba(255,152,0,0.5); }}
  .sheet-card.hit-2 {{ box-shadow: 0 0 0 3px #e91e63, 0 0 40px rgba(233,30,99,0.5); }}
  .sheet-card.hit-3 {{ box-shadow: 0 0 0 3px #ffeb3b, 0 0 40px rgba(255,235,59,0.5); }}
  #arc-layer {{ position:absolute; top:0; left:0; pointer-events:none;
                overflow:visible; z-index:100; }}
  /* === Not / kutu (annotation) katmanı — JS'te kurulur, kanvasla transform olur === */
  #anno-layer {{ position:absolute; top:0; left:0; overflow:visible; z-index:120; }}
  #anno-layer .anno {{ pointer-events:auto; cursor:move; }}
  #anno-layer text {{ font-family:'Consolas','Courier New',monospace;
                      user-select:none; -webkit-user-select:none; }}
  /* Kutu içi tıklamayı yutmasın: görünür rect tıklanmaz, geniş görünmez
     hit-rect'in yalnız KENARI tıklanır (ince kenarlıkta da seçilebilsin) */
  #anno-layer .anno-box rect {{ pointer-events:none; }}
  #anno-layer .anno-box rect.anno-hit {{ pointer-events:stroke; }}
  #anno-layer .anno-sel-rect {{ pointer-events:none; }}
  /* rect.anno-handle: .anno-box rect kuralıyla özgüllük eşit, SONRA geldiği
     için kazanır (tutamaçlar tıklanabilir kalır) */
  #anno-layer rect.anno-handle {{ pointer-events:all; }}
  #anno-layer .anno-handle[data-c="nw"],
  #anno-layer .anno-handle[data-c="se"] {{ cursor:nwse-resize; }}
  #anno-layer .anno-handle[data-c="ne"],
  #anno-layer .anno-handle[data-c="sw"] {{ cursor:nesw-resize; }}
  /* Araç aktifken: crosshair imleç, şema SVG'leri ve mevcut notlar tıklamaya kapalı
     (yeni not/kutu mevcutların üstüne de konabilsin) */
  #viewport.anno-mode {{ cursor:crosshair; }}
  #viewport.anno-mode .sheet-body svg {{ pointer-events:none; }}
  #viewport.anno-mode #anno-layer .anno {{ pointer-events:none; }}
  /* Yerinde yazma editörü (Foxit typewriter gibi) — canvas içinde, onunla ölçeklenir.
     white-space:pre → otomatik sarma YOK (SVG render'ı ile birebir; satır = Enter).
     Arka plan hafif saydam (not artık kutusuz çıplak yazı — v2.9.40); yazı
     rengi notun kendi rengiyle inline verilir. */
  #anno-editor {{ position:absolute; z-index:130; min-width:30px;
                  background:rgba(255,255,255,0.72); border:1px dashed #999;
                  border-radius:3px; padding:5px 8px; line-height:1.3;
                  outline:none; white-space:pre;
                  font-family:'Consolas','Courier New',monospace;
                  user-select:text; -webkit-user-select:text; }}
  /* Seçim mini araç çubuğu (ekran uzayında, seçili öğenin üstünde) */
  #anno-bar {{ position:absolute; z-index:600; display:flex; gap:2px;
               background:rgba(30,30,30,0.92); border:1px solid #444;
               border-radius:4px; padding:2px; }}
  #anno-bar button {{ background:none; border:1px solid transparent; color:#ccc;
                      width:22px; height:22px; font-family:inherit; font-size:13px;
                      line-height:1; cursor:pointer; border-radius:3px; padding:0; }}
  #anno-bar button:hover {{ border-color:#4ec9b0; color:#4ec9b0; }}
  #anno-bar input[type=color] {{ width:22px; height:22px; padding:1px;
                                 border:1px solid transparent; background:none;
                                 cursor:pointer; border-radius:3px; }}
  #anno-bar input[type=color]:hover {{ border-color:#4ec9b0; }}
  .clickable-net {{ cursor:pointer; }}
  .clickable-net:hover {{ fill:#ff6b35 !important; font-weight:bold; }}
  /* Block link - .SchDoc filename'ine tıklanınca o sayfaya gider */
  .block-link {{ cursor:pointer; }}
  .block-link:hover {{ fill:{inter_color} !important; font-weight:bold;
                       text-decoration:underline; }}
  /* Komponent designator'ları - tıklayınca popup açılır */
  .comp-designator {{ cursor:pointer; }}
  .comp-designator:hover {{ fill:#ffeb3b !important; font-weight:bold; }}
  /* Komponent detay popup'ı */
  /* Komponent detayı: sağda yüzen popup yerine sol sidebar'a dock edilmiş,
     katlanabilir (ok) + boyutlandırılabilir (üst tutamaç) panel. */
  #comp-popup {{ display:none; flex-direction:column; flex-shrink:0;
                  margin-top:8px; background:#1a1a1a; color:#ddd;
                  border:1px solid #3a3a3a; border-radius:4px;
                  height:300px; min-height:30px; max-height:62%;
                  font-size:11px; overflow:hidden; }}
  #comp-popup.open {{ display:flex; }}
  #comp-popup.collapsed {{ height:auto !important; }}
  #comp-popup.collapsed .popup-body,
  #comp-popup.collapsed #popup-resize {{ display:none; }}
  #popup-resize {{ height:7px; cursor:ns-resize; background:#262626;
                   flex-shrink:0; border-bottom:1px solid #333;
                   border-radius:4px 4px 0 0; }}
  #popup-resize:hover {{ background:{inter_color}; }}
  .popup-collapse {{ background:transparent; border:none; color:#bcd;
                     font-size:13px; cursor:pointer; padding:0 2px; line-height:1; }}
  .popup-collapse:hover {{ color:#fff; }}
  .popup-header {{ padding:8px 10px; background:#2a4a6a;
                    display:flex; align-items:center; gap:6px;
                    border-bottom:1px solid #1a3a5a; flex-shrink:0; }}
  .popup-header > span {{ flex:1; overflow:hidden; text-overflow:ellipsis;
                          white-space:nowrap; }}
  .popup-title {{ color:{inter_color}; font-weight:bold; font-size:13px;
                   font-family:'Consolas',monospace; }}
  .popup-sheet {{ color:#888; font-size:10px; margin-left:6px; }}
  .popup-close {{ background:transparent; border:none; color:#aaa;
                   font-size:18px; cursor:pointer; padding:0 4px;
                   line-height:1; }}
  .popup-close:hover {{ color:#fff; }}
  .popup-body {{ padding:8px 12px; overflow-y:auto; }}
  .popup-body::-webkit-scrollbar {{ width:6px; }}
  .popup-body::-webkit-scrollbar-thumb {{ background:#444; border-radius:3px; }}
  .popup-section-title {{ color:#666; font-size:10px; text-transform:uppercase;
                           letter-spacing:1px; margin:8px 0 4px; }}
  .popup-section-title:first-child {{ margin-top:0; }}
  .popup-row {{ display:flex; padding:3px 0; border-bottom:1px solid #2a2a2a;
                 line-height:1.4; align-items:flex-start; }}
  .popup-row:last-child {{ border-bottom:none; }}
  .popup-key {{ color:#888; min-width:110px; padding-right:8px;
                 user-select:text; }}
  .popup-val {{ color:#ddd; flex:1; word-break:break-word;
                 user-select:text; }}
  .popup-link {{ color:{inter_color}; text-decoration:none; }}
  .popup-link:hover {{ text-decoration:underline; color:#7fffd4; }}
  .popup-copy {{ background:transparent; border:none; color:#555;
                  font-size:13px; padding:0 4px; margin-left:6px;
                  cursor:pointer; line-height:1; flex-shrink:0;
                  font-family:inherit; }}
  .popup-copy:hover {{ color:{inter_color}; }}
  .popup-copy.copied {{ color:{inter_color}; }}
  /* Komponent highlight: designator text'i camgöbeği (PCB highlight ile uyumlu) */
  @keyframes comp-pulse {{
    0%, 100% {{ fill: #00e5ff; }}
    50% {{ fill: #7af6ff; }}
  }}
  .comp-highlight {{
    animation: comp-pulse 1.1s ease-in-out infinite;
    font-weight: bold !important;
    filter: drop-shadow(0 0 2px #00e5ff) drop-shadow(0 0 6px rgba(0,229,255,0.6));
  }}
  /* Seçili komponent kutusu (PCB'deki #hl-marker ile aynı stil). Çizgi/yazı
     kalınlığı JS'te 1/scale ile ölçeklenir (CSS transform zoom'da ekran-sabit). */
  #sch-hl-overlay {{ position:absolute; top:0; left:0; pointer-events:none;
                     overflow:visible; z-index:150; }}
  #sch-hl-overlay .hl-box {{ fill:none; stroke:#00e5ff; stroke-linejoin:round;
    animation:hlpulse 1.1s ease-in-out infinite; }}
  #sch-hl-overlay .hl-text {{ fill:#00e5ff; font-family:Consolas,monospace;
    font-weight:bold; paint-order:stroke; stroke:#04222a; stroke-linejoin:round; }}
  @keyframes hlpulse {{ 0%,100%{{stroke-opacity:0.5}} 50%{{stroke-opacity:1}} }}
  #toolbar {{ position:absolute; top:8px; right:8px; display:flex; gap:6px; z-index:200; }}
  #sheet-jump {{ max-width:150px; appearance:auto; font-family:inherit; }}
  /* Fare ile üzerine gelince bilgi balonu (komponent/net/block) */
  #svg-tip {{ position:fixed; display:none; background:#111; color:#ddd;
              border:1px solid {inter_color}; border-radius:4px; padding:5px 9px;
              font-size:11px; max-width:340px; z-index:900; pointer-events:none;
              line-height:1.5; box-shadow:0 4px 14px rgba(0,0,0,0.5);
              font-family:'Consolas',monospace; }}
  #svg-tip .tt-title {{ color:{inter_color}; font-weight:bold; }}
  #svg-tip .tt-hint {{ color:#666; }}
  /* Net tipi filtre çipleri (Nets sekmesi) */
  #type-chips {{ display:flex; gap:4px; margin-bottom:6px; }}
  #type-chips.hidden {{ display:none; }}
  .chip {{ flex:1; background:#1a1a1a; border:1px solid #333; color:#777;
           font-family:inherit; font-size:10px; padding:3px 0; cursor:pointer;
           border-radius:10px; }}
  .chip:hover {{ color:#ddd; }}
  .chip.active {{ border-color:{inter_color}; color:{inter_color}; }}
  .chip-power.active {{ border-color:#ff8a65; color:#ff8a65; }}
  .chip-ground.active {{ border-color:#81c784; color:#81c784; }}
  .empty-msg {{ color:#555; font-size:11px; padding:10px 6px; font-style:italic; }}
  .tool-btn {{ background:rgba(40,40,40,0.95); color:#ccc; border:1px solid #444;
               padding:5px 10px; cursor:pointer; font-family:inherit; font-size:11px;
               border-radius:3px; }}
  .tool-btn:hover {{ background:#333; color:{inter_color}; border-color:{inter_color}; }}
  .tool-btn.active {{ background:#2a4a6a; border-color:{inter_color}; color:{inter_color}; }}
  #zoom-info {{ position:absolute; bottom:8px; left:50%; transform:translateX(-50%);
                color:#666; font-size:11px; }}
  #brand {{ position:absolute; bottom:8px; right:12px; color:#444; font-size:10px; }}
  #current-net {{ position:absolute; top:8px; left:50%; transform:translateX(-50%);
                  color:{inter_color}; font-size:13px; font-weight:bold;
                  background:rgba(0,0,0,0.6); padding:4px 10px; border-radius:3px;
                  max-width:60%; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
  #detail-panel {{ position:absolute; bottom:0; left:0; right:0;
                   max-height:180px; background:rgba(20,20,20,0.97);
                   border-top:1px solid #444; overflow-y:auto; padding:8px 12px;
                   display:none; z-index:150; }}
  #detail-panel.open {{ display:block; }}
  .net-detail {{ margin-bottom:6px; }}
  .net-detail .net-name {{ font-weight:bold; font-size:12px; }}
  .net-detail .sheet-row {{ display:inline-block; padding:2px 8px; margin:2px;
                             background:#2a2a2a; cursor:pointer; border-radius:2px;
                             font-size:11px; color:#bbb; }}
  .net-detail .sheet-row:hover {{ background:#3a4a5a; color:{inter_color}; }}
  #shortcuts {{ position:absolute; bottom:8px; left:8px; color:#444; font-size:10px;
                background:rgba(0,0,0,0.4); padding:4px 8px; border-radius:3px; }}
  /* Renk picker'lar (toolbar) */
  .color-input {{ width:26px; height:26px; border:1px solid #444; border-radius:3px;
                   cursor:pointer; overflow:hidden; padding:0; background:transparent;
                   display:flex; align-items:center; justify-content:center; }}
  .color-input:hover {{ border-color:#888; }}
  .color-input input[type="color"] {{ width:200%; height:200%; border:none; padding:0;
                                       cursor:pointer; transform:translate(-25%, -25%); }}
  .toolbar-sep {{ width:1px; background:#444; margin:2px 4px; align-self:stretch; }}
  /* Kısayol modal */
  #shortcut-modal {{ position:fixed; inset:0; background:rgba(0,0,0,0.75); z-index:1000;
                     display:none; align-items:center; justify-content:center; }}
  #shortcut-modal.open {{ display:flex; }}
  .modal-content {{ background:#2a2a2a; color:#ddd; padding:24px 28px;
                     border-radius:6px; min-width:380px; max-width:520px;
                     box-shadow:0 8px 40px rgba(0,0,0,0.6); }}
  .modal-content h3 {{ margin:0 0 10px; color:{inter_color}; font-size:13px;
                        text-transform:uppercase; letter-spacing:1.5px; font-weight:normal; }}
  .modal-content h3:not(:first-child) {{ margin-top:16px; }}
  .modal-content table {{ width:100%; border-collapse:collapse; }}
  .modal-content td {{ padding:5px 8px; font-size:12px; vertical-align:middle; }}
  .modal-content td:first-child {{ width:140px; }}
  .modal-content kbd {{ background:#1a1a1a; padding:2px 7px; border-radius:3px;
                         border:1px solid #555; font-family:'Consolas',monospace;
                         font-size:11px; color:#fff; }}
  .modal-close {{ margin-top:12px; padding:6px 14px; background:#1a1a1a; color:#ddd;
                   border:1px solid #555; border-radius:3px; cursor:pointer;
                   font-family:inherit; }}
  .modal-close:hover {{ border-color:{inter_color}; color:{inter_color}; }}
</style>
</head><body>
<aside id="sidebar">
  <button id="sidebar-toggle" title="Paneli gizle ( B )">◂</button>
  <div class="stat">{len(sheets)} sheets · {len(net_list)} nets · {len(components)} comps</div>
  <div class="tabs">
    <button class="tab active" data-tab="nets">Nets</button>
    <button class="tab" data-tab="components">Comps</button>
  </div>
  <div id="type-chips">
    <button class="chip active" data-type="">Tümü</button>
    <button class="chip chip-power" data-type="power">Güç</button>
    <button class="chip chip-ground" data-type="ground">GND</button>
    <button class="chip" data-type="signal">Sinyal</button>
  </div>
  <div id="search-wrap" class="collapsed">
    <button id="search-toggle" title="Aramayı aç/kapat ( / )"><span id="search-caret">▸</span> Ara</button>
    <input id="search" placeholder="ara... ( / )">
  </div>
  <div id="nets-list" class="list-container"></div>
  <div id="comps-list" class="list-container hidden"></div>
  <div id="comp-popup">
    <div id="popup-resize" title="Sürükle: yeniden boyutlandır"></div>
    <div class="popup-header">
      <button class="popup-collapse" id="popup-collapse" title="Küçült / Büyüt">▾</button>
      <span><span class="popup-title" id="popup-title"></span><span class="popup-sheet" id="popup-sheet"></span></span>
      <button class="popup-close" id="popup-close" title="Kapat">×</button>
    </div>
    <div class="popup-body" id="popup-body"></div>
  </div>
</aside>
<div id="viewport">
  <div id="canvas">
    <svg id="arc-layer" width="{max_x}" height="{max_y}"></svg>
    {sheet_divs}
  </div>
  <div id="toolbar">
    <select id="sheet-jump" class="tool-btn" title="Sayfaya git"></select>
    <button class="tool-btn" id="zoom-in" title="Yaklaş ( + )">+</button>
    <button class="tool-btn" id="zoom-out" title="Uzaklaş ( − )">−</button>
    <button class="tool-btn" id="fit-all" title="Tüm sayfaları sığdır">Tümü</button>
    <button class="tool-btn active" id="lod-toggle"
            title="LOD: uzak zoom'da ve gezinirken sayfalar bitmap çizilir (Chromium'da akıcılık). Kapatınca her zaman canlı SVG.">LOD</button>
    <div class="toolbar-sep"></div>
    <label class="color-input" title="Sayfalar arası yay rengi">
      <input type="color" id="inter-color-picker" value="{inter_color}">
    </label>
    <label class="color-input" title="Sayfa içi eğri rengi">
      <input type="color" id="intra-color-picker" value="{intra_color}">
    </label>
    <div class="toolbar-sep"></div>
    <button class="tool-btn" id="anno-note"
            title="Not ekle: butona bas, şemada istediğin yere tıkla ve DOĞRUDAN yaz (dışına tıkla = bitir, Enter = yeni satır). Sonradan: çift tık düzenle · sürükle taşı · seç + Del sil · A−/A+ yazı boyutu">Not</button>
    <button class="tool-btn" id="anno-box"
            title="Kutu içine al: butona bas, sürükleyerek çerçeve çiz (Esc iptal). Sonradan: kenarına tıkla seç → sürükle taşı · köşe tutamaçlarıyla boyutlandır · Del sil · −/+ kenar kalınlığı">Kutu</button>
    <button class="tool-btn" id="anno-save"
            title="Not ve kutuları HTML dosyasının içine göm ve kaydet. Chromium'da AÇIK DOSYANIN ÜSTÜNE yazabilir (ilk kayıtta dosyayı seç; aynı oturumda sonrakiler sessiz). Firefox'ta kopya indirir. Paylaşınca/başka bilgisayarda da görünür">Kaydet</button>
    <div class="toolbar-sep"></div>
    <button class="tool-btn" id="shortcut-btn" title="Kısayollar (?)">?</button>
    <button class="tool-btn" id="export-png">PNG</button>
    <button class="tool-btn" id="reset-view">Reset</button>
    <button class="tool-btn" id="clear-sel">Clear</button>
  </div>
  <div id="current-net"></div>
  <div id="zoom-info">Zoom <span id="zoom-val">0.30x</span></div>
  <div id="shortcuts">Esc clear · / search · 0 reset · F fit · <kbd style="background:#1a1a1a;padding:1px 4px;border:1px solid #555;border-radius:2px;color:#aaa">?</kbd> tüm kısayollar</div>
  <div id="brand">altium_monkey</div>
  <div id="detail-panel"><div id="detail-content"></div></div>
</div>
<div id="svg-tip"></div>

<div id="shortcut-modal">
  <div class="modal-content">
    <h3>Klavye</h3>
    <table>
      <tr><td><kbd>Esc</kbd></td><td>Seçimi temizle</td></tr>
      <tr><td><kbd>/</kbd></td><td>Arama kutusuna git</td></tr>
      <tr><td><kbd>Enter</kbd></td><td>Aramada ilk sonucu seç</td></tr>
      <tr><td><kbd>B</kbd></td><td>Sol paneli gizle / göster</td></tr>
      <tr><td><kbd>0</kbd></td><td>Görünümü sıfırla</td></tr>
      <tr><td><kbd>F</kbd></td><td>Son sayfaya fit zoom</td></tr>
      <tr><td><kbd>+</kbd> / <kbd>-</kbd></td><td>Zoom in / out</td></tr>
      <tr><td><kbd>?</kbd></td><td>Bu pencereyi aç / kapat</td></tr>
    </table>
    <h3>Fare</h3>
    <table>
      <tr><td>Drag</td><td>Kanvası kaydır (pan)</td></tr>
      <tr><td>Wheel</td><td>Mouse altına zoom</td></tr>
      <tr><td>Sayfa kartına çift tık</td><td>O sayfayı ekrana sığdır</td></tr>
      <tr><td>Net adına tık (şema/sol panel)</td><td>Net seç, bağlantıları göster</td></tr>
      <tr><td>Shift + tık</td><td>Çoklu net karşılaştırma (max 4)</td></tr>
      <tr><td>Comps listesinde tık</td><td>Komponente zoom + pulse + detay popup</td></tr>
      <tr><td>Designator'a tık (şema)</td><td>Komponent detay popup'ı aç</td></tr>
      <tr><td>Block (.SchDoc) yazısına tık</td><td>O sayfaya navigate et</td></tr>
      <tr><td>Toolbar: Not / Kutu</td><td>Tıklanan yere doğrudan yazı yaz / alanı kutu içine al (Esc iptal)</td></tr>
      <tr><td>Not/kutuya tık + sürükle</td><td>Seç ve taşı · kutuda köşe tutamacı: boyutlandır</td></tr>
      <tr><td>Seçiliyken Del · mini bar −/+</td><td>Sil · yazı boyutu / kenar kalınlığı</td></tr>
      <tr><td>Nota çift tık</td><td>Yerinde düzenle (boş bırak = sil)</td></tr>
    </table>
    <h3>Renk Pickers</h3>
    <table>
      <tr><td>Toolbar'daki renkli kareler</td><td>Yay renklerini anlık değiştir</td></tr>
    </table>
    <button class="modal-close" id="close-modal-btn">Kapat</button>
  </div>
</div>

<script src="https://html2canvas.hertzen.com/dist/html2canvas.min.js"></script>
<!-- "Kaydet" ile indirilen kopyada notlar buraya JSON olarak gömülür
     (type=application/json → tarayıcı script olarak ÇALIŞTIRMAZ, veri taşır) -->
<script type="application/json" id="anno-embed">null</script>
<script>
const BUILD_STAMP = "{timestamp}";
const PROJECT_NAME = {json.dumps(project_name)};
const nets = {json.dumps(net_list)};
const components = {json.dumps(components)};
const sheetPos = {json.dumps(sheet_positions)};
const SCH_BOXES = {json.dumps(sch_boxes)};
const PCB = {json.dumps({
    "available": pcb.get("available", False),
    "components": pcb.get("components", {}),
    "board_w_mm": pcb.get("board_w_mm", 0),
    "board_h_mm": pcb.get("board_h_mm", 0),
    "pcb_name": pcb.get("pcb_name", ""),
})};
const NET_COLORS = ['{inter_color}', '#ff9800', '#e91e63', '#ffeb3b'];
let INTRA_COLOR = '{intra_color}';

document.querySelectorAll('.sheet-body svg').forEach(svg => {{
  svg.setAttribute('preserveAspectRatio', 'none');
}});

const viewport = document.getElementById('viewport');
const canvas = document.getElementById('canvas');
const zoomVal = document.getElementById('zoom-val');
const arcLayer = document.getElementById('arc-layer');
const currentNetEl = document.getElementById('current-net');
const detailPanel = document.getElementById('detail-panel');
const detailContent = document.getElementById('detail-content');

let tx = 40, ty = 40, scale = 0.30;
// Komponent vurgu kutusu için kalıcı overlay (arcLayer net render'larında
// temizlendiğinden ayrı katman). canvas ile birlikte transform olur.
const schHlOverlay = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
schHlOverlay.setAttribute('id', 'sch-hl-overlay');
schHlOverlay.setAttribute('width', arcLayer.getAttribute('width') || 10000);
schHlOverlay.setAttribute('height', arcLayer.getAttribute('height') || 10000);
canvas.appendChild(schHlOverlay);
let schMarkerBox = null;   // {{x,y,w,h,label}}

// === LOD: uzak zoom'da sayfa bitmap'leri ==================================
// Chromium CSS-scale edilen dev SVG katmanını HER zoom adımında CPU'da yeniden
// rasterize eder (Firefox/WebRender vektörü GPU'da çizer — akıcılık farkı
// buradan). Uzak zoom'da her sayfanın yerine bir kez üretilen bitmap (canvas)
// gösterilir: bitmap'i kaydırıp ölçeklemek compositor'da bedavaya yakındır.
// LOD_OFF üstüne yakınlaşınca canlı SVG'ye dönülür — tıklama/metin seçimi/hover
// zaten o zoom'da yapılır. Bitmap üretimi idle'da sayfa sayfa yapılır; üretim
// bitene kadar (ve üretilemeyen sayfalarda) canlı SVG kalır = eski davranış.
const LOD_ON = 0.85, LOD_OFF = 1.05;  // histerezis: girişte <ON, çıkışta >OFF
// Etkileşim (pan sürükleme / tekerlek zoom serisi) SIRASINDA bitmap, yakın
// zoom'da da gösterilir (harita uygulaması deseni: harekette yumuşak ama
// akıcı, durunca keskin canlı SVG). Üst sınır: çok aşırı zoom'da bitmap
// çirkinleşir + görünür alan zaten küçük olduğundan canlı SVG akıcıdır.
const LOD_MAX_I = 4;
// Bitmap çözünürlüğü: LOD aralığının tepesinde (scale≈1) ekran pikseliyle
// eşleşsin diye DPR kadar; alt sınır 1.25 (etkileşim bitmap'i okunur kalsın),
// bellek için 1.6 ile sınırlı (kart 700×470 → ~2-3MB/sayfa).
const LOD_RES = Math.min(1.6, Math.max(1.25, window.devicePixelRatio || 1));
let lodActive = false, lodReady = 0, lodFadeT = null, lodEnabled = true;
let panInteract = false, wheelInteract = false, wheelIdleT = null;
function updateLod() {{
  if (!lodReady) return;
  const rest = lodActive ? (scale < LOD_OFF) : (scale < LOD_ON);
  const want = lodEnabled &&
    (rest || ((panInteract || wheelInteract) && scale <= LOD_MAX_I));
  if (want === lodActive) return;
  lodActive = want;
  if (want) {{
    clearTimeout(lodFadeT); canvas.classList.remove('lod-fade');
    canvas.classList.add('lod');
  }} else {{
    // Bitmap'i hemen söndürme: SVG görünür olur, bitmap 160ms üstte kalır →
    // SVG karoları arkada rasterize edilir (bkz. .lod-fade CSS yorumu).
    canvas.classList.remove('lod');
    canvas.classList.add('lod-fade');
    clearTimeout(lodFadeT);
    lodFadeT = setTimeout(() => canvas.classList.remove('lod-fade'), 160);
  }}
}}
// Tekerlek zoom serisi: her event sayacı tazeler; 180ms sessizlik = seri bitti.
function lodWheelTouch() {{
  wheelInteract = true;
  updateLod();
  clearTimeout(wheelIdleT);
  wheelIdleT = setTimeout(() => {{ wheelInteract = false; updateLod(); }}, 180);
}}
// Toolbar LOD toggle'ı — kapatınca her zoom'da canlı SVG (tercih localStorage'da,
// restoreUi geri yükler). Bitmap'ler yine üretilir ki açınca anında çalışsın.
const lodBtn = document.getElementById('lod-toggle');
function setLodEnabled(on) {{
  lodEnabled = on;
  lodBtn.classList.toggle('active', on);
  updateLod();
  lsSet({{ lod: on }});
}}
lodBtn.addEventListener('click', () => setLodEnabled(!lodEnabled));
(function buildLods() {{
  const bodies = Array.from(document.querySelectorAll('.sheet-body'));
  const idle = window.requestIdleCallback
    ? (f => window.requestIdleCallback(f, {{ timeout: 800 }}))
    : (f => setTimeout(f, 150));
  let i = 0;
  function step() {{
    if (i >= bodies.length) return;
    const body = bodies[i++];
    const svgEl = body.querySelector('svg');
    if (!svgEl) {{ idle(step); return; }}
    const w = Math.max(1, Math.round(body.clientWidth * LOD_RES));
    const h = Math.max(1, Math.round(body.clientHeight * LOD_RES));
    const src = new XMLSerializer().serializeToString(svgEl);
    const img = new Image();
    let url = '', triedData = false;
    const done = ok => {{
      if (url) {{ URL.revokeObjectURL(url); url = ''; }}
      if (ok) {{
        try {{
          const cv = document.createElement('canvas');
          cv.width = w; cv.height = h;
          // preserveAspectRatio="none" → hedef boyuta gerilir, ekranla birebir
          cv.getContext('2d').drawImage(img, 0, 0, w, h);
          cv.className = 'lod-bitmap';
          body.appendChild(cv);
          body.classList.add('lod-ready');
          lodReady++; updateLod();
        }} catch (e) {{ /* bu sayfa canlı SVG'de kalır */ }}
      }}
      idle(step);
    }};
    img.onload = () => done(true);
    img.onerror = () => {{
      // blob: bazı ortamlarda engellenebilir → data: URI ile bir kez daha dene
      if (triedData) {{ done(false); return; }}
      triedData = true;
      if (url) {{ URL.revokeObjectURL(url); url = ''; }}
      img.src = 'data:image/svg+xml;charset=utf-8,' + encodeURIComponent(src);
    }};
    url = URL.createObjectURL(new Blob([src], {{ type: 'image/svg+xml;charset=utf-8' }}));
    img.src = url;
  }}
  idle(step);
}})();

function applyT() {{
  canvas.style.transform = `translate(${{tx}}px,${{ty}}px) scale(${{scale}})`;
  zoomVal.textContent = scale.toFixed(2) + 'x';
  updateSchMarkerMetrics();
  // Not/kutu seçim görselleri + mini bar (modül aşağıda kurulur; var hoisting
  // sayesinde ilk applyT çağrılarında typeof güvenle 'undefined' döner)
  if (typeof __annoUi === 'function') __annoUi();
  updateLod();
}}
applyT();
// Programatik görünüm değişimlerinde (fit/reset/zoom butonu) yumuşak geçiş.
// Fare tekerleği/pan doğrudan applyT kullanır (gecikme hissi olmasın).
function smoothT() {{
  canvas.style.transition = 'transform 0.35s ease';
  applyT();
  clearTimeout(smoothT._t);
  smoothT._t = setTimeout(() => {{ canvas.style.transition = 'none'; }}, 400);
}}

function classifyNet(name) {{
  const n = name.toUpperCase();
  if (/^A?GND/.test(n) || n === 'VSS' || n === 'AGND' || n === 'DGND' || /GND$/.test(n)) return 'ground';
  if (/^[+\\-]?\\d*\\.?\\d+V/.test(n) || /^V(CC|DD|IN|OUT|BAT|REF)/.test(n) || n === 'VCC' || n === 'VDD') return 'power';
  return 'signal';
}}

let panning = false, sx, sy, stx, sty, panMoved = false;
viewport.addEventListener('mousedown', e => {{
  if (e.target.closest('.tool-btn') || e.target.closest('#detail-panel')) return;
  if (annoTool) return;   // not/kutu aracı aktif — pan yerine araç çalışır
  // SVG metni üzerinde pan BAŞLATMA → tarayıcının native metin seçimi çalışsın
  // (PDF'teki gibi sürükleyip kopyalama). Boş alanda pan aynen devam eder.
  if (e.target.closest && e.target.closest('.sheet-body') && e.target.closest('text')) return;
  panning = true; panMoved = false; sx = e.clientX; sy = e.clientY; stx = tx; sty = ty;
  viewport.classList.add('grabbing');
}});
window.addEventListener('mousemove', e => {{
  if (!panning) return;
  if (Math.abs(e.clientX - sx) > 3 || Math.abs(e.clientY - sy) > 3) {{
    // Gerçek pan başladı: SVG hit-testing'i kapat (bkz. #viewport.panning CSS'i)
    // + etkileşim boyunca bitmap moduna geç (akıcı sürükleme).
    if (!panMoved) {{ panMoved = true; viewport.classList.add('panning');
                      svgTip.style.display = 'none';
                      panInteract = true; updateLod(); }}
  }}
  tx = stx + (e.clientX - sx);
  ty = sty + (e.clientY - sy);
  applyT();
}});
window.addEventListener('mouseup', () => {{
  panning = false; viewport.classList.remove('grabbing', 'panning');
  if (panInteract) {{ panInteract = false; updateLod(); }}
}});
// Tekerlek zoom'u rAF ile birleştirilir: Chromium her scale değişiminde görünür
// karoları yeniden rasterize eder; yüksek çözünürlüklü tekerlek/trackpad kare
// başına birden çok event üretebildiğinden çarpanlar wheelF'te biriktirilip
// kare başına TEK transform uygulanır (tek event/kare durumunda aynı matematik).
let wheelF = 1, wheelPend = false, wheelMx = 0, wheelMy = 0;
viewport.addEventListener('wheel', e => {{
  e.preventDefault();
  lodWheelTouch();   // zoom serisi boyunca bitmap modu (akıcı tekerlek)
  const r = viewport.getBoundingClientRect();
  wheelMx = e.clientX - r.left; wheelMy = e.clientY - r.top;
  wheelF *= e.deltaY < 0 ? 1.15 : 0.87;
  if (wheelPend) return;
  wheelPend = true;
  requestAnimationFrame(() => {{
    wheelPend = false;
    const old = scale;
    scale = Math.max(0.03, Math.min(8, scale * wheelF));
    wheelF = 1;
    tx = wheelMx - (wheelMx - tx) * (scale / old);
    ty = wheelMy - (wheelMy - ty) * (scale / old);
    applyT();
  }});
}}, {{ passive: false }});

// Boş alana tıklama = komponent seçimini (spotlight kutusunu) iptal et.
// Pan hareketi, metin seçimi, toolbar ve tıklanabilir öğeler hariç tutulur.
viewport.addEventListener('click', e => {{
  if (panMoved) return;                                    // pan bitişi, tıklama değil
  if (annoTool || annoJustDrew) return;                    // not/kutu yerleştiriliyor
  if (!schMarkerBox) return;                               // seçim yoksa iş yok
  if (e.target.closest('.tool-btn') || e.target.closest('#detail-panel')
      || e.target.closest('#toolbar')) return;
  const cl = e.target.classList;
  if (cl && (cl.contains('clickable-net') || cl.contains('block-link')
             || cl.contains('comp-designator'))) return;   // yeni seçim yapılıyor
  if (window.getSelection && String(window.getSelection()).length > 0) return; // metin kopyalama
  clearCompHighlight();
}});

function fitToSheet(sheetId) {{
  const sp = sheetPos[sheetId];
  if (!sp) return;
  const r = viewport.getBoundingClientRect();
  const padding = 0.9;
  scale = Math.min(r.width * padding / sp.w, r.height * padding / sp.h);
  tx = r.width / 2 - (sp.x + sp.w / 2) * scale;
  ty = r.height / 2 - (sp.y + sp.h / 2) * scale;
  smoothT();
}}
function resetView() {{ tx = 40; ty = 40; scale = 0.30; smoothT(); }}
// Görünüm merkezinde yakınlaş/uzaklaş (toolbar +/− butonları)
function zoomBy(f) {{
  const r = viewport.getBoundingClientRect();
  const mx = r.width / 2, my = r.height / 2, old = scale;
  scale = Math.max(0.03, Math.min(8, scale * f));
  tx = mx - (mx - tx) * (scale / old);
  ty = my - (my - ty) * (scale / old);
  smoothT();
}}
// Tüm sayfaları tek bakışta sığdır
function fitAll() {{
  const r = viewport.getBoundingClientRect();
  let x1 = 1e12, y1 = 1e12, x2 = -1e12, y2 = -1e12;
  Object.values(sheetPos).forEach(sp => {{
    x1 = Math.min(x1, sp.x); y1 = Math.min(y1, sp.y);
    x2 = Math.max(x2, sp.x + sp.w); y2 = Math.max(y2, sp.y + sp.h);
  }});
  if (x1 > x2) return;
  const w = x2 - x1, h = y2 - y1;
  scale = Math.min(r.width * 0.92 / w, r.height * 0.92 / h);
  tx = r.width / 2 - (x1 + w / 2) * scale;
  ty = r.height / 2 - (y1 + h / 2) * scale;
  smoothT();
}}

let lastFitSheetId = null;
document.querySelectorAll('.sheet-card').forEach(card => {{
  card.addEventListener('dblclick', e => {{
    e.stopPropagation();
    fitToSheet(card.dataset.sheetId);
    lastFitSheetId = card.dataset.sheetId;
  }});
}});

// === Komponent highlight: PCB ile uyumlu camgöbeği kutu + komponente yakınlaşma ===
let compHighlightTimeout = null;

function clearCompHighlight() {{
  document.querySelectorAll('.comp-highlight').forEach(el => el.classList.remove('comp-highlight'));
  schHlOverlay.innerHTML = '';
  schMarkerBox = null;
  if (compHighlightTimeout) {{ clearTimeout(compHighlightTimeout); compHighlightTimeout = null; }}
}}

function highlightComponent(designator, sheetId, focus = true) {{
  clearCompHighlight();
  const sheetCard = document.getElementById('sheet-' + sheetId);
  if (!sheetCard) {{ if (focus) {{ fitToSheet(sheetId); lastFitSheetId = sheetId; }} return; }}
  lastFitSheetId = sheetId;
  // Eşleşen designator text(ler)i bul (transform türetmek + text vurgusu için)
  const matches = [];
  sheetCard.querySelectorAll('svg text').forEach(t => {{
    if ((t.textContent || '').trim() === designator) matches.push(t);
  }});
  // TAM komponent kutusu: full_bounds (SVG viewBox) → kanvas. Dönüşümü
  // designator text'inin getBBox↔getBoundingClientRect eşlemesinden türet.
  let cbox = null;
  const fullBox = (SCH_BOXES[sheetId] || {{}})[designator];   // [x,y,w,h] SVG viewBox
  if (fullBox && matches.length) cbox = svgBoxToCanvas(matches[0], fullBox);
  if (!cbox && matches.length) cbox = textsCanvasBox(matches);  // fallback: text bbox
  if (!cbox) {{ if (focus) fitToSheet(sheetId); return; }}
  matches.forEach(el => el.classList.add('comp-highlight'));
  const pad = Math.max(cbox.w, cbox.h) * 0.12 + 5;
  schMarkerBox = {{ x: cbox.x - pad, y: cbox.y - pad,
                    w: cbox.w + 2 * pad, h: cbox.h + 2 * pad, label: designator }};
  // focus=false: kullanıcı zaten komponentin üstünde (şematikte tıkladı) —
  // görünümü ortalayıp uzaklaştırma, sadece kutuyu çiz.
  if (focus) focusCanvasBox(schMarkerBox.x, schMarkerBox.y, schMarkerBox.w, schMarkerBox.h);
  drawSchMarker();
  compHighlightTimeout = setTimeout(() => {{
    document.querySelectorAll('.comp-highlight').forEach(el => el.classList.remove('comp-highlight'));
  }}, 6000);
}}
// Bir designator text'inin SVG-viewBox bbox'ı ↔ kanvas konumu eşlemesinden
// SVG→kanvas dönüşümünü (sx,sy,offset) türetip fullBox'ı kanvasa çevir.
function svgBoxToCanvas(textEl, fullBox) {{
  let tb; try {{ tb = textEl.getBBox(); }} catch(e) {{ return null; }}
  if (!tb.width || !tb.height) return null;
  const vRect = viewport.getBoundingClientRect();
  const r = textEl.getBoundingClientRect();
  const tcx = (r.left - vRect.left - tx) / scale, tcy = (r.top - vRect.top - ty) / scale;
  const sx = (r.width / scale) / tb.width, sy = (r.height / scale) / tb.height;
  const offx = tcx - tb.x * sx, offy = tcy - tb.y * sy;
  return {{ x: fullBox[0]*sx + offx, y: fullBox[1]*sy + offy,
            w: fullBox[2]*sx, h: fullBox[3]*sy }};
}}
function textsCanvasBox(matches) {{
  const vRect = viewport.getBoundingClientRect();
  let x0=1e9,y0=1e9,x1=-1e9,y1=-1e9;
  matches.forEach(el => {{
    const bb = el.getBoundingClientRect();
    const L=(bb.left-vRect.left-tx)/scale, T=(bb.top-vRect.top-ty)/scale;
    const R=(bb.right-vRect.left-tx)/scale, B=(bb.bottom-vRect.top-ty)/scale;
    x0=Math.min(x0,L);y0=Math.min(y0,T);x1=Math.max(x1,R);y1=Math.max(y1,B);
  }});
  if (x0>x1) return null;
  return {{ x:x0, y:y0, w:x1-x0, h:y1-y0 }};
}}

function drawSchMarker() {{
  schHlOverlay.innerHTML = '';
  if (!schMarkerBox) return;
  const NS = 'http://www.w3.org/2000/svg', b = schMarkerBox;
  // Spotlight: her yeri karart, komponent kutusunu pencere olarak aç (evenodd)
  const M = 200000;
  const dim = document.createElementNS(NS, 'path');
  dim.setAttribute('d',
    `M${{-M}},${{-M}} H${{M}} V${{M}} H${{-M}} Z ` +
    `M${{b.x}},${{b.y}} v${{b.h}} h${{b.w}} v${{-b.h}} Z`);
  dim.setAttribute('fill', 'rgba(0,0,0,0.62)');
  dim.setAttribute('fill-rule', 'evenodd');
  schHlOverlay.appendChild(dim);
  // Net kutu
  const rect = document.createElementNS(NS, 'rect');
  rect.setAttribute('x', b.x); rect.setAttribute('y', b.y);
  rect.setAttribute('width', b.w); rect.setAttribute('height', b.h);
  rect.setAttribute('class', 'hl-box');
  schHlOverlay.appendChild(rect);
  const t = document.createElementNS(NS, 'text');
  t.setAttribute('class', 'hl-text'); t.textContent = b.label;
  schHlOverlay.appendChild(t);
  updateSchMarkerMetrics();
}}
// Çizgi/rx/yazı ekran-pikselinde sabit kalsın diye 1/scale ile ölçekle
function updateSchMarkerMetrics() {{
  if (!schMarkerBox) return;
  const k = 1 / scale, b = schMarkerBox;
  const rect = schHlOverlay.querySelector('.hl-box');
  if (rect) {{ rect.setAttribute('stroke-width', 1.6 * k); rect.setAttribute('rx', 3 * k); }}
  const t = schHlOverlay.querySelector('.hl-text');
  if (t) {{
    const fs = 13 * k;
    t.setAttribute('font-size', fs);
    t.setAttribute('stroke-width', 0.18 * fs);
    t.setAttribute('x', b.x);
    t.setAttribute('y', b.y - 4 * k);
  }}
}}
// Komponente yumuşak yakınlaş + ortala (PCB focusBox benzeri)
function focusCanvasBox(x, y, w, h) {{
  const r = viewport.getBoundingClientRect();
  if (!r.width || !r.height) return;
  const cx = x + w / 2, cy = y + h / 2;
  // kutu kısa kenarın ~%35'i kadar görünsün (çevresi bağlam olarak kalsın)
  let ns = (Math.min(r.width, r.height) * 0.35) / Math.max(w, h, 1);
  ns = Math.max(0.5, Math.min(2.2, ns));
  scale = ns;
  tx = r.width / 2 - cx * scale;
  ty = r.height / 2 - cy * scale;
  canvas.style.transition = 'transform 0.35s ease';
  applyT();
  clearTimeout(focusCanvasBox._t);
  focusCanvasBox._t = setTimeout(() => {{ canvas.style.transition = 'none'; }}, 400);
}}

document.querySelectorAll('.tab').forEach(b => {{
  b.onclick = () => {{
    document.querySelectorAll('.tab').forEach(x => x.classList.remove('active'));
    b.classList.add('active');
    const tab = b.dataset.tab;
    document.getElementById('nets-list').classList.toggle('hidden', tab !== 'nets');
    document.getElementById('comps-list').classList.toggle('hidden', tab !== 'components');
    document.getElementById('type-chips').classList.toggle('hidden', tab !== 'nets');
    document.getElementById('search').placeholder =
      tab === 'nets' ? 'net ara... ( / )' : 'komponent ara... ( / )';
    document.getElementById('search').value = '';
    renderActive('');
  }};
}});

function escHtml(s) {{
  return s.replace(/[&<>"']/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));
}}

let netTypeFilter = '';
function renderNets(filter='') {{
  const list = document.getElementById('nets-list');
  list.innerHTML = '';
  const q = filter.toLowerCase();
  const matched = nets.filter(n => n.name.toLowerCase().includes(q)
    && (!netTypeFilter || classifyNet(n.name) === netTypeFilter));
  matched.slice(0, 1500).forEach(net => {{
    const div = document.createElement('div');
    div.className = 'net-item ' + classifyNet(net.name);
    div.innerHTML = `${{escHtml(net.name)}} <span class="badge">${{net.count}}</span>`;
    div.dataset.netName = net.name;
    div.onclick = (e) => selectNet(net, div, e.shiftKey);
    list.appendChild(div);
  }});
  if (!matched.length) {{
    const m = document.createElement('div');
    m.className = 'empty-msg'; m.textContent = 'eşleşen net yok';
    list.appendChild(m);
  }}
}}
function renderComps(filter='') {{
  const list = document.getElementById('comps-list');
  list.innerHTML = '';
  const q = filter.toLowerCase();
  const matched = components.filter(c =>
    c.designator.toLowerCase().includes(q) || c.value.toLowerCase().includes(q)
  );
  matched.slice(0, 1500).forEach(comp => {{
    const div = document.createElement('div');
    div.className = 'comp-item';
    const val = comp.value ? ` <span style="color:#666">${{escHtml(comp.value)}}</span>` : '';
    div.innerHTML = `${{escHtml(comp.designator)}}${{val}} <span class="sheet-tag">${{escHtml(comp.sheet_name)}}</span>`;
    div.onclick = () => {{
      highlightComponent(comp.designator, comp.sheet_id);
      showCompPopup(comp);
      crossProbeOut(comp.designator);
    }};
    list.appendChild(div);
  }});
  if (!matched.length) {{
    const m = document.createElement('div');
    m.className = 'empty-msg'; m.textContent = 'eşleşen komponent yok';
    list.appendChild(m);
  }}
}}

// === Komponent detay popup ===
function linkify(text) {{
  const safe = escHtml(text);
  return safe.replace(/(https?:\\/\\/[^\\s<>"]+)/g,
    '<a href="$1" target="_blank" rel="noopener" class="popup-link">$1</a>');
}}

function renderPopupRow(k, v) {{
  const sv = String(v);
  return `<div class="popup-row">`
    + `<span class="popup-key">${{escHtml(k)}}</span>`
    + `<span class="popup-val" data-raw="${{escHtml(sv)}}">${{linkify(sv)}}</span>`
    + `<button class="popup-copy" title="Kopyala">⎘</button>`
    + `</div>`;
}}

function showCompPopup(comp) {{
  const popup = document.getElementById('comp-popup');
  if (!popup) return;
  setSidebarOpen(true);  // popup sidebar'a dock'lu — panel kapalıysa aç
  document.getElementById('popup-title').textContent = comp.designator;
  // Multi-part ise kaç parça/sayfa olduğunu göster
  const places = comp.placements || [];
  if (places.length > 1) {{
    document.getElementById('popup-sheet').textContent =
      ` · ${{places.length}} parça (multi-part)`;
  }} else {{
    document.getElementById('popup-sheet').textContent = ' · ' + comp.sheet_name;
  }}

  const body = document.getElementById('popup-body');
  const rows = [];
  if (comp.value) rows.push(['Comment', comp.value]);
  if (comp.description) rows.push(['Description', comp.description]);
  if (comp.footprint) rows.push(['Footprint', comp.footprint]);
  if (comp.library_reference) rows.push(['Library Ref', comp.library_reference]);
  if (comp.library_name) rows.push(['Library Name', comp.library_name]);
  // Multi-part: hangi sayfalarda göründüğü
  if (places.length > 1) {{
    const sheetNames = places.map(p => p.sheet_name).filter((v, i, a) => a.indexOf(v) === i);
    rows.push(['Parçalar', sheetNames.join(', ')]);
  }}

  let html = '';
  if (rows.length > 0) {{
    html += '<div class="popup-section-title">Standard</div>';
    rows.forEach(([k, v]) => {{ html += renderPopupRow(k, v); }});
  }}
  if (comp.parameters && Object.keys(comp.parameters).length > 0) {{
    html += '<div class="popup-section-title">Parameters</div>';
    Object.entries(comp.parameters).forEach(([k, v]) => {{ html += renderPopupRow(k, v); }});
  }}

  // === PCB cross-probe ===
  if (PCB.available && PCB.components[comp.designator]) {{
    const pc = PCB.components[comp.designator];
    html += '<div class="popup-section-title">PCB Konumu</div>';
    const ax = (pc.abs_x_mm !== undefined) ? pc.abs_x_mm : pc.x_mm;
    const ay = (pc.abs_y_mm !== undefined) ? pc.abs_y_mm : pc.y_mm;
    html += renderPopupRow('Konum (mm)', `X=${{ax}}  Y=${{ay}}`);
    html += renderPopupRow('Katman', pc.layer);
    if (pc.rotation) html += renderPopupRow('Dönüş', pc.rotation + '°');
    if (pc.footprint) html += renderPopupRow('Footprint (PCB)', pc.footprint);
    html += '<div id="pcb-map-container"></div>';
  }} else if (PCB.available) {{
    html += '<div class="popup-section-title">PCB Konumu</div>';
    html += '<div style="color:#888;padding:6px;font-size:11px">Bu komponent PCB\\'de bulunamadı.</div>';
  }}

  if (!html) html = '<div style="color:#666;padding:10px">Ek bilgi yok</div>';

  body.innerHTML = html;
  // PCB mini haritasını çiz (innerHTML sonrası)
  if (PCB.available && PCB.components[comp.designator]) {{
    drawPcbMap(comp.designator);
  }}
  popup.classList.add('open');
  popup.classList.remove('collapsed');
  document.getElementById('popup-collapse').textContent='▾';
}}

// PCB mini haritası: board + TÜM komponentler (yerleşim görüntüsü) +
// seçili komponent vurgulu. Tüm noktalar aynı koordinat çerçevesinde
// (board sol-üst orijin, mm) olduğundan hizalama garantilidir.
function drawPcbMap(designator) {{
  const container = document.getElementById('pcb-map-container');
  if (!container) return;
  const pc = PCB.components[designator];
  if (!pc) return;

  const W = PCB.board_w_mm, H = PCB.board_h_mm;
  if (!W || !H) {{ container.textContent = '(board boyutu yok)'; return; }}

  const cx = pc.x_mm, cy = pc.y_mm;
  const onTop = (pc.layer || 'TOP').toUpperCase() === 'TOP';
  const dotColor = onTop ? '#ff3030' : '#ff9800';
  const maxDim = 300;  // px
  const scale = maxDim / Math.max(W, H);
  const svgW = W * scale, svgH = H * scale;
  const r1 = Math.max(1.2, Math.min(W, H) / 45);   // seçili nokta
  const r2 = r1 * 2.4;                              // halka
  const rOther = Math.max(0.4, Math.min(W, H) / 130); // diğer komponentler

  // Tüm komponentleri yerleşim noktaları olarak çiz (seçili hariç).
  let dots = '';
  for (const [d, p] of Object.entries(PCB.components)) {{
    if (d === designator) continue;
    if (p.x_mm == null || p.y_mm == null) continue;
    const top = (p.layer || 'TOP').toUpperCase() === 'TOP';
    // TOP açık bakır tonu, BOTTOM soluk mavi → board'a bakar gibi
    const col = top ? '#5a8a6a' : '#3a5a78';
    dots += `<circle cx="${{p.x_mm}}" cy="${{p.y_mm}}" r="${{rOther.toFixed(2)}}" `
          + `fill="${{col}}" opacity="0.85"/>`;
  }}

  container.innerHTML = `
    <div style="margin-top:6px">
      <svg width="${{svgW.toFixed(0)}}" height="${{svgH.toFixed(0)}}" viewBox="0 0 ${{W}} ${{H}}"
           style="background:#07120c;border:1px solid #444;border-radius:3px;display:block">
        <rect x="0.3" y="0.3" width="${{(W-0.6).toFixed(2)}}" height="${{(H-0.6).toFixed(2)}}"
              fill="#0a1f14" stroke="#2a6f54" stroke-width="0.5" rx="1.5"/>
        <g>${{dots}}</g>
        <circle cx="${{cx}}" cy="${{cy}}" r="${{r2.toFixed(2)}}" fill="none"
                stroke="${{dotColor}}" stroke-width="0.4" opacity="0.6"/>
        <circle cx="${{cx}}" cy="${{cy}}" r="${{r1.toFixed(2)}}" fill="${{dotColor}}"
                stroke="#fff" stroke-width="0.35"/>
      </svg>
      <div style="font-size:10px;color:#888;margin-top:3px">
        ${{onTop ? '● TOP katmanı' : '● BOTTOM katmanı'}} ·
        ${{Object.keys(PCB.components).length}} komponent ·
        board ${{W.toFixed(0)}}×${{H.toFixed(0)}}mm
      </div>
    </div>`;
}}
document.getElementById('popup-close').onclick = () => {{
  document.getElementById('comp-popup').classList.remove('open');
}};
// Katla / aç (sol üstteki ok) — simge durumuna küçült
document.getElementById('popup-collapse').onclick = () => {{
  const pp=document.getElementById('comp-popup');
  const c=pp.classList.toggle('collapsed');
  document.getElementById('popup-collapse').textContent=c?'▸':'▾';
}};
// Üst tutamaçtan dikey boyutlandırma
(function(){{
  const pp=document.getElementById('comp-popup');
  const handle=document.getElementById('popup-resize');
  let rz=false, sy=0, sh=0;
  handle.addEventListener('mousedown', e => {{
    rz=true; sy=e.clientY; sh=pp.offsetHeight;
    document.body.style.userSelect='none'; e.preventDefault();
  }});
  window.addEventListener('mousemove', e => {{
    if(!rz) return;
    let h=sh+(sy-e.clientY);                 // yukarı sürükle → büyüt
    h=Math.max(60, Math.min(window.innerHeight*0.8, h));
    pp.style.height=h+'px';
  }});
  window.addEventListener('mouseup', () => {{ rz=false; document.body.style.userSelect=''; }});
}})();

// Kopya butonu için delegated event
document.getElementById('popup-body').addEventListener('click', e => {{
  if (!e.target.classList || !e.target.classList.contains('popup-copy')) return;
  const row = e.target.closest('.popup-row');
  if (!row) return;
  const valEl = row.querySelector('.popup-val');
  const text = (valEl && valEl.getAttribute('data-raw')) || (valEl && valEl.textContent) || '';
  if (!text) return;
  const btn = e.target;
  navigator.clipboard.writeText(text).then(() => {{
    btn.textContent = '✓';
    btn.classList.add('copied');
    setTimeout(() => {{ btn.textContent = '⎘'; btn.classList.remove('copied'); }}, 1200);
  }}).catch(err => {{
    console.warn('Kopyalama hatası:', err);
    btn.textContent = '✗';
    setTimeout(() => {{ btn.textContent = '⎘'; }}, 1200);
  }});
}});
function renderActive(filter) {{
  const isNets = !document.getElementById('nets-list').classList.contains('hidden');
  if (isNets) renderNets(filter); else renderComps(filter);
}}
renderNets();
renderComps();
document.getElementById('search').addEventListener('input', e => renderActive(e.target.value));

// === Katlanabilir sol panel ===
const sidebarEl = document.getElementById('sidebar');
const sidebarToggle = document.getElementById('sidebar-toggle');
function setSidebarOpen(open) {{
  sidebarEl.classList.toggle('collapsed', !open);
  sidebarToggle.textContent = open ? '◂' : '▸';
  sidebarToggle.title = (open ? 'Paneli gizle' : 'Paneli göster') + ' ( B )';
  if (typeof lsSet === 'function') lsSet({{ sidebar: open }});
}}
sidebarToggle.addEventListener('click', () =>
  setSidebarOpen(sidebarEl.classList.contains('collapsed')));

// === Katlanabilir arama bölümü ===
const searchWrap = document.getElementById('search-wrap');
const searchInput = document.getElementById('search');
function setSearchOpen(open) {{
  if (open) setSidebarOpen(true);  // panel kapalıysa önce aç
  searchWrap.classList.toggle('collapsed', !open);
  document.getElementById('search-caret').textContent = open ? '▾' : '▸';
  if (open) searchInput.focus();
  else {{
    searchInput.blur();
    // Kapatınca filtreyi de temizle — liste gizli filtreyle kafa karıştırmasın
    if (searchInput.value) {{ searchInput.value = ''; renderActive(''); }}
  }}
}}
document.getElementById('search-toggle').addEventListener('click', () =>
  setSearchOpen(searchWrap.classList.contains('collapsed')));

// Aramada Enter → görünen listedeki ilk sonucu seç
searchInput.addEventListener('keydown', e => {{
  if (e.key !== 'Enter') return;
  const list = document.querySelector('.list-container:not(.hidden)');
  const first = list && list.querySelector('.net-item, .comp-item');
  if (first) first.click();
}});

// === Net tipi filtre çipleri ===
document.querySelectorAll('#type-chips .chip').forEach(ch => {{
  ch.onclick = () => {{
    netTypeFilter = ch.dataset.type;
    document.querySelectorAll('#type-chips .chip').forEach(x =>
      x.classList.toggle('active', x === ch));
    renderNets(searchInput.value);
  }};
}});

// === Toolbar: sayfa seçici + zoom kontrolleri ===
const sheetJump = document.getElementById('sheet-jump');
sheetJump.innerHTML = '<option value="">Sayfa…</option>' +
  Object.entries(sheetPos).map(([id, sp]) =>
    `<option value="${{id}}">${{escHtml(sp.name)}}</option>`).join('');
sheetJump.addEventListener('change', () => {{
  if (!sheetJump.value) return;
  fitToSheet(sheetJump.value); lastFitSheetId = sheetJump.value;
  sheetJump.value = ''; sheetJump.blur();
}});
document.getElementById('zoom-in').onclick = () => zoomBy(1.35);
document.getElementById('zoom-out').onclick = () => zoomBy(1 / 1.35);
document.getElementById('fit-all').onclick = fitAll;

// === Hover bilgi balonu (komponent / net / block) ===
const svgTip = document.getElementById('svg-tip');
function moveTip(e) {{
  const pad = 14; let x = e.clientX + pad, y = e.clientY + pad;
  const r = svgTip.getBoundingClientRect();
  if (x + r.width > innerWidth - 8) x = e.clientX - r.width - pad;
  if (y + r.height > innerHeight - 8) y = e.clientY - r.height - pad;
  svgTip.style.left = x + 'px'; svgTip.style.top = y + 'px';
}}
document.addEventListener('mouseover', e => {{
  const t = e.target; if (!t.classList) return;
  let html = null;
  if (t.classList.contains('comp-designator')) {{
    const d = t.getAttribute('data-comp-designator');
    const sid = t.getAttribute('data-comp-sheet-id');
    const c = (compsBySheet[sid] || {{}})[d] || compByDesig[d];
    if (c) html = `<span class="tt-title">${{escHtml(c.designator)}}</span>`
      + (c.value ? ` · ${{escHtml(c.value)}}` : '')
      + (c.description ? `<br>${{escHtml(c.description)}}` : '')
      + `<br><span class="tt-hint">tıkla: detay + cross-probe</span>`;
  }} else if (t.classList.contains('clickable-net')) {{
    const n = t.getAttribute('data-net');
    const net = nets.find(x => x.name === n);
    if (net) {{
      const ty = classifyNet(n);
      html = `<span class="tt-title">${{escHtml(n)}}</span> · ${{net.count}} nokta · `
        + (ty === 'power' ? 'güç' : ty === 'ground' ? 'toprak' : 'sinyal')
        + `<br><span class="tt-hint">tıkla: bağlantı yayları · Shift+tık: karşılaştır</span>`;
    }}
  }} else if (t.classList.contains('block-link')) {{
    html = `<span class="tt-title">${{escHtml((t.textContent || '').trim())}}</span>`
      + `<br><span class="tt-hint">tıkla: sayfaya git</span>`;
  }}
  if (html) {{ svgTip.innerHTML = html; svgTip.style.display = 'block'; moveTip(e); }}
}});
document.addEventListener('mousemove', e => {{
  if (svgTip.style.display === 'block') moveTip(e);
}});
document.addEventListener('mouseout', e => {{
  const t = e.target;
  if (t.classList && (t.classList.contains('comp-designator')
      || t.classList.contains('clickable-net') || t.classList.contains('block-link')))
    svgTip.style.display = 'none';
}});

// === UI tercihlerini hatırla (panel durumu + renkler) ===
// file:// altında localStorage bazı tarayıcılarda kısıtlı olabilir → try/catch.
const LS_KEY = 'schviz-ui';
function lsGet() {{ try {{ return JSON.parse(localStorage.getItem(LS_KEY) || '{{}}'); }} catch (e) {{ return {{}}; }} }}
function lsSet(patch) {{ try {{
  localStorage.setItem(LS_KEY, JSON.stringify(Object.assign(lsGet(), patch)));
}} catch (e) {{}} }}
(function restoreUi() {{
  const st = lsGet();
  if (st.sidebar === false) setSidebarOpen(false);
  if (st.lod === false) setLodEnabled(false);
  if (st.inter) {{ NET_COLORS[0] = st.inter;
    document.getElementById('inter-color-picker').value = st.inter; }}
  if (st.intra) {{ INTRA_COLOR = st.intra;
    document.getElementById('intra-color-picker').value = st.intra; }}
}})();

let selectedNets = [];

function clearSelection() {{
  selectedNets = [];
  document.querySelectorAll('.net-item.active').forEach(e =>
    e.classList.remove('active', 'color-1', 'color-2', 'color-3'));
  arcLayer.innerHTML = '';
  document.querySelectorAll('.sheet-card').forEach(c =>
    c.classList.remove('hit', 'hit-1', 'hit-2', 'hit-3'));
  currentNetEl.textContent = '';
  detailPanel.classList.remove('open');
  clearCompHighlight();
}}

function selectNet(net, listEl, addToSelection = false) {{
  if (!addToSelection) {{
    selectedNets = [net];
  }} else {{
    const i = selectedNets.findIndex(n => n.name === net.name);
    if (i >= 0) selectedNets.splice(i, 1);
    else if (selectedNets.length < NET_COLORS.length) selectedNets.push(net);
  }}
  if (!listEl) {{
    listEl = [...document.querySelectorAll('#nets-list .net-item')].find(el =>
      el.dataset.netName === net.name);
    if (!listEl) {{
      document.getElementById('search').value = '';
      renderNets('');
      listEl = [...document.querySelectorAll('#nets-list .net-item')].find(el =>
        el.dataset.netName === net.name);
    }}
    if (listEl) listEl.scrollIntoView({{ behavior: 'smooth', block: 'nearest' }});
  }}
  renderAllSelections();
}}

function renderAllSelections() {{
  arcLayer.innerHTML = '';
  document.querySelectorAll('.sheet-card').forEach(c =>
    c.classList.remove('hit', 'hit-1', 'hit-2', 'hit-3'));
  document.querySelectorAll('.net-item.active').forEach(e =>
    e.classList.remove('active', 'color-1', 'color-2', 'color-3'));

  if (selectedNets.length === 0) {{
    currentNetEl.textContent = '';
    detailPanel.classList.remove('open');
    return;
  }}

  if (selectedNets.length === 1) {{
    currentNetEl.textContent = `${{selectedNets[0].name}}  (${{selectedNets[0].count}} pin)`;
  }} else {{
    currentNetEl.textContent = selectedNets.map(n => n.name).join('  ·  ');
  }}

  selectedNets.forEach((net, idx) => {{
    drawNetSelection(net, idx);
    const listEl = [...document.querySelectorAll('#nets-list .net-item')].find(el =>
      el.dataset.netName === net.name);
    if (listEl) {{
      listEl.classList.add('active');
      if (idx > 0) listEl.classList.add('color-' + idx);
    }}
  }});
  updateDetailPanel();
}}

function drawNetSelection(net, colorIdx) {{
  const color = NET_COLORS[colorIdx];
  const points = net.occurrences.map(o => {{
    const sp = sheetPos[o.sheet_id];
    const titleH = 30;
    const bodyH = sp.h - titleH;
    return {{
      sheet_id: o.sheet_id,
      x: sp.x + o.rx * sp.w,
      y: sp.y + titleH + o.ry * bodyH,
    }};
  }});

  const sheetAnchor = {{}};
  const extraPins = [];
  points.forEach(p => {{
    if (!sheetAnchor[p.sheet_id]) sheetAnchor[p.sheet_id] = p;
    else extraPins.push(p);
  }});

  const hitClass = colorIdx === 0 ? 'hit' : 'hit-' + colorIdx;
  Object.keys(sheetAnchor).forEach(id => {{
    const card = document.getElementById('sheet-' + id);
    if (card) card.classList.add(hitClass);
  }});

  extraPins.forEach(p => drawIntraConnection(sheetAnchor[p.sheet_id], p));

  const anchors = Object.values(sheetAnchor);
  for (let i = 0; i < anchors.length - 1; i++) {{
    for (let j = i + 1; j < anchors.length; j++) {{
      drawArc(anchors[i], anchors[j], color);
    }}
  }}
}}

function drawArc(a, b, color) {{
  const midX = (a.x + b.x) / 2;
  const dist = Math.hypot(b.x - a.x, b.y - a.y);
  const lift = Math.min(80 + dist * 0.35, 600);
  const cpY = Math.min(a.y, b.y) - lift;
  const p = document.createElementNS('http://www.w3.org/2000/svg', 'path');
  p.setAttribute('d', `M ${{a.x}} ${{a.y}} Q ${{midX}} ${{cpY}}, ${{b.x}} ${{b.y}}`);
  p.setAttribute('fill', 'none');
  p.setAttribute('stroke', color);
  p.setAttribute('stroke-width', '2.5');
  p.setAttribute('stroke-dasharray', '10 6');
  p.setAttribute('opacity', '0.85');
  p.style.filter = `drop-shadow(0 0 5px ${{color}})`;
  arcLayer.appendChild(p);
}}

function drawIntraConnection(a, b) {{
  const midX = (a.x + b.x) / 2;
  const midY = (a.y + b.y) / 2;
  const dx = b.x - a.x, dy = b.y - a.y;
  const dist = Math.hypot(dx, dy);
  if (dist < 1) return;
  const lift = Math.min(dist * 0.2, 60);
  const nx = -dy / dist;
  const ny = dx / dist;
  const cpX = midX + nx * lift;
  const cpY = midY + ny * lift;
  const p = document.createElementNS('http://www.w3.org/2000/svg', 'path');
  p.setAttribute('d', `M ${{a.x}} ${{a.y}} Q ${{cpX}} ${{cpY}}, ${{b.x}} ${{b.y}}`);
  p.setAttribute('fill', 'none');
  p.setAttribute('stroke', INTRA_COLOR);
  p.setAttribute('stroke-width', '2.5');
  p.setAttribute('opacity', '0.85');
  p.style.filter = `drop-shadow(0 0 4px ${{INTRA_COLOR}})`;
  arcLayer.appendChild(p);
}}

function updateDetailPanel() {{
  if (selectedNets.length === 0) {{ detailPanel.classList.remove('open'); return; }}
  detailPanel.classList.add('open');
  let html = '';
  selectedNets.forEach((net, idx) => {{
    const color = NET_COLORS[idx];
    const bySheet = {{}};
    net.occurrences.forEach(o => {{
      bySheet[o.sheet_id] = bySheet[o.sheet_id] || {{ name: o.sheet_name, count: 0 }};
      bySheet[o.sheet_id].count++;
    }});
    html += `<div class="net-detail">`;
    html += `<span class="net-name" style="color:${{color}}">${{escHtml(net.name)}}</span> · ${{net.count}} pin · `;
    Object.entries(bySheet).forEach(([sid, info]) => {{
      html += `<span class="sheet-row" data-sheet-id="${{sid}}">${{escHtml(info.name)}} <span style="color:${{color}}">×${{info.count}}</span></span>`;
    }});
    html += `</div>`;
  }});
  detailContent.innerHTML = html;
  detailContent.querySelectorAll('.sheet-row').forEach(el => {{
    el.addEventListener('click', () => fitToSheet(el.dataset.sheetId));
  }});
}}

const netNameSet = new Set(nets.map(n => n.name));

// SVG text'lerinde net adlarını tıklanabilir yap
document.querySelectorAll('.sheet-body svg text').forEach(t => {{
  const content = (t.textContent || '').trim();
  if (netNameSet.has(content)) {{
    t.classList.add('clickable-net');
    t.setAttribute('data-net', content);
  }}
}});

// Block linkleri: Python tarafından çıkarılan sheet_symbol verisinden
// Komponent designator'ları: Python tarafından çıkarılan components verisinden
// Komponentler artık multi-part birleştirilmiş (IC2A/IC2B/IC2C -> tek IC2).
const compsBySheet = {{}};   // sheet_id -> {{designator: comp}}
const compByDesig = {{}};    // designator -> comp (global, part eşleşmesi için)
components.forEach(c => {{
  compByDesig[c.designator] = c;
  // Komponent hangi sayfalarda görünüyorsa hepsine ekle (placements)
  const places = c.placements || [{{sheet_id: c.sheet_id}}];
  places.forEach(pl => {{
    const sid = pl.sheet_id;
    if (!compsBySheet[sid]) compsBySheet[sid] = {{}};
    compsBySheet[sid][c.designator] = c;
  }});
}});

// SVG text içeriğinden komponent designator'ı çöz.
// "IC2" doğrudan eşleşir; "IC2A"/"IC2B" gibi part suffix'li yazılarda
// sondaki harf(ler)i atıp tekrar dener (multi-part komponentler).
function resolveCompDesignator(content, compMap) {{
  if (compMap[content]) return content;            // doğrudan
  if (compByDesig[content]) return content;        // global doğrudan
  // Part suffix'i: harf+rakam tabanı + sonda tek/çift büyük harf (A, B, ... AA)
  const m = content.match(/^([A-Za-z]+\\d+)([A-Z]{{1,2}})$/);
  if (m) {{
    const base = m[1];
    if (compMap[base] || compByDesig[base]) return base;
  }}
  return null;
}}

Object.entries(sheetPos).forEach(([sheetId, sp]) => {{
  const card = document.getElementById('sheet-' + sheetId);
  if (!card) return;

  // Block linkleri
  const blocks = sp.blocks || [];
  const blockMap = {{}};
  blocks.forEach(b => {{
    if (b.designator) blockMap[b.designator] = b.target_id;
    if (b.filename) blockMap[b.filename] = b.target_id;
    if (b.target_name) blockMap[b.target_name] = b.target_id;
  }});

  // Komponent designator map'i (bu sayfadakiler)
  const compMap = compsBySheet[sheetId] || {{}};

  card.querySelectorAll('svg text').forEach(t => {{
    const content = (t.textContent || '').trim();
    if (!content) return;
    // Block link?
    if (blockMap[content]) {{
      t.classList.add('block-link');
      t.setAttribute('data-target-sheet-id', blockMap[content]);
      return;
    }}
    // Komponent designator? (multi-part suffix dahil)
    const desig = resolveCompDesignator(content, compMap);
    if (desig) {{
      t.classList.add('comp-designator');
      t.setAttribute('data-comp-designator', desig);
      t.setAttribute('data-comp-sheet-id', sheetId);
    }}
  }});
}});
document.querySelectorAll('.sheet-body').forEach(body => {{
  body.addEventListener('mousedown', e => {{
    if (e.target.classList && (
      e.target.classList.contains('clickable-net') ||
      e.target.classList.contains('block-link') ||
      e.target.classList.contains('comp-designator')
    )) {{
      e.stopPropagation();
    }}
  }});
  body.addEventListener('click', e => {{
    if (!e.target.classList) return;
    // Kullanıcı metin SEÇTİYSE (sürükleme ile) bu bir kopyalama jestidir —
    // net/block/komponent tıklama aksiyonunu tetikleme.
    if (window.getSelection && String(window.getSelection()).length > 0) return;
    if (e.target.classList.contains('clickable-net')) {{
      const netName = e.target.getAttribute('data-net');
      const net = nets.find(n => n.name === netName);
      if (net) selectNet(net, null, e.shiftKey);
    }} else if (e.target.classList.contains('block-link')) {{
      const targetId = e.target.getAttribute('data-target-sheet-id');
      if (targetId) {{
        fitToSheet(targetId);
        lastFitSheetId = targetId;
      }}
    }} else if (e.target.classList.contains('comp-designator')) {{
      const designator = e.target.getAttribute('data-comp-designator');
      const sheetId = e.target.getAttribute('data-comp-sheet-id');
      const comp = (compsBySheet[sheetId] || {{}})[designator] || compByDesig[designator];
      if (comp) {{
        // Tıklanan part hangi sayfadaysa orada highlight et (multi-part).
        // focus=false: kullanıcı zaten oraya bakıyor — görünümü kaydırma/uzaklaştırma.
        highlightComponent(comp.designator, sheetId || comp.sheet_id, false);
        showCompPopup(comp);
        crossProbeOut(comp.designator);
      }}
    }}
  }});
}});

document.addEventListener('keydown', e => {{
  if (document.activeElement === document.getElementById('search')) {{
    if (e.key === 'Escape') setSearchOpen(false);
    return;
  }}
  // Modal açıkken Esc kapatsın
  if (shortcutModal.classList.contains('open') && e.key === 'Escape') {{
    toggleShortcutModal();
    return;
  }}
  // Not/kutu aracı aktifken Esc yalnız araçtan çıkar; değilse önce seçimi bırakır
  if (e.key === 'Escape' && (annoTool || annoDrag)) {{ setAnnoTool(null); return; }}
  if (e.key === 'Escape' && annoSel != null) {{ annoSetSel(null); return; }}
  // Seçili not/kutu Del (veya Backspace) ile silinir — yazı alanları hariç
  if ((e.key === 'Delete' || e.key === 'Backspace') && annoSel != null) {{
    const ae = document.activeElement;
    if (!ae || (ae.tagName !== 'INPUT' && ae.tagName !== 'TEXTAREA'
                && ae.tagName !== 'SELECT' && !ae.isContentEditable)) {{
      e.preventDefault(); annoDelete(annoSel); return;
    }}
  }}
  if (e.key === '?') {{ e.preventDefault(); toggleShortcutModal(); }}
  else if (e.key === 'Escape') {{
    // Önce popup, sonra seçim
    const popup = document.getElementById('comp-popup');
    if (popup.classList.contains('open')) {{
      popup.classList.remove('open');
    }} else {{
      clearSelection();
    }}
  }}
  else if (e.key === '/') {{ e.preventDefault(); setSearchOpen(true); }}
  else if (e.key === '0') resetView();
  else if (e.key === 'b' || e.key === 'B') setSidebarOpen(sidebarEl.classList.contains('collapsed'));
  else if (e.key === 'f' || e.key === 'F') {{ if (lastFitSheetId) fitToSheet(lastFitSheetId); }}
  else if (e.key === '+' || e.key === '=') {{
    const r = viewport.getBoundingClientRect();
    const old = scale; scale = Math.min(8, scale * 1.2);
    tx = r.width/2 - (r.width/2 - tx) * (scale/old);
    ty = r.height/2 - (r.height/2 - ty) * (scale/old);
    applyT();
  }}
  else if (e.key === '-' || e.key === '_') {{
    const r = viewport.getBoundingClientRect();
    const old = scale; scale = Math.max(0.03, scale * 0.83);
    tx = r.width/2 - (r.width/2 - tx) * (scale/old);
    ty = r.height/2 - (r.height/2 - ty) * (scale/old);
    applyT();
  }}
}});

document.getElementById('reset-view').onclick = resetView;
document.getElementById('clear-sel').onclick = clearSelection;

// === Renk picker'lar - canlı güncelleme ===
document.getElementById('inter-color-picker').addEventListener('input', e => {{
  NET_COLORS[0] = e.target.value;
  lsSet({{ inter: e.target.value }});
  if (selectedNets.length > 0) renderAllSelections();
}});
document.getElementById('intra-color-picker').addEventListener('input', e => {{
  INTRA_COLOR = e.target.value;
  lsSet({{ intra: e.target.value }});
  if (selectedNets.length > 0) renderAllSelections();
}});

// === Kısayol modal ===
const shortcutModal = document.getElementById('shortcut-modal');
function toggleShortcutModal() {{ shortcutModal.classList.toggle('open'); }}
document.getElementById('shortcut-btn').onclick = toggleShortcutModal;
document.getElementById('close-modal-btn').onclick = toggleShortcutModal;
shortcutModal.addEventListener('click', e => {{
  if (e.target === shortcutModal) toggleShortcutModal();
}});
document.getElementById('export-png').onclick = async () => {{
  if (typeof html2canvas === 'undefined') {{ alert('html2canvas yüklenmedi'); return; }}
  const btn = document.getElementById('export-png');
  const orig = btn.textContent;
  btn.textContent = '...'; btn.disabled = true;
  try {{
    const canvas = await html2canvas(viewport, {{
      backgroundColor: '#1a1a1a', logging: false, useCORS: true,
    }});
    const a = document.createElement('a');
    a.href = canvas.toDataURL('image/png');
    a.download = `schematic_viz_${{BUILD_STAMP.replace(/:/g,'')}}.png`;
    a.click();
  }} catch (err) {{
    alert('Export hatası: ' + err.message);
  }} finally {{
    btn.textContent = orig; btn.disabled = false;
  }}
}};

// === Not / kutu (annotation) araçları =====================================
// Foxit tarzı kullanım: yerinde yazma (typewriter — prompt yok), tıkla-seç,
// sürükle-taşı, kutuda köşe tutamacından boyutlandır, seçiliyken Del ile sil,
// mini bar (−/+/×) ile yazı boyutu / kenar kalınlığı / silme.
// Veri KANVAS koordinatında tutulur (pan/zoom ile birlikte hareket eder),
// localStorage'a proje bazlı anahtarla otomatik kaydedilir. "Kaydet" butonu
// notlar gömülü, temizlenmiş bir HTML kopyası indirir — dosya paylaşılınca/
// başka makinede açılınca notlar gömülü veriden yüklenir. Yükleme önceliği:
// localStorage ile gömülü veriden HANGİSİ daha yeniyse o (ts karşılaştırması).
const annoLayer = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
annoLayer.setAttribute('id', 'anno-layer');
annoLayer.setAttribute('width', arcLayer.getAttribute('width') || 10000);
annoLayer.setAttribute('height', arcLayer.getAttribute('height') || 10000);
annoLayer.style.pointerEvents = 'none';   // yalnız .anno çocukları tıklanır (CSS)
canvas.appendChild(annoLayer);

// Yerinde yazma editörü — canvas İÇİNDE HTML div, transformla birlikte ölçeklenir
const annoEditor = document.createElement('div');
annoEditor.id = 'anno-editor';
annoEditor.setAttribute('contenteditable', 'plaintext-only');
annoEditor.style.display = 'none';
canvas.appendChild(annoEditor);
if (!annoEditor.isContentEditable)   // eski Firefox: plaintext-only desteklenmez
  annoEditor.setAttribute('contenteditable', 'true');

// Seçim mini araç çubuğu (ekran uzayında, seçili öğenin üstünde durur)
const annoBar = document.createElement('div');
annoBar.id = 'anno-bar';
annoBar.style.display = 'none';
annoBar.innerHTML =
    '<button data-act="minus" title="Yazı boyutu / kenar kalınlığı azalt">−</button>'
  + '<button data-act="plus" title="Yazı boyutu / kenar kalınlığı artır">+</button>'
  + '<input type="color" title="Renk (not yazısı / kutu kenarı)">'
  + '<button data-act="del" title="Sil (Del)">×</button>';
viewport.appendChild(annoBar);
const annoColorInp = annoBar.querySelector('input[type=color]');

const ANNO_NS = 'http://www.w3.org/2000/svg';
const LS_ANNO = 'schviz-anno:' + (PROJECT_NAME || 'proje');
let annotations = [];      // {{k:'note',id,x,y,text,fs}} | {{k:'box',id,x,y,w,h,sw}}
let annoTool = null;       // null | 'note' | 'box'
let annoDrag = null;       // kutu çizimi sürüyor: {{x0,y0,x1,y1}}
let annoJustDrew = false;  // kutu/taşıma biten click'i seçim-temizlemeden korur
let annoSel = null;        // seçili annotation id'si
let annoEditId = null;     // yerinde düzenlenen not id'si ('' = yeni not)
let annoEditPos = null;    // yeni notun kanvas konumu
let annoMove = null;       // taşıma/boyutlandırma sürükleme durumu
const annoMeasure = {{}};  // not id → render'da ölçülen {{w,h}} (seçim çerçevesi)
var __annoUi = null;       // applyT her karede çağırır (var: hoisting, TDZ yok)

function annoLoad() {{
  let emb = null, loc = null;
  try {{ emb = JSON.parse(document.getElementById('anno-embed').textContent); }} catch (e) {{}}
  try {{ loc = JSON.parse(localStorage.getItem(LS_ANNO) || 'null'); }} catch (e) {{}}
  const pick = (emb && loc) ? (((loc.ts || 0) >= (emb.ts || 0)) ? loc : emb) : (loc || emb);
  annotations = (pick && Array.isArray(pick.items)) ? pick.items : [];
}}
function annoStore() {{
  try {{ localStorage.setItem(LS_ANNO, JSON.stringify({{ ts: Date.now(), items: annotations }})); }}
  catch (e) {{}}   // file:// altında storage kısıtlıysa sessizce atla
}}

function annoRectEl(x, y, w, h) {{
  const r = document.createElementNS(ANNO_NS, 'rect');
  r.setAttribute('x', x); r.setAttribute('y', y);
  r.setAttribute('width', w); r.setAttribute('height', h);
  return r;
}}
// Seçim çerçevesi/bar için öğe sınırları (not boyutu render'da ölçülür)
function annoBounds(a) {{
  if (a.k === 'box') return {{ x: a.x, y: a.y, w: a.w, h: a.h }};
  const m = annoMeasure[a.id] || {{ w: 60, h: 28 }};
  return {{ x: a.x, y: a.y, w: m.w, h: m.h }};
}}
function annoRender() {{
  annoLayer.innerHTML = '';
  annotations.forEach(a => {{
    if (a.id === annoEditId) return;   // düzenlenen not şu an editörde
    const g = document.createElementNS(ANNO_NS, 'g');
    g.setAttribute('class', 'anno anno-' + a.k + (a.id === annoSel ? ' sel' : ''));
    g.setAttribute('data-id', String(a.id));
    const tip = document.createElementNS(ANNO_NS, 'title');
    tip.textContent = a.k === 'note'
      ? 'Sürükle: taşı · Çift tık: düzenle · Seç + Del: sil'
      : 'Kenardan sürükle: taşı · Köşe tutamacı: boyutlandır · Del: sil';
    g.appendChild(tip);
    if (a.k === 'box') {{
      const r = annoRectEl(a.x, a.y, a.w, a.h);
      r.setAttribute('rx', 3);
      r.setAttribute('fill', 'rgba(255,179,0,0.05)');
      r.setAttribute('stroke', a.color || '#ffb300');
      r.setAttribute('stroke-width', a.sw || 1.5);
      g.appendChild(r);
      // Geniş görünmez hit-kenarı: ince çizgide de kolayca tıklanıp seçilsin
      const hit = annoRectEl(a.x, a.y, a.w, a.h);
      hit.setAttribute('class', 'anno-hit');
      hit.setAttribute('fill', 'none');
      hit.setAttribute('stroke', 'rgba(0,0,0,0)');
      hit.setAttribute('stroke-width', Math.max(a.sw || 1.5, 6));
      g.appendChild(hit);
      annoLayer.appendChild(g);
    }} else {{
      const fs = a.fs || 14;
      const lines = String(a.text).split('\\n');
      const t = document.createElementNS(ANNO_NS, 'text');
      t.setAttribute('font-size', fs);
      t.setAttribute('fill', a.color || '#c62828');
      lines.forEach((ln, i) => {{
        const ts = document.createElementNS(ANNO_NS, 'tspan');
        ts.setAttribute('x', a.x + 8);
        ts.setAttribute('y', a.y + 5 + fs * 0.9 + i * fs * 1.3);
        ts.textContent = ln || ' ';
        t.appendChild(ts);
      }});
      g.appendChild(t);
      annoLayer.appendChild(g);            // getBBox için önce DOM'a girmeli
      let tw = 60; try {{ tw = t.getBBox().width; }} catch (e) {{}}
      const bw = Math.max(tw + 16, 30);
      const bh = lines.length * fs * 1.3 + 10;
      annoMeasure[a.id] = {{ w: bw, h: bh }};
      // Arka plan kutusu YOK (v2.9.40, kullanıcı isteği) — çıplak yazı.
      // Görünmez hit-rect tıklama/sürükleme yüzeyi verir: şeffaf ama
      // "painted" fill (rgba 0) pointer-events'i yakalar, fill:none yakalamaz.
      const hit = annoRectEl(a.x, a.y, bw, bh);
      hit.setAttribute('class', 'anno-hit');
      hit.setAttribute('fill', 'rgba(0,0,0,0)');
      g.insertBefore(hit, t);
    }}
    // Seçim görselleri: kesikli çerçeve + (kutuda) köşe tutamaçları.
    // Kalınlık/boyutları __annoUi ekran-sabit tutar (1/scale).
    if (a.id === annoSel) {{
      const b = annoBounds(a);
      const sr = annoRectEl(b.x - 3, b.y - 3, b.w + 6, b.h + 6);
      sr.setAttribute('class', 'anno-sel-rect');
      sr.setAttribute('fill', 'none');
      sr.setAttribute('stroke', '#26a0da');
      g.appendChild(sr);
      if (a.k === 'box')
        ['nw', 'ne', 'sw', 'se'].forEach(c => {{
          const hd = annoRectEl(0, 0, 0, 0);
          hd.setAttribute('class', 'anno-handle');
          hd.setAttribute('data-c', c);
          hd.setAttribute('fill', '#fff');
          hd.setAttribute('stroke', '#26a0da');
          g.appendChild(hd);
        }});
    }}
  }});
  if (__annoUi) __annoUi();
}}

const annoNoteBtn = document.getElementById('anno-note');
const annoBoxBtn = document.getElementById('anno-box');
function setAnnoTool(t) {{
  annoTool = t; annoDrag = null;
  annoNoteBtn.classList.toggle('active', t === 'note');
  annoBoxBtn.classList.toggle('active', t === 'box');
  viewport.classList.toggle('anno-mode', !!t);
  const tmp = annoLayer.querySelector('.anno-temp');
  if (tmp) tmp.remove();
  if (t) annoSetSel(null);   // araç açılırken mevcut seçim bırakılır
}}
function annoSetSel(id) {{
  if (annoSel === id) return;
  annoSel = id;
  annoRender();
}}
function annoDelete(id) {{
  const i = annotations.findIndex(x => String(x.id) === String(id));
  if (i < 0) return;
  annotations.splice(i, 1);
  delete annoMeasure[id];
  if (annoSel === id) annoSel = null;
  annoStore(); annoRender();
}}
// Seçim görselleri ekran-sabit kalınlıkta tutulur + mini bar seçili öğeyi izler.
// applyT her transform değişiminde çağırır (pan/zoom'da bar öğeyle gider).
__annoUi = function() {{
  const k = 1 / scale;
  annoLayer.querySelectorAll('.anno-sel-rect').forEach(r => {{
    r.setAttribute('stroke-width', 1.3 * k);
    r.setAttribute('stroke-dasharray', (4 * k) + ' ' + (3 * k));
  }});
  const a = annotations.find(x => x.id === annoSel);
  if (a && a.k === 'box') {{
    const hs = 7 * k, b = annoBounds(a);
    const pos = {{ nw: [b.x, b.y], ne: [b.x + b.w, b.y],
                   sw: [b.x, b.y + b.h], se: [b.x + b.w, b.y + b.h] }};
    annoLayer.querySelectorAll('.anno-handle').forEach(h => {{
      const p = pos[h.getAttribute('data-c')];
      if (!p) return;
      h.setAttribute('x', p[0] - hs / 2); h.setAttribute('y', p[1] - hs / 2);
      h.setAttribute('width', hs); h.setAttribute('height', hs);
      h.setAttribute('stroke-width', 1 * k);
    }});
  }}
  if (a && annoEditId === null) {{
    const b = annoBounds(a);
    annoColorInp.value = a.color || (a.k === 'note' ? '#c62828' : '#ffb300');
    annoBar.style.display = 'flex';
    annoBar.style.left = Math.max(2, Math.round(b.x * scale + tx)) + 'px';
    annoBar.style.top = Math.max(2, Math.round(b.y * scale + ty - 32)) + 'px';
  }} else {{
    annoBar.style.display = 'none';
  }}
}};
annoNoteBtn.onclick = () => setAnnoTool(annoTool === 'note' ? null : 'note');
annoBoxBtn.onclick = () => setAnnoTool(annoTool === 'box' ? null : 'box');

// Ekran (client) koordinatı → kanvas koordinatı
function annoCanvasXY(e) {{
  const r = viewport.getBoundingClientRect();
  return {{ x: (e.clientX - r.left - tx) / scale, y: (e.clientY - r.top - ty) / scale }};
}}
function annoId() {{ return Date.now().toString(36) + Math.random().toString(36).slice(2, 7); }}

// Not yerleştirme: araç aktifken tıklanan yerde DOĞRUDAN yazılır (typewriter)
viewport.addEventListener('click', e => {{
  if (annoTool !== 'note') return;
  if (e.target.closest('#toolbar') || e.target.closest('.tool-btn')
      || e.target.closest('#detail-panel')) return;
  const p = annoCanvasXY(e);
  setAnnoTool(null);
  annoOpenEditor(null, p);
}});

// Yerinde yazma editörünü aç: a = düzenlenecek not (null = yeni), pos = konum.
// Editör canvas içinde olduğundan zoom/pan sırasında da notla aynı yerde kalır.
function annoOpenEditor(a, pos) {{
  annoEditId = a ? a.id : '';
  annoEditPos = a ? {{ x: a.x, y: a.y }} : pos;
  const fs = (a && a.fs) || 14;
  annoEditor.style.left = annoEditPos.x + 'px';
  annoEditor.style.top = annoEditPos.y + 'px';
  annoEditor.style.fontSize = fs + 'px';
  annoEditor.style.color = (a && a.color) || '#c62828';
  annoEditor.dataset.fs = fs;
  annoEditor.innerText = a ? a.text : '';
  annoEditor.style.display = 'block';
  annoSetSel(null);
  annoRender();                          // düzenlenen notun SVG'si gizlenir
  annoEditor.focus();
  const rg = document.createRange();     // imleci metnin sonuna al
  rg.selectNodeContents(annoEditor); rg.collapse(false);
  const sl = window.getSelection(); sl.removeAllRanges(); sl.addRange(rg);
}}
// Bitirme: dışarı tıklama (blur) veya Esc. Boş metin = eklenmez / silinir.
function annoCommitEditor() {{
  if (annoEditId === null) return;
  const text = annoEditor.innerText
    .replace(/\\u00a0/g, ' ').replace(/[\\s\\n]+$/, '');
  const editing = annoEditId;
  annoEditId = null;
  annoEditor.style.display = 'none';
  if (editing === '') {{
    if (text.trim())
      annotations.push({{ k: 'note', id: annoId(), x: annoEditPos.x,
                          y: annoEditPos.y, text: text,
                          fs: parseFloat(annoEditor.dataset.fs) || 14 }});
  }} else {{
    const i = annotations.findIndex(x => String(x.id) === String(editing));
    if (i >= 0) {{
      if (text.trim()) annotations[i].text = text;
      else annotations.splice(i, 1);
    }}
  }}
  annoStore(); annoRender();
}}
annoEditor.addEventListener('blur', annoCommitEditor);
annoEditor.addEventListener('keydown', e => {{
  e.stopPropagation();                   // ana kısayollar (Del/B/0...) karışmasın
  if (e.key === 'Escape') {{ e.preventDefault(); annoEditor.blur(); }}
}});
annoEditor.addEventListener('mousedown', e => e.stopPropagation());

// Kutu çizimi: araç aktifken sürükle
viewport.addEventListener('mousedown', e => {{
  if (annoTool !== 'box' || e.button !== 0) return;
  if (e.target.closest('#toolbar') || e.target.closest('.tool-btn')
      || e.target.closest('#detail-panel')) return;
  e.preventDefault();
  const p = annoCanvasXY(e);
  annoDrag = {{ x0: p.x, y0: p.y, x1: p.x, y1: p.y }};
}});
function annoDragBox() {{
  return {{ x: Math.min(annoDrag.x0, annoDrag.x1), y: Math.min(annoDrag.y0, annoDrag.y1),
            w: Math.abs(annoDrag.x1 - annoDrag.x0), h: Math.abs(annoDrag.y1 - annoDrag.y0) }};
}}
window.addEventListener('mousemove', e => {{
  if (!annoDrag) return;
  const p = annoCanvasXY(e);
  annoDrag.x1 = p.x; annoDrag.y1 = p.y;
  let tmp = annoLayer.querySelector('.anno-temp');
  if (!tmp) {{
    tmp = document.createElementNS(ANNO_NS, 'rect');
    tmp.setAttribute('class', 'anno-temp');
    tmp.setAttribute('fill', 'rgba(255,179,0,0.08)');
    tmp.setAttribute('stroke', '#ffb300');
    tmp.setAttribute('stroke-dasharray', '6 4');
    tmp.setAttribute('stroke-width', 1.5);
    annoLayer.appendChild(tmp);
  }}
  const b = annoDragBox();
  tmp.setAttribute('x', b.x); tmp.setAttribute('y', b.y);
  tmp.setAttribute('width', b.w); tmp.setAttribute('height', b.h);
}});
window.addEventListener('mouseup', () => {{
  if (!annoDrag) return;
  const b = annoDragBox();
  setAnnoTool(null);   // annoDrag'ı ve geçici kutuyu da temizler
  // Ekranda >6px sürüklendiyse gerçek kutu (kazara tık değil)
  if (b.w * scale > 6 && b.h * scale > 6) {{
    const nid = annoId();
    annotations.push({{ k: 'box', id: nid, x: b.x, y: b.y,
                        w: b.w, h: b.h, sw: 1.5 }});
    annoStore();
    annoSel = nid;   // yeni kutu seçili gelsin (hemen taşı/boyutlandır/sil)
    annoRender();
  }}
  // mouseup'ı izleyen click seçim-temizleme sanılmasın (click bu görevden
  // ÖNCE dispatch edilir, setTimeout sonra çalışır → bayrak güvenle sıfırlanır)
  annoJustDrew = true;
  setTimeout(() => {{ annoJustDrew = false; }}, 0);
}});

// Çift tık: notu yerinde düzenle (typewriter editörüyle)
annoLayer.addEventListener('dblclick', e => {{
  const g = e.target.closest('.anno-note');
  if (!g) return;
  e.stopPropagation(); e.preventDefault();
  const a = annotations.find(x => String(x.id) === g.getAttribute('data-id'));
  if (a) annoOpenEditor(a, null);
}});

// Tıkla-seç + sürükle-taşı + (kutuda) köşe tutamacından boyutlandır.
// stopPropagation: viewport pan handler'ı bu mousedown'ı görmesin.
annoLayer.addEventListener('mousedown', e => {{
  if (annoTool || e.button !== 0) return;
  const hEl = e.target.closest('.anno-handle');
  const gEl = e.target.closest('.anno');
  if (!gEl) return;
  e.stopPropagation(); e.preventDefault();
  const a = annotations.find(x => String(x.id) === gEl.getAttribute('data-id'));
  if (!a) return;
  annoSetSel(a.id);
  const p = annoCanvasXY(e);
  annoMove = {{ a: a, mode: hEl ? hEl.getAttribute('data-c') : 'move',
                px: p.x, py: p.y, ox: a.x, oy: a.y,
                ow: a.w || 0, oh: a.h || 0, moved: false }};
}});
window.addEventListener('mousemove', e => {{
  if (!annoMove) return;
  const p = annoCanvasXY(e);
  const dx = p.x - annoMove.px, dy = p.y - annoMove.py;
  if (!annoMove.moved && Math.hypot(dx * scale, dy * scale) < 3) return;
  annoMove.moved = true;
  const a = annoMove.a, m = annoMove.mode;
  if (m === 'move') {{
    a.x = annoMove.ox + dx; a.y = annoMove.oy + dy;
  }} else {{
    let x1 = annoMove.ox, y1 = annoMove.oy;
    let x2 = annoMove.ox + annoMove.ow, y2 = annoMove.oy + annoMove.oh;
    if (m === 'nw') {{ x1 += dx; y1 += dy; }}
    else if (m === 'ne') {{ x2 += dx; y1 += dy; }}
    else if (m === 'sw') {{ x1 += dx; y2 += dy; }}
    else {{ x2 += dx; y2 += dy; }}
    a.x = Math.min(x1, x2); a.y = Math.min(y1, y2);
    a.w = Math.max(2, Math.abs(x2 - x1)); a.h = Math.max(2, Math.abs(y2 - y1));
  }}
  annoRender();
}});
window.addEventListener('mouseup', () => {{
  if (!annoMove) return;
  if (annoMove.moved) {{
    annoStore();
    annoJustDrew = true;                 // izleyen click seçimi/vurguyu bozmasın
    setTimeout(() => {{ annoJustDrew = false; }}, 0);
  }}
  annoMove = null;
}});

// Boş alana tık: seçimi bırak (pan/araç/taşıma sonrası click hariç)
viewport.addEventListener('click', e => {{
  if (annoTool || annoJustDrew || annoMove || panMoved) return;
  if (annoSel === null) return;
  if (e.target.closest('.anno') || e.target.closest('#anno-bar')
      || e.target.closest('#toolbar') || e.target.closest('.tool-btn')) return;
  annoSetSel(null);
}});

// Mini bar: − / + (notta yazı boyutu, kutuda kenar kalınlığı) ve × (sil)
annoBar.addEventListener('mousedown', e => e.stopPropagation());
annoBar.addEventListener('click', e => {{
  const act = e.target.getAttribute && e.target.getAttribute('data-act');
  if (!act) return;
  e.stopPropagation();
  const a = annotations.find(x => x.id === annoSel);
  if (!a) return;
  if (act === 'del') {{ annoDelete(annoSel); return; }}
  const d = act === 'plus' ? 1 : -1;
  if (a.k === 'note') a.fs = Math.max(4, Math.min(48, (a.fs || 14) + d * 2));
  else a.sw = Math.max(0.5, Math.min(8, (a.sw || 1.5) + d * 0.5));
  annoStore(); annoRender();
}});
// Renk seçici: not yazısının / kutu kenarının rengi (canlı önizleme)
annoColorInp.addEventListener('input', () => {{
  const a = annotations.find(x => x.id === annoSel);
  if (!a) return;
  a.color = annoColorInp.value;
  annoStore(); annoRender();
}});

// Kaydet: notlar gömülü HTML. Chromium'da File System Access API ile mevcut
// dosyanın ÜSTÜNE yazılabilir (ilk kayıtta dosya seçtirir — tarayıcı güvenliği
// sessiz öz-yazmaya izin vermez; handle oturum boyunca saklanır, sonraki
// kayıtlar diyalogsuz). API yok/engelliyse kopya indirme fallback'i.
// Canlı DOM klonlanır; script'lerin çalışma anında eklediği/değiştirdiği durum
// temizlenir ki kayıt ilk açılıştaki gibi başlasın (script'ler yeniden kurar).
function annoBuildHtml() {{
    const clone = document.documentElement.cloneNode(true);
    clone.querySelectorAll('.lod-bitmap').forEach(el => el.remove());
    clone.querySelectorAll('.lod-ready').forEach(el => el.classList.remove('lod-ready'));
    clone.querySelectorAll('#sch-hl-overlay, #anno-layer, #anno-editor, #anno-bar')
      .forEach(el => el.remove());
    clone.querySelectorAll('.comp-highlight').forEach(el => el.classList.remove('comp-highlight'));
    clone.querySelectorAll('.sheet-card').forEach(c =>
      c.classList.remove('hit', 'hit-1', 'hit-2', 'hit-3'));
    const cv = clone.querySelector('#canvas');
    if (cv) {{ cv.classList.remove('lod', 'lod-fade'); cv.removeAttribute('style'); }}
    const arc = clone.querySelector('#arc-layer'); if (arc) arc.innerHTML = '';
    const cn = clone.querySelector('#current-net'); if (cn) cn.textContent = '';
    const dp = clone.querySelector('#detail-panel'); if (dp) dp.classList.remove('open');
    const pp = clone.querySelector('#comp-popup');
    if (pp) {{ pp.classList.remove('open');
               const pb = pp.querySelector('#popup-body'); if (pb) pb.innerHTML = ''; }}
    const sb = clone.querySelector('#sidebar'); if (sb) sb.classList.remove('collapsed');
    const tp = clone.querySelector('#svg-tip');
    if (tp) {{ tp.removeAttribute('style'); tp.innerHTML = ''; }}
    const vp = clone.querySelector('#viewport');
    if (vp) vp.classList.remove('anno-mode', 'panning', 'grabbing');
    // Notları göm: JSON içindeki '<' kaçırılır ki not metninde script kapatma
    // etiketi geçse bile gömülü tag erken kapanmasın (u003c escape'i JSON'da
    // geçerlidir; HTML parser '<' görmez).
    const slot = clone.querySelector('#anno-embed');
    slot.textContent = JSON.stringify({{ ts: Date.now(), items: annotations }})
      .replace(/</g, '\\\\u003c');
    return '<!DOCTYPE html>\\n' + clone.outerHTML;
}}
let annoFileHandle = null;   // oturum boyunca: sonraki kayıtlar diyalogsuz
async function annoWriteViaHandle(html) {{
  if (!annoFileHandle) return false;
  try {{
    if (annoFileHandle.queryPermission) {{
      let p = await annoFileHandle.queryPermission({{ mode: 'readwrite' }});
      if (p !== 'granted' && annoFileHandle.requestPermission)
        p = await annoFileHandle.requestPermission({{ mode: 'readwrite' }});
      if (p !== 'granted') return false;
    }}
    const w = await annoFileHandle.createWritable();
    await w.write(html); await w.close();
    return true;
  }} catch (e) {{ annoFileHandle = null; return false; }}
}}
document.getElementById('anno-save').onclick = async () => {{
  const btn = document.getElementById('anno-save');
  btn.textContent = '...'; btn.disabled = true;
  const done = okTxt => {{
    btn.disabled = false;
    btn.textContent = okTxt || 'Kaydet';
    if (okTxt) setTimeout(() => {{ btn.textContent = 'Kaydet'; }}, 1500);
  }};
  let html;
  try {{ html = annoBuildHtml(); }}
  catch (err) {{ alert('Kaydetme hatası: ' + err.message); done(); return; }}
  // 1) Bu oturumda dosya zaten seçildiyse: sessiz üzerine yaz
  if (await annoWriteViaHandle(html)) {{ done('✓'); return; }}
  // 2) Chromium: kayıt diyaloğu — mevcut dosya seçilirse ÜSTÜNE yazılır.
  //    suggestedName = açık dosyanın adı (srcdoc iframe'de yoktur → _notlu).
  if (window.showSaveFilePicker) {{
    try {{
      let name = '';
      try {{ name = decodeURIComponent(location.pathname.split('/').pop() || ''); }}
      catch (e) {{}}
      if (!/\\.html?$/i.test(name))
        name = (PROJECT_NAME || 'schematic') + '_notlu.html';
      const fh = await window.showSaveFilePicker({{
        suggestedName: name,
        types: [{{ description: 'HTML', accept: {{ 'text/html': ['.html', '.htm'] }} }}],
      }});
      const w = await fh.createWritable();
      await w.write(html); await w.close();
      annoFileHandle = fh;
      done('✓'); return;
    }} catch (err) {{
      if (err && err.name === 'AbortError') {{ done(); return; }}  // vazgeçildi
      // NotAllowedError vb. (iframe/ortam kısıtı) → indirme fallback'ine düş
    }}
  }}
  // 3) Fallback (Firefox / engellenen ortam): kopya indir
  try {{
    const blob = new Blob([html], {{ type: 'text/html;charset=utf-8' }});
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = (PROJECT_NAME || 'schematic') + '_notlu.html';
    a.click();
    setTimeout(() => URL.revokeObjectURL(a.href), 5000);
    done('✓');
  }} catch (err) {{ alert('Kaydetme hatası: ' + err.message); done(); }}
}};

annoLoad();
annoRender();

// === Cross-probe köprüsü (birleşik görünüm için) ===
// Bir iframe içindeysek parent ile haberleş.
const IN_FRAME = window.parent && window.parent !== window;
function crossProbeOut(designator) {{
  if (IN_FRAME) {{
    window.parent.postMessage({{type:'xprobe', source:'sch', designator:designator}}, '*');
  }}
}}
// Parent'tan "şu komponenti göster" mesajı
window.addEventListener('message', ev => {{
  const d = ev.data;
  if (!d || d.type !== 'xprobe' || d.source === 'sch') return;
  const desig = d.designator;
  let comp = compByDesig[desig];
  if (!comp && desig) {{
    // PCB/3D'den FİZİKSEL kanal designator'ı gelmiş olabilir (hiyerarşik Repeat:
    // R103_diffI2C_1). Mantıksal tabana in: önce kanal indeksini at (U2_1→U2),
    // olmadıysa oda+indeks sonekini at (R103_diffI2C_1→R103) — v2.9.30'un
    // Excel taban-designator fallback'iyle aynı kural.
    const m1 = desig.match(/^(.+)_\\d+$/);
    if (m1 && compByDesig[m1[1]]) comp = compByDesig[m1[1]];
    if (!comp) {{
      const m2 = desig.match(/^([A-Za-z]+\\d+)_.+$/);
      if (m2 && compByDesig[m2[1]]) comp = compByDesig[m2[1]];
    }}
  }}
  if (comp) {{
    const sid = (comp.placements && comp.placements[0])
      ? comp.placements[0].sheet_id : comp.sheet_id;
    highlightComponent(comp.designator, sid);
    showCompPopup(comp);
  }}
}});
</script>
</body></html>
"""


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 3:
        print("Kullanım: python viewer.py <project.PrjPcb> <output.html>")
        sys.exit(1)
    generate_viewer(sys.argv[1], sys.argv[2])


# ============================================================================
# 3D görünüm için Three.js (r128 UMD) gömülü — offline/file:// çalışsın diye
# gzip+base64 olarak saklanır, üretimde açılıp HTML <script>e gömülür.
# ============================================================================
_THREE_MIN_GZ_B64 = (
    "H4sIACQKQWoC/9y9aXvbOLIw+v39FYrvnByRomRRkmVbNuObrbszk3RnOuktGY8fWoIldiRSTUK25UT3t9+qwk5StpPOzJn3ZBGBAlAoAIVCYSvs+v7/afiN"
    "/3eejFlaMHQ/zZbrPJnOeKPXDbvtXrcXNt7OcsY6vxeNxys+y/IC4715/ezX9kuRrv1iwlKeXCQsHzVevXgL4bv/58HFKh3zJEubPGDex53s/Hc25jtRxNdL"
    "ll002PUyy3nx8OHOKp2wiyRlk50HKnCRTVZzdsKaMpY32lHoDAaR6uFD8e3Ei8mJcDbf78h0O6eQ94g1mzyqy2Y6z87j+dtZUpwY54h/+lSw+YXXefvdj8+f"
    "Rx833qbJISBomiJBgVYFaxQ8T6BQR+MsLXiDRTth72AnSKOw2w2SqA+/OfyGQQG/vSCG336Qwe8gmMPvMBjD734wi0LWD1aYKgwm+OkFS/z0gwV+BsEFfvaC"
    "KX6GwSV+9oM1fg6Ca/wcBmfwCXvBFX4GwTl+9oJX+BkGb+DT6wZv8dMLnuOnHzzGzzB4iZ/94Meo39/fHwZP6bsfvKbvQfCMvofBi6i/dzDoBt/TNwye0LcX"
    "3NC3H3wT9fcHh73gB/oOg++iHhb/W/yEwU/46QUfot4AgD/jJwx+wU8v+D3q7QHwD/yEwa9RH+riN6y7MHiHn/3gr/jpBX/HzyD4G372gr/gZxhwHu0PD7oB"
    "40BIdzAIUuE4CBJoc4jSYAW0zzwuikbOP8aTyfNLYNaXScFZynLBnJdZMml0I+AsaObO2VwGAnM2SxDkBtnaaTn2kUaTvuenkBY/0ftTL2iHEthJgAmvf7ho"
    "Mu/hQwIsV8UMfJtZXNQQllw0t9Hm5Yyv8vRBuJUcEaEhEDxQVLXDBxVaNjlbZJfsiwjYlj3kcKRTY5beRxGTR6mV9RHRw6E2OsUS5RDkG3qbzSQpljEfz4gm"
    "7G/3poRVKelgf3fJgbIBOM6nwDgYXxeDdQoio+sdXWR5c84QI/bmtDNn6ZTPjthxcsRaLS99z04743g+F9KBe0caY7qazzebjcBZcGACjYxH3SN+3NsbHnHA"
    "UXDkkSY/DocnO92d0c6O1wIs2RsQLOm0GQ69I0wU8yjs9Qd7w31JZ8ajVzGfdV6/2A2B++cQftDdlaAjJacaY9401T7oHULP3O8d7vkUMY/TSbZoep+g99wW"
    "mN4WmNwSKDmwCYXs7e095KctLO6jRwcPwas84dD29QbCt9PeaclkjMKYTiaDGCUN9z4NBypCKfGw/zD9BAKZgtNy+tTJOjWpZbYJuRKb2sRJkugkHjTYT8sl"
    "y5/GBWt6G139M469KADWl32RKmgRXzdZIJxJ2kyBc6w0Ky56nqw7/l+sxbz/YibCpIS0GbZTz+et1LciLanPyFyhyzT5Q94OQegg73MTb2HHI5KW2VWzJ6gb"
    "s2TeJNc8m0I8wV8vv+/Z9F7chuFingHbb0FxGeeNKY9+IMWgcwE6xg1rfjw7W+YZz87ORtiJgmfPv+39+PjZKOMBfHrgHc15MMWOHXP2008vno3GPAD5vliO"
    "ZlAvq/E8mbA4fYUaRDZa8WARL1/CsB/nI1sjCdIgCXJNeNJq8jbz/GbeTrzdZgruTZCklyA+2EuWL0tpdTqOouQE4kPRmgx+R91NMMcEEx5MkKhypjqpbMew"
    "TVUCCkuznfoJ1EuwhK6/zNKpkzYKdUomksTnRZO4peczjwguFlnGZyD2ttN7DOR2R/xRlJ6EI1CLqNiiwMBEfrPf7vlco4Li/wlkgA7+DX3gvD2vFXYBLQqH"
    "FykfucqhwtcqM42SJT7WbSv0JIZv5ll8Ow4noZ3qzTJn8WRka3Iqpd/s7LWd1FgPjE3Y5EcC1KXSAwoMYU0Q0/y/euFgf3DQHw72vQDlNmgn+35swwOICJ1x"
    "V0OGm2DCpm+zH7dRlnEswuRt9oxN62PMIUZSvM6uWP7DxdurbLSEXgH914IseEBVa4EuOBSQ/30FXSlPAeU3ebZ4nWcgyp6D/l3fY+SgJsafcVaAWkvOIklB"
    "ty2a6W7PA/U2Fo4xQJqsBZ0KPDOAas+KQtrCM6EQ6VliSIKsBJ4FhkjPUXGVgE7QRCpA1O78+tuvOyPegSI0M38WzP0V/J8EmT/2js6hnT8cUbTf3v2moolg"
    "EbUc7d2v70w0hWlWifbru1KmC/i/rGb6q5XpMlBRK5n+ZmWqMDmZwqwlXs35COs9m7POVQwyf4fmJh2s+J94Mi9Gjc5tDdn0GiwdZ6sUgtmkEcMwk35Is6u0"
    "keUTnLXttHJQulC3JU35kotmzldjDn0RdBYQQF1QmVCzuo54QI51xDag7TSukgmfNQ0zUqRNoUO4SUjxZwwnmKUEa0ogg1SKNaTAqnF6eJkGcmC0N6CMxbnd"
    "LUpRuY7667ZYGPhbOdBO+TRbgGwmpRSJkhzJBUc2QBoKbMxq5EY4UrSW2pTP8uyqkbKrxvM8h4reIcW4kRSNbAVK3kUDJNGUYfNwVNI1RViNFiE1ZDhlU1Q4"
    "ZfoyIjbjOWRq2g6TET6HXyhXWeseqMHLdU196xjgXIv6hemZ09ZGXT9p1vH/zwzz6wH3Y0qvASzdyNL5uhGPx2zJoQQpa4BKvlpAPXUaP0EtYEyRrGg24qBx"
    "3vAaCZAOo0JnxxMUWVGQGm8kS9SyiG5pqj0kewvvtTTzKdemhLxaKy3m1Ax4Te1gNkymr0lOBPrMJtFX/WN1/oVViynvV7UQ866qtaI4Vdu2qrZtVS3E31K1"
    "bV21bd09XeTVqm27VdvWVbuAjpAs51U29S26fMOoKv4W2nxNm3LBhPYSFNNKzF0L/67BL2LXYy/lHe5Ct4yXAIDRIE+u+yQF7KnwdSBn52uYrvEOmzNsM71A"
    "IKsned899WFwft8/9VP4DE9VTSXvQxEyECH7p7ISkrTarfW8RgmBzrWnELlhUFSQDhIVzIm2oMKQbahUmI2K5gJ17W+SAC5NiuaIa68GOWC2Yq6VSLOzUo10"
    "W4YWklsyq8RyM3pJyw8iI2fRRaxLND2nQR0OSj99Cr0y49RnnHo4uyMVvLY8IkgWwimDFaLbAmeQtWgopA6LCVBIctBbJvVYRFAdGivEwfM2e8fyrIJNfI67"
    "JxXiRttLLT7VROtqIkVDyqagolWybzsjpvQqUZDxqnzBYUIKGODRzTjPiqIu1rqtY11vBJ+8+aNZiScGDxVVqGOKq5y5ffFHzpu3pPKgJ6ezmPM4fVmXHues"
    "shZbLgCSplm+iOfJTbl+XGFoczyy9Qa0kzkr5wME9JqqHttOlq9f4OoiRBizt1l58cIU0MR588cqBrW5ies0ddCStG1zS+KCZ636JQMhmvqpqaJnW8kw9YTY"
    "KnWFWD3URpVMcCvMqshOZaTa4PJEre6AglFVlaVENA0bKTDhsIZae1HA6DEoW4l8P7X1GRSkVAAJ3jCoybnDvoBArvJeP3wIUZVvvbmAic3jPI/XtCbSrWT7"
    "nulhC9ytUA5VPJOJovenbjpIoEZJkUANlZzyerK6uGD5Yw4j6/mKs1Jhzer2w4e3KlHZxQVNb+Kicc5Y2hAL7pMGZtHo1GXkaZUJK3OKUxZmJA8CflMAkGwc"
    "pMpjKfPs8UFN0jFqoufp6Mttbi1quFXmnftpu/CTlqWu5+AtQBPQ2opaL6kV0TLMlc42cLO55B1a9MNl+k5SyFqLHnTldHTtTkfl5FDpMdH7MOjSX/09DZRm"
    "qmTFo65pISamOXoKjUrTCIhmDSsXUHgbMNWZsryBK0ZFw2BsrEjZhXmprd+qiapYJwmKIA6yYK6aYhw5NKtKHoPSBRriGDSsKIFP7zSK4dM/hWntGPStKIfP"
    "3mmUwWd4GqXw2T+NCvgcnEZzUX0J7b7ydan+kZxKxahxWc7IHMGlaEPZVSGUIaEp/AQMaU3hB1w9dPXQ1UdXH10DdA3QtYeuPXQh6ahQMqQ+RQWSYQFS+JEi"
    "4Jrn8Zg/iYukIlCwJLieIVvqaTZfLYQKGXS9gN0SDMVNbwnueXpWb2IMnHqpV5axapmoi4EoCtUKFRl+DkXVULkxoHvq3T7LUHCiYMwKtaG0WYKcuHcaOWOq"
    "CbAVRtPMScSMJy/xQCEaOxatmonmm4t2H4sGnomWXAkWmIi2XopGXdB0IriIcDoRTCOaTlzSTCLACQUkxxkHJD+DDyS/gg8kP4fPwamq7ByZrvAXrdi/bGX+"
    "WZAjnxX+BQDWALgCwBABUwBcA+AcAMCdc0gyhiQzSjJAwAUA1gDAJPsImALgGgCYBNh4BUkmkGRJSfYQcAGANQAwyQECpgC4BsD5tjlgbXeyexDOBrGviO9Q"
    "fkP5Hcjvvvz25HdPfg9OzWyScZaDwh6n9g6f24RAB7YBtDnWOsz8sJ6hnbFFQOBjG8Tw2cPm5aJ5ObYIiCqrDZhf+OM282N/3k793B+DChP7WSsB97ydQGi2"
    "oY2SfxUhwGdjv2hD/sBqkHN77OfAbXM/h3EpA4Zj/gxoWgFFE9zixQ3ipVfurt2g9Fdt4y+icHepezcy3MxfBEhp1EyghGM/9QgATNKMYTCEIgtAn7gmQPqj"
    "5thnEJKJkD1KmwPNTACASSfkANZrplCGuQoBxmoWkBaqlgCkrMDwWCwz3Ev8SFvGR7cyFo+k8AGSmZDBfRxUEN6TUlqKoqGC70n5zIQs3j9VjAV6xfekPQpp"
    "WJY6FVHZUW3fsag2JXiRas3rju5BVU9CleuCcE061TYVk6qb5C7XBaAKpsJSDVPxqGqZHmCA8J8u3yJZF1A8d5xWlGVGXcpxC0OrS7lXEf/QiEEKLAkN18z8"
    "ojX3Y69VtHgAXDAPgBPQ0YR2LkAsQVDcYs4IXOBCXs1cvlwtqZQaqZQaqZQaKUkNKARJDfzu09fSB+0q1wXjMCSagnHUCt3OmgvJXQjJHQvJnQnJPReSe0yS"
    "W1GYUMP5eQurJKF2gkJj3STULH4MnnGACCJk8xajeAPyFeDDiPvkg0qCmKYPzGN+3zpqRdyncQgrSnj2Tqm2hOfglKqsBfTIaAPpoWj70qP1ET0luYeK5B7t"
    "OKSDHSCI2Ht+Ks/amMM60tGtTmQUlhSwpIAlBSxOpu9T6KXv0xY7tbmxZmpza2XRhCcV/YymO6noVayllDl0K3UO3UqhQ7dS6dCtlDp0K7UO3VqxK+0VNOv2"
    "CryOVRE2vd5mg7VxzY/W7uxAKnJmdnDGP8I0mifjBgivZzDz/+nHl/LE0O4/J+Af7SYdzgrQzjtFPvY8o1jmYxwxrKOB+nDhd29fvXwap5dx8VwQ5KaiY0GY"
    "ltMMACfxtYmAb47YvGDmiNk1bhRf82iSjcXS+RimF5zJBN+/ae7MOF+Odnevrq46V/1Olk93w8PDw93rGV/Md4KdMWWw43nBNe/QBlskvwgQG2iRcuhjTdc0"
    "ZXyapRxU7eZOb7LjHTm0v1jEU4a1d5J2liuuvcCeOFqO0s4kj68ILECBylXlBbo4ZKO2qZgIfdTrDg4+fWIyDnnrNxwINW1mdkw7jgRdODHDYQZmzDxr/L6c"
    "NqC3NJYsR2GOBcA5WgFIdwIko8MzxQc7Cabf/X3JpjtBZ+h5o5rQZTrdkfx2hZ1Y8NU5NPE1Z+mETg66+6HnvPPs+TePf3r59uzFq8ffPoeiW6BXj1+/fvH9"
    "t+IgaAgilj65OL1ZiGObsTiFmYnTm/MoxMOgrO99LFZL3LAN5LkYwZliL5eLPhLsJJOd4ONlPF+x0RVIm42cXa9WySTCM1/Cm8YLFu3sCA+VVG1LLJLlIl4W"
    "KDGEP17ioRM5fnSu8nj5JkqN5y1MTmXE6TfJHHRPmJhKTKkEFHLvLE2KjAO5azk/7VAbcZjXCjpwDzqN598IKB3yoQDseDDPJbdYLolQYFyS7iZLlLMlizU8"
    "VKNpZ8wQa018HAiTDESqIh9lx+MVntcCycBAiNgBhGAt60idMXol60pFteZkj+fLWRw9CGU558nyNx1tlS7j8YfH82SaYseOBgKM2/ATrGoxznXwhJFFYJZK"
    "wugE4YrcUhOTqx6C0k5ZnzG1ptZppG9t15wKk761W0t2XaqY0oe7uPcR5mJhgeahdbu+xJGcPg5Xiq/Lm1y59IFMl1e5ctk8y8XX5lwBelvmX27cZU7mxl3h"
    "aW55HPbm0lHL5bwEsDheHE51mkzUnPQ5nK+ChM9hfhUkfGX+56VGrnQEXgHZERVy4fPqOwgvQ7b0F14B2f2Hi299J+JlSKlTce1Uy75/ffPD97YaZ84Qf/q0"
    "U9ApWzPocxzRH7CHD82prg6Ol6ucFe+1hD01qkBNoB5xPy4Yj0n9+Cg7+WjQ2Qswq9HOW5FwR51kzHIN6wiidzYB4htpzAH2mpHuRoHk/pHdKQLBGKP3t/R5"
    "0OaJsWSkeqEBSj7xkYxULxNgEi+ZauTKEexwMmW5O54GopuMrL4j6kR3iUC14chp3EB3ypHbXwPdkUduHw9MVx2V+nFATDYyfBeUeXJUy7xBif9GdWy6cU6a"
    "GzmnuDCxYEf2ufaEGhr0w8QM5V5gGBLZViQr3ieSFz9qVZSUaFCQhTKdeN5HVjp4DgpJok6x8+NUTFToKkCBGpFkwBMmbiW84k0RSLSj4mSDPW+Dii3OLQGA"
    "Z99dwqKPxL7CF6zy+YhtNqkU+AKqVcVPn5q1fSlKYbZKVwHEYojal7NuBnwU/UnG2dnItQdkrJ8u5SzA7iHQIInuv6TAd9yzE0boeXjg4Lj76RN8HoWeOmKl"
    "uVqetZqNuDxcYm/6dq7tY36NlYp03D3pjkI7aDLC2yF6g6+E5L963gl3N80Bips227LdUKHWgvB1DeFvLcLFMRgHw7pCOEX6PMLXknBnUx+guMG0LVv7gJvo"
    "lHj/BuKGbXG6g04IpoxNCjFCYfM+oE7x8KGtTIFCbM6nv7J2z+vugOGcjeYYcgYGqMpTOjsYRo0tSJyJXw0WJ7weDWX0JOHAqyUEVsjJGe84E12oTxxmTj7S"
    "YCOkgJkwk/Ikr6lQPC+gidlITd/ExGykZnFSGFNUW7ejQWczqp26ScExavyUxudzhnO0guUJbUE31LC24wV4la48c5IjclCdPsGc49yd+ktUZur/pvacapDS"
    "hZ2w5ryqcNzouU2UbD3BelNzgvXmthOsV3UnWK/MCVbn9P89KPvcc61InUr8NU66visH3liBv5QDr772GVmC9kYyZwfaH8ks/6PO00qCnRpT9Do19e86eSu5"
    "QvLEZ5zDxZbW6SJLGb4CkX41Cv/MYd3B/+xh3c5Nyyobuq++4BAvYTE4/uTBXlHd4NVVDlSB9+rPH/uVxVU+UeA/dyR48D97JLhz07YasG014OccFSYsBsef"
    "PD4sGrDtNmBbN+BnHi6G5rJK6OsSfv6hY8Jk8Eh2Muru4I6jwkJ8yKMBV7hXu+3wcCEODxfiiHCBW9UJfMLeqZ+rGivEMeIC97MxzqGM09dxbiBOT8QZijhh"
    "V0Ya6EhXEKkvIu3LSKGMtKciffb5aRi0Hl/D1AmPEuIGp7lJY43kPXHBNaZ9NBCoR87uGp0cRFX1yqdANbk5Dll7cCJ5OgpVbXRVkbua4ZEbdpnFYLvMYq9d"
    "Zo6rOLT+KKfdZseWpoRif/NIXdfqdMMgjjoh7rGbfax5lOFW0Bg+dKIjw42bFXxCPNGRiRMdGZ5oWcCnhyc6MtzzmWKULl2p1hOAcXvlHRcPH2rArL1wAcv2"
    "BQBoTmZStSBVbKdqLVzAsnXhAuatSWva7gPQ23LgSdTUEVOXpI/UwQRIixfvoBKaE+G6jJpT4VpHRMzuILiOiApwnUWU/e5A8/ujDFTzR5cn/Lg4aQpFs7Pf"
    "3Q+7w/0DXOM3HmjX1GIN2mtd76YQ53o39UYZIMkEEgtBEnXLSBILSYbbt+tdaNjobDfxRpc1CBxykLtyC8ElIrjezSHe2a5aJqQNbdoNVzyGHHRpJWtetJee"
    "T78talefflvNVXsMbvzVHK/b6dI77nS7IcziLqNQH2YkJLuXis0Jj/LeRIRKea8i0+Fkq4d0KfArXz+AfN0wEL6dG88hQodd4cbXV7+2oElQYXUkqDCbhH/f"
    "dQeHRKDPiikHqxuvhmCg1op5pbTRf/MVCof4Sqxawiuxrv4vv46h6sAKKXGYFXL1Fa9wqIxNQClfE3D1dS99qJytkFLWVsjVf85FERzy6VNNdFNNZIokPtVE"
    "V9VEV198JQWIaztTS+m9uveFlZbUTnHaI9VTkBOfd0FF4xDTHIVGrMR80eWV2zD+mYstJcBNGXD1xVdf/udugODEEkcByQieNcNEkS8Z4l9yX0ROl3HYaeMY"
    "6af2tBnHGFTCv+R2CfpulO8GfVfKd/WFN0+QWDpbpWmk01VffitFHNRS3U8c1VK976vdWBn8i2+sUBNOcYmRGdGFgF/MnZY/dadEy30HeFV7++RN3e2TgVlk"
    "frv9HJCoWnViR+QhTmQJjzyNJfvGhC0hSM4Ai3FSFJk4s/JGnE/mSv1WgW9ZwfXxksuEXaFRu0qKNAJ94uNGnimQS+QY6ZzLrcUg1RvEqdyRTeWpiNTaLk2t"
    "PdVU7c2mYk82tTdPU70p6zm5yt29Ei0C6laMG+RWkxvmVJoKKh89MFxdDkLDYrXHEhQqc+zDQqKBJ5Z7NLWaUTC/OU6Q2uBPnxyvbFNgoXEydxJSZk4AGUKz"
    "AVaeavvDSmvDT1wv2U/CEUL6aSFDtZDcAtIMq3eBDNcGhGxksi9xGO1rvMEhhrqB3m4R7Ww25WTjRhGTANGiEXJt835dJv3q/CQx6q1lz+1j+spApUs6IV/l"
    "TJR7hNOuA7MlZ6qCi2+JXHk8R/m9UkvpKpD0VvmY38Wv/E6u5I5XLcN91tb95q0rin9h59++/BEkL8vfCsN6Wig/N0L57S1CWXhkgWDSNmdFNCg3A0WlOnTM"
    "+bmpuHLJUeN5DamvUNMS0eqpfnyf/cozvbN3pvcFz/TG4BnuWcrzzoVS5Jytxbph3axnjmDsoMQdkdozQzy0X87GEHEiFu//WIg4JnHRbPwRB40/zoMGd1fy"
    "02pUpCvxHFq/mce8fPcCl5rwflnS6oobZklL3jFLUG2akaOvzlWtovx9gTEn5AhxfRIdPVyhREf/VF0CMquDoCV18Q6lVKXmSosaByhKm1KLmnmYMqxJuVIp"
    "Jyrl0km5oJQzkMiLT58y+Kw+fZrDZ/Lp0/gB3kYS13iisB3rNePMX7Xm/qQ19pctvHGURuxR1D0JR206MNxmPh3qSR59v1qcA3s+f/3mxcsfvldrDfaaXuIF"
    "0hyXuPaeB8xPvSNu3fPwC283tyx1NWOCbBSy2E8xtwzI4q2Vn0NDzME1Adc4GoNrCa5ZNAPXAtfYoZKgMObCV7hrrVH6GZRsDiUbQ8mgVjPce5jjzxh/ZvCz"
    "2ZCKW9ckqjUU46jpi8VaZTZSdOBtxUTcViQmmismGksmgiIAj+AKt2SeiWKepWIeWwOPoW3G/qyV+ZM2mgSTxGYExuabtWNf88ScwJNW7K/aaMNLFgQaGGLN"
    "ALRqo00xcYLguqTZngmDWdd6u8GSAln6dIYbwU9BOJ3H4w9Nj1CUL/yeCRNaa4NifReKmzIKcejhxqC4uQvFVRmFOAZxZVBc3YZi69mIu6XgVqRqke6e2+Vn"
    "1yoflYvKo37APrP2zM+sTfMza9f87Epv222nUF6mExbalLndB/zTpwegpxUE9ipHBKryvOPgEduy7HoJ05gCLb0RuKHOZjZAEZ6BnwNFjbghJjt9iDcRpuBA"
    "lhuLvlg1aDEIagY1AKgYHAfPKKISJmgBMDMWAC3Df3EzUUb/cmHwLxNBE3AkwtBfRkHKrl+h7fqhUTxZ1ysfuyD0HH+p6xx9szaGLHXdz9EHsmviz3QbIAzj"
    "QVrXNt+X4m/X4G9V8L8j2382/nYN/tYX0/+O7B5+Cf770f/bu1/vUT/3xV+l/9d396mf+7avQ/9t1hJv6zR3mUcs1HmhB2ivWk1l7uraep/W3SZguz3bGkfq"
    "bhUI+eIntoTRPpIx2ndlLnYqLfcukqq7xVVjB6iM4LZwIi7a5uIebSEu4Mbifm0WkaGDubh3OxbXdGcRmTtAmwCtuDVDlWL1qGvUhM6epSesWiEoKbIcnd7e"
    "rpL311Fz3M4838j/Zt6eG/9N1CzaCfjFCWTIJH0Uw8T50czkJPfqxaZ8K23H7ZnJjNBb2UHmdmYJtPaulVnemoNfZxZvzydup04+SLZdLAfz2skX8slaY5XP"
    "FvwzwB9b+LEabPyCUlMSgVHnQPk5x2zv4paf0oQ7Z2LEDVWY62W8ybxWqO/gHrt6qtgbN+vcnWvvkeW78U7U4Bu1zQi6tkbWG3VEAgqaqjMSEL2r47bNcHtj"
    "DcMY3YpPh6HEUrHPnJzQTwvOPnMQoZ9WnH29W0pIg/LSujA95Vhvss+ImCMNyqQU7YwEbby45nnymvbbDGTUpKjZSDTY1ZQmtQ88HKnbBHq7MAzYbkmYqCli"
    "4t1pLEaZiDHmHOwYkNnvK7FTtLHcJc3Ij9qhqmHLfWO5t3FczbbRGe4bncldmjPcOTqTmzRnuHd0JvdoznD36Gzr9tGZ3O2xEYmvwSW+Bp3UZO/YRvpyzO6O"
    "j5gd1m4e0/H2kxrm79Z0EzwtgnMxIxE0mYbtFaGG4RWpRnVVxN41ppiTbZ9xjNAZhjWG+x0nrJkMbj1XWBdXHDDcHnx/uzsuVtf0TjnHj3dp1Fc4qmJYhp81"
    "jqoYNsbPVVk5SGFeXfhxC62T5H6mWzYheNbKfRiG/Llu35zgc7TP0E78WLcy2nVJIW4C8FzH3zogmE1DKYyYI4zk0okDlOuS5RdTrtXZQlELcs5VKLYTL14A"
    "eVgxQDWJgER0/py6PWYWH3dVt8D9Z6xDa0AxczrymRlamyo9jmAYHdkkBvEj6D7uRFZdacY6N7Vs6jWwpHCG6yF+TIsox1H9mg0uALGjUh6cbGBIn+m3fqqh"
    "1lTTTzTUmnSSTQ1n6moJl7tmyUTZ3D1iNrZXk+ZBjJM4ra02Q3woAfQKtMdjLMoRpFy4ArRzJUxWpjI19FpDkX1nWoquLOadaVm6ug+LllcgU7kMIdtZDogI"
    "r9sZPtNbw2e0N3ymN4fPaHf4TG8Pn9H+8JneID67Y4f4zN4iPnP3iM/cTeIzZ5d4e3nvsX185u4fn7kbyGfuDvLZLVvINUsgzg7vWWWL96yyx3tW2eTVRauM"
    "/OUyq9PLNZXxcbN57C7GGx4wS+8vty6933ZNyKxPlYa3iHbQm1Kk3ZiN7y0Xer7kHs+/8PrOv/CGzn/uXZx/77WbL7tv84W3avr/ObdqvvQmzVe7P/MVL8x8"
    "4RWZ/n/OFZkvvRbz1S7DfNkcwdTjF04Q7qrRcjynWu++l+P9mcs4m7rMqxXsuxXsuxXsG3bHCz1yE8GgQe1EbiJ8+lRr+daSHAbDn98+kPJj6cyDmk+5u+KK"
    "trQpTmmJ1K6F7TjcRJ/1/IG+07TtKlMurjLl4h0ENDCqllvXUS6uMOXimhNaElVhoKmKq0u5uN6U0xUoq4Fus6pYIt9dYvqC61q2Gddwt5mL+1K5uC+Vy/tS"
    "Od6X8kqFb+rSD0wpMGrv1PMLfWdD14Ms66GM1TexcN1U1chQ5tuV0QY62qbcxvctIdmsBmBMXTKj3bZ5lOG1MD+BmViKd5og19hnMLlN8GoT5k3GrBlMXADI"
    "wJlC1KRUB3OYQ6/8dt4a+7hs7Ld1uccipGgBMG/PIViVdSZC4hYAizYklPfQQCtFK163NPhAm9X5JcvnkxfizTmvHEUi0gv33maV3g95OeW2HCwiLCMaz5Kc"
    "6dfGvkq3GpS6h92tLHaqdCubiaorsZ/1vErnZtcoXJ97VfDfc+3o33S16D/y+tB/2KWg/33Xff61l3r+9Xd3/q+9ofPnL9t85u2aP3Ol5mtcovmq12b+19yS"
    "+Xdch1FvAn3RpEukvd+Mi+LeNd1yIlmbMC5c7r5UIlv7J2L7ROye4J4BDki4d7KmrZOb8ptmtLURm4dLinaqdkxuIrH5UTia2g+pWjSwD0aYXnRUt/lhtk/t"
    "w1PUU3GuuOsu+quV6BLHpZ6h4fU8Tm1l5kdunSLvVEnVqwTNH2kx7WJeoxeK4G3Z93xri9pztrarZg6sPoaV4ts1VK0i9VJ5z90HElmZyjEb5zO8ji52yf/N"
    "TzVpVRa7UundplbiJ1/r7aaybKOeqw5dvFnCFDuBZtlifV+HP4XxBU8OdPJ4kqyAkTvLGejGHZii87iKTkUXAsc5PSAeBfJ5pf9Yh5RcdYDeFPL1irWMaZ1E"
    "Utk/Xc+TdHJbeawY1RJRWQL13tWW6LUi1LefO5K0G5Hp208jVZ+BeZ0VSWXas21uw3B+rBf+cRqssmE42a1B/0aY/y/xoxNFvVQDQkUPZ4pX6yOGVsTktog9"
    "r15rvY5YpaKSGuo1otLYaG2AmSWIgc+8Tf0jPPdK38f0X3yJ86tc2/wKFzW/2tXM/r/7aua/4ibmZvOy7tKlMOpPPeJHcc/xJQ+eCtdjLjfyXpc38kQ8mJx3"
    "A/kfLdFLaBtB+sdTRrSTVJtCj4Hr6x+Zhlj6eICKKwAlkaFfOCHDOxHmltJvQr95RAQU4hPTR5tlzaIuvnejzLJmx/OjrBX1lWTAp3Ay8RROhsw5I0fv9Gh+"
    "zB4+bLJo7gXjY7EVOfaC2XGC5mOjmRfMH+XgzCnCowKcBUXAk5LNGCJsykWlV63oIL4pLcLEPR2nwBVG/lNFp5O3suQtq+CCUek8guTRDE8iSPbMvP/ZOnid"
    "JSCe9DWHRfyBPV8s8WSdLqDYYNZty47TI6Zf9bheQo94siY0eJHIXYvVgx3ZWn6cTtRNS6MMv+CKGcu6XGfPlesWI5MWmJb5GeC4e1gaucULBJU1MVPQwCmJ"
    "jv5VrkUamtFZIRideqVME1TBIJcD0Lk2zhviTY1RRQLn2jhvBM+KU5NFbQaY9tjk9emTwWPAawt8Y8A3tM9NrWsV3dh5Bg6tGwSeZPhWYcek9RqcLhni9jVO"
    "13IYKhNQdlHgSyEoK08X4oRbE5YRd/Z6JXm6cup4CykX7PgldFPKr0C1vUEpidWG/EFhcLtXlTPEMq5hArHeql4jFInNVLAiJ3Bf2EpNu/al1FuWlCG12SBv"
    "l5DoNC4q0xN5RzwKQev0cjfpQRg8CI01OpSP2YLxfH1kZvwoHFfzOU7LOue4epakU2iThw8ZdKjFEoT4EwOFnv29Ei52bALfsm2gjPaj3vw9TMGsyfB4lswn"
    "OUsdc+QwShtz5IzMkddJFHpByRGOgJfHSVqUWvcBrsy4HbJz/cj0VbKI7fbMztqEk9/uouQ34XSkQ2SMlVTTsNfH4vEIoZMK+WAkyrW8NUJ5q5hrFXNtxVxb"
    "MW9UzBsV88aKSWLkdZzHC8ZZXrPcIy7e39EnDYJbOibTHZPe3GyaZTAqsbfbNCW14YFZGKMC2xEdeGCWyqi8dkQHjkfUOW5ijbnbEg+avE4qC0eJEWrEtHCU"
    "GKIst1U8lzEMQTTdruxE0R6CZFf5gELwAqoSBvGadQr1igc2vZgJ+8ph5aRXiaS9R91F5MImFLl70qQHV6TfN9VCikkJjlXjjepSYEhdCtHAGrymDFsm3trE"
    "WwdpTQDWMma5Jag2jWAWDb4pZ3pj4t04CG58q8lGtWkwqDYNMV7AjvEEMYm1GK28p49sv9U2b/NEvEtuPUGghzLzHJx8SEUN5j8BR3yoH9ow6Ikb1IkJelOC"
    "nhP0mxJ0TNAfHOgNB5Re8J0D/IYDRi/41gE+4YDQE8+eRe9hMvUDLr/+gCtm4PkOPd9Jz7fo+RY9FEfEvQ6+k57v0POt9HyLnvYPMvK1DF8H30kPoaFIXXUz"
    "+8GD31FFf4JUAk1QXR7NAeoega6P+zN3l51/4JAfirf3P2NOP2Oe8HNzGlRSe97G7shfIGyt5PcUtUpBlzu1FQXNWucsKzsv+J2ppSaj14NwOFGaQEWS2Vpe"
    "7ZEiM6KUcNym77ldwHrBSEq8qLPnq0ikPL6w6A2sTler5tFuszuDINXPTCBMv1SjbmmGs1mpoym1SmQZu8hRT0/KB3hsHDrnT5+az/AFVBpYLSFtSU9LDJV0"
    "MC94hm+m3pmWxF5d2t6WtCSB78i3f3farfkO7LQ0xNy/vHt3p92a73BL2vuUd//utPX52uuxcu3gGVdHDKznTmunC9snG9WFUUyiwGo+pFRHB47CY/PaXYjD"
    "HmxW4Z7x6L1chvvczymoNmoJ73vheo3DjYLdaNc32vWDdn2nXd9q10/a9UG7ftauX5TrSD/98rtlJMS8r1pEXTo1JiRIu39UHEfxUUFrbr9wew06KNQMJo5y"
    "UHn0ns0vtK+Tg0Jig9YIunFAoDHgtiTudP3C8UnhVLvHUaLc2oo2io62dmXBPBh75nSI8HuPYq0+bPRrsoLMP3Q9/6rr5TfteqdrSCyk/rV+IRUGn7YyRSTf"
    "VZRLpVIcb1ktdZ6h89wUtQtnlZM0AoM1aT1J1QrX6A/eKa+7WYNGKlSTJOo6a2+5vfaW09pbYk78JEFap3vjSpy7fCZL4RjckdPQ6rrVba/xCTxmc2v70pKI"
    "cdzdurglszErIk4W8rLttnlyo3bSYdB6coKpph7GfYuyYeN0kLXt9HXzJGcPTMRrqUoycxpZ4rsIZz7bNjtsyGcYnexpAbJmTlXZ0e1Uim7l69ZYRUmsYfPa"
    "Teuj+2qTgvo/o08G6aMtbYzqtFALrQI6VwpZ5bSMSS1Wkp26CZitVbpNcufCoS5pCcWdK4ivqyuIpJdZuh0e+DejuZyV26SDDlpezbNKimdVbtPuZEPXKgJK"
    "FthNgEV8FYu94R9SPOPevEs9UFnULD/q/vnOnQk6BdSLhu/Kh1zYNvYwt1mNQCQrpKCmN7nd2aWhCotGyKV86mVX6UCq50epcxusrHv/Vp7XVlvt1nNdpvGq"
    "uzC/8pLcJqJ/0xTeFRs7jY5dp5rJqrC1MJkWFTTJE1FJlHzeXopUBf6uh/O/addftIszPdhrV6pdCXMVhZxtUxTkJiuOQKAzSKUhy5Op2V+dqAPk2/QGEd3V"
    "G3Qid7u1brx1kktfPRZuAGpq9plz+B/jNd6W+UxR65BSc9DQCExJ/WaeZR8eV6az5eJ4RkjLhJV73xu051jwW+uMrgxgRfydmwPPWcEKTmyuRzusqHvWUF36"
    "+y4sVySVLJveS2D6YJ1Vq8YMTffErneZeHS/tkhr2mK7wrPlENqbP0xUk9yGaoXn73x7eW8rJYNS2g1Zp0qMmn/n9yo0qxYaWOGOM3VYnDdsupA3iIX5vr9x"
    "Zze5dl8ahJDetZZLTw7T1jQexfubtiKR03qQrWwytH/ZLvURrL6/oF2HiNdyDEzM2jLkLzQ1A48ZAbUJBND8wnbhF2KCMQtWwSRY4vg4ftT10OhnVPhZOw5W"
    "8I3bGdqR9MfB7FFEoatHUXtJjmO0AmqsZY6P0ApmsMKfSTTzm7NW4a9aPT/2Wiu/iXYT0Jd5rbmwO7WKcmWEAecuIG+bmCD20Jxee0ZW0JoqyZFM0v78NMdA"
    "8EmzlAoqIKdkq2j2qHvSzkd6UmrmrXnQzjyYaNfg9kZYfkTbVTYjtqZ1kjUr5Nt0fDYZql4KUYjPqR0lX9A+93161aymVyUPHyYi8V+q529XIgHweTAx8xE1"
    "S0Hh+/dbNJ+ykPw7r+X5RAWArG+nfkoz5NIGk7A6m6vDu7hrrB8RMyIvbyfYt9I2Wn9MW3qeFh93Hz7MQD5hulGsBBUMMBmUYaQ8MVrPvmvPrEZ23rYxZklq"
    "NYkzW+JyN6dOqJaPK4tntmtne7KaT7ojq1rSqN10pDHmITP0WmZ3yDrqnKKp31RYaNe1IKmuzBer5Tqy2gYNUYjKVnWbunVbUxm3la1cHQ+6lW3Fmkr0YUgy"
    "edIUT1uLU6aeg+xIHe0Kd0vS+hoEbgW4hv5ZAd5A57foVcTNsUabadSUm73tFV6cmAPHqy1hCaFX2UowGctKh+fFEGGuAtYQsPZAtBcqsYLQA2slmIxlpcP5"
    "dvHpU/4oEY3VbOaP0k+f0gcRLs4CSSCxmsVx8ulT8iBKPDrAVng0kJw0Y4XtBrDdeP4M36+TO9MSAlTEZZiMZaVDKrJPn2JDRexSEQMVmUsFSNNE9WfNYoJ/"
    "E7zIsXUzntjzgZzNlHgD5I9JZvZKzWotc1RBhubYUweUIihh7lYeg/RMDNUFjvxVlSBhxOExWqVEQeeIuSIKhfVFtMOLFqa8UnA7JNNRG84qe7SCHQOtpmSR"
    "nDW7+aclijlR7Am7Ud0aoTuvx8MkHkAgUs9LqRFha64Xi22U4wgGdKn7YH0ou/HlVh7vxtAUty5xSJFXu8RhCK69Tbx1kixx2pNkKZtwkmyw2hGMJPqCCTNN"
    "cgt3kivnsuoUutleNhvMjus0UNei1MWYR90tm6SynkZAFmtYecK0qAG0T9HcAosnRcNgbKzIUAjDaZS5WmUbz1YiFvcIAqmmBovgQon9aeSUR7X59H33FFTQ"
    "6fvBacTgc3AapfAJe6dRgt9TUDun7/dOowI+h6dRjMA+Gm6fvkeT5/AZotH2Kd7Ljmb4HaC9/On7PlrLn77fR1v5UzQ5EC3wC6gu7rREeUdd1xr4pobGa+3W"
    "Bo5dYnPK1FlutkzeVq93oBHcKEVLuGj5Flwhunro6qGrj64+ugboGqBrD1176Bqia4iufXTtowvrF23pou1ccB0SZsokFLmIbCifkDIKKaeQsgopr5AyCym3"
    "kLKDX7NAUntz5R5FvTu76kWQ/j2ux9CRalGHfVEt0IZMVOVA1A4BeqL2qJIqbc2uQYbA4BEXSeW6T/0lF7xL2JVH1epDQ3q1YltozzpdXJutKhueGcFbiCkd"
    "WcHDI3gbMaXDMHjMBa+3p3TcpbZIyh7z/ZqLjtXH7B43hfLbIto3hYrbItbcFNKdwk9MtyC37BjkBgbqmn7h56ZnkFv2DXLvi6iya/iF6Rzk1t1D+EIRG7m1"
    "Kzm1K7m0KztEaFpO1a5teObjPWzjaBndqcdSMpPzxUb269q6fMXVvlgX21bDld19DEtwCUOHJXoBA8Nyx35jTuoC2diniQ5RZNYmCtBeUzTcCFTE4Ma3QWZH"
    "1OgZeKlJ6XkLarG5ZAE0RiqamJMtUWrCduxngi3yNicgNHvSSsmJzVr4mbaqTUb56wjKiKCMCJoTQXNFEEdbp4KkBI2eCpoKX1JVEJU4eBHhRJFkU7z5K+jJ"
    "W5xwVOhBI/6fS09b09MuVBVhgWNTTXGJJKwaSVW7UJVUT9Bvv35xi6GtWFk90BKtXNCTqQqCJmtx2e8gpuzMbUUNNWO1wd5toycjeuZET0b0zEv0YKlV/fiz"
    "ViroqTYXuYgUcqHYgOiJJAfre2YIwmcFvpSgtmJoAuA1PsgnL9EDeau6AaqpQql6xpIeMla60XJSiiYp3/6E5HLtI7l3yRf0wNYcZkjBmJmtA/eOb63uN2Fl"
    "u3IBTvwn9joozADBf4PPd8PXXqmdlSYyKUQQCGYlBGgWWC+m4rGqE0B13YpC1h6MELtw3isDAtuxVqVYEwYxYGKIjTvDYTnB1l0JF7TvRLhCCl2Da49C0XVI"
    "oejqUegNuIYUii5s38mfM2pnDSpfZtSOEIzZ3VbtdMQae9cm7P7Gri18rvU7JyPLTIOlrzDjyUujXSGU61jozplQjedSCx0LfXsm1OmV0JYnUi1dChV8ITTs"
    "C6k/T6WieinU8rXQuq+lRn0mNeWrCLkjOI+QNYJXEfJF8AY+mOtb/J4GzyPkjOBxhGwRvEQgIPwxQt4InkbIGMHriLgieIZfwPQCvhDpe/hArk8QCphu8Lun"
    "35VCM1ggUK5asf+2lfk/tub+iwCNYgHwHIDPAfgUgN8HuRjLXgHwMQBfA/BJQJbYAPoGoC8B+gygNwjFd6auWjNAugKkE0K6h8BzAD4H4FMAItJDBL4C4GMA"
    "vgYgIaV3qt4A9CVAnwEUkeI7Z4B0AUgvAOmUkA4ReA7A5wB8CkBEip1jCVgXgPUCsE4F1gFC3wD0JUCfARSxQlaXgHUNWK8B6xlh3UfgOQCfA/ApAAlriNBX"
    "AH0M0NcAJax7CH0D0JcAfQbQUp80965qdSxbj8XtFxT/4nsgv1DF0iG/e/J7qOB96VARhypAoQwVThVzXwUonKFAKmz64NWbRZLGuE+qhy23r0ApkGuhcyHX"
    "gkhHroUORVxb4Bc7EhdPs3DxNAsnrsW7wz1xc3iIHYkT107wO7DeOwNKm6BPZTjG+fi0Gegy/gq1NfgFvcWfoO0Uf+K18BAsxGXoazMKz/3CX6ExFX+CqqA/"
    "Rsv0/hjjYokxMqJkhAYjw5hPkTETNEM/l5GhVppopR6AiH/WYkQFYoYBF3NBO/YQWZyJka9L1laZuK2gm1tUEaexHauQizUPJmoHf0Sd4o+ED9E3PBXViT8S"
    "jvYdkI9F5dOvDNlH7/6prHn6lSGhyJxyp8k0dQ9zMtJM2+8xdHNzZf6kSYJLzBsoPzJjQtg5XXOR4TKUyTB1BkI/9nEn24WC7XqC7fqC6wYO1w0F1+0LpjsQ"
    "THdYYjp6s1FU3UJW1IWslGlELAACfOZfgCK68FegmS6I0WKATNozcE/xiT1/CvJ9ieHtMcVdYqxWgbFaY4wF6jTGugbJtqDwGeBaEgcWAJlAuhhwAWNBrLMI"
    "w7M2xs3aS2I6jLVqYawLVOf9Cxg6mH8JDLgGhrwG3j1TuyVX9eaAtv9VE8BzmH2bFx5wcAA5KLi0iaXP21gTOdREAuVKqXQzcE+hM0DpPIoMzduMZbTMz1sY"
    "GfR9qpYYI1OfkZGBB5pYi3k7lvgTehwCaxojq25OkYFR1uQAodscUxZY5znUY0Kd/4LqEbNghh7oMs0lZVHIJJgFRp6jeU4gnhl6oLfgxitEGxPxYyKeUcsW"
    "VGpm6IGueU2OQ8piRqkWlEVKMRfU+illMVNZ4PDULGS8mKhPKY8F5ZES9bGOHVJZYyJf5JAS+TMiP6Vyx4oi7FtnwtWndFD5xGsJ5bKiXFaE44JwXMh0AyoB"
    "1DhxY0IlyKgEGeVyQbmo2HtUAsBKXJtQGTMqY0YlWAlBSbFJotQamNm+zuCOiqkcFRM5KuZyMEzlYJjIwTCXY2Aqx8BEjYG5HPpSOfQlauiTdlFrTjreKoB8"
    "/MHxAV0hunro6qkB0cefFrKpjz8t5EGf5JEYKX38aSHb+PjTIp7w6feocpxHnwkQ51o8mo+9lUcwLRFdv2zN5WI1k8vWaWndz57c/Voxr4XrNbjXZb3d6h1t"
    "Xx9nQVvgT2V+23L67ctyYoTarL9Tbrfn9O5Lc6orya05Ed84Ew/bppNj2QqXQwF/QWwfE9tntLw2R6NwMFqBHlIhae4XrTSYiydyAjQLjMs3XYS0EAI9EcLH"
    "9KxOQXDxwM6YYsJ0x8encNPaMkgDUFuWlK32Va4taGYszrfzI6NkXHIhRqvZRpFLCLfpHfgEIL4lUojniWLxPBE9VoRml/MWPhNctAp87LOFh6FwDXBCR6Hw"
    "QNQMBnpcRLmgVaoprk/BEJ9BnDWtwFzTShTMz6BxriJcOj+HXy2TaDLfDNvNRWvqef6ZmL83J61r4cHhb9leCw+tqtCkvzlpQ4QrMcfH5CtKfiWm9c2L1qXw"
    "7IsUB4iltUYJSmsBzYv2pfCEXZV+4UlIKNLcongFib2iM2H3qWbSWPNIrMY3aZZKs1EsorUQr07j6Hg4IUAVDAtmrezHJsbBKRVKFMZCJZ9xt6YfeH6nmUdt"
    "Ovx+LWfFaF9LTIHRtpaY7mbWUT2z6R3u4tvR4W5BBzp0n8rMKgANMFlgQ8IKpFeB4HA0dyB7FciwAsHBa+xADisQGrDGZqeo9EJmhofb0YA17ulEBe7lRLHp"
    "ga9ZXuAuQHLJ3Cep9VnZot6cl7vFYKPxbNNeE3aRpPQo6YTAMb1AUiTTNIaaZZ3G6znDx07GMzb+QFvKk2xcmB2GuNSTs6jn57tN1qZDj+RO27Rf0GQt7smQ"
    "WdRMW4knw1ZRu1m0cvAWbXGyrgczsVx49ckvWk8N4ve0nhljfxrDRyx3xu9D8dnDPYMY+9cMgX0BlXGGMio9uR5TT5rAV0bal6GACc9hUPfqmob4IeezbJrH"
    "y1kyrn8cvFwRoa6GUFdCKMs4E9UBUn4lqoIEGtWCP3aK3PPdQndlodszp9Q9X5a7K8vdXtUWHKp2LMveri98V5Y9LJ+guMdur2V95Ygfh0OyuYK3Pd7zU7Q8"
    "h5ZWtHECfbuwYkNPoUkBTYpoUm26RfWqFCe6aYudOoZbqob0nKNulXkumtejNUJpXC8Vc3gyrZeKeTu9X5WKOTlrqY14dKuteHSrzXh0q+14dKsNeXSrLXnK"
    "S2/Kk09vy5NP75STT++Vk8/slqPPbM9vYJLv3LSV/d9cto31BYxMuAoWzJ07Fl3gUA0IAzSQGsx0qpV2TbRrqTEtmGtG74Jtew8LNKcL1nkmHkv6QWyl3OdJ"
    "dtp1iZL/Xc/bU6HKaAhIqESwRieq4F/w3L2sW2lKTL7AfsfLfvd9t8o8MBg4jwiKXOqts9EDcPbLhM6DhCUU5s34L3qhuqQw2UopqUiFWMiPxUJ+Jhbw52IB"
    "fywW8Gdi4X4lFu4nYuFevTrPIubUa/URemnXMsa5xAwPLwvbxOZ549g77hyKP+atVusNxfY4mJhn8ewAKIX9WLEVtArmJknX2/p+vSGtDbSNK7SNq7St7Xzi"
    "LaTBjMdQ5qSAkS3fRhrumteQBpStKpSt7qAM8tlaaw5t3Xr6oWYd0n77ta5BsdZmFdpmd7ToamutOe3Z3U6+06DvNGk3bq1lFcqyu3ltbh5lLDdb7UPaJW5w"
    "SMNN9RrSsNaKCm3FnbVWT1q8tRfYPWdteE095VenWNNZnFFjizrvNVhKBkdZTlp1Y5V+SLOrVIhyfPSPeZuqtAeZ/CAUtnnvJcTsLXtndg5j/63b+46dkBLt"
    "S0aoVB5yB6B6pVIdS6P3bvAImivfoHh63JLJFqz0PBdO6hxSLCIXeBz7q7yeKqpWwcSwemEb1K28oNq1X1B13k8F+a434HGVG6axzhgEIO/f/aCqHPUAsW4r"
    "g/PEunteHYO9kdTztoV/tcdTL1y9lBgAtNKgpARGNCIiVHElgYvovQggIRbQABCQ1AhonApI7p5KrXNae8p6ERcfolAoRw7o+Jh/6m5YGp+LxWwd9MkNezyf"
    "O7jaIVT5dFpK9E+VaJIUFYwPo/+vKYI9FV7G2t1w5lymBVZrGgScPt5mgzOjS2AfqdevtTZ+rXTw4Ezr5Vc69Fy7XmnXG53iraX400zguTU1CBHw2JkrwNTg"
    "JYs+YqOOduLJhE12NsGPGiJthu9sZMM8ZQ12zVk6KRo5L7VRsVqiDhwIc5wdsSLxOs8AysX6T7CTTHaCj5fxfMVGl6zV2ihboKtkEo25MhWcxjAN2dkRHiQk"
    "2hFI+88kcBnnoNtFeA9B3umXhkOxSwqUy+ip5syflh2p5x6p7QPncvoFyEtVg4k7czpipqM0m8poEBQ3dUUhCn5v4+EiUH18ViMiUymLVNq6qkvUGZyPS7n3"
    "O8KKv0imqxzZbwR9kKWrBdM+Ub98E6iDpvdLwDbBH5q0+yVJNwHtIN0vdrIJFtmEzX9O2JUYqkaSFwSTb4LUev3QDlvDvFhbtsJA1S0sENmWrYE/XvHsJzKG"
    "azHEq1JQBdH3jE0KmeyBWELpzOM1CjLMYipTXCZFAmVEOSjYMC74m1k8ya50qpyNWXLJStCLfFXw1eIpMDCb6OTAwCArhRyVoDhNFtSIhWHtguXPQO+JQCpn"
    "6RN2keXsR0qJgjpLH19AIxpA+WJObd1IfUVYDVY6kF0rHff8lxNkLVtT55R8KuIYnpKKAjKMV/eioz04mVR1GeMYYKtGpddByxhqnwMt4zDnvutTy/DgQbeS"
    "1DzXuSVtea5aweBWRBmLnFpvqD+rnVBbnbtm9WUst4A5UXitn22QOInrvxyx3Ugl3JW3yZ1ivCUtUYB+uzXmcyvmu1tjPqaY2hpOTYWtjcHMCiO6JZNFVTxN"
    "t6/XrGodoWyCp1LqMjlvXSJ/uyv+czf+u7vii0qYZ0Dg24xa11HCXalQkn7e5go/b7OXWek9HjfdmW2HwU7fUeeFykeN7SNJV7oRRleso9e81Bgtd73EUH9k"
    "iSfHqDmNnaAUdba8y1OmTBl7ehrj+CRnPEnxMpnO+AmUR5J7jiqX0iS8kQkA6LkJCO7X4c8YXevH+irf50lc6q7ZLShu73Smzj3xPDxdXa3cNwzNurxY3i/H"
    "MAbWEYsOrVhXV245MTOHmZ0rMkpxQ2SjRka+xjhO/5s3zlmDNM4G7VeRBtfILhoJL9j8orMTSGGPM35xFUfhOmmqO8OSOzBYuDpCY6WBQmmJ4gSzrSV2lqti"
    "JuKAEr+M+Xj2/BKNlLzEc9D3L0eacVwbSFJxUR6pL8U2pdhoyv5ss0hE21vG2eHRZU5AIbj+4cKcqmiH0uQ/365Pd4rlPBkzUG/Dam39yIwpIDxgUC6Ci8kq"
    "xsdaCqEcoG3bpLBqhs6iSwm32ufDl9fGlWcctwqPu6VYUOE3XWe3ojV6i4pto8cKLYlT9Sy6NNS25RUHacYRdx4Eoz1Zv5iUi2sF6kkYzb+4Zyf8HuTgvZLi"
    "pKyc2MzucGyVBsyhGcsvHx5ZO4HqzbEyX6THCe0PumceNG+kp51tGePpSr2mo6/qJxuklKrOvut6X4OCqgcLk4Iumvs/S7KVOfh9xytdiFpd+UuKYSG6syCP"
    "7y5IiWBrMoDjZPAKeVoXQj9o9yfoFzi+RhvcSvobsiZhSHeeOP8Sug2CP0+7ey/ULkP1DLi4WS1vk4uTl441rU0er8n8G8wVQTehZ99pFuKc2SnJ6lKPZuVe"
    "zLDDWtg05p/FjFkOgw9CtawrZ9Ky9x59vcxNhpqGxzBWF+I6W2koksqmFPc0NuKxQNapS+pt3OmyPbPufPaM+PYliK6Tl1bnP2Mu37wFP6jBnnn6x6qJkwp/"
    "lUdKdW3MilJ7haxmALRL7N25AgN18GfYoa76NtXuVbWrJFkCmVV2eMkdZOTrPsLlrpb5H632B8KKk3usulKz256oe89Ot+oomw3P/vrmh+/tTmbk5qdPOwVE"
    "Tqc74FkvGajOHLL4uDmiXhd9lK9TJawYfdwEC+wySTwnD4dJ1CoXAckingpXMYuX0vUBpCCHLNFjFtDAtwnSDmCNJ7R8hj0a1zoHnb1ArHcLsb0TTFkKs0Po"
    "5yO98twRpdnZmPkpEKutyOMKcFpjI/R9SqvbuM+k3VHa0VWDC8YI3KCZIbEQrpfEg0Ssfut18GBnR1lKQnUMDS+JxXINkg1aWoykiPbaZNeO56xQUtTSmiXG"
    "LgtqiqeXP0M7hrO2SfFKq50Qu/tAZ66XPGXWZgm0HMELdj5usAaw7jqCfZIL2S/UqiiZozJrpG4Y1KhcxbVWdGlDJlATcXtI7aitPrt41U6tJ/HWWrOWAUnx"
    "Qk4PJ69YMaPYYk/Dge8EiXjhUrYdOgGkppavLNpcmGIlM0mx4zzN5pmoWAdSE0nj8TThSJe1QpIy7RHW9WEI14/IRWg02fRZ9V6LCCwJbgVG4SQeFCscHR6q"
    "1Lg7ol8bqawgmIRaB+gRrZR6ni2wEvNuHIOhgNmTi/Q9HjlEu96EK0g8YSDAAqUgwcybTG8+JGlqNeE5TKdfZRNZIuULZIDVWsZvcZPZgiYFQMorPFKN+Svx"
    "FTihyLzK7aYjYeGVsSqR6VXqyQ020vn9ad1ETUWsDKtiFSUXr/sJ6Rw4KWDcRatj2h8xaYXBQG5JTfO6uqkiGWWzd/zsYQqaunZ6Sa86lBeAnBkmDGVaKnfE"
    "Ao+nGcCMIg4Rzu5MHRmVdDYbymY0cZApbaSmhg0M+FIwJjMNVzh9Dy+1FHa94r0WBKhxE8/8o1+MnWggptBsj+ecC5sJ8bhz4VCAe6LGeFkztbKOGIlYO1BT"
    "AVPzIHfDFEFo1jB2gwRtaGkwcwMEmWhvcF4KUBTTO75umNVOY72ElMpWjpIgNQO5MymQjSqbq5HAfMpaoAD+PppgpsTSUq0IZL9IdD5sI09Z8vuam8NlFuEJ"
    "UPc3b8OJwZ4+aoKorJbr1Wi9VSEDlF/Zz5fryh116NIFBPWbT9UdEZq1qNDqFMY8wFuv3rvh9tp8ZSjlFdCtkwW+JSAoD/moz1oKgLOhy5Wrsq/LLU/dDi93"
    "/XXbvdz1V/d+ue0rbfqS7gMDJ8wty2qQ0XO0Xl8SSbVi8WP1/VUURHolMlWHJ5zF183GPl9ROmSyfacd97qfukeJlIZtzri/1idbnmnXC+Faq5Pp37MtD7yL"
    "cy90Lkv0GVrqUOemlZndbe8MSOO1zjsDJo3efX6K0/uUyReKqge1JRqzqVXClDjHAL+n2I/TydNsOY/TOK++HHg3cW3LgLk0K+xk4uAuygemX5fMpjJPWg59"
    "VrH/46wela9D3lKcRO2F1J4Vd8rmFMAU0VhJFngsOiyr6Ta6slk6O6xiCcTNTdvdSNkU31Qpm1RSseQrSgqrii1v9G1/CMlKQ+3mtRzEVtLb7V5XrF63ta1r"
    "6GXYtb7gHUoyBz1qdBwMn/2ehWzFck236+n3pDF0y+YvTjru+aaEothNex+SrfenYTTncfM13QsuN5Axep04NsDr2wL8OfdOzAhJftsgeI4GwQW80nFdVvB2"
    "E20Y6LiL5qFDYYJXYq/alZdv0qhsLdvLokZLy57byA/SbeEwPNnvTOATt0BYii72qHuvp7WkxfHSy1pVXt+SZOyKyvtvKShGcRHcb03+vrytW650+speXYTZ"
    "9AuGuwPfW6feUAqprS+Hvtes+lZk7vCoFdpvptuFtBkwxNOC+Z2vR8kE7YhvGWFqDnlLomzryzIJbtZrGtRSlRJ6n22W+XtXl6DGNYrEE60+3GjXN9r1g3Z9"
    "p13fatdP2vVBu37Wrl9Krx/9fsfrR/KA6Uu1aB8rheRc3eMaR+mmQFV83NBMYZQLzdbJFrZW5tHFjpNMfgtXJ5qrk/LIHzyp2vxLpDrwhJkXVhL7MTAloB51"
    "T5Jypwh3rSchPG+UWC8SWmV+EufrcQbTENvEu0sL2pKHtqxQ/E2FYn3//YnQi4BwmPZKzw3DWa/0fMNw0ntjQsbKgyH4dAvax4nNEhW0Qn6vVjAFuqUhct0Q"
    "ZBZRjS3i1cZ2D2/PtPWO3wr0mxleMUYrErE/9vxVsESbL2M0owq+Iyd52J60l8EymOhqLr31WKe71rbED1AVP7DO9aMIRDw41spx3ULfMR7M1w35089lg+P3"
    "zCFTjIFOGMFor3cimrWZEwF1IQWGrOtCYgzB91UVdUkBGmrKv4nHMGuqlP9JlbNuqn3hibKqD8wizON7x11zadMta+yq6+fOW2WdsRrBa94+BQX6RTpRdhOr"
    "7aQw41OkDnJcmHDx8/fJ6XZ76LdK2DpdXeXciUsZd85L+XbG5qzK45zFzS1VLRJJZJVKpzx1YGfPd1rAfpT8VTJZfq5C4PZZjeGeGgFwXD2d1guXXlUg9olc"
    "JeU1sb9bGoGLL1BVRAnLD5B+ZiFF8jtL+D2zTqqU5481xIl34e2eXSqYHVZbOkYobPlRwkBB9Ull/K3P2f6O6rIr/eoLUZIRNoJS0JYGuo/6a54zEd3six/T"
    "s1r2S1/UO3LXo5UGeq40zLF6K+Xou9JoDILrWweUI+hDSWamxvLLd2Jk/UBj7rfaI142wSFlDr9qDNTzm6OfSxgTT79SIjH+TGP1t9ojXl9DjLPjaFzCmJjx"
    "FC1Zjn16CWVFBGSUaGyRUUTZbjNrj/XkVrwA6Aw03zF8B/yXEpn66dSJIvMXIHOpyPxFkLmkHCf4+JtLpk6+iCaghGQ+vSm3EPVEiZYWmXE0323O28vbyPyW"
    "BbFCehGN/WV74s8Q6QUhhaoQtLSRJoX4p1IL05teTYjr7eJvqwnRPZ1rUs31JyvXKRpPuWgtWiutMxbRwp8CSrTDeEcN1xeobgISO3OPmKYd5w7sXExF3HdW"
    "5cW8P3DBVO8J/IoD5j3vtf3xWffaXsndEQm8yKb6MtA5jGv4jLG04NIpkglTt4LwWCq7pn3bQt8pypYglfhaxReWT8VpWhWFcL7Jx1GvO7AgzwoOkD0L8hzX"
    "+/E51tRN+Hi+nMXWOWGVvBaukJQDJ2zJZ99A5UZ9C/AWJJcuPEF+yRNu7lYVnKXjZE7AV7h839vbc0II41546AB/ZBdRtxKtPn2czCPOHdi7WuBrmPGVgZJY"
    "WdHjebJcQuPRQFvYB6sh4IUaBLCC7RRiy8C0aEH+N9jwFgZsdbdqljkbJwU1l462zObraZb+cHEB47fG6EBhFIMupKrHCfopTXihQiYJvlOBrKjQxNii1GBd"
    "G5A9zYAz46mpB3NFIWETwQcqqHx3joNe+ipeLq37cNZNN833VMzu5ldr3h/JrgmtkkzTpvSNQd3krJlzE9EL7En66FcWkIEd6oF4TzFLn6yS+WRkXdvcBOqK"
    "HS75JzDMOoHjVcFBX8/RpNTiaTyesb+xtR3FeQbLxQQlfkObN6C6BlDnP6P8KExicUTSbOp7ZiuSOVuRaNwSd46t+alc2a+xJiYK2/jvnRZr7fx3Qx/BQP1g"
    "lUo7YtIwGE/SFcO95h3kRHFYSx0WK6PeaWmx1tpBuwoihW2kjO4SC2P1aIHsPAMccdroXMxj2lDD6GVb9VZYhEdwUkOXfVAcy2/OgJ/QG5pJQQLyRKw2pN5I"
    "QuX9o4cP6QSfuY+UqKFnJDFG6egeJVU1SfoVb8SNpRwb8DoIxmzoExQ7MLoE4pCB08z3Ox6nTsbVHn/bGCXuo9qKHtWccDMDjnXGTQH1GTfrYFvyL9gPB45K"
    "y2fd0jvPuqWVs25GHsqTleRWDU9Jxua4kwgEVfg7dt2sHJfJs9V0BrK6oGTaVwqsHLKBUs51Ku0rBaot8xn0A0kpuR1KCRJZgYZSYfVskYB8u1SHSJXXwaGA"
    "kRtlCyYch9ICVIaHD0NVoJowC6+GbomsCrpk4xXMc1VZpdctrgRGbpRtzVPMkjTRFa19pcByKrqXNM5iLnhB+UqBW1P96PBEFbwtuuJMFQDDmmJQCwTV8VZ0"
    "ZRc9BEXVyOZsEp32CrZkXZOVHbYtTztOdEvy26lQOxkVEnTAtvx1hGhbQjfnoJKWZgW1qSnEHL7Tx0wUlYsKWQtFx2JbkUGkjy0E6KngQGBkR6jFRJqTqTLl"
    "LWFT4MiNVItxjndZDUblLWFU4MiNVKlmFVLq+hWwKk5mlSWrKUhmSpHVZkjgUm4uTC06rhZLk5n0lbKT0MiJUskSAyz+0V7P3tk3WaVbuDl1mTjdzrs66K0e"
    "8xyQjmJRld7CzHhNch6P6eqPobMELVFbCo1qk1QotyNYxFXApahPkrioxkSoPpZWEWD5drmVV8RVfqeU0qOxycEGlbuvFRRVI9fmoIZDk4EFKeG3QqJK1Frs"
    "aog02C1ICbsVElWi1tOeXlpkk6dMMQEjO0KJO0qjaLY4h0FZamHkdgIqQ7xAWlI8XFhtxIomxy7mZCNZIbEB1Sg1yfGGPui+P+LqhcJgw2ojyoqc4tkbpxta"
    "kFKVWiFRJWq5mUr6UHIjmxocBuTeAjDrB+rorFpOKAXXYX/MoX7FAo7OyILVRvQCrUiq5SshhtValhNkXQMpNI2GOqKrstxFsdz1L3kfWi2AHYcURy2H2UH2"
    "tRdrfUwcRrbWy8oR8JaOWbVyF7FUEC2HuMtZKkis1pQWtoLUXsopLe3gm7P2ulJlpakUgRa0apfJTERDuwUwwbhSVlo5c9NW8lBAKxqtlpXW1Ezwu0r4u1IE"
    "sbJWXmtzTy4/fGhN2rhhUOVxg+wmd5a4KE1plaxrsWTNUlk1iVxC25ZgGzpaXqtiE6tuW6Jr7S5lV8mEz6xJmwUjzU76SoHlTj6Ji9kbJUaUxw0qJ5nGS51C"
    "up2AihxBVUD0a6MryMPiD6zTc3J9UeolerWx65WWG9Vxfr386AbbSEsrklY6s0pZF9Hhl8rqpWiz6qLmlug2squEBgy5lKF9pUBZYu1/qVrvUeime+m2chXu"
    "BTugFKUTvYziRKHZSxmfnrCUobfj+j1L0ioyhG6JbFfLIsuXs7e0RSkXUSxA6aYiBYlJnRVXAkpxC7ywpXhKeUpxrMVFimYvNpYj1N+BTLfcgTSL2WJssda2"
    "w3tfY0y3XmO07v4kzsWenPzyYk/5no6+bsMqN3HkdZvcXI7ZBHRSpLqWffeRkQDdzgJn/eUV3O3i+Fva8OLaae18cfrUbH9xx+tuhXHlqm6JcdtX2h3j2lna"
    "JOPaWbdXxl1/3c4Zd/11m2jc9W/dT+M1wPL+GrfUlNJOG7fUlPKWG7fVlPrNN15VNCp7cdxRNMq7ctxWNGo36HhF0ahu1zmKRs3Onato1Ozi1SgarubleM1F"
    "1tIW35G4NUnnu3F5W1qQMPcSjFkCYVlF2i7WlwaTqHuUoHg9arUSL32fnEYMftTdG9fuUWl3ccvWIq+AqluN3PZVdh255ansPnJbXy3tQ3LjrtuO5K5/+94k"
    "r4Nu3a/kNcDyFiY37vJeJjfu+l1NXoZs3eTkNcBbb5hZIwS3PF92+YtO3AW40pIVjgTXKzDGppa09Csj72w8NLVaf9bB3nMNdlJztW4n+AjV7Uh8ZSzD3rZt"
    "tcw21W8s+hijabFzPDQR7u0f9voH/SBOefLHil3NgKdG4XAwGPT394IYZNxouLfXF85FDKVmo4P+wcHecBDENzCoEYpBCJHPWTLFtGF42Bt2g/Ok+ANzGO7v"
    "d3uDQXA+j8cfRl38puMZm8TzRZZOKLzXHUBypAePBqDjMsmgb44Ou3t7vW4vOM+zq3QUdg96g14fUK3y+foqyyD1YO9w2OuHwTieME4ohr3hcK93EIxncc5z"
    "Bq1DBPf3egDKoOfEWML+/sHhYL8LI2Yez5GIwaC330NvejHPrlgucO0dhocHIYGLZP6BqN0DbME4TxZFBjRBun7YBUTrOJVVNYnzD6J2+4fkobD+3n6vT95p"
    "Np+wNEfye93D3qGMNQXBNArhz2E33JcQxlKokyHgl/5SjA+z+EMCaAb9fm9PoEGFIuXx6DDsHg4HIsdsnlwygW1v73D/8FBEhbKn1GT7g32oZwkbzxKgrNsd"
    "dLthj2A5mxC6ve6A/AW1HbR8v3swCEW6gsUiA2CGQ6g1AcTKpqoY7PcH/cG+gVJpseYGh3s2lLlQkL1/rLIEGnGvdzgQMMUcw8PDPaw7xlAoU+OEw0PMBCDF"
    "h7XI+DDcC6FLLijD4SHw0HBP+JnlzyZT2ea9brcPJQguQHk+zxPg2RArKBwMA+AM4BbVR4ATDqHS8GBBwWVT9Yb9g0EvuFiNZ0USE0XhIbDEFA8/nmd5hgwD"
    "vAb9YzrLCq5w9cMhRA2QMzAReACzxSeDfu8wRBAWAnIIsSlEnv3e/vBAuNdsDrwL9A66feg5ARVRxZ6BbFtP2JXssEDBLOOq3voH+4NukIDuF6fY2mF/sHew"
    "1xsQaJpRLfb7EOMyy9dUdiCwG0j229s/AJK7wTy+pOuzAAn7PeQMBYGaLWaUrt+H6p7HV6mg/gB4+XB/GMwZcBRw3sUFMhbWLciYgDY7RFeCvgQsPhAg2Wv3"
    "9odA1lDCsJOFULnA4YcCpCtQVQzItYMekkWh1N+gM/f60DElSHDw4QF0Og0qx1KVtncwGEoaVY8AIDRHTwJVl+iFg97BocxWMSYAuv2BzMV0if2DPkjevgNm"
    "ZTBnbC6rBYiAriXgupjQPOEBAhcow3oHXXJKfgFWwqbE5YmUqmRvCIJQiQ3NsiDsMygSys5h9yBYsEmyWlijADDNfr/XkwGy6+xJr5IivV6InC2hy1W+nDPo"
    "uCCjYcwRQF1L/cP9A+AFBdai46B7sL8PtSfhSxx5RYrhIASOEHAjKAbAm/2uii+EheDp7mA/3Id8k0lqGAsqALoWAFOOJ5cWOIL1woM9QJAUfJ3jSC4GMUya"
    "jcf4wIqE9A6DNL6Mf8+0TBgeDIFvAQhMA4MQMCDuuUAIiOK9PQSAJKY+2QeuJ98kj89H+93BwT4IMyOSQbRBhxd+Ih9kwmEfBlJVt4M+dABo+mU8Z5ao2Bvu"
    "7UNRBZiqCcRpD7qTAJl6At7pHUJbENiqpkH/AERNH8DLeB1DyZai43b394Mli8ez5erigsoKfyEay1coL4YHIPYD1TeGYRd4aDlfLXCM7g2GfUicXU2kkIW8"
    "YYyAnihZArlsH3oyiFwGNSyhwyGwBAy/svjASlAIaJC11Ad6MKbiucM8W8eiP0A/G+IwUcSTyZyJaNC60Bv2A91HQfhBdwZ/OlGYht0+pBwEhhm7ewDaR0Ax"
    "g25FVQClOAiKhKUp9BOIMNwHdgW94BJFHoj+HkoNp3+DZmIYGUrT7Q4lRHT2PrQpNKnVzxUklR157xDa0mH6vUEXctUiYDAEJQLqhaP462NnQQ8D+QhFOhyS"
    "TsqhMkEGAY+B6sKzRcwzkvr7MKYHVs/p7QHjDwM5wAIrwVB8MAyuZizmpNn1sURmANyHoUV4i0X2QSl/0AEsSTQ8hJFB+BU7Akd09web4B3oojNQCQv4Px91"
    "N8FfSwBzGurv5dd46cprM21FoRekYo2uTe7jcHd4wltDn16P9NNRetzZO2Hw6e32rYAmeNt45myjM/mbcwXzuAMq0GDvpAOM3z88xKr1+Ygu1AEnNzugaRxg"
    "nzwY+rzV6e5Bl+7tQ9mCXmfgGaR/KSMF9QL0ppOw1znsAcIQku75Gi0POsCfw6HXRvhGXHLkqXvJMai3eabMN+FxvRPz1JC0WYem7799IhOr11wUXcLQszi3"
    "p27BkoXwnRSEO8ut43AaM54XwhiVA3Nqm5TxN3w9Z/brAfrOv2tFXl3KnJrbmVynEfnoBOL9KdDFMmM7II+a/NGjcPgQJhDerj5gPCXwQQl6HjXB/ZAb0Map"
    "m22kMZU8NaS9eakSoamYaIWXR0K84TETr3MjR6I7FW5hHEQhliWWSI/QRpMxlHUcAdumfjNsMW+Utlg79VmQRz0/bSdHEgP0ijxIAt7Ce02KTAVUpzU0oE23"
    "n2xTIlYLfdQMy5r69ouwCUbz3m/mWYw2OXGTr+5wJrHOqEET7sZYWenAo5g7Ld7aaVwl8zlaHk+mKSjMeM6UzvuntGgT7f6z2TwZ5dPzT7Ni7sUn3j+azff/"
    "/Id36nv/8HY77JqN0U7BR7JxfWSMiYWneI3zfU+/FpjI5wEB1c5IueKdEbXO7j//UfjNf0xaHnyDLW6gQ3r9f3ROJNA7+YukIvc8lz3Ec/RJijwVUFW9wCtE"
    "SFrY9Rxm3Ba1V456vjVq34oKzfZ+IK8XHpUL+F9Osaq+Ly4mTnZriongcjFLUXvlqOdbo/atqE4xrafvgFNkE4Or3MRWof5cLeij3lYvwFJ7u/0hmmqrlI9o"
    "zqNKYRCuLTdYRepIMSIv0W2kqTbRK/7x/zTfP25/E7cv/jE5bXl/sbqCflAJm4AZG6GQsl+2A96x6engyshj3ux6LcsdhEOHV6vxQyt+WIp/XhO/Z8XvufGR"
    "yOFnExneg0g70/49iBxY8ffc+Bt7bFSbNnrwI3knjbiPjHUiA9QbRb+x97iy+BJXlZ4CuzY9/eiykbLOoMq80S0i9if5IiItAqN0/dwXZeU6rtr9qb9qTMaw"
    "9KCsI+Puh7El9G28WMRvM9xejEktiXplLEaxAXTMc2SEAE81+NwGnysw5SSyeJtRjionbdED2iXcBfXz6Ja80/q80/q8U503PUVQKmjZTkelJoQtOQdFqQRV"
    "FKUILorl+g0oKNsoyCNUXTu5LiJ5p7po5D2vqc43pPWUkf3FRfYXF9lfXGRUOIe4atFc2vXjleWqIWq2V4wg1qRWR9ZVAui4vijB8XE4/Kf2To+PD4zv/Pi4"
    "K5OqW0Bqg3WnS3/kPROF3lwXAungdQp6l6M9pAvTpP/dwyy97LodmeTWC+DODKhkwTVXloGm9l1hMYrG101hrCAozLgqILRJFgeZRDaPmkUrB0mHcrhAqxpe"
    "HHWDDG3VGR2UR3m7UHpVFs1JIeW7zbxVgMTbbfbaEO4F8kHmBhvFEczBEm+Xt5rpcXIyHDlvDzdSjJC0GUbo2QEJBoB+iwGDTbwbDbXo7cyiGG/CRxn8zqO5"
    "eBdA6Kyq2VDDa+60mqb5P8GosRPYoGkVdE4gb2eT0Z6VpcmXLGZgyDuQRe9YZ9aC2QB8ixZMBuA7b6k9wELF68woAoVKPpWPAjl9rGWJ15YlX1tGwEIysbVe"
    "tUdHwrnFHAENXltIg9fg2TLraum5jXadt/TEa3VelQya09TDeW1bUJRDp21bbpRDz9tGjFgPzbk5+lY9+VY9+aaeKubkKhhMepNapJ2zfFlTva0IBaAwKZV7"
    "PtO1hIJQgKcafE7g87YcTSWYUJv2q5lWQjZNJmvQT+2WBLCoOgUWLQpgUWcSTFlIvpVbjBa7cuX5KzNX1CZcsOhf8YesrQlIgZACIbmEzBEyR3s+JUODRlfd"
    "aoNKvxmMF5GgTOoQG10FV75z+ylh9zlfWUHv2amuFXrUV1cGvel770eBc+dN4KnzJDDwENHxZHVxwUAJBEF/vuKsvstBhf7aNDoMAX5rGu2FAO+a+iibMQU2"
    "Uc8d57vm+vPUcp+TWz90LCyCVyQRDkebDU873z9+9fwNaJYBTx0zXMLIOF7jteE5Xk23AdMy4DwKpT0tluqXZn8tmdYyT83aF+lZMXsSF8m4dKNe3P8jJRRm"
    "d3Kt0FwBsi5M6wsxVZA5fx5at1ysiKU7K2HlxL4dV13lMSB5tN6+4y3Oy+t3Qa0D7PpdUfdIeufwoHR0UN+0rjkzWBeCp//kMb+g/tyeG2wO1Uls7sm9UI9I"
    "4mJzk1c0fGrHsjlTvKSpjAehx2osNO27LLUW185tbcarMLsVufjWNiQvAartym1fqYG5djrNzKXDbWquXDVNzh1vffvzMqTMDdy4t7IFrwFu4RReAW1jGl6F"
    "lRiIa2cNI3HHKyTThrnyptL7jQnAVBneC3LhulQm+4ra1eyqtXzP47MctGNMjJezntvvApYE9qgRY6JGMctW8wk9cUhpJuLAF96rr5oCoSRKKUg4W9C5an3g"
    "Ch+FMMeqT9QiAEw0u/b9NBTvEQn8VJ0awjNL6nSVeKblR9zGiz4KXXOE504A+6gdlu0ryGOU6U/LeRZP9Avxtu0DXGhoWIeAbjn1g2exkJjKk3xEIb/FKHD5"
    "yKioKhLnwu0uKwiYV6pJrp12lXL5yEa5Cu3h0q5ILr6GWOtVUz9ycgxSP2LaZx/xUxaVdFhynB8l+sFNpP49b+HZP+lJwVN63BFylozpGmWj6hD7KYZIofrp"
    "BM4sjlLIg4vdMpV6HU8SSGveqA0lytLCLWYQS72iU6bEG8llI9voxA7OGdVY7QUMyt46jXJcr1HOqXGeb8q1Iuw49P4jCuvSAsW9JMC28l5yq7zXxrneUsj+"
    "f1Ah+/ct5MtthTTOmy3lHfwHlXdw3/K+uUd5jfPKKbpjzFcYgumpGYuSHV7pdS4t2vT7XHnaqZ1U4JoezrHSTikfrbL9CvMJDL/Gn7V3pDYD+p9PRHo7Eelt"
    "RLzDhSOkAn7W+HPjvnRbftv+PsRcq0eIxCQKMf//7L0Jdxs5rjD6V9SZ78yT7JKixVusVnIc20l8r7drO+lOp3N8ylLZrmmppClJjmW3/vsjwA1cqiR56WUm"
    "d27HJS4gCJIgCIKASvksUu5Uyi9FB8eVB+Bo+Xh+bkRtl9LzY4uPIhh2PSPe5fOh621vAXyF11X72E42Q6W/R8x8m+aX0ZIxp78C1J/ds7e/sIwKgd2cB/xy"
    "DRv4PGcDoDHQTfwyXxN1bOKXeZuo0yZ+mq+JBjbx07xNNL6S8Bk/f7aVUbYEZchDVVlXpTCaJATaL66H3IcCVCmMJjFt4iePT9InbEWlMEJxTeNUit/2eDhi"
    "+WjBmzdsyhSSjQdcSvEjNfD3stim1V14O0AWHGmJC+2YvMnPPvrAhncXlTZDn+DD2IGSuzctoXwq1rzjE2rk9wmFkjrcZ/JQ3UKAVzncAMY+FlX64l1wuebL"
    "RYb3++/FkXGUsotByODp0DydWlsfnk3xBBpqJZf/MMoVXTCQe8loQ8pA3HJK2GT154fxMc4C0l0QyHY37LFDrRdWe6FO1da8QK4XQygDynghVBp1L5DOYqhk"
    "QBk8skMDc0qhAUhtLXNm9eZvDkFlYH25IJi1FQuMsua6kv4MuUaai+bC7KJce1ltcvFi9KVKgxDW4HqRBh5cXk5K4GfuNSz+CF3O6TgkwiPgTdK6V8tlU30F"
    "evpv6s/Amc+bTkqg5+mm/gzIOG2S70BPpU39GZC5sUm+A0r+TfojoBTdpD+IPeokMfZfPEsmX0ZfmXSFdm23cC7iZDnn6q9hFHzjX9tRcKGUYwf863gUnKqv"
    "s8SMcbGrZ0M6MmaDVsnP6SX3NlnESy6f5u9F2FCRFbMj1y3Vq8tVMFTuQlFxuJWRLhSKJ1E3HIHLPKmwvkr74wHEcxQ3IqDxjpOrt33amEzl4Wqob9s0/CaU"
    "bBhDR+nY2ByfOl5NQdzag35YlyzYN5A7eKbe/23V5BtKiuhbEZfZa3z/9aaTbF4nJTQE3STFRtohvz6MWRKcIhmbSoBGzoWUUVaKeNwLpA++8A/pVuUVr8Oh"
    "r5bpVcJEL+x03sOQcZakwr7xYeS+J8VQjMRQRCqcNJJ3M5mC//UoTBHOsGhAYBMBKLAjx5Xccqqx5uGSpAZXJ3NFY+QcU00dhuqNCtXYpGavxcg5fEaVxAgO"
    "bztw1xC5CGVE1010+0UeUa/kiULU9J0k4chmNW14YaUtj4AC7IRoj2DMIyB7j32x3S/DmxFZicrdZ2/AGnurM4r+KnyZ+muJiFPikIhuYyJ6TjwHX36/RSfC"
    "n8zP6nxqjMp5YgD4nA3g81wAfskG8MsMACSQk3HGEmDORDbSnMQm9ANDZy1+QKckKw9Et9//bYta+X9LKiqNbUZCfBZzzAfqW2LEEZ3yQHeOiZZnMtgTB2b6"
    "Nq99ATGyaLC+iibcBWgcLkDjcAEaB09wFo9XXBI32Y6VbPrIlR5xWRsxayKu3P3+e9V84W/w3BeSL7wIYMH2wJyqUVJKfrfb97AGWr5lU3T3NNzrJQcZZfKk"
    "IGr59lTNs8DumT/feL9vyaXSvFYqWiO8Qfv/qKZVbu7eYdwsOCClrRo7RBV6YTIOuwXZK/ZxWylsdSFeLW7t3UkB7qpe9KLhtRl69UVh1C+8uAy7w+hF5f8L"
    "uFEfsipn4qAPaR5iFOTVQP1TCmTkUZYg/iuVmqbrbhXB1wL5zqcZBf8uckKNUOOmnEeMfkyaIxqzNWJbYPMgyYIlV6ZP4nlTPEuc8DUUv16cMLkQ/ngWUnQ7"
    "CJPO2wmPMnLGisyCFt4itPB2HmilzeKMQnNiJposCYN2D1q/Rbu9ARNSS81iPDwMD72EqNyWfv89J3uSn31XKun3KotP/22e2ikwWC9ZdwrX4U1UYK0VUJwe"
    "Vgpn11FB84mCWsBwOdGNf4M1wKa6XU1M+WnGlujlI3IvLXoFYbYU/uevy01kxxZkKEOs9rQ8haNC2AryDey3FR/eqsE3PyiYvewXZyGnT8JC+JpkR0i+NNnP"
    "WWyCL0+sAYzBV4OxAic5s4VTudwPErLXJ9z0OCaXhJG4JJT3GCneY5xlXFZxI0VlOhpz97YyPu+/x2EadRBXGBdNegiuMdSkT38cNlNNepbzJf0ahK1M+lrY"
    "DufDFu/VwFXdRU5v+DCAFDRnz6a+ycgDDbd0eMeYQXPZoFH6YZxQL13FDDk44HCswceyQMkD34XtSHiUAysXkXrGTzNDohPXR+rAOclJF1A8lIN+Gav4m5HI"
    "D2lG0vimRA9OJr1e+OiFEW0tXAuXYdyFkBsH4FuYMTJp4K+JMywUeR8KijUXOD6FflpgaLyg8ZH5xUFMOiKSUtUNkTDETogfoQoV+rLRJL0UZ0T9TpfIvCKP"
    "i7zDxFVVriyFpWClpEOLKXii1S7ocNpUOmeckLG/ENlfF9QVgvO29beAdi2zxvKjIz8GwmIj6MmPS/lxJcvcSGhaUSfPUNcVbW8cB40lxqrHVhJbnR0ribG6"
    "AUkaBnWo2LOSWMVLKykB8GBDfw1AxUcPPwZQmn+okK61l8Ve5XbpsjIpX7K/Pbj7j4fv4iSGUDrglPCKXxSNnZiGrI4bnqsTlAGGG586uOFwOk4ea9+FMw4A"
    "Hx8cGEfkZFfwHZHvRH23ZZkb+I7IdyK/hbq0RTQ/TZikt2Lisp7ftr5YOr1EZE6/lugcY/zgVu+2UXN5mRwWb0HXBSsIQZFqMThvXBZKI1YvZdO01ShNigle"
    "JQYJ3h/in/rXktC7nMsZ901+XCiNrjMHT0G0uCATJOXz74APxYV+5ANEbZ6rUKXnOE0ubOpfYBC7qFQygk0H33hsTrXDB8Tw/xtWgfEo/Vh9U65t1pr9LytL"
    "bPmds9MwfsIl6Tk7FvMfdfhxJ3402CqdPhehTwWhSwH/qqkvILfcCD6hr0m9QWTvBlf+I7x5Moy4VVLiq8JpSitgYCehxvdwRMY5eKdKQUPbULgQIR4jHoZM"
    "UiaKIhFyyKSi7tZ5qGb94F0Y1MqPofwI5UdffnQVj5UfkrVyUVeiMGYodJQ4Nv6x0xzDoNzL8Ini1cN4uQrxE9WvGmNm+led0cor8URBB1ye+rMYHxxmZPUY"
    "j6CxD4fw3OTaDIGZQhkejZax1tALKYH2+xlZrP1uRlYPAAKDaovgxvDRlR9qgDqs1C37j+20oK1S6QNWia0qWExGeo/BuGX/sT2SpU+duRC29OoI+erIousI"
    "BiTNzKxl05atq4Wpq6fkMuPBrA9t1oe20TfWZmZO3chpmkYGak27Ou5pL0ql5p/N2h+YTPfDSF3sSwlsUYmtwqEyifZKpPAAYgkGQQMBHNxWeGu+CNgZ24gM"
    "CweEbDtGt8mjmyj9ljKWxKTCfhpfxQmT+CQeQQF5JuSF4DsDfXZWeeQ2E+DHUdwdcqhGRhyh+MlEyC4bum4EV9JQqPKilHlZQUKLxRBaLCHx7/DSmNz7xF9L"
    "Mg6ckl3AZ6qSPq3SbEpLs5d+SxuwLEWMQalHqqHYS5Rz5HLf3tf7bEV0gTsGaF8Jb7tCuISiJmnujLrPUoeI6xnLvM6yrTvLP8XRrXekjebOQIV9BirsM1A/"
    "TUf9w36C913sPEfcrIgww66MT6zmU9MonrP7xLArkmfcJfnKVzzkpeRLqRIi5UqIENUPS7Elo8eYO/zSB+PY5EvI/kzJHTfb/YbcTwU5aKkd2FYQ5awHkyqb"
    "1uyGk2PYTaOwA0szKce8oAxOqMQlQGk3kZaRWEqTMXOSpzDJY60MYIdnUAfA5mzu22kwlPJe6lWb2Ssn1UC/wNRPweSZTrK4Jec6m2UxzjJ9RVEM4cUjQ2Oo"
    "IvdFdovgnngok+2b9Gx1huoqkbJNNdUwU00Fl72MMuq6N+GCXSDElgDviPStLgk36Jqz5UVItI0NSJxEZ/LIaIkgytgBDUczAxq6xmvCeuK3aDK03MFTd+2j"
    "TL/w5j25iu05tK6cdYYdzjEqadGUHY6gMfYHv7Xd7XRU6aD9gp7WEIHSdBGAq6Ap/WAj1lBJGCFwr8PRg80GRc3SdJ7dJBK7ifbXFH1tCnTILsF4eayi7mBu"
    "Sd1u30+bXJ/3Q80G7FuQlgbXXjxseQ2pXoIvySRzSSLCcjVaKJamQzo3YugGE6Tg9nyaarLbpjBxQDPmXsEl7xIemrMT4QrriWyP1cOSEgVCn6JbTjcxhUIN"
    "2rpzuOca8c1QqMZ1dLSAawhZDv8AiyPLSrbIebdh/ipfoj3WzOihlkRqIbFZZz+DI8o4Y4UlWonGbYYS4Tk+Kulz28i/PkbGRpS0YuC07vERWa2EqfejnM1o"
    "ZGxGCVAiZpvR6KslFsS2WJCIuQ56G90N71oCzd0026prNGMj8u1CEd2FIroLRXwX4hNBbkRCwxCIMxP7a21EapLTKdCk89qeHqF0+6+1m6P85dHPuJDrK0iB"
    "32bJTvHbMI3sFMuUTW9JwgCMu5Of2+X8rs9iWm60+jHvlrJh3OdfaRScJOLWMdhW1ozH6mtHfe2pr0P19VZ93amvd+rrSH19kE+Hg/fq66P6+k2V+2RZTP6k"
    "LSa3LS8GUmjk0mOUZHo1EOaO8mwmrczkHJNPDIQRDZnsxRnv7bWwQNfIXnLZHUfs8DmUM8qbmVVJ+CUqZYHfifHYEaYTD3idaYV0B47qL6jDuQpyjNSnTTf1"
    "yaeoj2DmQU2W5+Kdc9zXMpXNlJgcSyW5CFV4idokSVCQL8mX6tevtrFIDtnlZuKnGtsw7LOkkikSKlPwkPG4o/z+u3ArFVEO64wrMuRqKad1FBSnwnOgTRxB"
    "YsOykYYn8upLYPp7JzY7jRUYW7uKUpjTg346GgpFiRLOUVnhO/OREPPTKRMT2uHQPgPTwZdHODWrUvU7jW9/6qfdjqGgjcXpk1wBJo7tRJJphHiSiCj0Vh3M"
    "MYziUhlvqoImwTLKy1CAOklKBJMtATYtsZLg64w1tZ9IPxsgSZvAtxJlSJmYFmTYJquqGwQbFaOQahjmIb8LzVw9aSsRavOQffksRfqtJNNMBNTL3g0e1M0U"
    "2vgG9M5mSh0u98T+D9d7eosjEXzSkuMQIi6VqLnfoDXW5n4Dau7HctiaCHpMoBqYMgE1GFRX7wOxAXf4X30tHycyb3kgdl9RZrkjFP7C0lBrzONWyrXicAGo"
    "vkEf29G/6uyM3/pXws3ze8GITYkgDPpBN2gH10EcjEFtDfvAsHIZtiNEnbo9TsDdL88z+9ey+hsIk8ehtMO69/S/qns+1j1Ps3s8Nno8Jj3uGD0eZPQ4tnvM"
    "+ou3ATN6bHemQNl3OHPCpHTCpHTCpHzCDNiESeeYMGnOhBF5y+nsCaMJMrAJkgSMgAGQ7WHTIH3cNEipbjazH6nTj/gR/aAo8tEF5iXZVt5GJfeSp9udlJL2"
    "X4l+fBoM6Ywt3W/71cNp0C4xOTgj77rEJOOMvLE8dHSyhDw0crIiN7L/75Tu7/C2pcjvC5koTX8d0V9zH7g6cI0rtH/yRcReknFfxrp8mJXHuvw2K29cCsI3"
    "xbvEsTvYS0DL/s7NOMSMIzfjLWSUNv3A8B59G5ybZADFAse8gBc4FtjBAmyGbnMjrrsEhxo+3yU4svB5lJSmILaegm+lqANzFGQuEhgTzpQJfxAAbzGK7QBw"
    "s1Ovg2NP6pg1o46nA+0ZyJqn/H4bX6B0WzW0+oE4jm9iLUCcMf6QXHWjYsiqpMEPVVZr05fPQdZ/kEBYuUCIWN0SOQ43PwnJhuV/sl03UMFNog/CAjBsfhFG"
    "7N+KnxKl9mz/CPeCYfr77+3XbB6F6Rtoa/Nelt5sBwMwQdz8JNUkpaCPp4DN0XSq6LKdMHIy2rFTIx4LBqV7OLt/8E/NPkzp91l5bEp/zMobg8XQ+Kb1rwj2"
    "vo+fir8luu0PCQPK6gbC+Uwp6Gbj0M3BoZuDQ1fiUJ8TCWU1fB8yYl5sXgftzbF4Fb4pTtjmy63qtMlB86u2ooIt76vQbIrx+tZIXgcMpj+5vsP0w9l/6xP7"
    "rvXitVVjjIrtIvgqtRaAN8Nhq+Y9tzMh2HqtqFXurXt0q7Y5Cq4j8Eu3GQUYbXMzCTDjNLrqgfneZiwKqISUF1S/h1NDf9qM6WYWq+2T/06VP17+e6hVS+xM"
    "qWzlmJj85WtTW2RUtfnQwFjbAyYwXgZXwY1k1ZPW8OVlwES/l1fBOfuuB9/Ydz24aPXYvwetS7YDn7aulmsI/YxB31VvU7eUrZLYFYYsa/jjaXOot4KwNVy6"
    "LX9TReB2sf/jQbOvi4xb/aVJ+by5BVrB8VIcbOEF7VLKPtjh9CLo8s19q3IbbFUm7L+7UoCFq7xolRfsgSPvzTIYErjlr3la/+Wl+q6Vhy+vSsHZcqs2nZoX"
    "mVe4m5FDOTuJX5on8fFytHywNGKTin8V0WyCyYLLxYh9qZQhSRk1+7xpGI4hGIdwhSkMTSnYXW6tTUOtoewEu2ycgs5yazcYL7fOpoPii7sXwQs2OV/cvghY"
    "R9n/s32LiU3AgcHExiig8su8QA0LsAxZqMb+N8ISDJug7maXeX6ZF2ioAhNeSBSQs2vFzi+rAmVeYlXbN3F1d99r8OQ8nOqqh1NZplG8XDur3PhGlblmHaVC"
    "2s8JfRQmVDHkbo1lwj2amYP3tfiMXZ0B4Bdc2MLtTSqdqf7+O3wKJ0n0xwr/IRyO0R8N+kMUO+MhoPmP/xsDM00Y7qU3EW+USexi79o0zzApLSJ0fJsqZaov"
    "WxU5PvvIIV11JT/qN3L0pgtoiHeNzu0EmAngzcTUeu7/S9K656Gqf04CNDLZ/JxMBTv/nwXdyEIg3Ci1fMjyx+z6XmecxCAC6QQegJrXbb3Ao2AvZIeWUuH+"
    "1+TX0VX3/FhMxEKrwPaef/Enr3wAC0uFXr8TdT/F0TeVchO1V4rEJLtWqRZKzV+TqUDpMg1xE8hr8x0rg3OHNcrhMShBoar+sYF2M93DzuFKFoJ5y1I8OqD6"
    "KcMVa5+wc3mONYO7i0QcTbixZ9RnPY9vUN0z3GT5QJMd3EnZD9DlcIEE84ZIJzH79492WNpUDW447uolzp3VsimFrvK/AGerfQ3GN5tf2LEFPur8a0q8EFRV"
    "bQi6IZx1mnPlMIo60qBMOhroDrufhMNNvA/UKmC28on6XGutMo6f5qzdpNb9whEpePe/iArCSR2sqBmHT9d57zzee61pObISPKtlZPy01hdwA/WrZC5GzwWB"
    "yCo9redbMZmFF+GhNaV1HPBH+rQ1p/vI+OlMfU/ndW7JnV4j+stw7k1YNB9Pld6MTBAWzCCifNDZ6+g4EtsimgzG9/gIpxn/85+x3pzeRLSIMBV5MXoRcNch"
    "2gCihIY2001RnUeN81Zu68rKf7msJnZOf8WbOm1WmBVYVRsZVRuzq65kVF3JqyqEAH/VXmN21YxWe/5WrbKixHTq2CrJ5UeNQSK1Xo0ibPKYXMDmC6yAzUtc"
    "/qJsIYz5N1LzjyyJH5TDL53IZAkI1whyBbjQEFyNdiux+kJWIGhgptP/MY+SJh/Wh8r/y7wG9goh24wJpeGLwL5n2oObm2Ekr8BF7HlTmsjNNAEIfk4culCO"
    "rqIhuQjIayM3p+RvV1aw00u5aGbVMtpCh4WAgeGvcnYgHCQxj4RjVc8NiiP92JIrd6wunFf8UA1+qFmxcgiVKlE3wqO7MqVDjWg5+rLxlZ2NvrzCf2vVr8aD"
    "min1j4FwlPBa8WXNGDYrU18ITt3+4OSgLdmZD27LY4Vl+yo0bLKm/2dFmMAB1Evsf/US+z9bb7NaFYqbCmhu6lHDu+qOGd4D9LEfGQvwsn8jbS7u+v2elHdB"
    "FSidOV6yz1iWbo+HrZqQ/kKEKE00bpiETyy/LuNu7304vopajVWdciSs7ekUO7ZWgLLsmLV2EXf4l+A/wj+kEyP8o3sygn9pb0b8r9Glkfiweob30fCLK0c9"
    "cgpklmwCjPS3Q4kR+aEdsfTbYXcfeTMVYSoyzhIrwSp9QP1ZsfRy1FTkqC91RzzMbTgKExU1JZPQV2Zj2nYDYQAI1mhfeDtlDciNJBMVgLjLpG6cae+OPpHA"
    "YSZqWQ28VGPJsWPQfwJx1fKHoyi6pC7RyAAGNVVZouav/VJdzFm14WzAxpIPDFUNCvsVOVFeRgG10Ifxl6Y/OG3uoyS86EadTcY8L1lB7MtmDb85auwHf4Py"
    "82ZVfH1mX99EQaFGrUl/cgC1IoBiDBqVqMC3RlYqb4guVOEN9Ge5xknaZ7nYMY2fH1KSwhFqDWctYXA5RkjI3W38kEEn3aParCmblWFaN8Gil14Xl2bPNcY+"
    "60uRetCAg7uUBGmrzKrEplkw4EtsKEh/nO5opIZ6ePC1jh6WZrrMfovxWIpfsrlWVgmfl5KXYRAvtYZ8JFhuAj/4ILwMp4aJsWYleJkIxpcQfmwpfEkXq1hO"
    "WdIMuIwhu0UxDdLlmM3/qMzmiuKjcwk3vhJkl5x9SJPKsAq/auIcX4xeoFI541dDqdI5/5ezQadfyuRLI5XvBmpj0DlyrRMO4ZnLqjiue3dvkEVLJdKm3iMM"
    "tmSWkFuFOcQgof+vKTY4m7y2Lf1/SeuVlCZGcabdpooS4xPcxxdKdvgBfe2xJn+KLt7vQ84Jgxel/Mw/+62ghrWJni0IEiiWRrfQj6H9XNDbWiEcAldPOwV1"
    "DaUi0aSkXCsxnvj+b1L8fwlq9sGCsdINJ3B9xZUi+B3EjAmJS/1yDe4NYunZTfrEwst+6dOt0ynG2teA00TqayK1m0itJspOG6l+peC0MfS1MdRtYHj0odUE"
    "SzRbGOqXC04Loa+FkLZQrsGzXruJstVGqC/onDb6vjb6NqX6ThvQO9JEv6QicTpNdH1NdO0mup4mykYbXbkfcWmVigJsMmrvH+5JxnraQydq8IUYwogQeu3r"
    "uNthhYI2ExlvU7nBBNf8fThdDww0LQKPeuQFYlIZcS1URbz7ig7iQS8cDJuZOajQBaoYTSRAnpHAuhhBWBBfoZpRKPUXqhuFhv5CDaNQ6C+0YhQC+4nMXo29"
    "AFYNAF1fK9eQSKjblt6kI81VL0YuV8WrQWXaJM9oZuyrTYyuSOJVR5uprqkz+m/6m2cCEJ+L5/im+103HuzyMGxSMrxkSZ/ZGE7xzN8Lr+DpsumNFxIx3JXI"
    "Vm+DeuhXfjqNYvNwyvivUGbqE2qie3828u0ph+PeRQT3chBu7groW8rQXniZ/KaxRcCmGgICUo3hrVNk5e6ioNAfgPgxLJRelOBBAuOAnPbIClhK9PvvcoMW"
    "04W/WojFNQW+eeFacIZpGg5Oxd8zzLliQtUIlXrsFKK+QacYwnMZfB0ZVcIkHvZHKZOL2I8oaffBnLhktOpMUj0T7CwwdbKSTFAKFQpEJb4h35tXZk13KtUw"
    "eOcuKora4IUIAgKKGUAcBksAKC2IfhsZnCStXTNV0qKlP/OpMnfP6YCYReSYtci31rdKjfDm/Uh2elNohoHBT6cBVeZuvvg1gTtJ/r+bMJ2AZ4GbqN0o3JjK"
    "t6ZREAp4HAWDfhfzOnEawA9GspUCVzMV+PWngvHrSNqFKX1aoSjLymtWhFOFy9BCqXI7ucNLUQ1jaiBlXbSSlqy+FFpe5PWdLt76ihtfs8VfR/+Ik3Z33IkK"
    "P15EV3Fyzsn52l9EHCBIIQN3/HgRmLpzc0zEeBaGYW/QjdL6TkGNa/NBY6eRa/d7vX7yel4ainFVJCTjZtPXphmrWxc9+PiJVY1EDz7eFAlIu5Z1US4WQH2n"
    "qEkQaKjuzODUnQZcZP53UlwNVmGX5PLt/7CTDTyy2cSzAVuK7wifeBGodfRzUkzITaexehLzZsQax8S+YwXLyM1awLZefI+BwdtTBbuiesXv3lqRISv/lHCf"
    "KyFlDU3KMlqtCR7ldMqVOoWiFjdmQr/Qf5YqSvTDOx9dJwz0C5uKeiwY6Gc2JFHEtonCVEf50a9PXdGOOF9spj+uocdFV0pBQ3GUtzhoEaHeLZiWptPE3N+9"
    "e6ne6eOZco5jxC3iaIBuWOdAOcgNiLBxj24LRr//jmW5CoztzkoLlrBv+SxZ8XC1wXXfdDdFnFbP9td+05a5PhGXCkvyon4Qtn/b6jJZAyagVo2b3nNik3rw"
    "SpRKR3woY/mKcxibbyjD2PNw8jASDycPwee/+IjlRyo/huJD7L+DbghXkV/IMHydOtpLQ2HEq8gJHn6pflUWD+GX2lflIi38Uhc/EvjRED9i+LHyVT6+Yj9W"
    "xY9hyQx7ajpo4K2aFlNraCoFNlfqZgyLGb4ZqKNyWwHobwVfCcqLKaAh6yNQkPUO6Mf6BU+zWI/gJRbrC7y5Yr2AF1Zf1sBeNPmy/hWeUsElVof9eQUvWBK4"
    "xgIXiV9qDM4l/GWAruAvg3QDfxmoCfxdVQFOIyAvwx48fPYT9GEZluPgutwPeuVxMClfmp7mIhgBq/wyK7/Myi+z8stO+bpbnq2x5S4r32Hlr+zyDRcfVr7M"
    "ypdZ+bJTfsUtP2Tl26z8gJW/scuvuviw8sus/DIrv2yVx9HVD++4Bs0YVv1mldjDo9W9/QQxynyCmMZiVttPEPMs6aU1lPUKMY1LU5qYxkZcD2iK+2wg70RY"
    "onArW1mvrteqa+sbtdrG2urK+hrkPR4JTMxZDBwlthTKI4GJ5Z1ojT9RgVcw4KFAvxfgPoiT0o/yFSg7GYiP6tR8MbnAklfmM7D2Qa8+jCu3rVg6XL19XX0D"
    "dLit3G6OuK9yxkQrE11iokpMRIkJlLjTJe5UiTtR4i6InY4N49KPVU/PwBUXE+eGvNQCPUMKJi4FR7Sdpm5n0UtjZYzaj4v8JcqIX8CiFZ4wu4u50w1VlutY"
    "7hOuaokxsvq/x9FwtJWwPRiKvAMlKggGHBHhFZQEx/5BvFTSzjmKeXACDHYyDYaj/oCCYVMRqNK1KsRYgW3zYJoss/b7tC4bAzaTscQ2Gxy22+o82A3xobbq"
    "cTc2H0ArGacu9tSfovA3dtAVI3EP2gbalLJm4LqLlAlUN1HHeg8Cl1kt4bQmQDssVnUapFGvfxMZ2EWz4MDhGeGoEymHF5WaCXcTg1GGeEV4m4wf0CrPQAcm"
    "AZdNScsQ2BFftvmczqt7KtUUUeEX0SfgSIY8/zGSXyWwuwLeFgX3HIvNSKATCBdIqAW4mIwiuJzY5fswSxY7Mvpfk06qFNhpCXxyLUiloUZd2ccM30j0CBmI"
    "fB1Rd84Y/hx8fbTTKFTkBZn5Ik4EAuz4gKpH3kcQ9eSjhMiJW1nkruL6rdVafU2SM1XXGf1Lw73pG15u0y0gwqK9yVSV6Zfjm4WPiXiTGXUKQJ0Cx7TAVS+b"
    "BsTKi5LZHAnz9iY7DN+bBHFt1DbnwwjoWpCdcaDpwAR8SQJSSIqGiZuOTMcpVXdRNyi56lQ3slec7A2SW3WB82w277BAraQmfMhnet+Z5Gnl7eez3dPz492T"
    "89393YPdwzPPVJ/iwixtDt3VxRr7gc7bgDi+i4nvcRKf05ysKag3MdyneF/8Rs7c0/EFTt6UySTD0mYx8eSE4gp6aejpyFBlBwJ2aXNREPDYE3tRJLD4x7IE"
    "WpLgW+VaaVocSu6ClwmKZi3NkaYqUGc851M3v7kU7OqLPHIzH7cl9uO2eKouBkcv68Bw2L+h8Ty7xE4g5gu3bisEd6+tPvv3mtUL2Tkketlnx5AvcAj5AieQ"
    "L3D8cJzAt81XxqOlcdkWUbqm8JUsXZfT5kA+tCpHIKv2hF8Ufm94Kd5kvQzVd608etkv2a/C+p5XYaH5Kixa7uKbMPgrX4RFyzX1ayh+jZod4yVYh74Em5qP"
    "pDrzPZIazPlIqjfHI6lL/khKeNWPW/dhd3Ad9sLBuVRbbb74R3zZiRgTOd0939o//rB1sHUMirVOfHk5HkaomauEhSVDOYdgmEgSFG4+3hRKlavmr8k/QON1"
    "+SJQTYDfudntuMpPCdyGydjXyIWHsNjaPQNg8WUBtI0G5j/qIoUSyxy2w7RDYfez6XEkkLyEvaEQ9i5iVuSo3e6Oh1xBWjSo0pckqTOapIUyf+NTWOI5IC4k"
    "bLAnhWXIAF3mKI0uu+xMEnX2YTnCexbUDu7wLgDV7UaxGsNRviYpIqa7h58Yqqytf/5TZ5yebR3ubJ3sFEpcU8p70emPDkFBO4QLKlAPQkpReR8WZ5JAJ4BR"
    "yE6cgn6+yQFlYH06iNpw9QJoi+OtTFLoF3n7gdOvoKCUj0NR6aQ/vrpOwHsxb5iPmD1yOdPsKHuO9fkE03liiI2BIrOEXgRsvjAvR5gsg2+8Gvp+gWEs6nBy"
    "mjW57Qx/XqKqirgf0FXSBzaE79luBJiSqiLIiKorQ2/I6xOF9bBzOcRG6wU4AF/BgMsxeXuy865YEAw2KZDZEdipqR4J/sBNxPqN2iuFdlW9cGMTPmD/VCvV"
    "+voq/1pdr+PtTr0uBpHWrOm3cVhopb6K7+JWBJQVUQdLw/2AxmMJ2l1mIJpkfVZZjRZE6SoW0grcMLF/A7DVqQNuryr1DZbG538J1yWUWmb/TsRi5EcKRi1W"
    "nuNRQyxgCQNwKHv3DR7t/ZrwNgcgAI3DLq4ENXHO+nsp6C5AQHsX4q5ukxTfM+2Ik7dD8PZ4xCSNzOxO1A4nu7dcbcXHhLAExpWOP3w+3dve2t//fL59dHKy"
    "u312vr/3/sPZKecFAooA/46dCFhrjHTAr14WwNyUzeRvRRtJq1kcs5oYI3Z0s7AuvBbXeveca9jNMS7BGmGk1pwI2i9D6orVNsPKAl5S/GhKxs5qA5dCdxjl"
    "Iogsk3bMwlu6LgCCaFTLM/Bb5szfIZqDMt8KJG9jObjOYXGei23gfD8E24QRmUPipk5vdBJdAZSN+N7xyRGbAefHezDpSdGmbuXd+Wn7uhuzw6AFWrJgrBB4"
    "mMT+B9kiT7pkR6QkAm7Glxsj0GpldXVlvcGXHCtfLqxVXm00amu4ngQMuvDk+BuNY2EJfdnM8/XkXO0aO9EALogSh27vqsED2d58fWXsxenr4SfNyxgGwMxw"
    "mXH+zftNmgsYkqxqmf2hFGL1KDEwU7Gi9+fv3/98ftqLR9cOt0EBytvrfT8tzB6H9YJcrAhJ9IVnXgElEBSMD4Qhg/LLajjZDyACr87LlUwANwLApzkBfFIA"
    "9BpiK7AIqCwBuFIGXbb7aQo+6kA6+iMpdKMotLRgBy0Sf5oTwL5NIbYTS87OsFlmEIPC7vHp3v7RoUmtHaDWAtT5sAAd2GLs91S+rL1UwN5IaZkKx15exsq+"
    "lCAQYKlp8U0p4Vhd2UvaMfAD3KtZAvkVWCxCyLx2spSNF2GWGUyEU0fSIrVEXYR7HXYvQfA27DAMtCvasmK5oOR0SnGcC46wL/uRBazkAPmUDcQ+H9C5kVlJ"
    "9s2ptO+plIGmC4UzVwaAbGzW6NB9R/AHVj6DR4hpL7gAZeNisbCqYsGQota29g5n+Hv27w6ZqvXC/tn2OVjHmJPp0J5dc0vjPHP/49n56d4vuwyztRWxkJx8"
    "tpZ28fSqSsvV91IleWu+3ds6ZRU5NzFK5k4V1qlPdJDqhfENl/3p3A8Ea+M8zeSDWJ79s0Q6sKxQouQe31B2BkTeBh8BUYdfPb5jE9ARyZHQl+YCBZbL3x8X"
    "IYs2gWyUSagMG9iGLyt3SLki/hJiX1W69xCMCRDZ7VxF/Ol5NhY3NYcX1U3EbvleUMSiN3VjQk5YXngxLBZujeQQx2xjdaXxamMV941qZeXV2mptFX6BFL+y"
    "Wq+usf5MkCFPSN0LVrdRWamtr75aqWLdlUptrba2XofzkFseBipk1LggaaPraBSeD9kRFD5w5t0qMfsNq7KJU2qpEPP3OXweIJn5ZIDDHDvK1aLyOkr+5cIN"
    "HRAeokoRhBU2W7THAYyvcHYusPYw4Zgk9MJRo9DbS27sQsCetvv9tDP8UlgpfJWjJ4YXjrGkQI0VKBsp1cJXzcxu6lb5Rn55PJUozYJFF3VQw+lDigaFYxco"
    "w/tHMUT6ZNwoqpktWjxjwM/q+PusZu5VIAcfcvm3CLQ8VMv5DLpVlvgxup/JcySnagi6DaAtDCXoWcAa7YBlofko+5e3SkEiMm1NdsG/dHdaCBZ4sd3RMut+"
    "yShfyyxf85avZ5ave8s3Mss3vOU5/oS0NMOHuqdozSlazypad4o2soo2VFE+Xy32plRUxqwxiyy3crmj6mjg6crDgNUosPojgdUpsMYjgTUoMD24YuOPhuMu"
    "rIzcXc1puGRptxpFCSlXcj7fTW7itI/GhcU/XD5eUOhEoeIi7YAWK0PTKc74tqgtX/AbSoclhIXaQfyYCELBzb1JqAOIyjpshyNQXidXGUSTgfnawh6EBPt7"
    "EK1AdO+PR6JODA5aTzkORg6GjBUZc5F3/lsAV9T2al5c6ds3Bg8fPY7GcLjLUlooarsDJ/vMy+Tk94ZCB1qGsgR+eHMFtDKmyHKOvgp01+trtVdNCQAhCzSX"
    "OLyXqj60KxJ5p4wBBb7Ba/L9kQ4pZGFlBsJUebztxklyfN1nM3KvN2DDE7PBMHWE1Up91Tz560qOAmB4HScxn3l5SoCMI3tR10fqrEo5mSsthC4gIKWyOZMX"
    "yfmO9k+0CgmWRLJb8LTuLqzZh2dndf7lTtH+aeecmOkYkpk13+FZtPf+Z8VvzvoIUCrZncl7dXV7Yu8ySudcR90hV8MYBfFgVIULDrZG6/o4x2EaLZ71KTpO"
    "+xe0rIUAP+rUhQLTLLnMm5UNO7e+px92dw/hRkcu4O3rMO3GUdHZLXjCYf+DsQmw09aWUENxDaqqQUas3R/Wr1kJVpmNAvuX5DFGhXlwUuO8DEvj6be6vlGr"
    "rzLUFamhM8uqUbn2EUagcUH+gMdpKL9UON4z6P7p/DAaX0TdUVH26ZPu3r5FW7kAhH54BQEWWTmGB6sIx5P+PvbqE8TDy2I5p9dRlBRzNmOTe+zPx3JcTmPw"
    "k0NGV2vB69X5iWYKDqJzP9jHsOXC/jz8hZ3DPtiKY1su0pOMEkDrcvX4IEzQ/JQC+bmv57K8oR73Bvm3+G8/Hhxn3uNDdf9NPuSg934tYHQ+dG4n55ffOkV1"
    "t8fTT886qFR514EbmY83ytBBZU949sTM5i196MJRW7UHJ1ZtEyIQlIYyt0bNztvbOWsuCyQZBDZpWYsWnMlCcCYOHOPqm6EVIExyWYnzahClrNgF1xlspRdF"
    "MYfH6eU5O6AH5CdMP/wtyC7XKLiDJ48IyTiwQ8VpfNULf1ZHRj4eEjrDGCaSmTbxpN0RSxUC9zOFO/HAnXjgTnLhwhJVvaU5JzWifOGtB1DcqH5SJ4XY4pDd"
    "N2fX5U40Ugo/XiIA+HgtSolpYPY+DcEoBZ27cBD8ngXHAk0jGIhl8XMCP+vEusd+s4saRQVF9Ri0cNiQmidyVUuvrOf89YG5sAuHHw/Ot/f3jo/3Dt+fH+9v"
    "He6egjJQWXpgJW7xM4CKYWGcpP1u97zb7w/O0dwfWRnjRyDVgBvvVqHaZH9+LHw83Ds6tKFD1vKynm3YABBfoHnMX3KxUl8FDbghGdIcztnH6tkwVuV2Ngxn"
    "/uubYVcmFowPdUYdab3lxZPh76ENx+ii3+9yhNHcaJSOI4FrLpFsMmXSx9u0Tbh5SPfrSGO5GA3/+U/ZQQFqmt1BQUoxUrJFcxxsezF7Vrp7Tu7UNF5f0x4Z"
    "uw+3bLKo4wP7tTkDM2ky9hi8shqYFzYFhxrb3o38CQNIWwD5wKKmIaxuH+0fnXDTT26EZJivLDHmLG06/hF1vZXdapX06sKqSrHxDHA+SoSWK/PiY9E/GxFK"
    "86dFo/D772b63iHYfm7DgC6I6QJIOt7thenhY7EkcBsUrljMWjLESqQKmp3SThml7dZ0RWRFSy31pEIlGtQBRwqMLrwH7FRSaFRqK7XVV/W11cbqxqv1Vw1W"
    "WuXWC2uV+kajtrHaqK7X1l+tbqzR7PMPW/vvWN9W16vrr9Ya9bX1Vysbr9ZIGVOFUq0wUI3qq42NNfZ3/VV1PaNknRVlSNVWV16tsPK1jVcMO11WmnzUovIa"
    "UicB8shjgC6nDgbsmNbuMnGbCQE0bgGhsKjx7ToeRfCUlj8vgYpKP6ZOGSGIUMp0Eo/c1qEKfYjI+6bbpdtmYUrKN2aWt2usZNQQ3yCCYR0Job50WycQQrYe"
    "wivnorAtT2uyHm5zba4bkfcdDfZ/KDICuOv46nogD45h0jEh4n24cZlPy+P5vF55tfFqI8Dr2PWNSr3RYKdM9r3SWF/dqKyurDZwG6L1OkpwHN9UbvmJE+xy"
    "ggtAKygMgaX3+gyXziiAGW0c/BjfbMOj06Q4TEDsa2stBMyZD3vvP5wfw9w7ZRKFPloOmDgag0H3aXgZCbetguNQgsmb9RtBH2WmyaH0wtuGrxpeCXPDOTDt"
    "veGC+g3I5ghmTiyogM2gqRfgXJXREBLvjWHybqH90qwIFKIJpkT8q3jdY+kp78m9Jedbtg8WTELR74Z15wJPs9OmAndiGOITeMYzAhuu1GvodOvdgZtBqpDm"
    "HeWGRmBAZSNyN0RuCpXWQsyo7f3drZPtoy1t7I6+QdpskA5lXWUm2xSKmod5KHqwXyKpHxJ2CmdP17rT6JKq5LQuXA4dJShfFmWLGAAvKBg/RRmSCj+FgYB5"
    "IyQtlwXPIAUDDgtNszVMk12UNYAlo5FlXonq8MBJztElR77wWOylngqPuhxzD7KBVZWSEyLvYKE96SOgaKMBUUk0FlBhR+vPH4qyAYYfz/2ElwQvU3TQ5Mit"
    "YAItoUElqcW7jUYWll2FZdciceVFe1zFxv5KWwS+y/XwJ7DiHt6qi686fIn5gXVqbp2JqjNRdSa0Tt2tc6fq3Kk6d+Y85KiqqQbkCNOzvgxNvT/uxQnM0swt"
    "XXGfb/hEckgsGOq1+hosz/XaKn/Vsq6etVA5QFQMOEw8o4jphpzc8IMrnMMILmFNEN5DYXzR4jZ6xICQutgiXMdc1mMwxwLf0pAFZIM/t8hfLMER7jtWbUMu"
    "2P+lBAg1JwFHQ8iBLhwCxlAqjoPCjYx31R5fROfjm3PxjoxNVFf1y1+0nZ99PmaC/se3u+cfP+GZXoibAOLjp3O25R7Eg/3oJuoWNipVX4k4USVW/CXC27O4"
    "G8Fb/0J9dS0LiipTE0XE3VI0ehfCdLKcp5mqTSZPUNd0KF5YXtG0xi9EJlxWdtBCyUFhVMB8zki4o1qaOcpP5KtAbJg32TGKwzb0Bv/dZGedqlLvSEkto+pE"
    "Va1h1RVVdVrAiM45WN4thuWdaqqOTa0+CZaGwAcVDS14vcBjqVrjTfXbBfN6YXxDR5Fj0jKfGAn7VqF+1/0LDIyB4ZszRzFZRVyjjVpmG2UKIzB+3vnameS2"
    "U5+3nRm9ucttpTFXK4+l2crifZmPYvlDvRBxjBnKDWfhjCeNKqxbm4uYb4LbyMqK5PYsQgeqQcZU7sXwEtY+LomVpNheJhdD93Z78kxV9PHjsmgjINaBTOLg"
    "7YpqsoSnuss2kT/LZ1kSf6MUO6ZEXVFMvhpq4LWwZzcwrgFx2MTap7TiC35JTCGsVybjIGqDZRM/UeNRX2TBmHFr+nLh0uESr92FxaYGqyAbUswOi5dbhE1P"
    "ZQNg6sRriAsjUhGbEmT60beletquZxDLaZdjqmbBklVTb6hNE1U+Gjj2qG/yjEtZwFIEJ/Tkjgr4GNMrsRG+18MJfwb5Z/19XBXGC365IFAbUwLJzULODzp9"
    "JOhJJuiL9NFYl7NAP4ogvOM9fG7NRmrETh4jdnS9JNyVt6KKXLAiF1YRKeciiB7L70G+cfEsBbG0CouKptxUUcn2iqb1qnxyGDVr8PbBqAkp9fU1o2aNL1uj"
    "5gq8kzBqrqBtn1lzxWlxFTCrrho1V/GdhVlzFWa7UXMNjxlGxTW0pmkYPeitSWHWskU567PFW8wyrVU6r3iAL0IM+VKXft0CopHFz8sXYRCMZ6rI9pByvap4"
    "AsOL1PDc2at6NlqzmRVvMzVfMyvQTE02g0VWeDO1mc2septZ8TUDvLi3IpvBIqu8mZWZzax5m1n1NbMGzazKZrDIGm9mNUNu4PAk7+v2r+qgYK/Bqx3DKtaR"
    "EFhFQyRYkSt8hkTAk9HMe675JI6J9lw0njX3DI6esZUzcO/0fgnAnXwuIXBXREYJfZgHvxS27CP7R7pmiQlyT3znEdL1gXZFaPtJHGffkGlkagshs2xAtRpG"
    "bilb58ADjnHJwsayyhDBln0eSYgvE/V4iHopaWZcpekXO/w5S6OobtGob/WRC/+lssYBlUlPPPDoiUcPgUquieSamVwXyfh4I7MRhpN8QETTiSLZV4cTSEUD"
    "91bnxHi3v3d8frq3s7uT1X45t/UcRy+kmvb2UswMVm74hTH01tJhkIuvgbFupOxp22tPAU7Cu0wEA72NsuEjF8iyazt7p8f7W9voVi3Tko8C81v00RLasi+7"
    "zNs4HDazsZ0PUerkZ9mwqjSc+HCmTgQnqzvK/g99cVj9AEMsC2/Di0/Ui4fD+CbK9FO1e7B3err3aVfgjCxeVvK41pdZGi+kpFEDtZYtWtQSEk34WjQc9Udh"
    "d1dknggfOCCVO+AzephtCWp1051DBFsKXETtoFCtuANSS3w0Hg3GI+wpOJMgZcCfkwZk4Yj62JXCvoAiScR1sisFdPVvqXZ56G2py10pvA97vTCjakAs2aGY"
    "fNZlAkQ2gI8psBJQWN5Fm7XgwhRLhOQWRGOPmDw1BvyoPRcew5P3b+ekod4OzVbRhvzVyvrG2nr91cYa+EaQbxBX67VGfb1aW8HmMbVeWaHYyPrV9fXGq1eN"
    "jeqGLNgFmeY6THb/PQ67nk7CyaC6sspdE80iM3Tzgd3T7a3U1tbWhBaekXgVxEjd1VW7U2BEUF+gN9VGrVHdyO8P68buIsNF8eF6Ggl1qVBfXUWDkVp9gz+i"
    "rxmP6DXxoNGMxrIu9aXZAG9e9uVKd+vCsJC/3L0lIm07ihmJuMhtXfuz+mXEl2FbX1fSV0afX4o+I3hWVXwtyy6zApwIFokPZrMGhhf6WJ1JdU1vWSWX0Afz"
    "N6lT379VDo4kwcmvK0Jzg+oHmuQCykuCpDY/wipYGAfmQM0dRcD8QeBVNPoB+21RfOcJKc6tC0ThlwRVtaByR2DnDxyBHbJaJL4cHCX9jh4ncQTb0aR3BiqH"
    "Ljv4LyfIS9obRhH14Itf1OIxo73fv9of3xyo00a1Uq9XXwVcE4RNr9Q2VuBvrdbYgL9r6xuYvt6oveIe66p1kY/p9Vdrr7Alcw/HdjJ4DAroPw/OP5///PkX"
    "ZBMSrSXdP+w6LYRkVQlqZ4zKawH5l77CY82e4ENtnsC/mUzPgCk48PMl+XlHBnMftNxEXaBLTdDPEvArA/Q3ferel1pVmXfHHcaA9lWNuq6nF2DJmOHuYlQ9"
    "8o6sCNCrB3iNbUM1cAhZr6xXYS8u1yrrr16t4UejUWdj2KiwIWVjW16trK/X0ZVkQxatbmyw0V6trNXXNsgoY1szti1CQz6od6qXyyLlW9OdDua4VybETRwS"
    "T+wSQBu1QsjwGWM7UTxiYpW7VVjBkcIefm5gxXmATVhaWk5TU+JANoErvmptwyAD32QfQ/DKnhw3WcL5T0cn+zvHR6dUO4OBdc/6IF1TrUs8PEpH1/2rNBxc"
    "x22qeKE1rLBkuIjKaE3Gj8SoQRAKAjO9lpFeF+nE3M9W5OQggMFPlI18WfTuWDuCNS5GuGUJVFEqghxbMurtwNCnUBKDmcTB0c7u+cnuu/3d7TNhjqlaE1YW"
    "n6I2upLBH0VjEAIDI9UCvUL3QuKcIhNSIMuwvpxAmAwCmmsSjEbcJm6EfaOhfbC7rixEFJQV0KxZ517QuRUN9SYwMRlDElSoqmG4iyW/Jvrdl2Xb7jFRkeYK"
    "OVhQzZ9uhl6BmjQx4PAVSosSoiQ5djPseK+heG99VH4pl9xv93cPd8CI/uDj/tne8f5nDr0/Hl312dmYG6KKKx8jMTB/gvifaPv7QD00PWXcCIxs9ZjEN+A0"
    "O28INFJ7P8shmAejZ8Bga2fHj8Fyy2gOHtLlt+fVuQn+y58j5OlKFC82FWR86KmHa7uEXhVGJphUEo19c+ZCtNQzMPVF6/Yk9yhyaDk5Dx0iLN57l7z2WxXx"
    "2tjzWuXw6ORga9/NO/5wdPheDrl6Z2HtfL7VZG+NbnRP44GZpy+UtzqcwwDnY6Q+anqUosau/jBiPTmt5qPYXNSYj6x+il1PhnE77M58Zya987trRu7bJ1tk"
    "354fIbFpXkX8QcFx2r+I9oTJvvZAvqAfIrnSGXM3jUAWk10cZyUzhRiLgYgG/z2O0omQB9xtm6AD+zb9ObkjN2iipbPdn88+nuye7x/tnLNPI9zximA7nj17"
    "v99hpfW+LXESF6NFi16kWdMMMb+VB8FX3JFv8RKy1N9n7PQKAWrTMb94M6snVMYxBEJ6PWnLlzY8W9YxeyuOLOheyO73knebmxpGu/JFiyTrXO49MpeG0rvI"
    "LLjYCVN+Re0MoHGZPYQ3/awg9oQGONDf0lXUsuP+SntWHrKV11FNtDy4LAsFJm/QMsGRJt4mHKHKcWHZNoZeDnTi8p+FnMc9YBweeyoqK9wS8zCkgp7wCuQA"
    "iHIlFeRlP7xjWVr4eGWjNM+pysI2h0tTxD0MWkxRa61wQ0h3BZGh8s33ubn9iXEIXPCs9pws/4RQyyHK03D9uZv4yzN+OrVcvuXh5XOzcUsOmyW0ZsjbpuKm"
    "xTcqy7uBq6CQ2o5P2Ohi6ite509VYPlQcDr+xFosxwLmEQotsTx8mizetXl0WQYQU4nlATKfGss8IVz2r3zT8t3Re+655monGoyubY8a1OEAQMg6kgkw8lAj"
    "zu0CqAXDez4WAKRV0NF7xhmP64Y1ff9KOfDlr/ij2wHMU2yGhzFb8vwYoBpDfToKLRf6sNfvj66Howguk/pXh4xTBbwA/2sCkgSnxhmC5eFGbKcjCOG0ULda"
    "8hE6j1LEu0tDQWzmDkIGeS1liyJgllJElQPCNDPy3oksexpewRGQWmd5+/j+ZGtnL882S4KhdjXi7SyTA96LXHrgNAQ67Zjc8fxlRqkwPPzaFbQfN3SPrN/Q"
    "iLgq3Kso+9dQpWb2Uj8Y1YZRpJ/CCTPZJsnYKN+RWKZyi37S17kre2E7Ab82LRcmenymLwLsXuZVCgbrouZcWFzu6IY1l8yRgQfniSu4LI4bBlQtJxjJnASk"
    "sC88nupN9iyjXXKnmATuUUjObHd43uUBskybVuEoyPEkk6kDYXlKaTHQcoHj8sjSbJibqWsEWqJVpLfXlk9gUNMngP+v4RQioMsmJv7OQGjn97pD9CftlNNR"
    "u6zqW9l1QmkU1D3yOKS8wfn3Lu2TYIVydd7Ic6I/W0ydnaOPb/d3tdEqh/iWYeBzLS9BevPl6jOdcPAK+N2Ur8aRp0gXE6QAsv1z7SXD6sIynou2eGRLHhGQ"
    "8MSQpOubFi8IcpYmALoqVfvlziWVQY2HIOeHMRs7OkFs703ob+z4aO9QhSQUzsYWd/ZnQ7Ld1TFcj8H1wY4eQYoxukXAVO7ALiDaUDLoStwztilHvUlqVNxH"
    "kVmTSCp+aG3tDEZOdzU3DIetGOVrKW9+ErnanhwauBxZEttwfuBKIMt3emiP/+nx0dMMPwHkGf3TQT9r8Ics6/vY/xljv7N38iRDr+F4Rl4JjmE3YwJ0dInv"
    "8+DPmAcfdg/2nmQiEED2TPDurh/gnQCGTXEmxbWZZc8JW1nhJahnx1y8RWcHXZzSSjpGmRyDUW++kKI1eqJhWEbxTXR6HXb6YMpmnHIdicAuoDf8L4VX6M6U"
    "vxq8fh+RHm4RZ0o67KB4YHi93Y8uL+M2NDREKG6os0Rd7k30L7Y67/QvYvSmQuTYsOFZGRwSNzbW6vV1cUjCsstu4RoW5paT7ERZq62tragwZ9nV6v5qdzOq"
    "NfzVbmdUWzGqrdRfVVcgAO3tHIiu+qtO5kB2DaviEXelUauD74k7/A8idtdX1terG+oMmgVjPRvzWc1viGFUlYo8NhxrfsLD11GDxpTYmnruqXw35L6pPY8/"
    "f9N11NNdknN3cwpTnNnOAjO0lRp5kxgaiEmQrCOBSRLPAYF02EDQxzm0gVpmaHKhSteQlmBftPR+GX3Jli2Eq8Ida7c3bBo6Ph/qegeeGkoJB5JHkLBQEfH5"
    "IHjUXKKJovzMtuaLNAHxoDIPvTRAuyV/UI9JokFDMPGLO55KNoFpReHYkvo0n853YBRDe6yOccagDkzjKWtMba+H1m16O5x4hp60ZJwdHQTNAc87g6rh88J+"
    "0uHlnI2HnsPAw7IVrRcqe/VCeUNN1FMStHl3LNS4yrek9KBql/ZNPoJi1rRrS8/Pg3HSHo3FhFOqwrO+JraMj2dgFNBGOt5EmAxeNPXMLbqye+EHQ/1kx0jI"
    "PQqLiX0qj6gz5/UMBjbnZNcRf5Jouz800gZRMu5dpKFMNxeGxpQeq+3emasi52yuJr0H7nMuCdXIX2JFCOfPEAqOUV2eOb3NBgR199zJ/exJOK9JWTHS9K5a"
    "Wnz0R7uXl/ySlN7QOXVp02SSBLpBbbTgW+MUoJ6vGYucoLX08BVPaeVJM9b7zL3KvWr39dJWVGcDvgy7w8jrwEOzDBSatk52t/x844SB3Uqj0OUdhlYhi5lA"
    "OLif4o64xCSJHyKhJLeXP7nGGbXPa82snLpRy0Qzpb8E83D6+bU5lxpB0ME6dc8W+Ya/TbZtCl2l/XHS2c4QBu0m3NO8jSKJ1ZyvGFBsztfG3KzQNlHM118p"
    "0JUst35QQqh/forEywrugZBD1t5ffScDvKLXjRDa0rblMASe1ogCZp7TRMZ5wmt+ZJwpDDsjoUcZ9Q1HtWfsJzuqRWnMTnk98dFE58r4WTECnLQKduCSpgnY"
    "ujk1LdbB4wsN6yJ6/25/6+z89MOWvPAx6jjeZsSyMPCmLuM1esLNO27VJ7vnfJs+h4qZMTcJN1sw4qaXjjKMreXh3rzXzj6CZpglzKuwzVTGPtUh1ridd+7m"
    "aXUeilBknu/zi+5iwT/JSiRMMRu3PfPa3x5Ai2x/+LDNNlJ4LCF0iBQxh6VJpjmpjYIW0URpHyl1NUmHc61c2j/a0cgVSr+OitWS1shihFS92HU81EcxFFLU"
    "DNuoYxf7ypyqeLYtHY3VW1A+VGs5b9eaVt/+MG7mIV0mT5PbPKUNjWZq08OXJ/vr5ZBzxCZ+OJ/MmSQLLTtTEJgVYjiLQ2ZpJ6U51t+Wf2aAl9bufvje6NTG"
    "UOsB9a9RT/opiTO9VMhcjfk83zslH8f5n2oi/iX4v+7MA3YBX+XF9wLxilBzy2OR9JjNAK9GuOUuG7+wC/OIWL9y9oiuRNDfA3c2joFU5YjrWCmByp74shGg"
    "fFjFM3UUbep8CWOPBgUec7TEP7gZmzO5HQDqBQHvRMCdea2y2jl1l1suRjNbg4gX2SX0GzZxByyN1D/tnX0W7jD9WzAefrgi4GDr570D1I7tbn/c3zphTHD3"
    "3bu9bbBP5WGx60Xr7TnSyxrjwDu0ylB1Hkx2dt9tfdw/82OyQJOGP0sjkpPCQkVzQudB4rvpL0IHxE20TXtVe8q219PokmE1q9Kd4MzO82OjhfOTo4/vPxzu"
    "np7mtKWxz2hUFSCtT6zWvXTTu7UnG3z/5pNSLMXMImRVzRqXjKU1EwNjdfkwyAn4iPHuzZkNkdmVjAs/XOPgrGfa8vrIZrY54qMpC5r99oqWvgUhlOzGCrDS"
    "KOQ8OggrCkkDK0Ca2pHyuE21UiNRJ3PZAfg3lIxeobrz4WRrMEj7t3M9oKWyp+k1LbflZbWX5fMr3PTQWSIvjQ2B+yceUI3kEOTq6n5mOlvFagn8Uod5LmcR"
    "oUKOfnNBgStTHlhI3DLi8FFzbRKZz4zN57fpdpTHGebk+oIFBSG8wjCoUMnVPjulZ+ilveWJypoiIzmGWcG9MEsN3pklEjQN/0VtBr2fdriBjnpkrVN5+DRF"
    "k2XS47KhZrfq1cx65cx6Tov17JrL+TUb2bh6auowKftn2+cQlYxateATZvcdKPdHXzM9FOM1QkAjpfBidbdY3SjGfZHvJTfSfxx9XFxk7YD8WQ3gAyRQIxc+"
    "4NlFgB92LlS5k3W/yfewhtf5yzQaJvg6uphxyANv4nWIJFIo0kcpIKP7y5d4jQl5rj/jbErm95LCaAlHZBccxqHw4IzLQAVnB+IFZBrnt0zOakbD/lPKImig"
    "878a9z/nYEPuxCyG7OHDT6V/eRou/PfXvvhkGiXBtHdm9MyK6Tqri/Lisr1HuymbybaZnvt+yICcdUPkyjx0QyBnDK9clCvuyq7MvcBNhP1YmDqp9+9/zlBG"
    "VdTKyx6f+c6JeVJ/xiNgi5wyGo33/KUk3jkIJEU8Az4w0hy13SnI0GqzyDn9+y7vtRmGzFXcw4WohHW6gRhvS5+hewvNAMl75tZTKuqUFtTTLtCVp7kA8+wO"
    "j1OFPs2e8PyKUJsgamZkUcRDj2xKmcG6T/4MUgpKzt6bPi2yNznPdktZfNoOlI6c2qbI3Oz6fDe5idN+AoqK4p/Crd0NXZCv+QfviItvG1YxfhhwGQ1R2sRg"
    "qnbaDkeAVXLle1gM5Xrj7iieXazNhNhE2KtEHUNuMZaxETYZq5vz4ACaG6rm/HPiIWw6cHocOH0j/dHv6LOkesnHiy4pl13I+S4KvAsIxnCpQGhntzMvRBuZ"
    "pczRmtONgiROLqD5bqIkA/QUVnoms4ajfXrAFVZuVUk6q66zeyitYL83GI8imX/UbnfHw9gwOtHnn0+OXlA8HFHVsn3BWcpDzdBRxYfQ2dRz4XFP7GWM4W24"
    "2CureOZl4WrPriw9oAutstQjy4d1i/uVKDuGEFmOJZ7KgYTRHr3DMvbLrM2FXsU41hh5PhX+gQbIvlccDMF//lM7HlMzm6tXvE8i/G57wWrk6Cfub5ZBtNoS"
    "2UqJqwHzp48EvHoLaRxEntwzgW7QeObAX382F/ReMOvN8ixqscTDk6P9/d2d8/2jo+PzvcOd3Z/dTigqluw+CCq2HDoa/fFadIddJnpe0GjYhkF0YL5RxQX6"
    "Rr2p4YmCFPzHQTgQL2htXNjRSxSASLiZ+RB9LTMTBLkxy74hzaNiKr/NbXQuxj1b5RZ5B6412ZqtqcisVAuhlsdM05LAOWg0F3ufXfQ9TclfrZ63GnOvVd2Q"
    "vVQVVDEBhubvJ1moOT4kVHP0pYK9SufxMvGca9RDvZKFvlqhVsrzLVC5NqFBe2laSNgr05/NF6Y/T63LU9WaXpZ/pRVlvk/NX1AzH3/OvbxUq/bqspsQg9jx"
    "Jj/JWst22mE36nlHai+8xZ18POcydKlc8vdMLUZ/xvOvSdKuvTT9KNkrNLcUX6i5RdR63bEx+WsuW581QNbiVacyPvw5VgDNhy8jGyF7MRnt2BfdxgRz8S7a"
    "tgrPQFZKNutEysnGlY0XXUN94lW32O8+FvVwRuo/yPkaZxpF31s0Fekib4wf5tjGg/cTuLWZZuOLQ2h7GM0eTKV00aM5ayBddanfgZ97BO+Fg6EVdiNrXmU4"
    "pHyYl01qUgIOKulMfLRrzYVuMT0oZN1lmvPGrehewvmDmRjcD+LB72yd7FjJOZ7EzbXtrDxPEBPKhlQcIsuxqndaZmKdOV9n4nXiPilb4CLtxPSTb3XFe4Xh"
    "Lo5HoebcG+Rq5GfR212RLHnOBemmF41Lpjk3n8X4kSejSK68jLkZ+G62FsKKkad/1QG30Bfjy7w4QftH73d2j88+vP34zpqpVi4EMuA9Ef6u0ek0JDKWuTc8"
    "jlKYaiMmAhZaeEUC8p/0jA0ui++YbMWjgdyo6njz2hUOrN+OL99tc39UGf2YFfToAZ2x/P5SVDy+tjXmvkyDCnld8PhXJ8jZe4aJthlcyovWTMRyvH67FHDW"
    "HenNE3REz4WW0IKzWaN8Dn+THbJmmIxxE9Nk7sSqWBik/X9xEV+6tXIv9kT4hDlraxGIIgfu4FSAbrbzHJ/u7R8dBmYHIOitUPPbU71Mzho2ZLaV+ugw9Y1I"
    "pktv6s17BLu/FT6jjm+E5HMAbMIo1nNkCZLNy5vXci1SoGkimO2lO9NBd4/4fldARjHjjXlMwB8dbmv/+MOWioimzUaL7I/ykgaxlFH6Q9Iz6ZhzrttAXDaa"
    "yRN+3VGq3E78NvOU/JKUHuIrW1Kbli75FZSM1wqylzY0CL1uNBx2B9fhgWy9cpVF51kcdx5iy4FFQ1lC7Vyi5c+GzE671WRPaQ/lcyJz4YiAWtZbo5ZOaZqI"
    "7p5t7ZNXOXqVHcgK1mDL5ANzydktylWkwFQuspDPWVQWdh56EnQo/H46uOZipI+9HxydHH/goRbxJNC/AGYpbutgzkL1szBlcuLbcMgkxMvuOJIivlF4WRTm"
    "P+FqlNRV9bj1+qzatezatdm169m167NrN7JrN4j3GEHZERbKFAOQvGdbJ+93+UHL3JxziSsPbr5RygFE8d0Qvc2UDjLqrah61sZEejxHZ0kcg5lTiZZdNsrO"
    "MZOyK88xkbIrzzGPsivPnEazRjgb9Eo2aP1eJLv6anb11Tmqr2VXX5uj+np29fWMaSe4l20uIULXhO1oh/hOwzNKPxkx5gsS8xvc6jeVbCYD6lh+IvDtRedW"
    "qYv4M2bD2KFyi49ZPRmTrIw708z+sjOhDUyyGphkNTCZ0YAvmEg77Q+HRehcgAiUzAe+mfVujPBXmc6qrarwYIQOSLY59tnW4fvdwzOiAxuFyVWEUTwoGmci"
    "1VCXXcT+wm9V+hxetjUI+eVDHv5Hm9Pf/tJUT2aIV6LDp8db29mRoM2HwqQUsQJvMKK8PZSvkyTygcbMiTyZFWRMOfA1nuMrkyF35SktKdD06O3/7G7b/YFG"
    "1JwgwhJPoy+lMX55HeVweWxSi3Pv2DvHygXyztCZVY+dnL5VINHGg+MSpasV8DBrdLXlZzg4fABJoBrLgP2Tl4foqVEzbzF5VzPMmCWOhHN2VhUGbE8fpxdC"
    "itpKLyB0KGU4OoAoQAos5mvr9bqZwb4Nchut5jba+dC5nZxffusUS762zf0iX542JqwrTKuxaVqh2+rWMFjnl8w1YRya6KyyFI8/SIoVinRk5dWxzpyDnRRy"
    "+AmZmZ5Rx/Rows6L4IaTWziP08tzQFz8FhPAswebb4X/XbU2VQGWbqcyaeIm3dnPuP5ds/ZQG97EhTfJhFcvDEeAoNi9P95UhiM7v8bzJ778RgEWtaKNiSgj"
    "7AAsHvkO/O9aUDi0+lI1izCC/rtqFgF1rIC0BLjiy09RDxJqlVta/K1dfGIXn1iee0fSpwO+YzgLCtxnB/x4GzB4JTuQL0x87q8Y6raEb+I3+HfTmgpL0nH9"
    "8N/wvgRqWEGjCY86QzNwzvAAWWyoxPB/qzMmRoZkaZU7+1Hni0Bp3M8zxEffJYl5iXhoP3O3TFbzGjF2yoxV6LZ44GwSFjIzdgsDkNg2LAiz9w+3/+5GYmLs"
    "7CguCN/WUrDYvHOjZLQya7exRySb99u+Xlz2Txv2K6myvbnkQKM+W2aBnbFBufPC3akyRl6SahC2f2NHJBHqEH7xcmf9k/dvi97I6pbNvCMooVJcRdEkMSLG"
    "CcBncM/6vBHnsdjVhQWcx/Zg6SKcMJ/jDCCvyLnRMYP6cSB5Un11rVJ4yf6sVpq01Edsfaf/LdElV3nJtUpTQkQ8ACDX2w3VPoNgl+ifwPOjRAGtiDY1KBuH"
    "l1isSBsEtTQBw3E/vY4vRydwHbiBNyoaa2wGgcIVBI7alvNK4oY4OF0pCD/SrFkRsxjexNAuM9Yv1ZhpZXL3rVBuFVLBaggmTR0eRQAQg8BHiDetxnzrrI8o"
    "mmO+onETsHDTuQks0pXUNOK9rX8Iu5eytzjPs/p4A5p+2VHcU2DY0R3ADWj7VdZEZxnhTjgY1n/w9sJK8alVDeAH/HOH6d9o+jeNbt2gAKJd9Hecm+HdSK8O"
    "uiWO6p1I/qaTJVXEIDM2+stZn77ksAlOCjpPYhK0KrcSL8PUwrLIqzNkoALLfMmS8LPMS1OU+jYuZ31g9r84KHXxKmS7Gw8ejBgBgZICxQnOjDzqsU0tclX4"
    "HMQSaCwLopXw8CmoVsTPsqTjkixDsRxY6GXRjwlYjyKewJNgV3DQk23QgWb7RxrhW7hBN44653gv4+63xye7Bx/3z/aO9/d2d/jVjjf29lLLjLsd0o2K3+DS"
    "0LwrJKSuWu9E6ei42IPNde8QbIy29w4xHrcBIE64F3910te5ZKM2qvT6naj7ScdasuqQ214Qf+xLaKP0i6ATj675E02Hgjt7Zx92TwTSnpjlqqobudw4Fesm"
    "/JKR0Y54takga9f/tjf4qzTunJMnaWwMOkXTUuR2Yp5pONzzIewn54w7EiO9+qpmpmXzJ/1FYqk4oNAzIRce7FwGMyvD7IZ1SOEdX3aq2WcO9RbQdxFpeZ8E"
    "UlEXbXqa2gKlvoikXvDIGSF1XQHyXcxqUV5EKjDGPbGBfLbwPFvcTX1S7ouAm23nA1em8tJsMPfpga9xv126H8xXy7qG781Z9uT5MDJihAnzeTs0ijJyb/oy"
    "uIScmc0t35vEtVS9YBjXS9uSpkmkhV5qDPM7bBp35r+/8g2T9aQnA4B/gJwHOjNq26GQ/kpjMuNt2nBG59xhyHmy6hsH+9ljFgj/SLhPGGfVd+Kt/ZmDYdbS"
    "Lyxzst+pXHskZz4IHs6iju1+QnH47X6PsUx4CqjZHIiEXDULhk9SEdvmJekjDv64HQMvidzAdzIj24mEDZZENOqYoKQqCdGQ0vhizMNQkrWN3VW4Och4TkWk"
    "dVmbt67b5h38dHogyFvMbNAhhiGv9NVj/BYx18OqHdIhusNaHSUImnFtwrRzPlSvJCnJC4EBvHJrBNaiFX9oCfNCa0V0dLAvCbNsw7TmLVurKmgOKHirlSr8"
    "n4XKBMUhI6FkL4D+5Qj283MmxV6EF3EXvEm3NHx2ZlDfyxrTJf1ZanJ4GZDaXTaWcO7w5kNU2IY4mlQrr1ZlAnoaVoI+wqeDK4Bi1wmFA38jFjT6lEVMWwXb"
    "npP0pao1J7lA5vAfOUPpYzeXgQWcyw41f7VlbzXZ9EQmpVFt87JFASi7U5p2x4PkWhz3AhuPk3fpeDga9z5Fbbj8xsSiAfK28LrFaWem/ohomakTb9mJKKsI"
    "j5GlVdMF8VrRwKVE0KwXLnnGWTQcaUTrpIrZ4J23QQJENmnBNaLh0eJqXIwbfyXN8tcrx9vv9D2+ZKVRFyYEPwHVpS/Pl7nbVee2infhqja/LvFtfaLCxK4w"
    "mVHhFu68lhdpwa4wswXwwwo9eQlnNQcc5k4ycm8bWLeWURdzJ3auWivKRR8aR9s7rF625koCRQ5/LHsLHHSCqjlzTpUKy48DjQvjeUAzaj0f6HqAI/ZsBKk/"
    "C9aN5wON2opno/WTgn4ewj5X72vPSlhgHc82iRvPN4kbzzWJgYs+G0Fqz8fmfKAl4JJyNveyUFsnMoBlr+Ru3uenR+/OnmgHL7To5upsYUb2xDzJ4jMXs/9m"
    "gUt4TsWvtsY3ahsWmPCbWSLisyJlVh6chMsWH75nwuHroUPKENHD+Oh1TqCJ+fZoaIo+uXBQ/TwvauXsngYFDfPX0dwQhX47G6wB9RJOoo/D3kvYRyPvh/qE"
    "uOOkKD8l8ghR9GAe7CePwB5I9KTIA8BH4r741JmnB4ssdD19yrN7gewWjQaNvi3eizyyPUFX5h2VjP7gWPn2vldzb32fTg80CLUlKE3cIsKlbk9aihGIj5Fa"
    "Pcb3pvZmqP1PaVVme3wRnfU/fhL3m0pzqDj9Z8uYNbwYgvNnDGV34zOMPOtvM5ji6fNLFRfvE1ii6B+TgP+lxqgIGuLTazBSN2klE2+5fHJQdKk+c9ANkxDN"
    "YbSMILykdnv94WhLvmCtYfRtDcZT9iiJlAtkWptqRESfXrdIHept5hKtWO7gPqJS1VPq1xFHtAJS0Qq2cKNkomkB5kpBw7/NhC+GIb5KfgYJif3lNjVa1iHt"
    "AB5LovCyoCP+ymp3Mke7n0m7k6x2b432PvNfdtGJgeJnPtRetaTcfGr1VXFHjmGq+NDrvamxjtnrq65O3XaF+cxKTLMU9WvpvW4xVmCewF20BN0lkb8i9+WS"
    "x+XOWZ8YaVi6U9NQeoA2z10Ml1p0KoOdkd0hAyvdH19BVF+zJnyKWP7gqdPYMQ1xHQyac2ogjfdGmaecvFLGhoCj0r+8HKIludgXCzX+lt7UAVIeQ0410hBp"
    "TllcM20kyrJonI0YY6tUAQlm7IMFhPwswJPnAnw7uX0ujJ8U8DP1/va5xuu5AN/ePtd4ZQCeQ27TcpRYRo8dV0eSmmaa8HiegD/UgKcXjlZ8tjtgK/fdfOcv"
    "Yr6Dg0Qtd/TofDfe+cOMd3AUDLsdPQzfTXf+XNMdl1E+ikeCEJa9OGVu7qwR4X0BnZ/6abejHmSJh3nKe49iL4Yd9SEJryjdZ9FwmgTyMY3E6s7tzE3gkU6g"
    "JTjH7bqLGYPwzfi9LAzHXfos5TpWrthzOSiQZ5t5Dolb2ZscZC9lEjTfL+5iHP2xLu6fi+Rex/L5tPY6a295Nqmnpm7Oont0qI/noq8/tEY+gb1RKmSYjucg"
    "sc1Ch7/Z9uKW1dUBK1I0IqM6ZlEPFlDnlLmaz8jH5vBp7/H5Lkiw1LL8uD/Uh/ss7+25ftsf6rH9KXlepqzWfFaO6Iv98dDhyg2DkRkAwx/6YuGgF0/KIXME"
    "ruazctCMmEmLjsgikYOyYwZlRAtaOE5QboSg3NhACwwqCchq3qdMGav+LU4uwmHkFXb/d+/wUL64g5PLRT+BV3Y/c0/7b/kvdNQKYPaSTnSrdPa0wue8ChNP"
    "hV/yKtx5KvyUV+Gb8ZYO0hP1lG5Gt42D20WcdJTjl8y8PS6c083r7dHh7vnZ7s9nH092zePLNRvnAdHaQ3fOuBbIOujAoiG55DyF7Vudtx+aunce/4JzBNvz"
    "VxwrzFv+QLJY+FcgXftaDRN1vKw04X6A+ymrVng5fz00p+Hasaw6PgubOWtM0JwUNXATx3qGH6Zr5ns8AitQ9+sI4FYCCAoToyMcTn0ROLVsOI1F4NSz4aws"
    "Aqfhg6PWF3dXBo/ya6xunf3XCKCBkn0NAIVN58iO70q6aNMYXQcebP18DitEqysWndEUUxM2WzXFuKQA52BqSbCSSczgD/wgzQp/wnLoY04yAn6LlfW2mdQF"
    "d0/yHbQOiiFzllua7y7RtvgPHnFU3Lm6dT7n1Jlk1Pklp85dRp2fcup8c9xtwnWcwzFFJYBbEnd4xnBk+sF19ikoLsZATl2TriITruwICRkCktL5JSe65Of8"
    "kne65C/5Jb/pkj/ZJVuZxFIzzdqbDB+5cm4ZFXgSLSfMzhTxs70M8Vpnyq/iDPBn0rWhA99edSIYg+8Jtsw7ZUsY7nHNd9cy0rXz7FoFfjC4oQRmPbq221Cm"
    "nCK9klL3m57Swk+ppzs576ZN1H3vbRWyBPiIzRUGeOD4HSB+DUHsYGCP2dowwkVQFwQjnHMIZ4YTAtqi2x30hisD8epowio0b1iSr7pC+lxLO2gwvRsTtHZv"
    "B/0hF4hQI8odsJ9RvF0XB/JGzQWDsZq5T3zp4egkihN4ZjYDpgrA5kfOE46Y1wBLAu4kgds7LCuwJYLE0WAU95j40tlm/esnj8WlreIG3BZpHKVAwCnjm8KV"
    "guE0B8MnS7RBOFir1CW9pPgjTCOyytQq69hFBn0NXeBwN0ysjNndk5OzraRztHP2Lh5JGy7iC6hRCIGxYAM3HFx9ZXV9Yw3NNfhzyFfV1ca69qB7oSpUK682"
    "Guv1V8CHsO5Ko/5qtVYViNUbG9WNGu13yLp0QZDb2t49fRd3e3F75jCAVIKOIaHOXjIYjxgPlA5etR4fkFp9tV57hZN/fa3K35nUN1ZgTGipxurK6gZkvqpu"
    "NFbgo9ZobDSUcaAay5WNOn+uUVtdW4MP1uf1db7QS00PdkfjUSZ6v45qlbXqClgalVmL1Xp1A7+q1UZ93cCQJa42aiwbqrC/tQYvuF5fX7OQxOTG2rqAtFZd"
    "xUrQ/bqJZ/ZkZiNTrawZU9qgtF7NuoA5t9okRAYFoclBYWSsYTp1t8fDUb+XMzMK1PdIs8CO2Sj89OIhvEHNDBFydrJ1eHqwd3q6d3QodoNRfxR2z0hlK3YE"
    "hUv8C9KAH3bT2TuRp313N7JaJA2Nb3whKoqmG1mIAAbuUH8gyZ9Ozz/tnrDz8fnR4f5n5d9U33nWoVueljyy4MdPVHAxIRNTNA7POKF42qO6k5mRMxCtTIxg"
    "XLIinMDLCyd6CYCrz46yJILY+QJ/HGknxlbX6r5WCOoPaSMc8WfokXzKX29mNe0Qs+6nZv2xOGGsPk71uo/sdS/d8VKGMZ/s1lU0Odr2zh56qNr1ufElYWSV"
    "ePrNugvKPStmesL6dWTDcdxhfcu54bUro2OszJqSQhdh+7ertD9OOjg3N1942ER9B+Q1d03d9OMOG/c4KVLHf4zOnnhIDAgVz6l0mhURicZD+kectLvjTlT4"
    "0ScsvzaLREm732H5Q1pgavQV5sPmC0+fcniD2935GYERh0rNkIEKOTAJ+Ct9OU0YtmAkJsaEdu0GGH+73+v1E5OlvLbl7v4gbPMAlKSfTDTDS0HqfV2DxzYZ"
    "7xMB8Fg2Be8bbwbvhBeGXjmgzWG5ofvla888YEW21fZtOqKD3VL36IlmBPbXmQszaIS095LDrAULwGPWMVCOdukKtac5ekcWbMvqjOmH77U1uTB8mh3jjHUV"
    "3cmo+XRZwJhx58db2/8LR8pWq9CoV6tubBxNcu0m16EEwY07szXSbGHCyHTEGCMXvRnmF8kMZGiUandjcdrthkk0dEqanOBDfHV9zAYsBsnol598az8btDEF"
    "kSMaAcjk2jc2hOzRcAKOWauA3PdRkmasA0XPvPxRNBxlFfCFv3ytr/wvpbNBuPWp4nMWm5hfql/B6a6dWvuqnBXPpIjFNISXWXUqL1M02LFV0Eu7x86AXvNC"
    "dzz7UuDGGKhFxhlK7joxxV4jqxMP2UxqR0Bay8bXXDn+uFdGGd8FWf7q8ZTxrh1V7hErRwnZ1iSzLjJfW1ITk86O99mp72D3UAd9tjinoWF+bRdxg7E5RRw1"
    "9eus5ZbDqd1AXZ6ekrsJK9eeCfPuBv7h9Rawx5YUssfS2lbu5L4inEvB4pDbi1AYSjFaCybcx3h0GaUQa4qIo+amAyLgjoDrZF4aeZ5tm0q5f629auYuJNdK"
    "4Vm2mefbILRzNIjdKh5nmaMBPrrtgVcBOrFekQMoG8MvVJRk0J0CFIbW9WCC97DhcHRV1J7OgpE70/kRk+6vwvhnMPXvTPvPY9qW0mIBjp1rkyqu6tgcj/49"
    "xuNI9ol/V5RY6OyYfS7i3qb1sYgGMrHOTOo+vS7w+QgvvSXGH2+IuSS9e89QOsh+BBrYn6h/UHT/2584WWfA238nHF5HtuJIjDbuRc1snYSZDpCEHZR9eTgK"
    "pTccK/w6DBWRBHLZLqrvc3bmy/7V858xH3iGRM07WnEZXQ40bdgm+VqRkNosD9th2jG8LTQK/fHoqs/akvaX5EKxmS9LiETnRDX7dGgPhJFpI2QceNOrC9+C"
    "FTYJtGZQsE7KT7VyaQGYJxlZeXEYnCXDOYA51WXkFq1+16E5PBLv41aBRyRQa+CBMscTnCUd/kZ7VxDuNwpLNkncCeZlY4/Z9J/tmGXSn3LYXjS8ZhJV3H4Y"
    "g5WGHHaMXWPXsUKpzZg/GbErFuK0uScs557sac9fYX+GLhH4SH6RfAW8Uy6HTrmq9nl3pky7pL/Q/vW8G0rOYdWz1zyhKtRn4YZFxH1I1OEbWmr+bBXMfMui"
    "Z64f7k2ivETFjV+/C8UJfSCEW0M0ljl4KVcnkVYobpU44dL2Dh8osHM0QGp52UgGswUIJqRS95JRlAy10lrd1M9uj1hbZWq8+bo2BmEG4KUMMcMrIeXDWvyW"
    "6+8nvJgHXDtGkr1dPU7/bVkxZHJVn3TxrPLNH6FMX1zvom0rXs8tCM3Q1XCbCI/s5NGtzFLPzFTO6HkVXYZsCv7Z+pvHCXoP1t04c3xuEbEb9i5Y0nxCIo8o"
    "jTZfN3nCoyklIot6l/YTVyG0J5ihzM0IAu+B9zZs/9Z08yRAkf3gW+fvsuqPcpz/YhLtxbBzOXT7InghruxFJGB/PDNvGfsR+3++KP2HyaK4eFEhtStmHbyc"
    "Fcd2wnD+9tK9f3lZRbw8cB5pV7zbYLz0HRtGxhQhpLvJZAubDpNcVKS2mbZ/j1UcZkHB+u3Jzrtz8ft8n+9ORUfcNs8wM2lltpJFKL1RAZXMbSabRDZsc78z"
    "X13nVFyg6+xoZDnOaM480cxzNLGGujBrLvhX7PcTjStdmVfB+1sHb3dPzv5uctLznLoW3c//5FOa13vi3/Mc9zSHMvck+Pgj21/+oDbXUSz/NJd3UBMrgHIP"
    "t3uWd7o5znq9cNQWrFgzo4Ots204rj/wAlZf+3PoDuuC4NqGecvf717h0aex/DPIxbg3yK/Pl8h3Lf0fLscL3iRz1XaUV4a1N9QiF/hc3IlTy2yFLgp6BLq1"
    "yuFBSYCo3Inn0mWVcmseoDAKZ9ofDlWdoHBLbWIw6pDw+dFnouVtIFpDv8GQMtEpKGdWKyuvVk3baqnnk4xDXRdwBuAxpOEZRrBVsyz/Zd0J0CIl542e26Bj"
    "sajkbq/s64jVSxTef47lwLyiqtgcTElVjvEfzdJzRM18+W9eK8j/TG3+31PY884eNn+U0xLCER2vtjn3ev+ppvUZQh6jmLGtQLDY3o1rvwmLfdQX6lW91M+O"
    "jg5nSoHzXgB817c/l779Kg07MWVuDxE4Fzxyi1ycNI9Tpf9Xirrfteh/hvQ9l6bdN8Pzi+QiY5fxIGMXYfvW67+a8va/RubF8TYlXr4N/oXk3TzV6rzS7ned"
    "6X+4zjRrSs4vQv9XvU2dLSfPr8d9gB52cN0X7EjzneMPR+DH5YHyNw8QIi6dHbl8eB2zoYiGw+8S+3+DhcyzCP98zv4VpP/vljbfzwgLWdr84ccIzd//G84R"
    "3vrK224mAFLiuxnJjHMKn1DmQUUIDP9hJ5W/ton+94PM94PMf+xBJs8g5UGHnMkwbos9VbMtcIKys3Wyo2zYjj98Pt3b3tpHiKLQye67/d3ts71Pe2efafr2"
    "/u7WyfbR1hlNpM5qXSfiD7rISPvjq2vvgakXsa1q5lEKO2biZXvGo95zDd6LdW0CWOiJA8CN7VoP6xpEMiu2u1GYttlXMyfvhHTeAs09mO7uGv0RwSmjiPbj"
    "Cbel7OgDZgURXKDpy3sbj0iuGW3g+9n3+c6+uS6nZxxL5z3uLnq+lnzpwe485jsuz2zljzsxq6WdU0axvHxQiv99P30/9+nbu4lIZ2muO/iWvaVkPMv4C5/l"
    "jTmYhRydgM974NfLZh6A2aUfpkvIChSQpXIg0tZ3rcNjtQ4Za89SJiy1Cj0MMKjcBtYqTAZ31+YyupyBUDUnUTdkglu0P+7FCTRYfHRPSsBq0P23WhuOUdR/"
    "kaZErANTWaJPHf85gun3O+PvqpY/V9WSswB0SDg/FJmvQpSrhWBWkebsMjCdAgzG6aKGCOvqLp/vKqDnvcsmgo5mtYdHJwegzcnRjuhQKWTueQK4vP14cOyG"
    "VxET7fR4a3v3nLfmi2/j4e5amfEX5fIL+5T+T7CnXOhs8ijR3isFAc35YKND5yJ5EUMPiXq6m5KFnO7fp/VTCS9/twgGiwsR3wWC7wLBMwkEz8uGZssaJJrJ"
    "gJ0tha7gAd5HH+GNWTCNUdzuRn9Tp81/AVfLBhUfrA501X3/Nf6axfz3+mr2+SqXDpz/s3ww/xmulTHQDCM++jNvSWobm9Pp3i+7W2dnu4cft86Ueu+i3+8W"
    "4uFxlIJVGajpIIoi/c2d3zOhkTeKkU+JO3zhbd0EUSoY+CyB9yHuifqlyUHvct4y/hEeKezjHT8N+hi4DJL7ZMbET2uy+kinbs4c9nIjpIHmsBBtmocPs1wz"
    "EX+7z8ia9HDN4T41j01k6d1yFvZiwQbnnY1/A+80C6pLhoM0HkXPIBD9mc5LvstJTx4O628sIelZ7hN70v4otCK2oVuSNqscpY+IfZXH0P7AEz53SnJjR1GG"
    "U5OKpMwHCv246H9qdK7WlST46wi/Krc0LBpObhKJ50uhWvhauQ3ctIkn7U5tRxz0JB90zQO65gFdM0BLRcIfLmiJAB/WyfTOdQdTL4RMhGAHXjJWRkBlBqEo"
    "5iX420GXORCRtCT84ughQmg4syMztraVhqPY7g+LahkgIAsPVqrMBNZkRqmJt4UJCrvJzBaW58CDt2Cc78H5p6+jZmxqZ8SWyFg0/4DIJ9NgHLfuOQPZvBc8"
    "cPP+JuyOo80k+lYYJcXa6trqxnpjpTQNBOuXBWrTgPFEVXzc7TJ4OkI3hTMZQVY9Mw/Z+4EFbBqQFyub9/KHW4ybq23es792ZnDJyLJrpJcZ3tQslHSHJach"
    "DsgJDLfMqbzagK7eHsSD/egm6sr0KmsbjTg278O+0zImqXgMupVpIG0GN+9l2AazphPMgVYm9jCsx+KHSxNx3bF5Dx9OC5B4CuuSQlb3H5v3/NOpxpONijCC"
    "N6NiLaiVGAhLnwdzSic44GimhY2R9zYOh5Tm1Php8179colALaA279UvtyBxFbJ5L3+4xdjy2bxn/+yYw1JfjcqrmHsYhSnpBEt5pxPqUQOTUMiwV9na+vp6"
    "vbZaktODdTjsXQAaKKEYdb58FYWO0/5FRFNVyL6wuy+gyMyAcZsBW/pxNIRhEeU276dBm8OeTt36/HSWBYUL9Xx0puJQxZXNZhpYDI3Jb0ZYON87Tcqz4CCj"
    "SzIfeCUtMhz0R7ndlT0M5LYF3wYNZDRSTo+ECYKI7yBKxr2LNBQ/O1E7nHC0VZt/AImgLS9taIZNFNSpPYAqoouUIlMK7bn7KxK2w16Uhnw9mWm4ohROXroY"
    "OTZhroFpDq6jNJp/iQx/m2xLal0xhpN0tvWigXJbaRQuTuxvcWd0DR/XEdTl4Lqj9nnN2hBYUt3iRVxn+qBde4iU1j8t1uts6t7dOWern4qTzYOQ45Kktb1U"
    "VoMK442BlMP0bvAE2MJcb91jRJ3Ne3HaGm5+TopfxnGFC0dMTqoQYQR+cqEDvlAEgA+5r8M34/NfSwGXtWAmsi5dxxUzdE8gj4O+ApDHRp77tX06vIj4INBU"
    "JYaBkiesAauybSmjN9TxbkZ/aNwUNm/hgeaz9UeIPfCpxBn4YQkmC3VdiaH2NK7VNlY2alBAeo2QJRrVTILpB6oZ5NJPohlgxoI7YdrJo9ezEYiKWfCbSlOL"
    "EVBBolxGQiNrmQvwPrE3k5jEhjWTnsTeexqAI6M8cj4xEYlk+STrTflhyuit8ssFnBHclOb1dcHVci8h0o0oA0/iIzUDU+JiW+9mNq48PZuhkptDTzPkXp0x"
    "eQwIm0cO7KO9G8povyRJhQNWkr2XDFYUWg9+VmhnkL4GIBBko2iNjKdVBJHVIM/kbfERf4rpce/s4j5ykOsBD2bEEk2LDzZqPD17NhCFqqcJcqnATsBh+zcu"
    "yZFm7nPUF6P6jiWB2a1rkFkYkBIcC3i453ZTcPT5yIpv/zLawzzekgyJTnurorbP6JgRTt3TjBHmXp8cTt6/3VpgLgf3aXTJpHJW81jJyHoM9tkYwMOQHXUu"
    "0cdsTyo7a/uoRXHLXCS0jJiQeJZwOyP4uOQdbc/JvuqTcn2okUtJ3+TVN8zTafNynPA49IO4OAqiIAniIC2x9pPhqDBsqaab3WhUCIN+0G5Vg+sWjG8wZp8d"
    "/CRgAErpPmFL/pKNwbCCPakMo9E2vNUqjippMKpcsf8uWGtpaZpGo3Ga3F+JAvw0JMEVS/8/e2/C1UaSLAr/FVl3HrfKpGUtgEFyNQez2MywXRbb3ZinU0gF"
    "qrFUpa4qsRj0319E5F4L4J6+fed97zvdRpWZkXtkZkRkZMQDT6+lMHTlIFCf13IfUqzDyVxoYMCmTsoGMFyy1A2k5EtKHehScyAZjIGXqYIAmXD4VHLEEjZh"
    "V+4DDsu196rpeV7SCNMTYDqC9cRYn7QUetcLC9eQfMrtsy8sONde0IDWOdeu2+ODfeNljbuE3Xs3mHAS0CumhQUz5Li9+4WFuj8chih8r0Ol97jGwySOSLw1"
    "hmbux0NeAVbsMvwLcNfrvCtd0RIaRACbOtes5bIr6ILLYHb8WRbTeDw+XrkLC1ljIKZNJVBOZkRs4VFgRpwA1TUIoe5rbAdWBruH6Prj43VD3H9Bq8buukM3"
    "PTh+MUDHhHFfYIDh5/cIpYHYPAz9PXIeIuDbu/UPanCxYODLgyT0x3Wm1tTXyBnyPash43KrRCabkfm1ImHsaOB5h7Dw+AF7GqRZ95UIfKGjBkIo3HvVmrsu"
    "i2HyYqBPk3s4MAFVgg3pKN6p83Oq/jTQ7IYA4uhDAN0IjgkNPRP3WeQ+ZKMwhVFF+cQXvN+HRTe9l1ufE5lJgMuHl3hX0ODah0dcsnDvxAhFA8nqnHquM1yV"
    "JauGqpMD2+DADdqQ5nOXhY3ZdIgP82IaAFlqeQbvuhREy9olmI1FiMP9KAiG6Y6CXH/TgmkZIa4vLMzwB2eXL6IhBLMG3r3uc9x7fDQ63KCSzqjVsA6gDCCt"
    "dXbY46y8gI7QUuDUrjInNmbP6Alrwn+0/tSC00tfIbwPAd9E+EHotFm7Gt0rUV1vN08gvAH0DNobkGXI33wZ8vsvQf5yZPQNZJy8GBOBsCpBQ5+aUsQxBQ2z"
    "Tfv3tVgnG7CPcWTAmeOliKvR8pIMWo+XSOvPkeW5P4GU/p+BlL4x9EaDDaScz+fqxJ6og18e+wkUDeN95CeAf5DX6SyttWEXTr0QEPlLcPlxr72OBXX5GVY/"
    "3D4Rl4F9P0n8+35MswrT6xtZHh8xzyvPS1nsPczZ2Bs61ByiLQaeQUWMnEDNsVEnEA1hNPxMNW1gRQDWTfOR0BqI1x2cVRXGsbJQXCE6X+CQqAMaqsA7v2AR"
    "/gnhTw9QwMHOZF6zl71PetniohucZxdAKEX8J6SfnqJ6+CwRkYBSXmDyJzxwG9Id4oQWFWwIavGk3YAFkX85DoZGXMR8GdgKb8I0TtJuyPg8dDOdSJJyGK/g"
    "TtDoBuUmO5V5g4ZVo+pXgB0B7ODaC73gfdQLoIvZeQCdMlAKB+gK0KppjNoVEk2assxVAZgyaBS6BYgyaBQ61guvnPQ8uvBajLZR+ISdNBPZxcRRJj6rEeAh"
    "gcN5BL+AgdAOx0QFicgbBx/3tvthxMn2IcfmtO665wZ0/caoQbSp3i2LpeLqF9RxqtpLjAG/djQW5UcjKhsNNQ0hTEPiRXIaQsC0EKYhOg+xcwH80HgA+1E+"
    "ICHsFgDkmdNz4+AeADs8AyrffXjVhJLMtbvcai/hyC0swNcyfsEKMvtM2rawWZjFdG2QXRsGIXQD7mE4cIcdQNljaD9sCGzkDBoch10DEqAexmqH47zIuCFW"
    "jwyq9ePBaSSWG9D6s6k+RhIGPAy7Z3ecjO8DJKKWr6dFgVKDZbwv6X1VB+1gYy8+Dxrh8KKnTnjsxxj3OpHi6c3uPLJBB9Rl3BZ5ijfgoCNvcO4bcCOAG8HO"
    "ae1kgyTw7Z0MN7JCLG5kcBpigd7IFZtQbTR37mEcoGk49AG1I7CGnvUtklOOQ2QuzhR2wMwIUuMT2OoQacW+UgthB9aji1siHirwg+Ouuhi4vGWvmhgd6ELx"
    "7MilwfHoYzR96LRkcXEueTyjUQezCWLu4+OgQXsg9hd7D2d0f2GhvI+4X/5sz0Les4d5L9UZYVB5e2GEeXs9Hnb54ZAybLXZXC9iudZ7CRNNlw2fo38utXET"
    "YhrI3xvodYLMYjhcWBiopeJ5MxGjVwvQJECHGNkwE9N5ZjyoM2SAHcA8zhUPvCu3z/0gRWylZCbO/ruFhUhSZncMCIuVDo2988pYlmIPggl+1fLMbQjgCjU8"
    "PsZm1AeSO3wUrQcWVjDAz+3wHFt6cAoKfhyWs4ndAw+Jqkzvxg7SdUgtkMqyiv+MBGBq4EaAuDGQuBHCcg4I3cNfYP9Vx+FYxPJFQIiusCltqLd4Q9j30waQ"
    "2hOUGsOhGVG3UtdaQL4LWbMwmgWqL74QyFA/svspLHCMu4eOHAXJ9phkZ3Sw0mDCJg2c/I0cTtU7vSw4BlMLUkgcBkCdpo346go22V6Es2xMSqHAdQcGAEhW"
    "FN8fmeUzLQ5o9Cf+nSxiE7iRDAUExVivWMzrCEhwSAL2a4KnHSccedVIza60YbtjN9AEGBAgfv3XIzZ7PeKLqZYW8UmVzBue/jkNLxbzOv25hjeRxKJGw9TV"
    "JVpzbqVOW6nebQlTYPlYQG5u580hTgrrRiAO7laIOD0YgMUmCY3goyU/2vKjgx8l7U4xtoxgwNKWYMUDnbuyBP2pBGsZYK2Vari2AddpV8N1DLil1ZJxJFHX"
    "M8NIMH9sFH96oFiHt7fVNqddbRkjvTpH+e0kctPbMBuMHEk1AqwP+ds5Eq19hcgVub1LIB6+9wimk4PplMAs5WCWbBixS+aAWgII2NE5UMNzSZGZp0VhgDqM"
    "T8GdK8bUJfksbDvde/67xSs7Qa3Z7h0pAMZpYIoPgD7Kn96xnuIYCZJcsrHzBSXJkTtz8BCXJBPjTGQN43riG/PNxTdWQQo/sCmmAbU0PbySp5YlhjYxC7Ih"
    "bSgPK6PBREz+t7cZaym0+kiwrWajc8dfrFsSW5gJfYpEn3JrJpQJRTJr5oQlzQ6NZousc2R1Q+PM7k4F66yiuhMm2KSzaJZaLPW1watdlQhJNFlCFGHaI2kU"
    "4B9K4D1zNFIPr/0wNckJcAMsEDi1xL8lCj11UoZxTBFJcNigRqyZXR4iqS4nVBQToYlEEGJL2ADHO3HHQKoNvLquS51x9Z7cS8bP0UkVJXBOV142jEX9NZpk"
    "HK94HDSCJIGZrJ9+Ot7ebtDA8VXNhdpB0q3NUrRzwQEqCLoazE1t5CfDWz8JasM4SGtRDIM/m07jBLDtjjQ+YMbKO9CoA1d1PriAYeZDpgfah4EGftTgz0Mp"
    "XyeeW0unEhQo4TkxCq9HU3E+QJikZlxQepQEgxDbsYOUW+Z0llc6HdaBP6sukNIi8Zcm7m9PZmqXZBKjK6rvBV59EgzD2WRaFzyPCnvEz72oXe/+SLvsTOuq"
    "3m59HN9Cc+TxV59FwthB/ZWHx198VeNLh08+TPsmLH+YPai4JietEujxsbrAzXgyheX7onIrYPldp6fPTt3JdeO7KyeA9zL2Esd3e/ErLtiXWA94GllIr9C9"
    "DghYN5AXiHu+AuosZkSDBP6wQULUWDMlxGbBIR5f+0mYjSbhgO7fPkjqvii67TSRU8nHL6+sQPysEN95t+aScDlXTPMdUFzTCsnwpBC/0ll6h9eKJfGrLrsu"
    "i4d6b7zRL01276WPj0Fj5KdctCy8WvfpQVtdyi0e5O7bTfGed9+/24jCNM6SeHpvnvUmFRTKjSnELZGGUtSz/fVU1xOOoU19XxYXDoA/1Ecp3yCfy9AL832M"
    "GvsbX/unkPHsGD15f+1vHOyeHJ4eHx792odYSdEp8XAtnIueqRXYTZjGP5+VY0F3jG9MxIVT2h1giEuBVNzIgCD9oxlG4A3bROhPDzHCOjlVKWfy8mlCcdw8"
    "Udq9wtCOuC9SMNfiHkpVfcNoHiWgir/n8bmG3iws3GO5J/5kOoZwul7AnNZqx8XXHFqUGZoCejw4aUVHXDTIhadARaco75MSPbxyOwgYv3++z9jYMzRZmHEX"
    "g5db+igYkCiSIHGjQHGfuFSKmH2FE8K+zYJGNJsc0bMqL+QhYo1TriBuytBHcPjklTE4UZytSwF8tymkiaRjgIc/FxzA18ATDWEkyU345QvKGQ0p1eLS65QE"
    "ycYF8W6EN0xBD+81swPxeocuvxKXObIQlKDxNrwPXJIY4rjt4AR22vIuxc3dG4Q9kuD1FhfhVF30llyfX5Th9YHb8KfT8T2vagkYgRjv2rjco5HFvMgBHMZs"
    "cB4tdi48zEv7eTaXQz5g+VszKQM0xz0tG3c2mJt3id6Y01s6V1NF5DJSNFKZ9q18qoXFOP5yxmBfe3yknejxMZHLPPFgPDyc8BTZ31BBM1/QffQuX7yX8Iy9"
    "LaWbQX6LJkDhdCkHbDFAVgEERCpR7yYhSgQkb/MMNjn5CpD3H04EHmX2Ho4DHikqlMcATD1trxrpZtBpkjbKUUgXFl5N3XRdtL0LTetZgsx0vdmFFeItvc6E"
    "vHWimkRN52X35NwngMcjZwa9gEXs2hdxKOUFpMvcBMWsxN7kCoPspdM7XLfRoNvM4cWil5l3qfe083CMp3t94H33/aneL4RsWaIlyupgNUuFmKQbCD0YHYfS"
    "AF2BsbPhPRy30dKLgDGYxDfB9g298gLCIYKdsS4Y3zrwEz3JtPDTC/hydSaiFlAgLj3xqixpiHyOa2pnGdcmtMFAtkSrNehNKpEtJ2EidhBmm3iSBxLY42kL"
    "LLscAsA8LlBxG+I0ZaoEd26UGk7864DKXFhIG/yZyy9NvcoI8zhtdUqj4rhiN49CR2Z421a3HnHjKoknUm/Rj65RL1/0BiYpwc0aNcpoJ8rw0yrcx4Hyh8Pq"
    "IWeRE5f1SXacbl9lKJmXiSksJDLx7M464R7mJoJFNskTnEdSXlDD7x5nZqQ0iIuB6l+2P3zc63OlXtHmeldQMNuSp3JKwdzHxzzY/uFv/ReCAtg/dk/LoQ35"
    "0nPkVklbnyPQypv987lED56lB43O8N4OgPsAEicF5lRmTDvZoHrcqzI8OQM/kcmai8p8L+rI9Cb5yZ7wHH+gVSJjTtBYqDlSKw9XAVBfodzcYEeyZFfGAkWa"
    "DqK4GMkCygxNGI5r3NQWF0lKVqXrVC0a2B3KeJt89MgfXz2f1ueW7jWIfO3TR772hozfpzo14Lc9fbpH7M/CyCi8XI0neuLirLQnVosKo2P2ilgcW+9WStoh"
    "0DMmgwsznuera/XFbLFuiIIsHhsZ6sDcS/sloj18qFl+gPu6hbFxAHNcifnV7MJCIA5kR8QUpc3G1aJrQOtYoFJcIMKfPth9Jf1MzmMSlkoxQUqnKlBCYzre"
    "RQVjl0n9JlLPDBsVEmhI5UpxT9ytLixIwXDxaouu9iZxci/vkcMgffNGD3ps0jGkPZXxkWKJddfekI9chUyGbv3FaIdaXylsEDr2UN1MKMhZvEdo6iwBG7Lo"
    "dfSFLLAfi80LrBg+Wnhvjx/tCyCrprN0RHdteCedAobM7ev2RNWbPFPv286bVqHuzIOaAQiqheqhTlllBowfrxY2n7lEN0BI5xp3sl9Wlpc7y+vDqDuKXCdi"
    "LcQUUb/n53AgIxywUYBU5VHaWUbfGRQqv9PnuiuPj6iKUU3xIEUktFyAKSmZ/8VFXHmMi1tLVnxka1kUbisCKaklvRG6K3PVZQEQzGiabKMyf6irCeVliOAQ"
    "YAaULlUGU0RaeyWVzWmv+iI1H/QdQtnupQYfqV6zi4TmcseIcGbEzL2PtG4org8uDoqNXVCWaW5fty+4mUDe7iW3E5C6S7olZjKS1qQwgPtdTmHg+VsNAZgK"
    "gXvwOv4DdxtjNlB3GwPzbmPEZvxuY+RlbMZvJmSNJbcbo2dvN6rKsO83Ri+836DBlNvmX33NMTqf4TXHGEY9fB3DCOpxHzN61aKR6NLmJ7jeaZMN/PGYeN4k"
    "BAZpTPyveF/ZZHjCw+9cimP5eu8+6AWPOaUcr6ne0ARSzzXlsjV8NnJM17ewbfB7XJMFalBjFheBHaPmwHINGqpBFOJNok9qlNcs2WX4ChFsjygLShUMUG2p"
    "a5S66IWvnextx7rhbnVF8SK1nbsjt1LftKzUtpmamSnNrmy/mSSp2KdQ6yru1s6i71F8G9UQZ8k4WLcurtPV1O6HlsghO29evAngj4Y4sSH2/WzU8C9TJzhv"
    "XbhvVDDDoM51mmNBWVQUAa66ZWrRq7TBkiL0ecaaFxKB8lOm9MDkeZbwPZ5z4DAA4xk+8UPVLK1wud7sjuVeztUng/OUKCPzznnkPoyKDRtQw0a6YSKrN5qX"
    "AarOY44ejilsQjho3hhvzUeNFJaqcxK6pf2HUhYWMCdkWKcLbSxgxH9JTRxLEgDASEiIgxnaQKCrhJONne3+7sHp9sftY5WjCScOr3lfiX1mnm8OHUlQcsel"
    "IrTYUAJz4W8pMJfL0gBPhSJkvn+2LiRQN9gt38Px6SVw8lX0Y2HBX3dmWKepbOfUjeYDee9CAbPz5ALh0ifgGAIBjVxRHu+hKG9YXZ6CY0Mqj9Q1fTZd9HyY"
    "GWgt0Ukpyraqm4xlF97B5GCwpdVl6eZWl6VhxIsDV5CPEy+1cEA6RVtvdVtvplzcLy9tHBdHgTQZYWswG/kB9iy18upswl+vvSCjXq5ih1IbybZBwHBJbynj"
    "lVp3OpLxCp5hkFI89QTlG+RV36SWUWArc6HuYiELV/OaV2xVmZZDhoKQ4ecWEU3qLY7UrkDSWwn/KWbsimtjRXeOUcCIdDqckC5KIIuKtaQcmyGmVHcfL/Kf"
    "INpTgyrIcuMjqF45SllxlAoZ+dtQns9l41KhZpITag7DhjSs4dnPoiFFCjLUYyv2oHw+GuY+VFzBWk/Rw3JZtuMqy28a5OBpi3V5wEJB5Ky55EW36XLSaJoZ"
    "vf9CSxk/b8OkNxj7aVrbCDlFOUxrlxnH5GQ2yGBXz/idaeC14HxvwWmOL61nU7zNptfF/PYLRfPeA+oDdzNhiiqQlqgiYYoiFGQ+wO6QeNSbiogwsiNuE396"
    "7M144BoRFh/ghWhAgq6xKB5fiP6qQrMIDZxvoDlN7LInonOXgfONENXYsxjZGVhPW9BgIfFvbxHFAlBiSPb+HxmSvcoh6WzRaNC2dhzSor3M2Cb/2gjZEf/a"
    "C9kW/wpCtovkHjugvx/CIkGIWrs/SuLXXLZTEr3k6v3/UGljablNkyvSv/eaj4+h0ouqZfq26zXnE1NvN4ST26IBU3yGUawzcRnB4sUb3Z8A56kuoPF+1niF"
    "hVKbZi8U94uhmyx6EcvOwwsjgzo0aqk+8z4JkpuUxji1SueQUM0VwpeWqiqyZVjR+7AXIQ0N2fmDNuOWB7JJ4Y2u8KOo0C4wyBeIpVFZOueZyMlzHYSovql1"
    "N2EMOd2/GxkX/4zAvMgcK3Oc6PEZnEvjcTwARBQIB0RE5miJr27C99B+kUOIPPAHo6AXAQ6IR0uZPCNaMCoIAcdeAtkYwnjm+83P1eVZN2eNO5fy0iecd0he"
    "4/f946OurW3V1riDf/eyTghRJozi19tQPsx9BLVLMYZR0I3d7o8c0CCVvvyrDcfvNv/+YXaiU9IJ+PejpCOMSsC0XkEBHagxXXFiVHxtVHz5RMX47xr+XaqK"
    "E1nxtaz48vmh7Dw/lP/8M4cSvzv8+9bs3VLFsMK/26eGllFxCPRsV5ee7+rvlV2lLUDcBqXW5hi6T1XK6UMbYeHcMWq3mh2qEnZCImnpYcvTZe2EorDQ7MrX"
    "/6audH6+Kz+qulIo60dpV379b+rK0s935UNVVwplfSjtym8/uT+Hz+zPf68uT/Td2oDD8hVgFPhfP1Vg5/kC//FTBS49X+DffnIIZ8+NYZb81BjOnm9i8FMl"
    "dl5QYvRTJS69oMQwKRCJelmhomM5yRHyIU4qsRSIQoTxEmSXU/8qOAkyxTk4wePjMSpq6nYkf0k7VBs61IYjuw3pXz8WqD+MLdmyW+L/taMheDlsyKbdkFgg"
    "nEEq2iilQcc53AT2IxCXZPjwtu1WEW6RUcjg6UI6bhXJYhYyerqQJbeKGDALmb28kIoD2Sxt+HRpa+6z56tZ2vTp0oBnfPaMM4ubFCY5rJrkqzxouxL0Og/a"
    "qQS9yYMuVYLeF9o6q4S9KzS2GrZfaG017G2hudWwl4WVLFlIWMdnnDDoVQx8UtTKDUkrt2xHPc8u+KaKajDGZdZfUT/fxagFW4UWnCSW0bVw6GVM1YHcv9rZ"
    "UAYib9VJIm5fuvOLyExcPy632itdwfl+D8WV4vLKypKM/GxELsvIL0akyv5PHflOZf/diFTZvxqRKvuvIhLtvXRVclMm/2ZU+U6nt2T63430VZ3elun/ZaSv"
    "6fSOTP+Hrl61828yz0p7TfUoS3SkggyMSNWjKNHdlE1aaa2pz7b+7DRXVJtWVZtDo4A1nWtN53onQRMNutrU6canUa3qdKqbvWqU2lGfrZbOpZrlJ/O5w1+T"
    "m/eyfwBBYZuFhUS77R/C2DgpwdhxUoKxg6QEY0dJCcbOkhKMHSYlGDtNnsTYSfI0xl4lT2PsdfI0xt4kRYy9T0ow9q4MY/tlGHv7r2Ls5R9Cw/1ShNpOSN/V"
    "RqY0+F2hz8Sfeg/z+WliSJL5zdAmoVmZdpRB6lvvMk1RLKpGKdmolH0iMagxuCC9lWAucQcBqs+xbbNhJahdQppC99RhkXhNsnwnRKWJeMJkvMsi0XJq3nwK"
    "xQG66RTC6I3Ee+t8u110nW8X7rrz7fzxW8Ndf3utJdx76iCG6rkeYIAsOYyv0KsL9KwcF2nqBpqLZIluKvZgI2mM/VSodTV7Pd1qSAnuggGy/b5ngpHEPPZS"
    "vKmXqhf1C3xLnJ63L9gAfjokSkfV0vjRaxq2TwaPj/XzOjdn5S+2PTK1Bh2LDJB1nDU4R2OG3aW7sFMZEloz8wd+TEfU+diQOKMplYzmHfAydhmVjSaqIvsl"
    "zqYczHJs7Wl9PHxNyBWS9KNCoM2X37VWc/Lr9xFJrx+svBvkB09cgJPmW++Y2mTci+8hdyPMiNEs8VsiU4ZyVDKf3H4Xv7/DN1Xy5CMTzkFyEs+SQUAGNKA2"
    "1IwPxxIaJjWcb74A8S2rj2J8tIifs1+Jidok+2H5og+n3LFS5bKipyf6He7CgnXQmOXOprCghyXNtBdk8MSCDGBBomUiofnzqoWX6+bVl9DzkHX74qkkSrWw"
    "CYAvX8JsVBwvU2c5Z41PXc7gNUdiIgoqFocXvQTaQlY58OacVre+GormtOi2yKqYRIndpOTY/So26/P6HtewZ3WnRq2vufULvuv/qmDS448fyiD+riAAYLsM"
    "4r9MiH0NwWrv0G+mBPtHJVhrxYT7mwm3ZcK1ly3A3xTgR38y8U1IejDg1D5u7O9v9Hc2Nk8Pj9ExpMwJx1tHD058vTe7yXVMatVJWqTqJYE0aFI7i9SzgZr0"
    "ytqtw3bDyoffWM8HJetZG1TIbTQdlyVmOir3QQ9gzaNO4sQhJWJYM/U6bajCyENdbAhm08Vlf+16XCzNrdUXo8X6t6i+mCyWHsqNdDoOMwdB8kprarkpk6bZ"
    "Ymuxjo8tyLaMemj7zziMeAlz0yCG2Kwsad2HHOO/a+xydfLrSi85yEnqkhzl2kNN1AW9OW9eLNZPYzEXi3iRsljv1eZ1XcmPf60SXvRpLCorq2EnMW5I5cO6"
    "QKzXVjfyJK7YGqEQfxyEESr51m1NUkjB/XSC9t42IWsc1W1TSwCwsbl9shOOJ+HASlvGtM1ZmsWTeoUu6cvQ3TAtDBiPEmbVjbkevI4xeMJDfW6GcHpUSY4E"
    "sQfwMNHPrup1JPaMm/LEVl9tJAG5enDeHpzt97d2j/t7ux8/nZ68vebPurfChHtPc23Ik6PD0xzoiXL2l4M93t487W8cb2/kMhxb7tlymY4Odw/yNRxpz3k5"
    "6E/b+7s54E/BJCyFVZ3sn3za2Dr8UuireH5d2eN8vpOcx8HqnuRzHuWd9xmr+eOTM7W5t3t0tHvwsX+0t3GwrQrctF6ZGw05O9g9PHhZrjcqzny7LfUez4Du"
    "/t/ntW/ZxWvlvXbxvXP+7fbbsPH2YtH95e31RB+63w1cNPpwlrDPpiT5c25XGYUFs1ZuNkri2xonVUl9e9Pnb9GSABbiTVDT3nRhE12s/6JMm2ArVAe+QAf+"
    "Y4rKXH5tFiXxeNwfx/H0/Ft6sbgOm3TtmwNkBZwwtW9ezfk2XHS/9SDw3vj+tvhtsfbNrX17gH6n304go+use9/m7rc5jOk/y2vAV4RJ9i1dhEq+pa+/OfAH"
    "KoKIEL48+MdrSF/3MAX+vS+Jw7rxB+OM6udQTFmdQTQ0eaLfy6fjnwn7NdHY8iVhX83Z+ZpoqrH0sLf3vZJ21DiJXUvvo8y/q4VpbRig4RX0w9yoHdGbudoM"
    "/lWOm8qq7fj8qptl3AMnefVXONj12Yte0QPAazy0svcqFLl0ECeLpG8q1sy3cznmFzCr9XPxEPKibq2r48O9ve2t/t7h4VF/92Br+yvA6idGiXGtm2hjBnVl"
    "dQYL1TaQoHgiy3rfomoIQJqeIlaUzSwDZj1Y9IBk+A9u0qn2Cbae/hHswrsnsAnUu6Y5q+pc+9tbu7h7GfnICJWdCVVsrWx7h1+MPPhCTd9DJ3keSZgZEHai"
    "uIUB4bsgFc8NItu5wgwibE1MwXoOLbmIGGW+3e5vHPVPfz3a7n/YONndlCNX49q+ysvpKbBe68U8R5s79W77paD9k8Od03q3U4THgSrk+HyyT0OEdiKmZR3Y"
    "PvisgDfPgLXokYobd/khTVDKML5EE2RSwsVRabe0DIPAGXPAQRlg/+xzXT3ypjZOnmjj/uHWNhz1O3tw2uPUv7ClabEBsqTjDV6S3YarJ9rwYW/7YAvPt4PD"
    "g6qxAsb+EvBLVN/sluTeP9s73T3a+7Wee69UArn7NUeGlgBtbG3l+kD2upAf2/FRB/aX5roV7rbYjRcZFvLrJQ/qz6Fr8tnaln6Wjg/+eZfx3uXsM4aFp3H8"
    "RMscuHSmsHspzWpMKOpbY+zV2KdlRt406lLZua7QG7ihrfX6f+j3cx/3+lUP5mvdGjcWCTtJHX0jqVxou4rsbWGV5Ta4SPE+EU/+tgsZ823AJ/K4TXAzAXbN"
    "5rgl/i0vPy0t30gvqwGfiYk3+Ni7hLtbK3SP71PiNmzvcEtPUXmvbPh4WFY1H3xlJGAcD61OXjS4zQznEI51k4UE5LsvlSCfm4ZBIxSsZKZqcCTEPiFKkUna"
    "Upc7PnEl8DfUFiGsKmOX3aF5KZK+CSLB4cb4++yWXaIJ4HE6/szf6kJnxatdLNhMIVYb+tbDlXHs3/Jhkj581p2+d35f0W/WF+w2WmJ0+nRgQeytd37DKvPc"
    "mnluRR63i/X8hvQkUyOAu/r2cf9gY3+bGs0n5wDFx/cskq85YA2tqyxnJ/hE6+R042ATtghCGROQHn5UQfc3D/cOj0UewWKmtp03nfXz9vHptjKTd0K5VKIl"
    "+6kvXusUfEf2ATbRE+rQxL/7EKPZqqgBRNpOjPqn6FjPbuHOoeyIBbR9N21rQAACBD5qC8iJP7ULgZ1TJPElYqfyrbUCoL444Qnk/K+QlzgwnduPCxAbh0bh"
    "wgFssQn7uycnu5+3NajYXm2wD2f7RxokkpuqDXRweLy/sVcChkPHrYXYO7XOfvjh73DQnhxtbD5TTOmOr8sBnPq4fVBRkPkOyG755t72xvHm4cZpCbD5+qci"
    "V//48OzjJ0Cvk5L8B+WDpTPnm2n6qxS9fm5ZYIFbuyfAB29u70P/dWHSp3Sh/pOj7c2zvY1jDZpUdrSke8qtcgF4f/t0Y88Glq7Sc/i5d/RpQwPl3jTZsKfH"
    "GwcnhKqHBzqLsB3JUSI3HAITLEjah3Jw5u4jrI9ja8ugeIst2LObHCAQmWLTuBGzdRiN7w2Yzyd9sYkdHgBNxoENwsTYW/Y2uIhle0tO5vcwigob78k/dg8O"
    "9LYLu5WFJhoYd0C5dcppNN585qbx8Pjo0+nG8cft0xMTWL37pfeFFU1XBXDclgUM4xmc6CfhMBhq2K3DM6At+ye7up/4VikHtbO3e2TBKHZkmztCyo2J5Eye"
    "ha4vDgVI+APfywbRjG7jcuXt/ra9cXq6fXC2ccq5R9qZS4m73D59+HFr++j004eznSdz8XelT1CE5WUi+cRPQaHWVJv42RK99heWOXulaZ/D4LYifZrE/+RS"
    "sgqAm6q8nVpkmAQ1E7kEGEiIxJeeJ83kyzge18L0MMlGSFBNR+EAk/8jvIJe56kLVv+WKUswvEX201TKCkMZXlWVIVa9XRK10XqrahWUA5wa3cgliffv+YR2"
    "bXaT65XcpAoNWaqJgy7fl5owbO3kt6WaW1bKQHdjXJq5LFvHzFYcRmtnKMls7CnN3jMArecA2s8BdDhAvn1y48G08ux8IxMtfAqk9TxI+3kQ2U58ZfEEMO/U"
    "Uu9ZkOXnQVaeB3knWyWnuWS+1flSgl94InG1kKrUL/Sq1cIl4DwquRTkYX6OHyHiAr2N6j2SFsTp9skpwSuARccI/K8WCiLqjWbdfZqJ+CuZBNjMBsXU081/"
    "hYuYPsNeFBOu/n++4/9FvuP/HlaBrCXkCb7t7YO/mpN4fHxGyvFvwVwAQTUMOS9pF/fxeGNr1+IWX8iH/FvT8dMkmMzGWTgdh8GQhlsDHR1vC2n47vaWNeJS"
    "GDy+34wTtPnML9uNrJ9+Pdnd3Njb+xXm65iUAPgN/b8hLwDn3BPi2kiLa6MXiWvlHid8U0AhFVxHOWPwh2h/7lnF0DMx2oFsLKDEkeR5i7CjkILCqnYfL0b7"
    "8p6tPMdO4tSNMG0jhv9sKZnJRuQKxsD63dNP28eyJVCv1ARL7VrFuf8BqoFfGN5grDSTeNq2yOhapADPgJ8leTA6l02c5piNfxay8eh8NuOcprw6XCxAp+VK"
    "URQDFiEDhfwywcr8g3JwbarDWTadZZQP4WMKKmicB1wGR/7guz0VuAz6cAIDvfqRKD8TjrC1mvJEJzjfE2eEv5/gF9VmR95H8TXDtBn+QtqMx3wUXyPvd8o3"
    "w98ZNk/7I33FEa1EwL+w4Fx6+nKg02zWghQawvreueqRJqfDyCBRb7hPlRoMjBGbyfdR8gu6qvvHNfv6SGAXysHCrbsJfEyeEXkco6FJvJPnBP10QAgtmeOK"
    "TLL863F/R0JbWY1WCxja2GBTscJPdq6QhheE1antraMk/qcM4XcZDOx32IhM7X5VBdmAIqa8TS8q8yOc0QYUBquqzoHKqPLKy8otosWtNAW4710u3i7O2Il3"
    "lDg+4/6/Lhf7i4DfpzquzfZJ6dRHi7z+YCTUyu/YCbplycWdygcAfCkgs9hUdumQi1v3yYkjj1PK8HesyUrB3a5wcmXbLKwspC4lNegrC+1sfpe3dneoGT8M"
    "LmfXjcEoGHznTSZ1rFQbYfaNNwBSWfZOqt6ygKfburQnKjkqSz5VirtcQfwVuRwi49UkVPVLXx3ckTLwkus+hNo1UeAd4KScMOFgHk2q86hTVpdnT51Pb7mp"
    "UKXnJHSaCKJbqzNqBddNA2Yc614WkRVNWwYwrllsj1adcV1nMjDBNTjXn1c3LSvJcVG9uktlBKjwjCj1+Oigwya3l8jXP8PQv47iNAsHqfeQzCIiFbuhtCwL"
    "BXUz20TaA9BhZHoWFtBdtz/Pm0ej9Eim387n7hynbpttSPUbX1hcFGhPC8GKOXWlfS5lD9F0wWOaDEb03obObNNLlk2czTs4oLbnqgRtbPOpMjagjA3vSb/l"
    "lc9b2i3XfOPUS96H6DY9p7DOn7fw1kA+OFHx/UpPWHBSDTVeuITGqwbRrw3Rr2GQZkl8b/aoYIFeLl16IKOG2FzQWJL0Rs5HQhSPLfMs6ZF8tbYFHTNePv4j"
    "uPdEIvr8PA0nMMwtu+Q7HjSRyDsRhtYs1PFOKdawNqLUxtBsvFYdO79gsWGaG8iKsMqN3wySSryhsSHEF9ywsSm3em8A0rYz8UKt9yZ2kyvvAY1HUl2SUOnW"
    "h/wsppRQydlloog5/vhhg8Mor2ACgotReNoHH5gsnXSJQZ6y56PB10ynjXkETz0axdG1TptikKecxnRNIRIyCPH4E6GyY2YS6j6iRB6qSt8n4lqncmK7zpCM"
    "Le8Fpmz56Sgw6hxSuM5IMzo16qJwnZ0IdlbGc/YW4qcw7cYopxSuz9k1kG9q1uDIl+gCn+V3xJBg0891VtciCysgSbMJ9rNucCQU4t2v20wHRHB2Qn2gLpwK"
    "FMC4EhcEJQ9gfBrAJFTETJrTsEMGrBAb1lldyeuwzyVSP4guleJBvCmhM4OmDM6MN/PmrutxJrQgDUKJXYYpDYOgIZzBjgvpF9bFFfvq8mTnEicV5HIkFZSS"
    "KxVxdoPflpwIwld4Gte5qJoHUVKNX1rwgx2wr0Jxkkp3IgQVV9KEGlythtdgoWHdsvabyqC4S4ZgUWiDE2o82OBB48kEj9CvNHhYP5PgYftRhl2meKSQL9mK"
    "zj+GwB7nRFBmFKrGIpqZgoR6hWhJTjdeRiAaaZkazYeQnPFm2E8adJz5pAHLMPhdwkwhrqAmBgGC5GSj9QutTX+jNFDVW1JU6FNe3NbxDZh66CbSDOc6fHLX"
    "nedoO6JgRJlyIe8k8QQN2daGcfSfGenOc8lUjZtyTmt+KpmYtFE7g3ToW5jUpA81vLWeAsrdmyr1ZAqap6uGu93A+8oCw7eJonxS07q8z67ZPdAn4oDue/d4"
    "sQP8M/q1Kzth1u9RtBImMVlU5d4LLoVJ50RIXR4fb122712dJ/SO/gIYrTso7gQXETfcvF7x9O47Ejpx1LjEFYacwkzaM20120s9xa0Mgf4no/xAIAD55mRv"
    "2k337RKa2qfoCXB+EVO2RJUvWfWEb/0FvolrJ6I5NdiD8I2fyLtYr1H7GrVToHlqH4/OpE8KhAoX6zgpTbcbzudAqwn/pKds23Dbk1jK+Eik5HzcOgaEyyb5"
    "LM8hX8Oe8Dozclf5W56Y/pZh/jRvOAzP9y96pzBBllr/NkTYZCD31nLKHQGZgEm5/v9GmcfCgmvjmEmt5e4+04Rtl2OXzeKc5nmabSaeJgC4+GIlgirObSdl"
    "MixmSH+gECNkuCGeME1h8LLu8sbKWY4GqQBbWBA4cpezKF5O+XSnzKZ8uiL7xvqNsyH3Bbeb5QSMDCif7qtX5JORGVRQ98ahOEA6ooEEDH4ymyoSkBgFexAt"
    "fAC+ZJpC6l4uLFxKp4/MJpYg96XMxuklzIzyQuX10/PGj49mcOAySUZRs2SA5YkrapqMBAYqljnoixkkFsUaYVZCflFpRrzLBDVGmcU3U3QZxaoQK6PRui1C"
    "NgWEByorpdq6zRJIy9I71mZGVBh0t8DMlDKj7RawimY5GpDAcnEsyVdqRjCTLOSYZUQwg6KkRCPMJM3Ip1EEmEFXUoIRFsbmqRz8KliTx5RcHBPkKCxz8cVs"
    "wrObmPflSe5ClZk0bDexLlGZSdDK3cYEWFi4U04SzG/TEdyAOx1Yon2jEqABTNQE/ZczRSRDZ/myfnxM9MsSozMYMIYbg3KQ8dvAfgyac4phcxYxbOKjFT7O"
    "ZSx7wFLAKYzKTZTLLJq/+z/TO6vRuRY+PhYXh8uAtoKp6DPOmwCKILEl+JNuf2GhD6fBDg8yg1khXDXCLMe6QEG5GFbOynRHTLIymEd8Liyc/NJkkrHpnrA8"
    "Y9OdMZOx6VpOiFJmcjkyTQSZyd10fRiRhNPxvvRQxGyGB2BIbGCkau4HElMIGGk25wPpiYgwYDS3BOkjCBhpOS7JbqGUW0xLm6qzUIPLgPN8lWh/GajN+3QN"
    "99qshAvqFtx0M8UGIaEjv1mejesaLxwb/OnRcGHhWr2aYRaTZ0ETsWWaREjU3XEwXM/Mi2R0kVbODkKJFSmsyBx3kxI1B6bYyW6i9dyYwVjSy09cEcOAKQ5T"
    "HLkUaV2YaiM/1kUq7u5mmJXc0UADSmJZ2Qu/bqJ1FVIsW4caxnM7Vnxe91TOKwnEyl7HPVmnBmOVShRPFZDmYFm1Jkc3Rm0M9Hlet5/4AZv01Os9I59wH2w8"
    "3SvLm1fqyFVcfH4HhQzITohgmzaFYBzP/7J4R7i/zINrp4+WcSTy9a5eXAqDRzoGPSPLOJtrwet1kWByN65xxxhIbsbN+fjUCdKPqlGaSCK/uvIeLihjerSl"
    "KW765to0fSOLO7+GgqAkWVGJDoNIMN7J6psR2ajSocbMdIHLnSNLOX+pm9Grc+6dk/t1jyxvo8S9Qkrk/YZag7BHaUPKqXAvilbMZJR2vcH8AT0NlXd1hmfP"
    "iEuQQsuBKt6RKgeqge3/zUfzQIQO6u4FrWM8hF7AFhdDfQMjTL/N85dcIXBG3DkMPciPGF5DMV8450Uvxuhbki6RCu3NpN/QN28yXZPROL6LHV6hl1UfnZb4"
    "57Irb1oXWEs8dfgtMt1d0TJQDix9w+LTPxLHlWbrDAdmZR52lTsVEgEYpt1yjlUe5iwj0/w47IAV2E/0S2aVlEmHzoFbdHsprBfIesjzealXriznlUtbps+b"
    "dbkGGnF6mMCCoaWog+tm2hszpSsfDetsRnjdSn1jpeGZyQebconvdRXbCIdvAiPQRXOJfBlDiPIY4XUr9Y2VBjl/EPwPgPoBaT8ghoDCoWFWP/0fHg3ZRvgH"
    "GF3axiDNvZMWDn+Ur+Rz9Md4jo6uH6Dfb1pz0705Xlr6LOY+d6UHzehCeZLkqOQXUXa07oyowBBaI1j/bsgkq9ZNmBztri9XUHcm5+7xMWXGEMXM7HZoDgL7"
    "0R1z0O5gzrBt3giOkxHe81LVI6ER7oXwKav3EgjIBng+BORFr9WEkTFPXgxBo2LPbsao8cMbywwon4kWF9lISp3jqf/7DFUSiDea+glKjBOgpELLvS35YRL7"
    "DZoRVJ+wycD21s1ZQUxxYuTcDjzfiu45Qm3GqHIdBs/lO+UAt4coHYVX2Z9UrCiNSoazNTTbax6ikelqSjrYVsRCIM4H4Vs5gil0uYUK/OTO1qSSvwypWRVh"
    "NbEiLOdWgtMUcd95DH2j5pydyxn4pbWwIJynZo+Pf0tcNSGYkvCU4PExSy0PllGqLXs8t/VrOz/SSxenQ5BSi9x1h3tHpCXMyLIm5ECfo243/EVcNUSuEuTb"
    "0CJR2JZ0u4nKcB7Cki/d94OqfT9MHU1HPMyrzjHTWVLGzdQqN2w82LPt4nHzwmTMpL6l+U1ihOpdOPMUE0qWYfcyRoId4TJxbhgvqSsWk/JJJTCZ7clymNRq"
    "QC/XQBhtxuTbOgDO8jLxeWgYDPx7dHBtVKlZ4NI6q6ooKwpZ83QKvGrwRN/T7/ebZpmIydFws2pELJEAlWi1KN/ckT+++kK+CY2IT8JJIcUoakxMJqzlCLCE"
    "EDg1jZWm1sFIxF4qmP31VrfpvkFLNnaU4SAktVWXsO4whV3DQNS/HBVLUIwLBD6EPqIHD3BZjxV17A/DWdptaXkCCiSFt9CnsOlPKF5EbpIu/kHgJwqOR+1g"
    "TNCpmFYH7ZA+CI1lqBA2pVH3wZAK7dG+AyQDdwOvgyja0SEpiNIxKHrSIS58skVNqUg40gIlGXWi5EYYA8zJ5BIF3d3zJoP/LpCYuAy6QNAYLc0FefbSSJSH"
    "l8fjMwdMws7JX12QJc3KR8isciTM773TzRa/PDZi2jyGhhWBDblaLihqs2JkdTjM8DvPGW9d48biGjRQ/Hzgy1tqBXM/mLAB+PSRot9i/iEXG5xEQHCp5Zby"
    "Q8wHmiWGf+OCE3CrSjjh6Sij+eI6sQPIhRTPDP4N4d8U/k3g3xX8u4YSUn7YpqmbYzTTSkYTWXt2A2ce7Xnsnvw6k3gku2d3KDUQGzLre4EQ7qF1IP5FjzfM"
    "gLw/pInhDGyYbnDUozXr+oveTSN5fc9i/LiGjzF+XL6+V64RMQ8BH+E4uM+MkT8ckoPh4eeAXM1ic6A3wdVVOMBqqYf3rlV6fuPSO2UkmUsyK8avRwbx9N4R"
    "I+Q2hHjxHiv1E8cYLyQn9Haty5QDxCJ+b4/FyzdbuEsBe3AJP+odl96/0D69Cqh0vplBWkIf5vsv3NU8sn5P9zlJUTp9PsCNKykVW2NavzwNFwsmG5NN7q4n"
    "i4tzKwMCwZkO0eaAqyOhaqTlKYsoz7VesHj5MMoJRH1f4mQ85Jbb1czcFObknmQPAnFRKbUhyBWu7DGIUyjQj67H5CrAIGAK6a+dFrLKAsLlMg0gS0icBr//"
    "hjOud9TzGZ9qa9PFyH4ukiZ3Vpzca5pchMTUjM1ys2pRTiUz+5JpUuQUbXaN5ddBgzw/064n0jl1JXZDAuFOoREiUTc450Ns4zDXRk0u/FmL3EAtY3v8d0cM"
    "maKpHHSWQIEGvi/LAeyY6VeQnJi3Vucjjln2CYux/Xws4daoiFtXhFsEislARucmLkfql6KXpPSfnj2aM4MJkNBGVFmehO7+zqfYuCk0bj4kI3GBUkZ/fCQm"
    "P+M3BmgRUV4VkBp6nz/cqyOD2jDpGG8WNuC3v7OH781bzEpt26lt8a6npBJcGS+s6dPG3k5lRZQI9VQ/htEabmd0A1jLYtJFtNZ/o7aPl9nRdY3y1Iy7H1QP"
    "SxqCAkXng74RbF14sRFsX3hjQWDdeAn2etS7aRTIae4t5aZhkNQoTsMoTVZDzAxjbNIaYocYq8lriJliTCmJDYkTkWiS2RB9JaINUhtir/HlTcm9tTcQu64M"
    "z4z5kHFDuXxkxEigoQxPy45nmTipONfNdGNxyuirwko2U/RBIWOv82eKmVBBO1Q0wU68KjmXdNllWDBgNgqMmDX/M1aY/CGzZn7KqqZ9wopzfsUKE459Fsyf"
    "l6RAms4Zkf1oSSonMdM+5In4Z4KcV1Jii8jZjbDUQBHyAdH8ymFK8H7YCzQhP0TmnSSCwyfJW5tUiy56mY6oIL2GNumVVoJxleQ8oWYUP7sEFsiKIukoXqOp"
    "FtNTYeMkGJYTj4IqGWMHniEdh/kWKXh/Cvu9cI6J9f47jsU4NxYVJJdBBQ3+1DGJG+EQtmUgfBy8PqSTM5fbN5Qu7nke2PsTf5Adxxl/4OaXUnrD5ym9oUHp"
    "mSVYzYxzufOJg9wYlpGEkhQJ/9TRC3M1V9I0gtYY/YEF+TwOmRCchgSCEKYTKS6U7eMjwm5iiLHjEtGiEDcydROm1K5ztzMVNzO8lgdSsk3J71w3FFI2EUy4"
    "Wm7ajcQWKjSAzJufBiU4oWuB5PZaBYfxTkj+7FB+Q8D2NbfUeOAAQnRkmeFWEMb4jFNzQ3/u9iKE4Wpy4NS8v+CaJiEQbVySJEadu2WGXOcp3mAkvwiCN9Q3"
    "GHl4kUwtTSFPqrKgT6+09A4jyt1hDMZ+mtYGKafbhmnta8DnP5mRWAW9as2mAb56pvecKAb26oV3kXX5aFVrQnmddrMpPCAKBULvlXg4aqoF6kif3zsJF5xC"
    "8dKIyqlJVqSQVEi+UDUTiJ0SLbqFlXGF7x5U7SoGXy7SBiXLuIp1w6UjNLxFmtOuqF1u0Dg1RGTJeGS2opg9Mpn6LBmhzNajVKNFbF5uuDKt9Vw6aFlBFbti"
    "BLNiXMWYZoWo/CBn+rtytLOSSP5aeD5IDQ96/L2RhX1oLoAj8uiPIHLuGa/AZWgLbNgQL3dgKffl76eB8ZIZFaKYUUHn3xX3/wV8LoyJSC4muCXDlFnB4phl"
    "Zuj/++tjPh+V4HUOGRG1tWf6VDrK46Yy6IAOe9J1orzWkh/8guIkYzEe3mP8M0A9qRH5CpVK2yQmmnkPzW6LtbpN1u624eTgOmSR8yCfRT2cbOwf7W33jzdO"
    "t7vtt6uM5AdmZOvtKmpNCAU8cUPXn8Ky7D6Qrzi6HUC9rDQez+huVcbz6z7GpVkydgmIFOvRVp2uLCc+avvVHr5F37LrcV8iXM0jozyOMkXLai10ltj7Fs3r"
    "+ddeytxW6k+m4yBpb9WM5kIWw1pXu6YbbKSQBKbGWwzR2j/VlO/vv3yLCs3leSaBj61tNpo9HZn+PvOTYNgvTeSOLrzaLMLC0bjAaUwboKON4ThmD1jN0faG"
    "4mTYuLuvubW3Rk/QGySvIU4AmNcTQh1vYNS4TywcPvxa9GrGNIuuKBuznw6Pd387PDjd2OsfbZyc8LRvGQ0cSklJ4ZnPj9n89icg3H+m+Ys0F04tZDg40PzX"
    "YvCrOob/0WhC+82GNO5UsjXoebB7qCEXsZgrKA8hSyaztrKSP3EK5RjAADCYmJ8cA6ytuuvUvtcWFDdRi99z/CNQk35e1/LLn/LksNgKVmQR6J+hO5kbypNk"
    "jp3zjayTfnjPLGtaXg0HlVDqNMaxdcRAYQamCnfFbjBHZ0dDoeXr9iZS3bmRw2WvpSyBkGe6qHeFB6bS4ne0MSVGN8CRU3Cyff6mxeD/xjLriF/424HfC5d1"
    "XFeJPIkqj5wrNkQnLng86G3/zrFcD0uP4c6N2xsqheSGgTvcRS/3CC2vZpkBqjHFhKSDwIQi1FIQ4qqBuHTrJS5lPoJq6YIOHz2REjDX/+M6+pwvRffLdJXP"
    "huyGPmEinu0Alqw6MXlRJyZ/sBMv78BEdkBzh/2CE9v375uPwfv3rcfo/ft2j9sOism9cU4dFG018eN7kOKJazw7ASaqZT+qyvTTrAAwGQv0UPtOt+T2RS0Z"
    "P92SEbTkqXrHxXovHWXJB9ViOZUi9AxHXh+ICqnEb1Hv9MJAPqvWQhpoyci7NTPlSCPDlfpQapO/avWoqDBnD83JpIk0bSwKDX7mopRI56k09QjKFSMpKw1s"
    "4wVQq2iL8lSBhs2qnctqnZUajzUKq92GsD8rHWxZYi3Fbse1Kx9OHHI+C43yRoAACVNtst6Ni7cMY2/Ir0HHMVCu8lXZtnzpJRo+wGdRXFizsEDOyClKP0CT"
    "lhOURGvcmM3CIQsAEj+EPvUA9VaLCvsY7UVyGCPtS11h4lju0gxTEd/GXiIVssaNmzANocFo04l/sbHBdYYG1zmmJ10eOu+L18XT91BsOfjwa90MdEOC7j4L"
    "NzvnkEBdm2Nlj5xI04NWGEUBYb7TEzDW071xGdsclrHNY7TspwDUNytdZhQ5ruA/YBbGJWxfuYgysUWUY5v9w26a7J4PEHrz2IdDTm0c8s1RJOdVKCeS7nNj"
    "7N/DoMCZkKItER5yEadEH8gQdYqjwb+4vSkOoW/nMQ029SC8CU6EfhNhB8K9QkeUsOnMJpuAAsHw8TEkVQCaivSQ9KudyHVR5JhzIKJulCk4AM46Lbnf4XrY"
    "aqjUcwN1tkeoXqiVtbHjRFCgXhURFomrtblCrrWdqlsjfs2k9MjH7we9sb41gpTz8QVsycn5QNVAfgvoKgk2rpEad1XHJRIhbATzw7l6Rrw7TFav9KRM+UkZ"
    "knL8AMWoUiCeVJWd/KGy+Sks/P/CYdMYjMLxEHLldOHGeV24fWeMKmIS5+bEu4unrko+4s+y+IxmBO1SCpFGMExlXMsQJrWkjAQb7FnPigyUvpd1GCgtEnRl"
    "wp/QvVmZAc83drH1img5ACWWUgAVDPuEaEVjB/3bopFeI34/nAIJtBfcBGNIwjtHkt0DiQmL/cMYeYHo2mlCUkO8rRQqNJC8SURTi9F/JgTRMgiBj2/hOMQ0"
    "CJ0MwjSNEx5p2FZ8Ed7OvIHYgy0n0jMO9szJWmcDVkc7PVEsuK0GNwiahdEsmMvJmBUnY2ZNhszRExKymSQ+5VrGxzLXuEvCxkwvT7PUIc2jRO0QzgTvEnPZ"
    "GQDc/TJCUwH3v4xwM6Iw/KaNO9OY0ejtpHGHChZ3HqS8hgBTpfA4TLwXWe/zWUnT5R7A7iHrvZGV4vAFHX9jQgkwBDPr1DB3TLUCDMXzSRjtkGHpLlqQuVbf"
    "SJT7WXd73ptxmSdQmqe4i99ha1jm8oYoO1VkpnJAP4t1reRU52DIGpQWIRSY+F56lPPg5Lhz/YSGyilt+NRo+PTPbPhzjSsyJ7Mcc6LYUkIxPHvQ4s9mPIvQ"
    "LpKtOHtlK/5aOeghJ107IQYB1jBEB0ABxsM/RPgWl+2NzORjX3nr1Rk3wBEIJcrTyQmbCJ7pseoxrDyNKvMyfHp87Lwy8Onx8c6Zwd7JZrlNd35f2IULYzaG"
    "6maucXs35MJTg5HW9kUt1W7jtoxrYCj7wpEQqApLxaTsbApfT4TaNipua13wfT/9bjxqwte8ZKgX+ifVEBEExor+w3GM+EXnXjz4Hgyti8nAyyiJdlzz7Woo"
    "aSci52LUWHvt+SzEP+lrJLYi8TyWIF32ils6CH6fAe8AzUImiSMYCVU0oNjhgNwh2W2Q2W+iYPT5UDA6Cpw3LTEEeMnsV40oi3gemZUPph40Mtxgdjxb/+G0"
    "19prbneH/9Iw2GMb4v0m0OtybIeczYSxDdEStLg+3gHw/GOYEJ+K0lcgvUlrL9KiIMznLLfaru052k5dc22X0XZqx0rt5FKXrdSlXOqSlbqcS121UldyqStW"
    "6rtc6juZOgyufDibCs3itFuhsUCuzn8WT92HhJurlsjGhaABkrtBJYKVIQuiV/xT6IU3I/jj85+Y/wgBxeBZDCScIixcaXIshN+XoGEKp/8gHD+HiLD4fb45"
    "wPAknE9GzULfLIRGn4PiS1HsGXTL5wUeTguv2X1eXMyLG/Nm6eIOpwIQBoUessBwhO7PT+rAntQTXjpO6+Anp/UFkwQz3+PUIsoQpHwJvmf8m8h3cUUq3rVe"
    "8Z9r/nPDf8ST2Dv+08dct56wVEk/+/znhP+cmhv+tjAXruyEd5ZXVlpclLGBJe0pDb/jPOi7tQ4s1zfocvtY2VSo00mEurx7HvpKCUiM7Lz931yt1vk2dN82"
    "grtg4By7560Ll214e794LbebL+dwGkSQYfukjlt6rjCVWFVgm3dhk3f5CJ8Q8l5smSccqohHN34qNLlUkGttuWz3J4C1nPtAIq15pp4BibDKWd4lIJdRO52c"
    "jYvbSwd5RPR1IMMBIzk00GFqwEOIRLugLbb2rr1aldrkqabmZS94H5K6JeXYBXoQb2mixQB6tbLUXCV+R3zCwdBiiXrqnwqO9AOOoOrhD1wr5AZlfB5c0Hrh"
    "3CCuFIwCLtMQKe8QeMsEH4apDd9y5x/OO8vLnQvvwMFfRn+AC4PopebKO4rHD4Z/19iKUGDkPBunV1qozqf5ONTl0xAuEycvh6EdqOOyM+Tc2HeEhvTVVZid"
    "jwAs0OXQeziPLrqd9rt3SzC8LfpaxYGmr7U5sULu4XmrCS3HqGWGgSUeWKF3W8b7CSTpyDrPJXKifSDTJ/4dcG5CRIfSYlFW1tjfPUCXTLI8iNj4ihGumJNP"
    "3kO72cSrbmhXC/62u+9WVuF3qfvuHcS2mvC7AuHVLra+3VyB3zb8dgBuDX6XIYz51rrY7HbzHfx2jGn+SMSTkCxw6xDItSsag0ivITT5h9NpLrWR20bWeBn3"
    "atfHDT+EjS8GHhA2vjHaWgi5Ojru5je0f9MwbP/OrbidBLDIUWp0iI/XD8/9C7zhCmGfQ/KNq6yTjjo/CO75QXCnC8I5VYV8Qu2xT+cp/onxzxiKu/IS2D1T"
    "2DNj2DDHLuyv+DaQdkT1fAM6OH18xLOgT/2c8FPnhtO7+VY7hBzY0gh3ZDgq8qQXElfF9rUYjj79LRBbChqFEQViq1gWIj/Oab6spQpoxBP4+66ZJ5mqX1GQ"
    "s4ZubRf2vHE4rF0KSQo6FQkEZfWibiNqvqTjCPdE16kTa9Wd5Z38s7pnn7v2gcsxaMAlg8S14JLY0UuiZWyFZzg6t4LKCNbRonAcZSTHakMGNFxrxSyj99TA"
    "KOA7FkBLcN2R+xUiLNqSJbMSUOYAmkUFwDYFjEa7JHoVazIiOm3XJWKQCrwEiljV+FkeZ1Rjp73ahl3zhC+JU0WITePx/XUcHV5dcQMWIcJEQGyEVC7Ppgv9"
    "gt1QMi8aDK/TWVtdWtx+AzvxpqLDfBLqqVMRknA4BHkrxHNdYfIg5abuusApcsKtG88ZP5i6P5g4cro7DM9ZEmjx/DlTV9IoibJMPKB7noHLRnBSveL3Pvyw"
    "NgqhvASBT8sAorPSbK6Jvo3OOyutlSZeEjH6MuIBCuORAscivx6XtwyGXzSlWDUVyRSROisxiqXtF83U0Go4HNcZoDGe2swQk3Y/cq5A2PCx34O06V0XXhut"
    "C7zpCoQUt2UtBdALaAhfRXjcwh+eIteYEExmlrUZPIO7Hx0NxXLbrgyfJAP5uZVmeShuQ1GDWmGAl+Gi4UWTTDBYRhFNIl8Rjd8imngjEf0FvXsIIkXEE5ry"
    "+J52uSOQleJ7sSodVhCaQYuN/CYkRgkCR7TRYK2YCh0HV8xKMvIB16ST/HCsAX/LBUktwmWf8wud5cLc5l0+9iwKM6BmxUtAbsQy3oxvggQo0XXcU9baK3yX"
    "wA/OVCpLlmfEqomtqvudeLowCrgRk/wC2UfvSOjEYCxBELf3aV1AxiOzWd3PzJbhFwQ1neYqCWroFy3kGXtR9wsz6PXyPWRzYeGLcgh2dL5p3AtzC3cPKB8U"
    "NjmZND4gHRxhBg+xgN/L0LINpWRW0yIW0wBxH2APUpkCpnNEZA2qtNFasGq0km8UhglGLFFtQKpObimVq8HwgHDRhCFRNY+A+gfxBJZailb5FCdi2aRL7h9w"
    "nRSh+OsU2Hr85HqGOpAp0ME+0huZvjGpOtfRw9h8zrIn6sz+W2rqPFFT50+pKeX4a60E2k+3pEw04DJRAeigfBxl4ygXJ5n4lnhRjEeQlI4Xi9vNFafk6IXy"
    "do3yCgITzfpx+kiHOelhhIlXM+A55WEWsLpmZ1hp2hlwN2EV5LpNZDetCINGbzKRqAXcr5pM/q8uM7isWUrNmSmzfaXDSnzKLKlhy2W2dG2pvba0tvKuvbbs"
    "2ps3SmahhtJ02M2B4iWyl/9RtUgxFo+xiENWpDfztJzokE2IEa1GY1ZKi/Dd4JVUgCkDA7pHgJWnrvJUTDYIFZlDb+988ARuv0SoY6DuS8BJOmfIkf5ySR2s"
    "T2lI51l51YshSe+QLtV99RXLL/Ox2DQ1PduhZJefFDHp3wi3dmMvMTTxUxgiCuMVvdRlvLNAKOqYR52Q7nqKY2w8RcMz84ht4aUV7ppbXn0Wcf3WYf2Vh6dM"
    "fFVDDB0kQRBtUu8UxY4F5dKIkUbxyyZgOzTCqbeHdWO31T3etcx9bq2XFYYQ3WE8oE1bSPK2x/Qu4uDEqY+ybNp9+/b29rZx22nEyfXb1tra2tu7UTYZoxcw"
    "KqRusEMHcpCFoUCvhXIlJ+PT9wuwWJmYul9Cukv3wrfc849/J6GYBEFzBe9bwrwDl9SUDd6n0/09OodEw4FqqkkXLCXJj49VhfAxqS7FSi8vhir6EGZ0R28V"
    "YKRoRdBg/SrrakUAlnqhk7wW44B3ADwoJbOK6DqCoTvydh3EY0mBx160zmO6R0B9c2WzFBYDz+35UhQaF7CHzIpTA2H6cNXRPeMLPC2JVUCOli4Bq1DnHVbE"
    "sAZ78aTm1BdFVxbrd/jNW7JYd1FFElJTivchAl0vxYIbrg/9zK+H5ODrBW2gdqMF7S3IJhsUplBFXLsMr59oBVaaKWtyGok/GI+upplES0BXCojJmBtC5JxZ"
    "3SAK8NDlCjywmAOk45UiwyuycmFHXFsyZhZx55v5oshUjrSU2e/DkoF4UhDiyiTj+LrtqNWERRitPOQ6XVreGUsDg6Hh4CrKmyKMDDuEwpLuM/OB3iwm00wa"
    "J4ni6E1wF6aZNklCmnuRP65xJY7af9YXo8X6f9a521pYAxJTV5aaHcFiLLfaK+IQTr1Op4MEERAgLStu2SWxvxXXgl0EynlXWs7Sauddvpyl1aVOvpz26tIa"
    "L2e1qpyVknLaxXKWV6AcaisML+AGdUV8Yg712aHYx0dD1E6UmzCkzy3OAAIbzls/GXibeST7xZ8J/7lax/uULvxZm5uSccOarbCX0ENFTLTVvX1D9uKQLgNi"
    "pi6eL9fZR5eVmc4PDXtOWrkYMPU2uLwe7wLzLJXjpN1vSYUpIBGBbB9vAloGR0kL6o1+BjY6FhALC5vSdniEZ/8EoJN7yaelb97kRJU/3cez8j4qVpAlqr+4"
    "bYtlSdqDSjVQraM01z9c//YI5CHcXiCFMiKHHW6IhnIdNmUMCSkVU+3GtW/PVsTtGa/bpFQTWb8RSfy3SiASnyfo5vO68mUYoFhGT6krvKTWf7lKo4B9FIXx"
    "N4UmXa4LK21MRS6jXOKSzIY817xChnwvf6awQgY8FsRyiYpLp2xxIEp8N23gfjZNTUjUzoRptvzSK3Pm4AtF2wapvvc2RQFCSLtJBD3ZScikSjWcTg5nY7jd"
    "HjTrhX0UQTx+ZJJef9AYPO/tLcb9CVJl4iffkUCJkxpvR+1yBqQYJyLSmqbpONry4xJPRhLm4KS45iOd3+CcxW71/rwmhJGsCU7EeVTGsC6SvQlTekWXzEl+"
    "E7HuKfDwl/681VYdmKPvW6OvktYttTZOQqzQOwxqc07x+Vfyo87K250U2k034UFh/wfmPbwLxidZnASh03m3tITG5tEt0K9liS0WmuZvhOA7B9dpvQMw/uJz"
    "YxxekyvUstI6TN2b+1zACUtg05TlYSPpEQIOwHnzoiwdWW4NgSJPA9ogVoEDP7/IaYmukJboCN/j+I+P4/XxusydXfCPro7oHjg6wIAXf9VkA1fZsxpBjcDZ"
    "f3Bm7uNjDOw96jjjs4cMsnESDC99zVhSCQXW/9AR7yyAWtshSDZlEy4IvsY1+FVMYIjvJP289xnei4drDzsCBC/RwzmFDuWhxjCjlXjXZEZLtA5QbRuHT4VO"
    "5Yuh6XpUKmOlVq0tIi+Kxsw4W5koI0ZoHw3G3+3+HDkLFNcQNgjlgrWmq5aPhiVJC+wIv/+gWDyVHbfudiNTb+WZNsJAy3bO50GO3r9Wfl7othaGOCyMrzkH"
    "MFVjfBxTUn0Tqqf54S2gT6sRFEPteMHUhTR1OVztlfd7sQVVh6Le0K40VD3H7lU2XLbv5U17qi28Ytn2J4Z9/sNBjF9Y2FH4PxP9mCluPdC7KT6IE75nw0Yc"
    "yXcM+htNKTk+eQdC71I/sXX6+a1TatX803s4H110W82ltXfsfHbR7XSa71rsfIhfK0urc/Y7QEwvkCNYZecT+IBq2PkVfayw82tKWmPnNxSzzM7v6eOdoVnz"
    "1Ym4MA0QzF938ppcEWlytdk/z9PGbeJPTy6K2l4cpqNgTvEldnv13Rq/oe4sr67gYROh6CifE+BWdenHlaU32e/n+OxLvB+ohGtxOMmZo+Gnyj7RcFZ3RyT/"
    "XE9EJjFY3G6mGBUMPD7+BJUR4puarDaNb4OkFl/Vstu4IRN5+TU/Gloxp7V0FM/Gw9plIB+y8io2x0ANn8bbw+vgS8J9BZLwpGKwPznGYLvVg01wcrBd7HdO"
    "SmJHXL9IHvSi/qtiq3qMlnmDNBNAQKnx6D2y8MpjyZ6qeNSrnNUp07ME0vejMI2zJJ6Gg7qryS+DqX8qA56wqbw0vRTqBcFzdm7zD9liHEdRyH51ISV2bGVJ"
    "0AjVrvtfWkj4YPtTlEINZgnqOWyodH6rZk74FUy43ziF3p4db/dR+W/jYPfk8PT48OhXUgxUvt/NeuDw487VddHAMiB7U123ZxZgPjb5lV9na0UdUypB1oSN"
    "MD4tjNCg/RPyjzzFWtSDLfJdaJVUteg3SUxzXwRIv/dCmzBsb5GeLTQP02HzwM5bEJ0tnoibjMt+jvL2/42p7rFnGf/jxvpiQizauV7RZsgDpzLwtIQVzle8"
    "cSDs/yDpZdQ0UsQzGyPtfOcCPf7BGRCtPCyjlYn6nQB1XKSYb8oo5iG7cntfYbxDNpIdvDfoNXzugtNqSHfcG29lqdlm8fqNp7QmLtc7K81Wu6sibtdhcNea"
    "OuJkvbPcXF6B8721uqajL/W+Ke7gc1smaaOjfJaMbwrbM5KiTQJyp5hy4W0b9305FFD2xsICthT1YEmpA2pEdVPj+xam7QXb9lka1M6iFPAjGJ6MgL5GR7a4"
    "8crI3YhHIdtMY8WHt2aOG28cdbtfNkFW0/fMpjs3HtBThNGi3Sc/325oYntp1W6luLN+urEnpY3FFWvSqsjiN9kNGwhSc6BJZsAxfs2slH9zOwVyAPfKUsbC"
    "wsji1Vjg3effY0+8e2Qni03IoAkT0YSJ3YSJ4BPC/CUI3hDniel7m4ep/VRnB7ymQqFNawCKPPkLe/086zmsZD2fGKPJX8F3zqaYQZ1DBbbzhXP43HQVcUyc"
    "V25kaAY5dHCVz+SAS7R/akatwy9XE56Cf05Nf+JKoSH9U1cEdaGs9ciUjogpxbMmPwg8x88wpIpW+TsKD/lrVnk+J+oeZGDsXGMlTxpZsfx0nMHpOM6fjgM8"
    "FGnqkFR9fCR84VZhzMmNoeuzopgkURML5fAd0Mb2inwGvCCLSvWQMy7N5tEKy0U6viXhxOjYzVFSqGkUPaVSpMf2v6RWOknaMYsl9Mc8LW5emHr6QdwQvIqk"
    "uhSPkfoWdPQTSaFF85F1c9SjuySb3iAKWJEZUAoRGirqljQ1iNhwtcLv3/C+SxoE4c1Fag7G3bhAER1ApZtITEKkrtDFvVBJCSJbSaaeNSeFsWoyaPvaCpPj"
    "pvaq3PgVhs8esp/pW8CI6voj/eMkxx/oY6fTzPXRmG19Q6nXYKBWpm/F8pUZw8oM8isTVrzb+1eGJf4jQ1KWqXJd5NbSP5zCFRbdzirLW6V3pPwC1VwihA2p"
    "m42S+LZGpgY52SzcCZigJGPQJzOcv1h8jfevxjOkwMm/Mu5QOLYleNtVdW1bqJpWq5JtkPGxQbEiKfRQDeIyhCd2uFeJ1Z3Hx1xEnjEpNCwxmm2PzGQG8wBN"
    "9COlB4VyGE7tmIXWpVqLXbVbuDTPN41un7iiE5pOoK8KKKEE5XnyCODuYaqK0wdGdVn5Q0gCWaYwmvjYwAbQbK7UG3i639xIjVWFZr/cJ48osRci5ZBitfJG"
    "s7Q4IGr3ivN7Fn2P4tuoZk0tz1DPbVSF2vkuJWufz0vv+81NOkVL/8UbffN+LARCJ3y/0guB8KrC61JlhrBCmQHtJkrhjbXFuOy/nIoMsLnhazhBpL28EeUt"
    "+Lnqed29F5IXfzN0gmIUiogdx9ivzY1nXYnijlEbmWu0ut0mKQxkGRGtZNVRGPFGa4VK5oXPY8rc3n6X+lzZL974CRmuVLvt1k6TexRHCBWy+mK2WFeczgwf"
    "4dRuRyE6wIJW1D4enckNL63F0fgeMgBZ930RLYDNpREwfRtHr3jMdn5HxxjS/LpCYu9zMY74G89+PlZQmXiJQsO61Buouu0pueBHPiop3vbkW9nZ+h9qILFf"
    "L2ggHnbeFxU7m5oI6BWVrhJD6Wqsj/WB3DyRVX1CWHsGoM8LawcGb6Q8SZWJcMXWPXqGqABmxwaoWHBsSPrnFluLtsbKGGs29T5Azx8fY5Lt43t0QzCBISG6"
    "usRpFd/7dNCJA2P7JUq9/EusP9lxTnIcf/ygJA05SqNROyOvc2juWUsjYDb8IV2UABs9LtkRza3d1FEbl+ujqZkzNz59fpTWUJqFzVAiFhuNqtD6qsg+rtTs"
    "qtrNK4nY6qI0r6VJ9qSUkE8MQj7JE/IhaXMD8R68mHgfAB0uPYAGSoBQdeaMn9WcewHTtrLEnh0P9jQnwAKbydP4UFCZqz5zq/Mwelj8HGdPuPiCZfbUrlAb"
    "ANFMRxnQzzOU89HyUxJ4MguK+hxFfYBB4VJH6ugkbOqWKof8vXTdoEgrkBPD1STc3g8HS9F6D0klltit0kMjDEPz+64rZyiWYVbcBtfpROnSwVc9qlYWfqOd"
    "jxfbJx9QzavlhtRudlY6kBkNIisfMDVaGVpgUeNEinnPDxOCiVGykRiZWnGGcm1BE1G4MNErU8jM5KGFw/wD2E581gBnh6vP1azq9FqnWety0l0duXm+xO4B"
    "vjdgOw4OQ+EhD8uDCiuxZrcqloKXswz33IEq8EmTDrwpqSdHvscvXFsrndWlXn7D8B+99vIKbiSWeIinoNKCfm1jOA4uf/YXP7srRhXPCeNS1gHfdYaZCY1P"
    "dfAxm/z1hUGl6KmHiM/XWtXev2Rrk+SifxWcmNS4TdNmBitjlo62oskGQ/bU7Z1kNAq1dGvDOPrPjLiOnGTFTxUt1KBLvwzQKanJRVabJvEUTuJ7g+hhyDDh"
    "m1O9Fsn0AAvKe0mEcXU/84tU9jX4I33FwszelgmT/kiXg5Iuf+Fd1kzpJM17RIi06U9hEkVQNpYpAzLHQSrpaC5J6mHjExuMbDVb+MpHPjKpwfHRWhMJq7mE"
    "dlMkrFkJnZXOskhomglQhczQysW3ZXv6RmRHAHdywEsS+NaIXJaRl0bkiozcV8+01vF9EbBjUnuoXHmnzi0Fo57ZetQgFy/chTWAd/kVMbWtbXVkBfZ7WeOp"
    "EflORm4bkauihKVcCWKw28t2fKspC9kwgNW47anRR2G4KGI1V7QY0LY9XSvtpSWe0Gna89iRM99p5RLkRHba+aJEtzqdfMKabOoxf0m1yX+O+M8WHUtqVr5s"
    "f/i419c3tGqK0k42qCszzpGsgCx2yfJVZGPzcP/oePvkZHurDxxV/6Rzutnf+nraQn0pCb5ZBb5RDn/0PHzHhN96Hn4Z4ecCfpcPyAH/+cB/frxweKY3ybPj"
    "s1s5Pkefj6FBSx+Ojj63+rv7H2WOg2dytAs5PlR3uqqSH89l0bXgUAGzsLZi4tjzgxNkg5a9sHO92T7dbPUpe1cOGDVth0/CITlbeEk1Ri0uv5PEUqo6uIoV"
    "q5V8WDkOHK6/vbEpn9t23q3SW8rs8RE/19Qn33j5Z0t/tvVnR38u6c9l/bmiP9/pT11bS9fW1rW1dW1LOnbJiNVtWNJtWNJtWNJtWNJtWNJtWNJtWNJtWNa1"
    "LevalnVty1Tb+gsm0E+tGVzP+I7f7awsrbVzhZjKp7IoOGf7l9PyMjJSMAuRNVhqd/ONITJalmej6tnBye7HA8CE3YPTfnupv2ogqiuNBEnPt1faYeg/Itth"
    "qHd+kfMZyu2Yp14GtIXtupEYvk1K1s5Ir3XZm7Yz0lJfpB/Rh0d9Pr+2i6ZoKpS7CkuF8aM6vrmrz0Vd92muAiq4z0mrY//ecBjav05C04EoHObRkBvWBdhP"
    "EDiZkn+IB2NLlHb+CRgVI42cwS101ShLuDbZMHxntMxk6bXHjv0neWlB4yBGZBhNZxmZ7fEepmE0GKFxtVetuWsAYatPZT+faroaDNV+Y3jMTqjoJ3qiYQrd"
    "0UkjP+Ua3J+DcTwIs/syoLENYXqZtYvaiK5n46fL8nMgvDA3D4dD9hGw4KnRQixRA8VRxhwjjHlieCi5MDIU+8SgUPpT4yELqBwKAnh6FBBkjuJ6NGFCUnzj"
    "zkrak80jTH6g8/mZlZGPndGip8H5mjLQvgBO/mIhFtZ4hC5wjIf30uWshhf7gwYPhnWGKmjdbJ6r+YklYcxeae+c0oku7ZlTuvhFr8TrXNuxbcEeuHoemWsl"
    "i41GsbFRLVEJCwt1UeGbyzFq76PtFDhJUn748NaEY0AS2mRcbZnCJztzdOYc4bv3rJFZewxarZT+wVBcI9YCWoWacHPVIbfLiOJzkUg6+hxsGOAZiAX7yrEc"
    "mhSKM198puhRGDV27fWwDhlKVlATctiA3JhYPrvrdv3yFRjml42oqbDUsKocqKwrFy0rq1qtNtrepKiCPEYJBM4fnJHk/BhOM8G106uaBsci7rTadW0T0X/H"
    "M4SmK2SRbRdiLA6Y85B/HMBhfaGz872tl5VvaeZOVlqQl0G8PxxyN9pcFFgKaBqsToo4k5XgTFLEmUTjTKJxJpE4k/AKj8nZpZcJr5cYL/shWzGXAhHV2DoZ"
    "kn9zBQdtkLzJwmn9ghl9qWej2eRSxOO7buUVcSj8up3GRvPwhqjRbLMB/G0u98bGgd6Q5/nCgv9LvDhYd0pTK5BF7HEEBRRWnSFaBMMI1nWX4wgPML5qu7TV"
    "AD6+qmrCey9+M6ArnNJGNJ9rBHQ+yV7aDKHJwSeB1CSu5WlMqo/mtqNSjB0nIfumBexJSrAnLmJPrLEn1tgTa+zJ7zhx+Y4Tl+44SXHHict3nKS448QVO05c"
    "vuMkJTtOeREtV5nmNzZ4P78gQjXEMQ1xLjlRydwxYi45FUea4C3uNP2fZDZvgbJWwQIo30KYlYXK75pwismPwJb2h1Enj51vyIBXXTjG6EkHaefo1g/3MnRc"
    "P6Ovf0S9mXSZKJwNtNClkjTmJ30a8SKGMs8wl6dt+oCy80y98xkbXqAFP/TGkvYmxfomheKog5a1P/3M98a8WBpxPRKxLuNZMkA7EmSgpmQ1ZtysKid3zDxz"
    "Q0vpHmjeEV5sb/uDkePkZOANg9AK3DkqEyi3wLZ5wrBRMD2tLC3mHVKVOOZz2QlMdjx1uL2hI9QaQrdytOlF5ZuNoFxwzzN7dGeOWGL2O/9Af6CeE6DFB+4T"
    "Cu9eByXP2jNh2mVY8rwdbaOI1PPgAhB3pGwSkQWLl9Kk0RzHV9lWgj7lmgDHamUDKK1YfUXtJVXPy/0+5ieDIoUZuyQej00Hj5l0O0RebvMelbmJdlwY9yl3"
    "eBvgdV8JzzovqQWZtH+9JoPV05Ugo/+vF22IC7TilbEcTvBM4datrcqALOFvmW3Er9TW+3q870f+Nd72bfoRagIN4HS9DmqGnkeNDjChqjdVZRqXfLgAhStb"
    "ajQ+nbOaFf9ZzVIuc2spVlRDBHyiZdf5lpn6gmI2fA17wjeBEqBEd1UC+el9NKiZvUSNXOioOtQekhKFthRW5AAomhukCKqSBdXzJAyRZ1UQv8+C4EfwXPoz"
    "1XCgp+qhtPvyNNorU75X8tmrszt5LAvGQpiv1N7A4dwmPz5Z4y7BV39AQ8EGAnTkrR/CMgKi63vw9VinOFq36sGHyffHoY8kovpmZN69K8y8C4cMQsNduWVQ"
    "NovZVekS66ZzpAhgnX49puvNPTx0SUciBPrHUuggCtd5uPTTgKC64znqePEOAFcRwKimOax0YjqytDnPhB9hMDslZ1jzmTOMzyruwWJs+lLYcys+eoZveTqb"
    "hYQKvXIY/pM5MSij0EOHmVhwyByYyUxlY1mJo2a7aO6tGU5XUoITPoMRPbjk11xjE/Jk7A35z4z/ZNy98QTdG0PSFaVcUQI6PL5C64kEsrBwTd8QS/qUJRNH"
    "mIFmI7o8C0eZHQpDtjnSKiLlmpckURB1dbi3CBZBipBq9y6dCRoWs6mFyKQWLp2ItLTcXsnwT2zn3xLCTjTiNT+SaX4ka6C18SCJeIBzJIbLhwrnzmH+iWSI"
    "toAMH6W8VWhSXJ5tbdrkebZ1kwBEKVS/wsF5YPfxtgKswrl33+SRb8njIUr6Lf+vjYCb+01R7vVEqu8l562li7cO/DQv0PVLbMUsttAcGHyv4efb5HwZKDwe"
    "fiPDIwyvinATHSs7KYQRPuU2vPzXIzaFvzN0J/XWeTNanBFavX4DlJiFAi+fTSZkHGNI+epcWeHf0CVyHrdeXOxPrl4xL9eevzhhN14Mf++94ZsrdudNF53w"
    "zRX68xm/jt/evL5m+96Af/Wy4qzgZn8E9U0DUlx37tkdu2T77BqOozmsKuCOXFiVhXzC9XM+Hrgl7uBw32TuTmhPjMMebcAbUTghvn0vBv4hx8GgZjI/t9DT"
    "b5CQJMHXvKsWe42JpUsZ8gyJ3l0a6lTolbE4galdK73W4CMNsZ5eeXpbEVFAQOajYOVGqLyTe2ojC6EXN1p1EF/BpLpTZIY+wQNrCik9/2XCEL84B2aG4kT4"
    "iuWVbpLRJTKT5uG1XXjhmmRi7Xq+qpjTlJgux2A6S0eOrzzXRz/BtCktpgE5sPdwU4Z5EoJ0POyh3P2FhX2uB+W6iiS0sMY6rvY9+XRFPGQwaUvTrn0/RXAV"
    "DLihm6ARA32AchmSzMLWJcIs4g4gOLt+dTWDRUkgwqkwT0XaIZiEQBXcEGMvv21QGeuqIx1JH2DLddIukCVRioIgFnHH5bRnT1WzyJF2xImsfZEuAwpIRpDP"
    "ZFjUKEqSwEZYwRtx5mElLGcH0Q0a5ucvWwMRFHlD4KLQ3M22FYmWJGaXgXqHGDb69MhvR0Guv0HPhQGs2asxbTrm0JuRHCbxabaOce5NMDO+p7V0seFhzuS3"
    "Ni6ciEFVSaLEhNvTTljag4MVl4UcMxlQVcsIptPUzBWArDn1YzVrsTVlMS+Ofotl2dECOdYTjhBdawYp1gh3ORE7BvIXj14JkYsDqAhfP4xlugpByuVsMpXx"
    "4htik3h2PULhsEwyIyB9EmT+2Ew3I7oaSylNBrp6LcgkIwypJMgaxL7qiBlhph9Y3SlGm7DHRstJgF2R5jILi5IKldpEm3uQzlAK9zL45Mgk8BwUSMxuTtWe"
    "b2wdMrerMGg9lUjRtXA1NfDOaGyKSRWNTb0019i0pLFpWWPbpa1NVWv1rhvJbVbhiF7GMgaKVBii9zsZw2z8oiVkRhTLExugiXV85RsRxXpELgPpKJMRVnmM"
    "ONIS5yuDwMW3AhVhxlOI47XSKIbc40XkOE8VYoC+9t60CAPU0iQoFVIFqhgmU41S5GFkJLhWxWV5ouCaW6BmhY2D2pCLUy3JxTMb0h6GQkoO+kPop6XAmIAc"
    "nXVc8Umjz+JumotXbiXxRAnS0U58fRaFiNZpNy9av4qv6ZmVOZiBJAJQ3Qkyo11PgEPeVlQZcD6WoneMWGRquyLX9t20TRbbAGbLajA+uhANZaKJ0jVjWTvD"
    "/8Pel3CnbW0L/xXHq88fggMGnKEFK16ZmzZTYydp4+XlJ4Ns1IBEJOEY2/z3bw9nlAQmafruve/19sbozPM+e++zB/aGgbfwS8j6MMiigcq/h4IKkddTiS+C"
    "yQkQEzq5xunCHXZYOA55xXFw8qClcoEOQFU7BwnhpKsbOcN3Xr2rcjusG7HiUI/1a/vVD61jmheOqU4TeemYhoVjGqpjmlcf037oHNO8dEytdJEvO6Zh+Ziq"
    "hvNVx7QfVhzTfMkxLeUV+dJjWnGAC7mtY1o+v6V98WaUxGc3bIxcYzPu1KhYeqwaRXFkXRzaq0tokkQnbKLT3H82zX/4pgESPx4G6bC0b8yumgPMw2mSGdwN"
    "FSvHChXUh0uSKEd6iAfqDDrGTn1bwGvCiiTUkhuFYYzrQB+F/YxRuAo2RktLYEeU65dbdBkym1enrOhnscJXzp4sRFdsv6ocnsgrKix3QicV9u2KVs0Glpw4"
    "OrpJrCu3I4UbVIMqxFUWNfdaLy7sSNhog2B64wU3oWzUIH/qdjj4D6T5N4M0j9mZ5w3L+h8yFsm7/w8dDllikG956qHC3YulZIQ5iHerkTvIuIoU9HRVymLF"
    "FSdSgagbpvGfg/xvsvlRcM8lgioEpkqcXU3U5QXWsGYVLwzSgU08DuAC1ziJ7fRJtgHp6P7UdFpG4ANSksMMVyU2gOSRnlTlS5HOQiF7oG9IxLUS6WGTmd84"
    "UGzZ7hsG6nGhP607dWQGT/T9Zl9udLNZPOu8yLM2qfwck9LzHLLu6LfnlE6d7A5Hbn1mW34zs22hiWma3v1pGuXhEpzy26dWCbAaBFRJtH632YzlbMZVsxlX"
    "z2asZjOumM147dmM3dmUMzkKhskXcxylk227XLRs6iIVNlWFqTl1NTSdzzyRV2E41DLwjm/jL5ktbaiHXMtRgfBq4UmnyXvKe3KvwkjbX3YCrI28tbJ8PlYw"
    "bO5vnoyTwadNkS9QJMX0DieJhFb29FePhTjtPLSSCLVZDCf1tQyclMVBy/9SKMdOk3I5mKpEdAKnXiXnQ3VroZ/ErsO4NojCITk3wOrKsWJs10ziZOl5+DgN"
    "vkTx2UPt0q8yQQycsuis5I2++PdKMb3NYXgaQPObYmQXPA2i8fPTl8GfSfomTMlCFOR+FJzDWmLbq9LpSGkH4NYL9xTFlydooYsfIpOJ3BV+KJ8mw5PZmX81"
    "GIWDT7xxyWhk1rulDOoFcMgeIamjJfZ1DLEdy9GEopajpRl9nZAlaf76BF+HMx03GEfkmubNOACyDztP0SSf/UimPSkIlp4Fk0kg5SG7HJXM8uksfxIPkiGK"
    "Tf3OsVNJg4/nj5I0hXZf4LNEpivKkxiZK9iG3y5FPbmYJhlanuvIuT0l8XJ2pIZ1nEOhOfw75nX4AsiQ9J4uCxxIGW/xQAl7Y9kXnP2ttnPzSNu5EW/8jnjM"
    "6c/tml7JCsh8zFvxyBMPizFU9yWKEHCRp5QhiCj+Nfb3Z5P4jBKzULxT4mKf/KuTYPDpLE1m8ZAP9Wlyxh9hfB6lCTkB4YjkPExTQM0UzJNAINsfhHGIO8l2"
    "UllUjTzee9Pr0DPnB/Gn+Cz+EB/FL+I38av4QeRwkeUiBiidizQXWS6CXCS5GOdikAuYoFkuhrmY5n5smjhFL5WxMfm+rgyEFsCTek5aA0H2N1tYHV+gJ3gF"
    "eK9Y0DCSIoapFi7MhBFMDEQZ7PQSUQlSemNRhBoDsQoA9EYLtoFclsUke0ASOo+TDMDORY7mR8VNmdF5UoLqfeKYC8gFm1oX1SGX6MIlQh/wG15MYRPg5gjG"
    "TY48ImOHCJFP1YtfR5kawtedbBSdorQjrCP8hiK322Jzthjv7VlGbelng24645Za9pztEc2TWboh5WWHG4GWM22hO4X1KkJpYn2bTOk9h4EkLMwgQg4M2+eD"
    "S35Falmm+CpFmdiXUdzrCP4MLuBzqsrBgVh4iwFKeSJawFNQcABTabYJnaNPwiwjxziWP/QzmN6rD3S4LyLoqyf+ZP08DIgPOOMfWlEc5bU/PThUlDbJOA1i"
    "PlPE0ET8QREnXNVHVnBJxS/0MZXZPsNB/hNP6B+e+I0l3SMUTvuVvsfUMlQ1lK3JnvyKcT9Q1DFF/QoQYZiTJSZaNYr8AaM9gAMUd8D9yDh0HtU+khUjDPyW"
    "1k4BmCBoGSIAgZ3FCccZ5oo4EFNAKuGMsxqNMeHgNKIaPiMwSjwAQBQ7yyA2pwEMOOZUDuAPjBtx3Bc77g8UhjpLgwnAm1x/i1PA6KYBacxGcOX9CRGk4kTi"
    "vP4HCEojTZj6EYIsVIZnNvMBLp7C8WHcdeoHFCTd/s/wFcWnif/H4sxI5s25XxfYe5gyAzQvSAGaMByECI8ZRUGrqnLXjZOzJXtOws2NFwBc0H4Uiilbwk25"
    "xFDXq+StBDqyoo6WGf+DEIm3aARYhDBQpV8C+CgEAk0JAO4Z5LYBb0A4A9Zg6uNECKsiH+8TragSCqciPxZuRX4kZEV+aiHuecFiHxukD2/wTf4lt9kz3lWt"
    "yv7fR2nuV28WQwvRq3iFtpV3FSJXCvCtLHzDpVCBFmgdD5XePyrH0hCHiliE5Fyk/jx3NGVgJSoUIab5opjLSO9X5l8i5y+rge7D/cXJsHmyKsPTH2wLJXB9"
    "oaETKoCWyxEtxjh1c1dVLDfUN9Utb8Bi9ahIio7PSMarYtxvjMJIVb5cObGTOstv/NwomESXYQ3QN7xvreYwusKfmtEmKtuPK5wvrAWZJDVvI4a7RHsECzbe"
    "h4g5d9FUXBBvwO4lGhJN3zE0zaXmHXbL0vopdIm5OfPc0VfYW8PQ4KMA7ddJFR/soVTpef92YxieRwPyAuno9/Rqb2HGHvmxssjJT6ukvVnL6288baHTTogx"
    "gdyARyx0SBQul88bm9OLTaEiZeGYYo2kp5aRRRQbMUxrgRzs7fus1rBY5V9Zuvob8QgmQM6FtYyrOs5rqib7DUlWVkx4tGzCoxVzZ0ES9gCpcvz1iStUuGza"
    "bi+btn2aNmL5HFj9XNVBmf2VNbEV2Q3PM8pkH/Ze0frkrQtoEllvl2h51evJaFlCfNYy07UD1VRRYtesrg04AJ/MCnqCbp8f2me6lLuqyw+ru/yw2OWMa6s9"
    "UA19TY8PAPRWgNbLUmfdjDC8z4XE2qWfW2N8PQ0+z8L94kI+1sLaaOUe2YaszlPK+Ny3b0HDBfnr21bXVd6xFL1sv+ax3q9JXqzJGnlVZ72rJHfTWsEUVghq"
    "EqqhzCsMmEjXisWxWqcstWLjpYJW45S2onF6bfeLWkWKvjdGSfLr6xxN5UXXbB/YEyYtvL4OZRqZCDYp8fV1rEqhiWBAjqV+euR0oDR9JgkIXHQ+inxdqwDz"
    "wpYUgNztYgHFJVteBEt5yzULlmCeK9gAN+YvcQKiXLVcQ7LJCny0vn+zvnM709AOzAuByt446o3iZX5jTtJOPYB8T3JpD0C7EcFDx/fd88kkHEaI2rsvJ8Oc"
    "aGFHNVXplHw0LjhGQabedTNy56bUqfAJQweQb8AM+YfSQj0eVyjLj7Zckh80qRx/Li317pxLzM4p9+x8aU7aq5xZ6oso3ZCKIlrHglhhzsiLA0VWByowybI7"
    "t3+62xWxZZ8G0ykNDbjLdKNrxn4wdu60b/9IO4HJL91iLXJrQsWy8IKT6TURmjRZxI5AS7t4KACzKM3skq7yFC/rKKfe3E2rlupOcobqLtIyLune7HxZ12bn"
    "N3dLlq7u0uxcdKu6o/bKkh7Jh6/qTlHizf0ydVR3jdJLk8WgAut6F6OpcXtnYl2IJVPbWe02opZQzSyWCBx8aLdA9rF/HCHzv4TlsIyyUocm8wyfPG0kLZWi"
    "Fvhu62ppQvcnURzEgGnutkXiP2csCD2D9wkZUWzxGnrxIP77GI0hoFEobWUmblkMSrW7FRfaJwVD2/TW4PqafuRw+QLWPlTYOJedRK2OxMzv9IkPGwEiT8pK"
    "Exzo2P8BT/0HFWWWLUaXn124KaPWJEmnI1bRyNB9PIXlSYOrc5wbbTmczITWjpwDQUyETgHFmEc/9Ad53zL6M/J/lT44xdAf5WKIxZ7j9NRGGjJNlUWg8Z4c"
    "Wk+OHo3k0D54izQkq87XZ2hWxoqlnBB7pg0L7WUqa68tzu1ombfX2cYHHS23OxFnnrjwtXutqZg0TsVZ49xrdsSxydcWF815o0NvCEjjH7N1CLl/vOL879Vo"
    "k6C8xgckrmpWGkYSxVV/j2CdJuZlMkTzP17PhG4bn9GpFP3wpJMONlZHdfQdfDT3O4Sk2y3n3IyqYz88IxRsz263p1JR39FK6dod2pHmuFItAmLlbHMdLLuw"
    "teUMQ47guXQ2OKQZG8rTq2Kz2lwcQy95g/eNe1CrIJ/0Z2EyCfN0bng+evXwAHLeR7SJ4haq5j2347x+dcvKB6ZKxUjtCQOVq6NxCaPw01wiDmLIjHappU1K"
    "1e/Z6JXLyMtpooGqJjYXm3rKkaAJZQgR1iFpvVK2GoO9IMtZmAGnFhM5BKmktzrkM8lPn04nrNZDYxCIBEEJfBE0QnOOBG+VhV60L+S+tMXOS9ubvBbjI1uu"
    "nQRijJBqtKyUfZIXjFW9RKaxwdx0/IGOJ/MYUtH3Sa70uHXGB4YSxcPHDiWUbTEFE8ldb+XEepQkDQV7bJYJV/p1Cn/0hqMFev2YT7TNOsYFk8Aw9NzspNr8"
    "NYvmFOczw0NCf0yzLJ9NHsHUhcPr66fsNCrDB3XOiGt+FW1tvVtiVKGgxU/EF2e5XXtmmbbQo8nRwkJudoS2fri1NWP1a5x1vADetS6Vhxh7ABrjZtqXn/+9"
    "b+6i0K2y7WXdM7sHegapBwhRUAaEgRh/MYjySG4JJvlTFMcMekgs/hPa0UpiVtAHeP6H3AwtdX1aeeQ8Ed/NKVco5YnV68cTg+v39y5g6TSnntGFR9Um2P9G"
    "cT7z22SgUp7vbDfoZ+Y1HVIOsyMymhHoFugaP+onW1tJ5V5J5EoFaqOss6WMWnWl8ZIMkDVtvCQj/X6AB9IbFoIEA1De5kU3JtrlHwkv7IWtSuEG3V5K7Wnh"
    "ghTaS82UQMphekRu3hJaUZgd5KPwrSQtDSI6sJfpGetF5I+dpr4vrzXLFrplX1WbmVGdiaEzkelMvBv1Y9u4QX4Yo10HG+RF1l1S5jJGOqZ4cyB7sxZ54lFe"
    "CwjZTcRYDDy1ijI6ltGWoNuj3MW3CTAk8cMQBiGhQu3UyoCHGpCDMTaobIMUrQ/FFXZKXOMlnjBiyFQH3MWvrHCt1AqWWAav9IQ+x6sMd2deRPVTQn/JZ2mt"
    "rD0m8Qbnuj+tZk0ogxMsq8ize1pBzlBHcMpyQZLpQGif5spRnzOjZi3e6N1Phrf0rndpn0jzPARaWKLnZdboRos6KoIfoXmnoikdGa3ccCb45o0PZwGCPwRz"
    "6ImNM4mMOgHHQebhd8tHwWAU/hrOawmTDegtVD+ERi1LCsnPl6lhha2SsFKE2qwoTp+ciUjq3fq/SYOYHETyxiqntNaR4GLZ1uWuSvFdd6DNhZoO+wMexAho"
    "NqZ1jCllmPoRe69vDfjtQs4ANDhCUxU82WRG5r10cOr7gZJLeoyrCOTWiDdHoqVN5XQqhVxGNeCozQBm1hJxKoN08h4x2srRIywZDIj9rB6R4SCji1Xcz2Mx"
    "Qhlz3YxpUe6YmW9S+3ynOjKxfOu+Db640QCDpAE4JQZ4fV2bFWUCs1zV7Qk19Ijf6aU8X+VjR6VyMXekqA1sYh1dUBNd3Gac4ooQmzhbFliiirygC1yRytUN"
    "3BHBppu1oOcR7TmIsfS99UmT6YASUYWwbidhIcsU4yDDkCAGNBSMuYFCPiu9IjePc0UZzgAl4UzklQ1ggp1eXSUm67qw9gdpGFTWpxJx8PnguLMk+cXBo47M"
    "0l2epYtYJWKElU1RipOjuvOUrns/QuXd6ShMw8pKMdmdabWZpjdNM1klroqma+7GwphLrsSyJp20Qt6KRorJaq6WVe8mFnNXNFBKNyyikQPvPDHxH0Gp8POH"
    "KB+9x0pqUwyKmRZ1L4HckQXZEL77EwCs+sYkgGNRx/o1IC7KHIeFCKEZDpxqAsiQBirAhXJhOY7zPVc0AsIJncuOhXySw4pPaliVHTRjeW7RyJX3v/il7HRd"
    "v4Sk8grNyK7rupdv4Gup39PCBPWOlc0XKCbnLJEXc6wv5gzRBKkEwsNSrOtI49X2t81Wla8gt4n/tjRDCwjnCWk3DeTyxngjuoiPFmh9rXRSfgaAD3O471nC"
    "YOhuCTVToiH6C+xnxA9lW5OSA8Jy+7c6/VhhSsTV1Y7M9wbuPTCouC0QfXCxrevrQWFy0Tg7YjUF1hpcSB1q0OzGPfR63ivlvL6mLeLkHMhVQdPu19cxbgfs"
    "Ifyg6hNXZDCn8obGbpZ3vo83PERzECss7nqVw46Ezti7XHKWsQtejzpCfyjCcRSv553Rsymy1B2AwPzyGT6N+G8klu1x5glK0UvRfy2pOPenBSB0AXVqbAj2"
    "zefWzMjmTRWGiATYBHtHVl7PyQUn7hwYyRdI+oLHbEiJANaur/cRTyKkcU7SJgzccrFZNLi3KfIKI3x/oihmkAJYnESDx7b/3EJ1Y7T8wslPH22K7naNOKgo"
    "yElaqmjZkmJevOqi8XUp07bv54WBlFDAuBrHiisRsngp5qVAgzl2c9JNY9pYcUoKgpP2ENfjrQD99U0dLiGccZUNnNUDLM6dV1qlKHud5iPcR9NRNNgUEs10"
    "oyX/QPz7DMTE2Xhz3MqQAwcQoTzQc02fb1ZaJ/UWxDHRFVzNWeaHkR6sIhKb+MApK/HE6gyy2k1juzrSjD3iH9uivycJACyM/bNFflzfE0BS3nvZiaZ0B0Z5"
    "lVdydorLTxPZ5zSv3a6HkocDiIU/oRsoNg9MsbhtaHMkNZ9iaztdyY+vx3V8SZFSWiekwcSMEnN3Y6koQpc+8N8TcYIcAzunHwmnj37qhklSMF4UVsfKsOnm"
    "F7/Ima7MjLVtlhuQ3I7SCsnCqq+b3mJxHqQbL8WBxOpqZ3i1AK4bRuchby+U+3RjEOAWMhWzFLvsJG6KYoWegPuvCI8rNMg2xWmVYhmS1y7Rd+Cfi9pL/8Kr"
    "IP5swfQD8dIm+opJZXLvphyKpCnms8mi5WnLSlvk4NKkZWUL5F8xuURfORk8kSIaRvhJTEYbCjbVahcCDabqpKItM0iPxRvUs3uEWttw3Ia4uAOHUhAXsMur"
    "7jpJ9scVasCwyDfVKOIl6sPckqMBXoKXgM1DCm5V/iqdwgLXE3OWGKGFIjYfFfPb4er6rbptlux08SRfaXjZuzoBZOIkZw0GsTmLh+FpFJPHOdTDSE43vgCo"
    "xlczrklJ6nPsOkZ6T1AJZF7RDaODlu/pl8ie9faIpar8M1QIqi3NpqXULBmVwssx3Q0icriFWjjysHu0ynG8EjBV/Gavh/7fpY94FuXQ4qQbgDVvnIRhvKH8"
    "05C/+JLPHc92FR87XbGse5ronW/pIStwoMTjjf2TXoTsXkVO81avQslgREJXYkO2qO7NCnaym70Nxi5RNQEdlATxhhImQId6XJAbgP5oWvFMPTyveDC2Tevr"
    "HRhK3wKIulZlo/rmWpNJwpuCMgYR9fPc+FaohZ5nMEVJ+Fe8xRCT4Pr6GKVzlBQDOp9iBMWIM0z4uQ4O9rMqBxFFIkSEVfibeKrQ8TeF/Pjc+7O/VB1cvEaq"
    "kLpSoUkufkbzajM/Uv2f6v7PVP+n3P+ZJ6T0QBuV6YySOmYF6BNBHjlvTjK+V2Kw9lg8l+mvt7agTyfhWRTLe81in1Q9nPQD/TaUEoOgJK5RfoQLncagsG5K"
    "Li45X4pPE6P3Jh2jUqR6orKtbBjPPtbBR8eDuncz2hWRGk3mz9AGxudZKAL4zI38fF+Zpr/f3tp6m9cyHlbgxgYcK2WwjmGr/uL45+AuvIymQNnVjpExJY8B"
    "brMsmEzHTkbIUr2z3ScxavOzFGnM2AQFif/jwyiSrRVpL4PsUyGNuUZuGqp+JeP5WRK/Pj3FCcb7Wj0ISuVKZgR5tqEAPEPku2zoT/QM7U0OVaDZOWI+2lTm"
    "m/lTk296OC3kM4L7D8iNAppAf1rtCum8mJdnm6yRV2Sf216WzMxX5Dy2PUVV5UR9TvRb0PaujuEyPvdDMfdjAbtUM4+0IuTxMcmlW54TtrZ+4SPheoZjDo12"
    "e3urI9DOg0OyoViP5Dj2iSZGqVZJfuw8ZtrUiuo+piPqkX1q9LGgpB5WdK6fK313nHm7h3s1uKXQy1pKLKrI1zmX7Om9YjNWvqHVZC8QB8pHh368Fw9UlFTK"
    "8cQLXwdwvzOptYZykVhDnQcqv2ROF1LStpuLnbudu+2iRpOjLCRPj63A8wLwd/PmLudBGwjvT3Pbd4ZeLdnWzt323dti53b77k8NgFhq9hRdGntaPilb3gb6"
    "lbm+blc3xQ6h7NbKrUBhGMNCCyMHQ3ttSUs0qxRIFoGU9rq+vpVXmUv/egymqu2eixRKvKZUnpMRp8HTlazc+lLiq3L3q6MtfYYmfnIYAKKWoPZW9ZZJvL6x"
    "uBGYkysSP2iRLYxcjOGT9Lyh6QSqfgJXMhkmIuc0iXdL2XDQQgi1nTv3dn7y/s45jOKNt88ePthI0o0IzyoZxqD3F0nLbHDvW5rHNPAR53sJ8/IBRfNrm09+"
    "PzimS+aYp+R4FIxPj4nFtOldX/+pZrm7tbWshMzMuGgN67+4vrYmZ+z51ZPzIzSA2U+gO6ah62vZ0Osn+8dyJUyHZBrrVVd24/p68HdP+rs4i85geh/Oc/JJ"
    "uGIBcM/g9MPhvbPD80D2kGyfi1BklvFe9PbC+z7gLeEubEOSrG5GW1sxxcUYx7qwTbiapjl1m7taU4fa2ZTCWQU07PaX5kI1toFOeNy5QfsxSDZZw6LJQoll"
    "9BSKxk5gQtA+kpExkxiZvH2Oq046C8QtO7g5mjlTotHTuZXhIFHMRRfwISKgADHxPKfJl1pXNGMEw5aCMVBxk+BMqiSjsnFWlcqLgcmAmZqpDiXQ8Pq/tMwT"
    "J9wagPaz9h90FmKfYyV4mdy5swPLB4iq1HtF2Eyo3izGgcsKtJqZLI5xy8YJdGrbWCJyRkMmiez+u52PVecB+jnRuJNLI4rliKa4MfZRkS+q7dy7fRt9CaKj"
    "nD8qU9FbrjFYNKdntXLGnc494k9Ng8GnB2M4cCwvFbqo096U4PX+7MSZzsiZzEAkQg0aXcmyFwKUSwKkOQuHVl0DO3J5rUBgEg6LntfkvDpRamrdWG77ph5b"
    "vdWeqmI4IzEQ1rnEnTMCAG4c1fN122bncQVmgFsHYHnZnlIlWF3Od6luEBkc8UYSj+cbJ+EG6l2xSSWG/vqquqJJ7WWCZxJQT/IBnCzwsRKnBq5ka39GetsO"
    "3GjetiQeRz6diri4Z+/once1CHf0yN/p/njvJ5LuJhQpqsbXv8980FxksylirJlEjJw+bgTxsBwvO4Ez5p5KfqtRA7nz4927i8ojGq08otGaRzQqHVEtqVe6"
    "9Hc6t5H+LMXDZP8ItVcVuIuSNhXxd9CGT1VF9whrL3TzNkr7FqOxVVE5qrv4ABjFrYuqxDsycV5Z4z2Zekmp1jmHvTUCYBS2kDuFAOSSXhkvWhdN2VijI2Pm"
    "MmauYy5lzCXEjMUA5RKrBjlbMshh9SCn1cObLBnYGUOjlGVgboRGo+WgCLli5YvLuyrcLvmqW5DYHcTnsBkDBUOJnzUPykhMLyqZ/cfHdMCOHz95f/D69Yv9"
    "4+OtrXJcwUEuPnU+mmV5MuHwZnJCBvc2xdUwzAN0AQxdRQnrwTjIso2TbIPsbw2zjS/Z1eIkQ85lnhCSWAS2QLv3udTLjG/ydIZGMoin0b0TNu9I7f0YVUA2"
    "NxWQRx1sNt1QU8x/6U/HD6EbSRxaZhJRljirmZJOfm+RJ7/sv35lTMuxP2DpvWdTUJGeKU0exkM0kCtr6NnVAa720h2wrMiMdL88UsAV/E64s/ZYyXGvNECK"
    "rnrjqjHvl8fMHotkqRUjXzVqrKNnajtVIfTvu1jslwZvBn5gNsaj0JkD7wpuBny9kEZD0UTYJrEd5SwYU5q85SnSFmA3sSicbkJF5RMrybJUdqv9LzkuxFHC"
    "NyplIBMnoaWjFTs3t4aP4vOFCbGTW3IfWGWtWVKFXcl/K1RRnN58a3pmKaIiW3GeVZnS/JezmtqKy+KYpaPEotlqPy9FUUa1tS3xEp5cHd+3bJHekkx+Z5pD"
    "qXHkTLWby9RmZqIwv3ZFzrQX8i2riqdfV0FLIONVESgRLhYH7sGjw2OO3pMSzJGwJkA0SttPy3FdFGBh/X9jaW1P6UVthz1p3XeGdjL9UBaXDw6oLu5fJcS2"
    "h4ysat7syBtNyQ2qGmYRarOrk5/E7+gp/xEQ0TjPrsfXDJ9yjVwCru4tqY5t191oYM532Ddbg8F0mPsrz5+bgeeDwC1/t5x54zjPniBpKMGZwVx+2LOU869p"
    "+oFWHsrrvlVaxHU0yE/fth9iVDw0ufrRbkoKwqbbh3kjOkLL7xSIIdC3hraQ1qWQ8CmNWJme8mT36ERqO3v82Ir5WMgxYyUeK8K/WlhKPla9zKloHeM6K5hQ"
    "TqEdYFfgVH64pNRRqRtLc9pdysbRAJXoW8pXtIIQtOomX8XC39QMPuiTjVXkVVmLZaBNy+xMvTWg1EJt/OJ+LB2IvAjfSray/sNXiuk9A8touVpodZyu1ndR"
    "rCX1SnV40LUrrKangYuQPLYl7QnCeapXnRAxwSvYs1YT0JwnLrQlQepxGJwruwkEeWlbPcikRXAJiV+UILE0EFfC/JAVoCCzEqlXsJmhKz7yUTESIILkIUv2"
    "RwtkVxJYqrnbCetkeLVg2RWcxooslFAFb00WW0AsXzia07mxHc4vk4W2++Fu3A9RqzhrXfjqGfR3fImHmLmO+UPGXOqYjzKm0Jx+IP39D8iBGS6oJirsOVCQ"
    "Srqaq5W9/Z4dLTT3Fb3V7uIfKwHDv7e7le19RX8xk4PSFrbUYV43MXyaGtaGPpL7G+v54y/U0+jYNX38KzV17Zo+/JWadnRNtCj5t87RglYw//apWdCC598+"
    "I1jBh2+vYOeI9skfCv1Rtfiry4lSM432kXbDYUfD6sd6zXDTaij79zXlRsOuiewefDCM539BJ9xo2IdpAc2zLWPlldbGSzecNnGlyLfexiP4QA8AJM2H2TH3"
    "hnwo01piG1+i8XhjGEa6RpWFemlpChwe9S1IB8DNhnTGLEPshyvnsaoSdZ9SPbk0k1GYqbgRHnkLi6GSxbVVKKLnXtXFq1lXZeFOUXFWGV8rRxewtooch6b7"
    "BuO6OZt9NcvdgCweQFiytcpXjdnevvb4Dc76l/abInt7G/vEP4gu/5duuitVoFeY5JtQVoru5cJMfq+wGIt/981osUOuKvBrvRt6t9piyTTRU5rbgpBcCXub"
    "rpqkxYsb8HzdD8NqeWu4nL+7XM68ms3pKCBUcX07d+/du9fFxyDJA5vaXEzp5s2KUs7mFI8FjcJDR8N4xtHKV5QlZWp8SkndAzR2UWSQ2OxJz+qnEo1T3t1U"
    "J8nHXaGXxq9doa/GaV51p/NiDF9hpAD8KOu/LbC/nEk11NgbRY2Jx/rruf56lUmz7+Kh/rrMlGunpzrfa/31s/56pku801+f1JfcGu+XMsAZJi7dHZvmuD3K"
    "vKtHXO+TWEOtkhLbYbN1R+C/Nv2nAx0OyO8OxcpcnSNPhIp9kYs7Xv9RZgxJHmL+LmTrClQJ4CRj53JTmdzcVLdHKHYEGWeuyDw7t7JBheytQRIjrFYOA9Us"
    "X15EmxNJ/ofeyhysmSOnukYDAhAazNEonPQOrrQGWe1ga8sVEfp/fOfwXPc2Nt9yWdzslH+TqeCNPMHXfKSKo3gjQVN2GCVb2gjOgohkuamarPX/PNhkrm4s"
    "OT2tWZxsqchzmekzxEoPruUlyn+TDaeqstqUU6nF59kyrd2KxtjCHlUeZW+gxmnI0tByNtk6oL1YxcO6tQVTURB3bT4n8lHLEjvlFTBgKWSR9pX/SinBNEgy"
    "FKiXAk4Z2qP0jDC9tS36H7LaUxpsTR8JHL7IYHUIHxeQ4zXnWJ7hZ52hIv0ZJ6LBXTj99N3B70/qu8PCGSh2SbbilN79AYw1PoMt8TQDoALQBO32vsk8y24t"
    "ineaDjarO/BOdaCDUlorGvmZ2uFGlHJMoCQJ5fQlsoIkjc6iGF+3SDHnIKlhz5JddsB8fZ3cJwX26+uQUZkrlROdsqEuY+9NpogDMTvv/UmWsd+9h2qEHu+z"
    "DDoPEyWUv43TACpgp3j05CFfy1bfRbZDTVp1+SDAAX22pAafC1Ysm4d8o2j7Hh8y19LaK5jl2Ql7ccjIRR4qpsndDDBHb3C062Zc2O7VHiI7JqtD+Ytmij9z"
    "uGBac5++LxqUMvd6DyUUeJVpfwTIkIccPtaAYlP0hT4jHAbYZeYt3lddgObi+1NfVJ8LDMnf132VPZZ2HVijoW3fUi9eP94UrMbT4lfUN9rlFkEUcTXGUlnv"
    "CiACAo2TMaFuZJimd3i0EGQJtHfFEbfai4VX8Uqrd4HzUgqXh7HU2OKGHHuGaOdF2zOMyZ6hQeMPc+nbE9aShlaL1Wub2rzmEGgi7sanyoWuTr65hAyqgpMM"
    "bZo6UE92mUAdddy1AwsAthbuohFY0w96+rFeFab0qBHBjWuOYaiPkJ7M4VChbgvj6IW76XJznMXGvLy6TxPtQt1R3bdGQV4UlUqNUb/EYcV+R0RmLeLdCIeW"
    "74aHsT202BoaJDU7R3JBHP+RzgWPsvKmD1bzf66+56wr0SzJEuj3J5oFlPzS8mwAMLB7tFhog6E3zlJnVTfda/vz1w/nz8wexOfM285bl0ky0TcrrkyIgpty"
    "10vzoUgSRLBgqVmwaDclOe3Q3oq0EyHGLJOpoSMoa6liarMCqEAl+l2zsmRnscbLvkFHbAVR84hugxTy3yHjeWmQA1B1Om2AEvmxASiRDVAgBQFKoU55Qcrj"
    "GKlEIkn1cY30lC40mAkX0lbrHxp+f8yUF9xf9NdvOvVXRa5I6P6Dge4f4rIMAAN4/fCriA5jwFfJ36BBD0AK/c0gz9Gs5NBOoG2nyKRCtERBZeoaJKVuKtef"
    "xTrVhW5ivCXNlnNqPVmmMqUpEt+YG2boiAUsMQmTz1BioXolLWsUow6jUw7fLwezMWST7aN2aeiX8fLKkYarhxd6gOyRJLC3YDcwbqdbHLnQnA1c3w+h1Ii9"
    "smnI/Vz1qsq8V6bLFSzmSvcAxmAuIC3sCOX3GjlXmnPoDw5dcugjh75w6AOEtFGWzjYCvniEuy1+QSetBnceoFOd7fZeXiQlIq+XK7Sb/F6IsKUY77EoetpS"
    "sNleMY1TVCUJs+sVZFFbc2/lvisur14n9GH+zTVa295UWCWZbZ3j3sa7OA0HyVmMnK0N1Rg5vLVb9xZkQ0Y9A7rG+pxdhYaS7Y3S/5i1UGemwBirRcUdRPwE"
    "dJj7y5oFeMthiT/WKKE92+QFe97FueM9UuMNY4N3gOm3XZj+i3TZOZmiyQnSjpW+KSKjgPKxnOnXrMrWMpk5OkyP7K0hOFpBB0iV/qWJZTBkoqP2m6QQ/sgK"
    "Y/sVSIbI01zdcOXItamnxQ8FosHsF8PPzIOvktp8CKPYhFs6cGrGWEOMhIG6LOLAva6iwL2j0IF9CH8kRCtIi9FEIuiWEj0mVs2jei5zTDMZTqltoKkylmQu"
    "JKnDZuibHSPdDTCVfzQMNd1ScNTuTV+hqW53iry7zt16rq0rsDV1GWKitnyVWD4OOOctq9BVNWjgU9zbeDVDu2RodiPi+jawdxsT2b2NYRKyTuAEZVw3gglC"
    "esxOw2xtVs269YiC6JI1FZnBm0I6Y6XCjCvxroAtWjHaq3IZrtRvr93slS30BdvP5AT0Dc2EWOBWxlZDcrFkANh3eRl/dZ+cZvtkcM/2mOR2TYOM/MjCA/56"
    "mzXbaAp/yYO8V1P9kfiHSq6eIJ1Zs0dC1zJgr1idgxHpxGGIams4paEB8mHr8wzZNzEHMoSWnqG+1j2ZSozGPpjqgrPAQVEc0yKM2FmLdqwAxMse/rHH0ouD"
    "fhiUb4SUqCQYKT57SRt0AoAAAHQpCowedyNH4go9GFbI/EeBtWfLW9NDSh6X8OH8FQCzJbJEzkZRIkUuEsAbJTxi/woktoYvuaoni4X2SHjlCFibqdSC7CbK"
    "9nZYBZ1hV8Plz1SexszpSmBSquTGwRlI2ZmDHEV81GdHFLgKxsdTusTskgU5EwaVpygLzvp97949f9zbRPsM8vUusK8qhgzpUpihoB4hOVLE0T3j8ZHnsqCU"
    "JyRLJtTs+StAzAJ6Br2SotG92607/H68qcaxKaRKU5KaSElNby4EdbwH17DdD+TY9XM592oV+mVQYu8ofQm6ZyhceoZoOXJn5qgdw+aPTQ53IjN9kKzpUqR0"
    "qrGPTH8FAWIaCYc/xBIbGQcriGeUXCqSz8qIlU0Rs9BI8YIHyge1/pxS7BvV4CEsXx4r7MPydYOMkJspabc3ihR2Yy0VDqcb6ni6fStkKmhulOXhiclI3g0l"
    "ny60t3apfpZnEDv1nAAV9/DGonJ4XBZxp8LjX/GBSaPckS06iSYvAvP66JA2sDGs9wL79cCw+a0snuXbpp8C2CGHNop1qEeVCjQmm1VcCDGlJM5Tnp8FGGNG"
    "FgRegckdBEu53EGAt7px2fZ86KdC8blpSEK+4sTeQtfjt0nnwllB9ZRaWrzqHbPsCOzUzdx7Ygd1WPSZXb0xsvLGqCxYsS0kbW/5YUS1EnNVaY/ARptL6sQp"
    "9yQLQC3GQUEmxLLybcimQfB1YiDoTMuxBLyGJIj2TOh3TMQgmPqbpJ60aSL/TKLYjbW9USKj9rtKf5iO5ebb7WKuvgqdzPVnRT9zJyjf6wbuepQm0tCdo0Cx"
    "SWf6a6jvgCl/paGY8Ncv6onsdCkJnEtBDCk4MQiWru2mcMUb8sIzZGjrULlbdI2VWPaaabVnoJnUQJpMZzk5yFRvGFkRXdasHWmKqOAT0jbHTM5YbVvOVRyZ"
    "2D9s25Rhx2IdKj76KKhk8cD0NDserNuyVLjIAMQh973ZOaLvhg91WU8fswCNNLuCKGNr+CySMokR+fakVasqBBCnrFU9fb2NRxTP5nHI5AEMnt4uGD+Mk7hJ"
    "UxUON9zZRBlEapLcJmpHDJW2bFb0AFrYgEv5DCj6grEF3RCZ/eQ4twuWCVBHqH/5baq3VFS+XRFbmKLRgoz7m4/QMHACKZntWNYSdUAWGACoKD7bJ0vEaBBE"
    "jvKhk4C2MgM+AMUylOIwvyKKSoNhNMsacOtJV5aOiEQm650YWYihbCCyKNlpoIAdKgG6zQwDy9Xxdo3vIiJHWxcNKzC3A5fe9g7aoAnqAfqyY4g0UB8jDa00"
    "sJL3rutdttvrKD931V5bI+UtGamSVX6SmaFpyCjLJa909ovyJdoFbCRRPZnWUPd5swPUVgDUlj80jvwi/RwQ6G908juuPNHkuW6wJCkg9uvUnO79zweJnI4a"
    "Wo6YiZF3P/Gg6TyKZ2F/VsEPrXgsHS59+515/aEWfBkuFXwZSsGXkcKMb2hX0Kr0YkvqBb+ec2xZCIaA0g3rE5n1SVetTyTXZ9n8x8vnP5bOmb//CkSrViDS"
    "KxAtXYHo714BvhfiNYG0Ap3fCy5Xo7Bfc2lbVzThUsYfPFzPUoTnUzgnIRUCKkaWQjUTHsZwjx85FrWVaqCFnz2PTwGVxGsJaeti6uOI7eUDanK16Be4UMuZ"
    "T4fhEbGbrq/3od/xGXZzecO8O9reitYP4yMfNTr5aBUmR05xwfmJsby7fOGr1smw0uUW+KYdsFiclhFeg+Oeacz2PHAFveYGiz0N1hUFsC+azX8V0nhUEL5w"
    "cUa/652tQAvPl6Y1OhJpJDWGvXbPII+kUoU45JmDQp5/RxRSzeq/HpW8oSd/B0q5mJf3sOqGoaEvvnXHogOCzcXiotwI+S/QDRx/JZHOfq6/p64Gykkr6n2Z"
    "nsbfSqqvq6hBHWUljK9R0FiDfj92V8mdYwPYvmhC/UQT6i8VoS72C8DuYF2S/ThYsdL/bkT7X6bEeFjfmRZ7uZQWe1mmxV6uS4u9NLTYlwpa7GQFLfblL9Ji"
    "q6mpVFNTwY3UVHoTNTW+GVsHXH1sv1yhd3ZJSvX3qy+3AKmnJ0FtHz5EglY3BW/V9UiI1HQqWN6plDq1tAex6kFc1YNqg8Vye/6DNf/fw5rl0v+teLNS7IB9"
    "WbAQrzRfTlyKljolpb2S3djMqLxrTkhrOguznHLqArj54yLokyBpDMe3msSNqZ3xbiqp3PH9lOz0SSiYFWndsTCF3wbznvF/h5YziQyOJX0blvVqAqRmD6ru"
    "XoMfPTC36EleZbdGGx4eW+hYIUGZ5UPLtUaJL9gLegdyh0Xx02iMKnw6NdvLemfqljwrpqZ7qUotmOg0bkQHdK77m2n4GZCk/H00DBOyna0sKW1GgIuiV5Ol"
    "OYwnrQ2SJCkIYIhVRUOPDWKWJDQkNmbNJHH0yA6ykezxlgiwUD4WcF9rZMFwTtZD78PZ/PnB+yfHj969ffvk1cHx4wcHD9SLXVGuZPHA3RXUghLE0Hvjxfp7"
    "QwzESMzUDqFNaFIwH6YKM0Bf2kUOlV3keKF2Cq+yxMPIsq/GiMtbYbF44Y6kZAzbUtv+flvdnc/F4m2hE0F8HmSlDjz6mvkkjZfawByJwd6g9wAdNMCqDuDn"
    "hQc4XvKFttwThrPkhlY2K30mbExmsLdOwo0QKEq4ZKX5Zcz4lDMkTuQ+9G8QjTkNSDotLBNDqz41Xov9Y8s6gEx4QQn7nrC2gBnQAIY3rtoAudoA4WIpMEBA"
    "Ml0KSJK9RKXevFseuQtlT5hZpzdmnZ4UhEKAkAv9HwWLH3XrBI/fPK+kLx5F6WAcqntK0hlTZeM5868YSYYZyCRZ3EP8KcyDfcTDejEHWDS+Fy1EaDC5HREa"
    "L1mARBAmwQIu+Kfw0HCeq5tFis4HJtQBhJhDrFXdtyQb0Ox2P931QxRvEKOGv6Ng1cyPG+l2WI/6YzRDVdcKu3DIx2iGqq6VdiFGNg5ZMRH+XRa7MIBaatnh"
    "6GgbWTbbXYiYU0Sjo6NkNyErptryj51+jp3Et8VU2txA1g/KfCtCmjXcU8vd4TK99klcy0haoSInq1fofMGyfEr5HfIkoovCiryxHt+wschIMGysHwE/Z29M"
    "sKhtWNMbdtp8jDhAeuNeO0imsN34+2GS58nEAsAUH4wVj6YXyRQdkYpkGsZPoKVhL7N3auDs1GSh0SC6nyPb5URUcFCRGo8usG9H+GeGf4bIEmRv44o7MMH4"
    "Uz/e7lLCGYoCq8v7vBYb2m0qtHmGiULjsMS5rmrORvTiPQA74kIFOr1mp1/YVfLFWm3W0/oFnp+ZClNoaJ8fMW00ZCvH/rSgc+BHBYnc7aieNALluBgPUKgn"
    "CA8PnPIJHI15PRUTOBHY/KR1CeFYyD5BMiZhdLFfqJLculOPG9Az1EWG77R+gaGhkqG7wATq88LtaaGjaSOHbXncyPtyrgZcQyQiOGeh11NhCEXIrTsHYLEY"
    "o3bDszSZTWtn4lzoae564qzhny9uWaZftcCf5GzbCzfXCwdQImzm3nasJ3YMaeNdP+2PTX9pL5354+1UzP2zOhVp5DevBcoq0nqMzQqkaLdfrw5s13Naj7E4"
    "hxltnsHsnsInLslELQnkwFSMxkVADZQxrMjEM4Znanqt5BoINJSg1iUUneYZAkUKweJ4C+k4cuBVrpNriSh1ScMJkIaH+RGsH3wBPMXvVH/DHzgunAe++7JR"
    "REYy3QXWjhfzhn/XXdQ57jNYy/kCXSkizkoWhZAEPCcVuVB+dixAybB4sB4sHq0Ji2drwOKhDYufG1j8OFgCi3+U6rHkGq8NsLgEh9vCJTMdsAyUwTrXv0KA"
    "igA4LgLgyALAqQ2AMwcABws1xlfL7xvyItRh/zqVzNHxfBQO0yReMQLk06E0I4wB7h/6CuXtAp2XptCjRRFRMWA7MGb7VB6EII52CSvK+1JyODuMjoz6LhrC"
    "UI/A4zCd4rbdTpHFF1bGA2HejIpwYExHCCvWD0VbWxGKhwPaGbj1JCLfHhfPYFpxBrv1WtqkZ3jnLNrOl7a7Xj/8r67vt/dqSQ11Mg7jBlrooQAezlgF6NPr"
    "LctWKuVZnJCEONQSLTJqkp7JMa6RpLN6nt+ph/0YsbrDqNE+EjGic/CJr2ZoThQ+u0em8MAsYbQLZ53ZvRfoARJgCn6iUBf79oZY/rhUydF2d1ueqAaa9NHV"
    "jiy+OhtVyIO4W0P9zibU6Vl3R17k2aDuNn+k6hZRiwOnuJ/taoH0DFFaGP5hRiOFDULfMNRIfXeP0PpKUOMjnnsLRGHKbYfFloBCQAaqVk3AlkKY1RSWC9oK"
    "0TYIrRx6UKHPLn5aN0RRAxa5tlTYR98rXNhHFyxcGD4vFyh7U7pU83X6Rvq8snO56VyuO6fFN3BpnGUDuFhDi+vWOjWxX4ZZBR2uI2MeulhHNrxnL7qiTSK4"
    "9QJvcR6kG2G/tnQUQk21KK93oPC+RH2MFQGkxj4DOD6E8c/M+GcN/ycxhLsNJwFv6/RwhtOAPx3+6R4pXVIM7XDkbf65gzbBddpdjrzHPz8eIaWDadnhEOvE"
    "HzqtOrLLkWhxa6wjb3MkVh2pFyUy7RHyT+y1htF5NAzl3thRaPQUlgf5kECbQHtwOU3hEgewBe3AMaXAGAK34cROAUp4NWe/uHDNaJLR3GhbDGSlFGY/IwOk"
    "MPv4gZNh6FOGCLYgGcf00/utn7a2gt1WB9XPdltdggRYYcPHt3oT0+WYyMTcphjodc1b1NbDHtI1sYdUKph6a2ARGWIR7M9sT7L66NHsPdyG4QUbgsbXfdfU"
    "oI5Xd/NDcze/qsI/2gas1ToNc5bueEgMo/583OeL+7DZEfb/zR/6tb/tCP6vLZqRaMb8iz8yhIHlKW0noemkREficEd0OuKe2IF/nTvwA386FPhJdO7h7z1x"
    "V+A/TLgnbosf8Qf+tPEX/t6FAHzjD/zpwk8XoyGqS52GrlMGSr6LSV2B+aCVu/gH6u1iUocK/ghdQBt0mAAlxE/48xP2EYMdSLktOrexbvrThp+2oJSfxB38"
    "uUM97+BfGsQdyv4T/qXf29Rt6txt/L2NOeDfT0eiKLzwOBmGg+BG3EpjhxKNwqcY3g+XWvznqf56rb9+5q8/lbr1s1Uo4DKThE+GZ2G2onP6cfkBWh1D1l1o"
    "+HHhXtjrGNfi+oHtRl+mTqvfSyTFwbzQSyZsldtahBJpuiSvE9FNPtiYNkE/mRSsgi0I1tI9+ZTbk6+lyHcj5hoy3w43g02xeQL/BptHkgcndWLQld3Vgngc"
    "BTw0QDAL93AK2OAhXPLyBRgdWB52TBAlmpLDrhXRJcwQi+ScFVlflAXTpBfCADDzk14gBr3hwv85YONEVa+6mcCq8NZalorX13BpKt6UP6OWVc4Qr/YaiKIB"
    "9u6/f7iiOScVFUIJYm8hCpHzqshLjPxvqKVTrCWoqiWoqiUwtXSLtQyrahlW1TI0tbSPkCcPXdrawr8U6FKgywGYRXeFd1yGA63lf+3ARTpgknyAvs4zWJ3D"
    "MUQgP5c+wyNkh0GP48XxD1cpND7FUIqhePHf/SmaqhxtbY0Op0d7MN2tYZLXMCAvH2+XlIur+A0yDmYR1nsO/wA1E1iUlXF7Q6r6+hqqGx75V/TQ2O4lxE3A"
    "7w58Q5+5nd5rTSzhuyOOXWJuVAuKBY5Qg50nQFUWqprihY/J/ctlOw8O6dNlabEeyyUO5hJHcxlYQ3yK0U8x+ilGL9ZBHoiloKDuu6DgYdS7Avzz+rqr8WLr"
    "CRwWM9pDg1v1uKff7Jn++IQ0bxtIiRhApJFhOSSpgVvZ9TU+Ll3kyElBx5XnCmwG7E5cvbSJoZiSH1DEjUq+Tw05Lbmo+lWJgA7zMgwVBLAng50JVPNhdlSP"
    "AIJBVLOzx7RQPdJjAFhG/ceqIjTKJ/DlZ0A9Rl3qFoDdKCbfKABReeZ/CGoDj58VAMtCZ/YfUSeSCLHUEGIN75cAEOiMyLDYf48UV0wVGxt4CxofrjYaFZLS"
    "Bj+24UReJf4ASFOCviP46NgmvmPJE/Njb0aCEqhkQewvMdtNyLf6DEDaLno0G/tD2DP3Bzgair0PJ6s2gtj+1KC3g2YiRs0xuvrEG2+619me9tpKDflDUMP5"
    "jmmaAOUOFobj8Smw3QTIpcGFTGEiy8uoVEwN/RoikdOMYPJiJF8jL20AJDkMjpo5TJ5Xh29ctAZEEaER+Jmav3ShK77f9ngjhKaiwB8k0HGsRnAlqIVBkjyc"
    "l5q9j0WaS7Or3bq1FSVAgwS8gjiFEMTeyBiYFD0n7wONi9zSBgbyfghwB8hKtqtKtoj6wwQzxciPi9Reu76+BU1FIlJNkRlZiKHjI3SCF/n8pb3hQpcixARC"
    "X+VFnJ5zn6Rh8Kkf0+Pul1E0Dmtw1tE4lXH0FFpmREvSJvZY+rdgOrKtrWXLS+OSQnkp8Ubgr/8rHAek89HHKuWmEwXd/Ojzr0ipqxjEX4A5/CH7m5LwUV+W"
    "kVnpVVZVw6oQFiuDjkvxHdrvYPfYtmMupBtGaWEnQEZCn+nGoNFAQhwQITRHgjwwPEQJxfKkfkQCENlhvJ39QT+530YZmDbslr5Hb7lQBBHHMa7q9XXcuoTL"
    "C+D1HpwPlBOPZUVJswlID0BZq/IxxIlsL1Oz0sv9VI81gy6nsJ7RIrOnYlD3u3K+gvsdICtTDw8J87sNxPWZf9DPqTZSpaedhHca2tKl5RjLWOjD50BX0vsz"
    "QBP3UoslaUXbsacUsXMnNObQKCHpUX/MlQ3kh7b/A+Th2EOo610Fe8htC/Zw+/m/BzU8SB7dTdZm7Hq9LhsZ/sPZpV7vg1VCZu54fdr5Fv+Q+m/Jf9FYcStE"
    "ahbwSCaSwL8PdCvv+lsdaXODc3FWmkXamVRPn816JjXlTTiG3xh+I/iNkLXFJwBONbSgtr3c+G5bevfLmLYZwecKNrMcB8qZBO44+AA7dSvRMejNLuBPe/gb"
    "8G8PfoEG4HDGYWQzt+aQc76HvwH/QsoccnI44zCsLtRxX9Z5v1DnfbvOEdZ5X9Z5v1DnfbvOGYIO3LaSCTPEMCINkgXDr6o8AR/FRI7+I6/NdGtr2rq878+2"
    "tibwfzx/Q16lqV41yKKPwNZWnmg4pXBLhU9O4XeqVm/KMz6FmPLqYf3+VHXJammyXkvyFZRbmnBLE4gptzTxOfbjwh3v/8Ag9Qi56cL8/g+Nunw8fg8UQsvP"
    "LXQd6ecYdU4i6xTjhYsQZGsrxQ91xVJMopISxA5Sj61MsWkXG9pFTijTsC/ivxJLgN74aPdONi9BdSSvNr6HAYZF1stBAcpdSavsOCgmxYIiPMpRxMiCR0Er"
    "YmspgNAjEgPgm4om/piDGstBNFUjOXDzvceDl8jgB0wzHWGBqQ+Yw0QucrnOC4UZqduIhmhG9TFw/ZK1LpoAM036LwaHCh08UppnBqzNqF9eELdjrigDv9nZ"
    "bkvcKt2FO3YOC3tf3rUYUF+IWLXmFkO+ddGopU2Mq9dkposmupvfVsE5pZLINZGh+X1GjsjtoU+6yBL5xXwK0e/rOK5EJ/AUoUbDxa5qcC/uyfiFQhAUysb4"
    "GhFWnmVYmggnMouvNDc0iM+Q+EEJI/jLEwTEFlokBdwhozm6zw9Z8Oe+DxRDdEu+bOE53R3sRb0AVhapNAwFPVxneasRGmxMhPPEbdcimjEB5wUNWmKe3dn1"
    "9Qj6N0OROr4fEBe6IKIQmvotIJLX84j4i6GDI88TFYNP9FbNCPf3DE95nJDuXB+2bIjvTrRlbarLQgF+czcfABt5feZ4fpFIxZc/xAEYDsB/XMeuBWZ+DcoO"
    "0pDVfXunc/fOnXt372whavPjnR9/+mnn9p2fKNS90717+05nh9M6d+91Oj/JwE733t17QO00Y6+eetf57u6PHv3c5p8u/3TgB06E00rotBI6rYR2K6FqJWxG"
    "1ErIrYTcSsithNQK/DFj/SHQ6LSP71QEfRDJgU17gSrNF+o9FJCeXbk14IyiweDQXsTQgXSxaQF2W0FmVM9p2gw87HLiNWF68DuDb7gOUDKDwhGlxTofp8U6"
    "b1PWEXGaaTVMClCITiCRRQgteVeY8K0qQKSpOANlY1OPDOsq4kITaaJ3KHHW1RWH57O0/fXM3erw9keCJJFfCR8AlyzTnYQrRyN+qY8P540Q4VoXrsIaemkI"
    "8ex2iXJDvOyWhlX3syLAhF7vuuARp5kPvwMmG7QlADTdiiqPs9kJkR5PxWEEWEGD5GiYI6Bc1ajt7PKQ3ldnl7MrYHC/bV0/ceJ6ckRObTOXcJ+uItg08Jfe"
    "5jFuji/AVgVRcedckNFThJutOX3OTeY0KSPsQVLTnUD2kQkjZRzIMJ4HdCiWWGGYAM0BuUUCSni5JtfXtVtk3xc5CBlWFfNcyehMRUdOdMDRAKqxYQhQbKJi"
    "qTvwPzOYrDB1GwgGDCspp7vhwkNQADeKfi+14+dO/rm8SyDeyc/xpuEgsf133m/vdXr5bnuv2elZJzpJloJ2DcP35ETnGrHkfSRytcMgqqczyQtB7kJJE4fO"
    "XTBOXP1GZLjOsOlIWfL25KP+DDcmrADTh/xMw7imzxu278AiP5TbHsEuxygeANyTjEQiI4FjIiGZrYjBUkwmIku0pWIbqp7SiurTuFdLZUWSFaPql1gzhVLV"
    "foqsCxWXqjiRWtIvtHS5VVitCkNYKiInQhJy5H/PZvVIss6TK8AZ6EuyRGR2a8fogbHKWqSUMC6UFuxcmSrkCdUqzjytOnhpfVusJp3VDmuOcUe+i0+TqwxV"
    "4gcbQRoWeA8WJz22uKKRHzZRTA9N8YX9yEdjfHGDLLS2Luo52uEGaIU/GILIuVw5lI5dyNai7NE4GXz6EGWhdW6mSUt2A/avzJmTtyWyWLw/Cqahu5cPUSLi"
    "EN9zDo/6E2LmnDJqJ18AjBIfakw9CQaj2iQpWic33mxgLJEkkgDuNcibjWLHU8UYYR4S3pGoUrG6wBLl2FFC8oEUekBh+R0LSFqYH/W/vADh/S4A7sMQHYKE"
    "n2co14DMd49QgGRas/bUqTzqruyRcRBj/GSSixhyRGaCc0/uirNk5TN21Rv2BeQYrhLAzHDx8Ik9mWJXM1JBkc7po0zaG/T28h4+S9mq2NYC36SfiROD7j7N"
    "dWAmVCptmIfz1mCWnhujT4Vwr9OFe83KDUdnirnot8fstrGdYYjaLXvyt9dpt4VSJ6Ib9wQdlzyJ0VHTEJFSOyxGdkWUcjCKBp/iMMMm3YjeXSBBitnRrvqe"
    "9d0bNbtiWMr2mlx07jmhXltMy/WZaXHCvR2tHxC2Ql70NwEeDruKd++fKRO4e06od55YWqwttn2+xC6w3FKusEFP2UsfBdnGSRjCAQpREXeIjgNp5tGA+ljX"
    "zWAA5eHn4kIciy+IaE6gxTN/gs/m+9NgEA5ZTRO1PL+gHuIAnzzOIYOUNXoKbaCTIGgHMtFL3FxJgFyoj2P54YkBoC9T0udpCykApyDGCWxanLVgkMs2A+7h"
    "S/+kRSdEZtyH8AhmhEzc3wK4aAPMl0BZvvRftgDYo4p/zStYad9fapR8Hw9XoTZCbDEBJk3XqF9iD3zIXoLBL8W+J574L7+uYewzRAyCHNU59SF9UETZrq+r"
    "JVZcGNPbOA8HRpU5vIiyfBN5XEqGuCDRyd7l2LoHduoF9EZC9rf+geq57tUjh0+HdJ8lBk1IuMDHCEK9UfCR0XGUP2FkHZY/qAeNpJ7AJgjqg2ZSJ1aI5kjM"
    "vPvsKqD15M3+8xevXxl1LyN/NkKOsgmO6+PGoD7Ap1DE/5PtmcCDOG8E8HXq14g+GGwPm1MPWkTKoDGG0MSrj4H8kb3w4NKeNoL6aZOZUxPoJH7PlcUjP6pH"
    "jbSeYnfPdv2uZ6nbnueon+D1M6tXZ9tdts0jeX5wxILC2PbGhQhEkPDhuhfsNotZizE6r9QTOYvhqAJQNcGBp/KIfA9IumYC4wpE5swkIIIRRKZ+4iZsdz3H"
    "LTkOcTsT6XamtsqbsgOGJ2rvxIQORSgHpLa9iPEdDtEi6dUIrmGP3jtDIjgh8AbP26PaExTueAJ3L/yJNF7xWAkSPBev/DfqyHzlKX9e7LNx2ofYPqFw3Gcy"
    "bYz9FYjKRaxkyt1MfeUkFD1kcp8RAyID+/AnhT4/ZgziuQedfaU6+7wotj8t6v9MoUMjo8oYKl0PWA3o38yoNNopjeGSZShPwgOe3DckPuP1n+pXr2Zc9OZw"
    "40w+xh+7TGo7QEztMik0TDP0XDesHhSwYQVWH/ozdyxQzwu37cHeg9pLrOQV/nkIux978WWvJi2Pn0uJowxxwSKwQx6BmKuMJ9HKrEDpHXPWM8qAYPKCfyDp"
    "ae0YBnAMAzhGiaWemcl2SS00qVDOeOEqZfCwQhxWKIcVVg8rL/c1XjKsyqz2sPI1hqWeQ8fbCRq3NkObNmFw93GNms1/iy08bvzP7uEve3oXN84OEyRE5kJ+"
    "XEBfemaTU9csmubSYfTpWxZfgprNGCfVMqoq37ABOAFt2Sa4o4o0O0Xr7EmjW5+WLbO/qCO74kWdBP76P9fCRtQAsq2Ryr+ZwJjM6eRT7X9AigpoLepQf8UW"
    "xvKa8qfe1TPM+gxzPTPKtfh8x53e3gGK41Srxh8k03fvWdmvuQP/uvAP+viuFpAQKP525G/XpmJ+5ubwfc1uMLOapqBGUKz2E6v9/WgYfgjGY9mJoHkX/kEX"
    "oSMBdySRHUlkR5LDnUK4a+JNB5+Rv3VJNx/u1FHDAHUprIhOMcIZ4TtXbUqLsyFvqkJxNbdGiEjKwLsyJwE2QAGwvnV3yAGehNe1GMbSgFWHodFPG37wiZD2"
    "1VfV03bq6VI9rlmoivKhKh8eoi7SIeom4ewv1isDbb4A1BLLyY8ufUDLsdEZzYWZqGaOALtWqbJkTacC3Je1JygEGjYMgHLP3/5S5wgEcS6RWULllfGmFR1D"
    "YaD1xESjNXRGUtIZWaEtUvbs8iS2LFVwKvoFG7PxGMWsKcqk0sMKMzaQu1BkZtgXIQnwaH5FZN+JEYlJ9lVN6sWeXLMo67h2Ui69tliEtEWJ0yum5LHY8W4u"
    "5d3Vc+Q9XZ6NbLTIylGVo3qexGbOE//KAXK9ovQdAksFHUM4/zHyYugDRRkT+oxQnpQ+MG5AnymaKaAPVJfmrh5KJJ00gOU3GYuR3wMx8o4Wogz1KnqFMFVB"
    "TdmvxPRrLD+7qjuR6g51cSY/uyjjKns71b1Fyow+u2hIAT8z1JOnD0w+l59dNa4NTRomzZG322p39tRQA9Fp2gPsNGc6NITQRIdOIXTuHfUO9cTYJUdOyalT"
    "8oxLLhbKvG3yF1SmLGWpmBSISHOpw5pL8ocilqTGdmrTST4Sh23WEmqT3g/muAd/70mVJUokdSDWJrpN4TarNLEmVEegjhIqEqEuUpf0le6KHdKD2hE/UgqW"
    "7nLhu1z2Ryr7E6oqlRWNngMG+I16RmyadRXL1e901zKC8wLSVvFgyYZZwQbOdBQpCzjwWbJ/I9WnEaEd5eQP3XRhmU2cznZYYRTHdeLDZm7UFhqhhZt6Uo/E"
    "zBiDMLwPxK5HXkGj1kINibFNtnGIq12fsVkc4mmjARwVP6w0joPmcFLoMlrBibdrFsap7OaUrOBIvjq161I7VqccEB82Yu2EEgW80Gh/Q0ckVqDRgckD+NKX"
    "OBDKG4w1RsR2sRbK8eV3M7ZjXZ/BTdcnsTKsnegYTrXsdzJ1Jj0GldTERWYUiUM9N/Udh8ufOJOKUyoSfFSJYM3iwwRVYyNcNvhEPXJ8DINPVCOnDEGDsqSU"
    "BQOoYUyZMIDKwIiSsDPYjBhbjtUQbkHXQ7JK3JSuDiWXBLepa4W4S32wj28GoxJWdhRMJNU2sSRI4K9LOpMIvtrwjyEhqmp2pSbkjpBqlgjeysDq9SD/K7Dq"
    "S7LS0EW1iQuuOY0GK5rE+xmh01hat8jyYPAp68ULLU4nX4JccBM276zp7UOq5qv9tLuzzHhmqbe9jaf6mQ5tvMUJnJVkGJ3ON4IN3kA7G0GG5gjT4YYelFGK"
    "BMSg0Sna2YhtD31TP7JM7UggF7o+FuNtoO5ho07FQIMECZ3gHwCzqJncRxsXOcoTQbYRKmVlsxO1xxE58nqY3KhKHkG9npiaSvAbcw2rKxHTRjmZKxm3BmmS"
    "ZSpyJobu4VphpywidXmXEIod1pINgA1Hpj4FIj/lX7JsQkwAGQpMqB8p0z/o+0TRpSyLvnDhavR9Vd/Xhr/qrJ0sP2t+607Bck4bhrwaQ3gbxWcrDiC6q07f"
    "qoOfQM9USBnLM+ZyEHEwtnIs+zipYx8nW4jYNqUX22p4yoDeISLYkuk98vO+eghh+1O1yFTQIQkkdZIB3S5gF665qUJs7HohTxv5dlzP+kO4Kyw+Hkrxw4Ux"
    "Mvy7SJvtg6yYCP8ucYfbVvWmaFUP0rdDtqA3RaN6kFeFlYEpKcG+GDX82Wo7YHmdPfpUHQRDOzZCVIBEnIK2eUK/XYlDBGZrj/XxyhiHKOz0YL2dnqy508dr"
    "7PSBvdNf3oABV+5neo5cR+rAfdwPF/0qCRK2hcQy7Ix74Osrqv2USPsxWjHW6npl/I+wvzHLI0hPAIr3gSgc3MdBw0+oDcs0UG6oUYuHNy69Gof8aozqSvar"
    "8QjC/GpMnS489NIj2cAfWK+8BVb0aCkTmZSFycpA1evxqPR6rA5w1eMx3h5f1zB22no4dooOlhZFRW8N6fWThbo15anNnPRi5bOllc/YyBsx4vBNkRhx+OJI"
    "jLigHzu23BBnLZ63eL3zFq1rVGVdYyrfg/nFqg/rsL5il/VVdhqN7qLD5ayvsJL1pZVCl/KsNFjZT242L3tXsO9kfXdKA3SrblL2qrAO/kymdw3oKdqZiw0J"
    "Hlkk+EqTc651WodQ9+zrtmunxZ6nNS+0JGsGm1bT88rEo2MktIxFHyKb6xAZXGQj1HpnRltPQ7xnh2aNzzDXuT/cjrV5SYAjQ7JOlu2hocztsDekd+bEV1MO"
    "QGXuNzHJW4kRJ4QRj+DqbVp2cQG3raeeub+zxnkdrreRsZWLuXTspW1Bd0nZKW/AESoHonAHWzug18WRi9dK45EzyDlDBUHMeSoVUhtzYrKJM4n7orFJiRec"
    "eSsRXiYYQpteYAjHxuJS+Y1E0YAt0xE7VX0j95Rkp/Pr6+x+G4D20KDAgBnUUC0rbnaur5NdtR10npQ4EAXwNVwPfE3XBF+TNcDXqY0uHHwFXV0yymRMNgE5"
    "3ZW09Q6R0EhSk/WiMtl8ACf9r9DNT0yXz5IShnO1MJwUFFZlH9q34uvrW+g75SlEKEGcSpoVDWsbahXLGzJ0I2IhqSDeUO6cN5JTac4Ha0apOXbj0zcuLjUH"
    "nWAqKTmgSyLAe1jI0Rb5Y6C2pz56d9qiKPyoJRhJS9GN8jttr1QABRlNXgz5P5ZzSRlKk1FG+Cinx1sgKi2kmapNtToPbrgoWrclpXXXvikq74aDJJ1l62yR"
    "fHYSVpkihXhAmFKLvgrSQS/VpJQC6KJg7rnvyLoW6KtlfJGCvUY/JgMhBZAe2SB9CtA8QivJfgYElJyJepfgMBDZoYGxE88zgSmD4CU5ENhOGRyHJmaiqRYH"
    "9A5cQ+hoas81hD6tYGw4YDqpANOSoIPB6W8YYEnKJS5IuXQA7/cjV8wFjTN7dRgpCVpxqIYulykmcGNgmXR+ZaAxBsCcaPw0JaWzAgjO1gPBwZogOFkDBI9t"
    "EPxinRNz97Y0sY7KXDvLj8uvAJ++4sgUT0iFQfUp4FCfLf7DDYemgh+xgocoOQ8vjNVztJbOH2fFg4WWOwLcN42G9o0z9wM4O6l1es5rc1IvhEYAytXmjVa7"
    "IyOmiEJYu3mKWc5s3jXFTFyW26k4w1xO3EScYj77JJzZoTIjpdHILUsOcO51j2FBm9ZJTkkstW4Z8yaAMES95ax+Bj/T+gSFOxAIDFtzjp1T7JwP/rB1ybGX"
    "FHupj2kB73IOdoGvOK442BLHCrZj/Q0jcXiMX3my+eTWQnwlSvXp5XBmTrM+2fKkK4aMPN0WeyaxxGHOq57N9Twbn9E4y6QKD0hwPVeG1MnMHtqiufCjeq3b"
    "GHv11p16Rq8gKiKAKHoJicx6JZjv34wp9HYlU+jubcuFA1z3lcAF4MWqJ9GA3KIUwYllWbsEV8hD1RBuY33b5pWaAighxi4xcnQGiI6gzKewDJZivPziaC3Y"
    "6FvfGkxJKJOoD211l1lCym2j5AjZPh6IbDP8pkkNtxfbXiTG0oO8Rg+hxrWW7thhilVY3YGINfwtGFjh+FlAyGE5YegnsFez+hhARFpHFmsCWxXDcwrPIXxJ"
    "4UsKX6LVh7Jx/4QLYmY8WRf+AIUT6wnb74DQnEJkeQ9ClxS6VH4EbLN8C0fmrKArRoOcIPtvwrzBdC/stVdY9pXOUgpkbGQerfG5nN6eI+WTQD6KFMW1Kp2w"
    "fE9oVcBL+lMDrQwNLN9JCoaBGV5M/1V+Br6RqYWOPIGQKXKRpraIlMgVNHq0HBotM+76IUrDU6zXwKBvMNZaquX7GmwNETbEZFMVmZpXC2PL3bjeLLiXzpa4"
    "lw5URjbohdAFud4ZMn38RPIByQzg4RW5wOy1BRt3Va4xle9W9hnfXhwVucWJzZW10BNIYdueITvXtIqlaJGsYfxcZ6yb6fLud+yjk/qBsvlK8EqHaiFZEvX6"
    "KPhpeRtFPi8KgRruG5k103wq5BRvis0G5jGOvqLD5Ai1LeBHmvzs9DArG+rs9jB3hYHPyJzwCNnc6TJTnbwWHW37R8lppwgkbyjVrSzl+KiMq3dBwb94zPO+"
    "vVPyocJsrR2brQVUAsCiaMmQYpzmil7p2xELExD7r52lVQRLBrYO6ArZTikazn+TKFelp2kYwlV0dXxMMOf4mD1DPkwuNEfmc4zBgjYjRLqOy3pvAhlTyInx"
    "loeT3nMOF3JhbME9Ve+xiSvkhpQq+9i9h258oRSkOraje88goqCad5aIavVNSKiQlOvNnehCGUh0BNt6FzKikA+iy4ItvWM7tlAC0iokPb7YsYUSmDYOrHUY"
    "RBxRyIfRJacuvVeBFVsoAWn243zvJKFwIRfEOk+evZcyopAPo50Hit6+iinkhPgKvmbvwIkulKFEi9v4hMOFXBhrM8J6D2REIZ+KtjkAvRdWZCE/JllIfe8t"
    "hwu5ILZ0a/YeJQtPirc+Tr7OVT3M8jD5sspVvXLwC7MWZ4BMAJqvXc5/J+/yi8XjxHGp6PbKOFV8bkb3S1w9utwVDAm+YF1hqke4WDx32yplMc29Wnsyh+Fp"
    "FIdAeO0fPHj1+MHbx73NzYXdkZdhNtoHKmkYpMNVs925e+/evW7njqwXcIyzEXOSpe/KMA/GFNFWri6nliWMMXKpX1ZGPY9hJAD156qqIHEzUriUK5xEWRad"
    "h8XtoOKXFnArP5lNpuUYVGoLVUFGkt1MOu4AZ7FtR3JZKeXdwUdvXokom46DQYhkrVuXneI0bCc8jAI9tXry3Xr0EhSmbzwdBW5UGJ9XRZSmjM4zUUNvA/jr"
    "t376kRO+qKOu/ZHqmBew3+gJ1q9KGUCzm2RufrMi9U+git3k7FMUxwCXdTu2h243UsrL6sjTcZDjAbJLn5Nk7YHiD6jorwQXKw7VMkhiHZjcfBePTm6+zSHK"
    "8W/hFOX6c9lZystx9unK+bfygOWFCPf4qMGpsFc+XbkdWnYq83KceyJz9VU8lrn5Lh7Q3HxXndPcDZdOrRqbFbXk9ObFmGVHOS/HLTnceSmq4rTnTrDi3OdO"
    "sAABcv3pwIFcflTDgrwYUw0b8mJMEVLk5nspyMgrIpdAkbwUtQyg5OW4AnDJ9WcFkMmdYAXAyZ1gGfjkdqgSDuWFCIl8vHIRgqqr2uAED60H+uSrcQLx5uc/"
    "9p8/evCiCj94M5pn0QCPjYsfjMMgHSRBru4lHeFeLjr6rYaB7WUpS0q+WnWzFnK5VcDCj4FkjM5xK7fuCElB8hy8SROYl5x91ovNKEk3xRWscs9iAkqbgJ1G"
    "63a9VKO3Xes0KxMWIrPryaXZMac7o7zWbd1h7t82e79BIbnFQvHgRmEYW2MhTJeAZhKrKbTj3JF/h0utYleYRc/Nd8Xy505w6UbIKyJX7Y28On7pRsgrIpfv"
    "LX1xV6R5Ffspd4Ii5xXbsxavZr6vrxlT9VQrFCtdqa1c6twJVq97XoyRMORhGYYUj7OBIZdfSaRxbUl89hWUQzYNB/j2o5M7nR9v/9jRWz6KI9oZO//QEX83"
    "HaGW4uvphUEyOYFl0gSJfSj+t9IO34mnoGddgwEZLp2A3Hz/Qw38Qw2Uz2xuh74F1VfHOFdfa9xx/+D+fw33Xywuq65k6xI19/HTr+fzHbx+/aoKhz9IECtY"
    "+5Z2b90zFMoo3Tf/3MXf8S6uuHX/Ey7K73UpFq42e8Plduifm+//7M239Gr7v3nNLBZPy/eIDeXNNfL6G8g6RS07N8Z/ImD9Gih6mljo/r+Uevjn8LuH/3ud"
    "8X8h2ve6fFzdQ2YO7M/fcGBfBGjDN/9mHO/fFpv7h0vxP4x8/Z/Arv6ho//9EZyfyxCzAOUMyHz29aTyywcHjx68qSKWoXqYi68CpVjAlsX430iQ/mtRoqXr"
    "tgKu5QMJ2nK9twvA7h9Ea00q61+IOT0rwwH3hBow8M6AgUGwBuaEwOxxkI3CovBb5pwRyEH60DuSPRFMKfhtGzmTS5VZy6MayPWn21KuvuSUvHOnpDwMnBIU"
    "mv50k9C0K9LYe4xyq2mUhzrmbSZKooi95yQHa8f8Egu2lqNjjgNR9dzYe5iIKlGG3qtElHihvUuOtAnb3lOOc7Hn3muOLdwQvZ85+jFq0OvIQcaRESvn6/gR"
    "xz8MoMc6MoxFecv1niWiPOu9dxzrVjAgmx/8/Xu4UMLz7xP/iow17qMBvl7R2rHyAIwGgBB+DLXTJrqDWs7mRhVNqqsWCm0mIN6Le8oUjOfBJ7vEwtoXAoqj"
    "0AlVuqTtWzkaQ0C/V1ZbaAlgL+9txuQzZNP3cQsmpxth6+EfB0/2j988eXv85MWTl09eHVBH0eFYj+3VmB1LHZH6QdAXe4COzIScA0pggWOYjPdR+KWGzi1v"
    "1XJtXwF68DjIA0zz0Bhx/ms4J3TldQq71Km06PFLsI8+nt6woJKL1hZI6SVGo0u5Nt/cypI0r9V0vfaSoRWqJlraWaB5mIXArOXR8Uwbx8WyMxLFcJbXcnnG"
    "lnLafXTziMrCmWWs8DA7qoeu/RbZ/dhLD4NG4wjdxjWgX9obmkBoC/ASVa9K9prZzHvqo2k/dIHGfgT09sq2trReTXYYHfU9zJZCM6jGZJKkvpU09EU5dTo6"
    "nfQg5Bo0CjxvmFxxXrOdA+NuG8B3NIHbks0+tYLpdDwnRRNPqD5I16Kmt9J+mNN0K0+oxTWb0/nJT6s6WDc2+TVjqa0eA2ym2clgHE0rjGv7O21bW0z6ZeoD"
    "zoGkQ2jp3Rc0SUmaZPApU3pejo6XSkRVr8QPUS2LLrt9qf6tjQkWffzhqKqqROeWlAY1stOjW7VsN0ZDOX6E3siViTCVqWiSKKENPVDZzunqPczrCexsb7HA"
    "iRobxTdZjw9w1AZ5tbFQXbROGioqcX2l/AOd5BYIlOcKb6Gmyg+kMh661r55qpP71hzLUbePSGevKsFbo0q7XDaKTvNas1O3HGYDnQlIy+NZGrCUm8gWYhJ8"
    "Ch/BznowHEZApLoXEnsAFBHtsWiXfCpTQBuljt1uoGm27ajvavyl0LnQAliqlyFqEUYt2ld4E7yCDYsbY/MkScabPrmx3YQpB5SQQh7UkEfxLLR8hMnWgUYY"
    "WmA5t/054zHw0Y0pfOCt5rSH9QK8dkBXYBriBW1rTe2ocBD6UWuQhnDLI58hnSbjIEadUyv0NMANM38Z5qNk+OzFwdNHs5NosD8dA75Aiz3e3lF2Bttanzwo"
    "thP8xXYG/gjbURYDI+ecNtnR4hBnIdtVibjptKJvApth3Ez6QzwhBn2pRfJ0kBmnhQK1eKZlJTOrkll93MCK8sZ6VVmeM0qjr6Fj8QS2J1aVt0IsG6BeJdot"
    "dStG/ko2G0sNJ7qA0bT35mcskMaod0jb66pGCEHukUIjn3/X7AXCgD9nZ9iOpy+FoTLEOMWbxQZ/RedcDoTN66PGoF/Vjwe5drP0m07KnsKVXQvUHIViKAAl"
    "0EG+c/SEjZrd+qAAQHMCoKrIYQhws+kPEV9ZLPRpOYGuD18mw9D/jBrZkr76kJQNUCOi4Gp0v5FanJlysns8CAaAJpOKsWHHmqUwpq2ivahHaKODAGnjjcFk"
    "Og6Z2FJ+e8/V4VD+ewGywZk9yyzOxePwNIDG9mXKsX+1WOiNYi9F9TCkU97SUMgiZUwq0Idxs3PUz3tXYY+wpqwf966iHt1u+S7eawaDixvdfp802I1eMgXz"
    "3dQ7gf39aSPSiKZxI1eeSTng4BQ6+yQeHtfQz1wuUtrVMeFeVB1uLjRai92FpT8S2CNuKFxkxjEVR8Vk7v6+r/qS9y2fHTjI3YCc7KGRicDGTONmaVypUrZf"
    "ug1OQqggJLuLx7U2eR7vl3sfoTdpaL7ZxHkWVu9wADG6a2Clbdjeme3jqpHdv3+fXAUC/rWX+bC70FLvgup0Fk/89U5by/nNy5eiMc2leSOEe7BxH41QVQCL"
    "QIFI+1M0eRA25iGnY/cW6H1Vbf+auRXtA3N9XXlYiKWxb07ecenEOFBVnhP7qKJJUueoktXwqERpsfUj9IUGxEzayI+M5VFnUAhv0uQLuY98woYbkJZEp7TB"
    "SUaWezcmdP9teovSlHlw+D/Y/BN7NX0npThy4aTqhVtdSILOPw1rahkUlZqiMih3zBey8PcGnZI35fY7TsiVsBvH+V6ht/JCPieuDAqhS/DFlk8/CQ7BqHqf"
    "FuW5K1KtK+AlLDG6XEOvl+h9Ay3aJPCTHRVQrOxLlA9GLCjvbtKW1TNoMwCU4n0PqoXquvWwGTO86lPCh16qbTijgzU/bGCTTfiDxjg555DHLusgOKc7ktzc"
    "EZgU3Q282hLoRtwM3W6gRmwCcCcCUAmtt4ttQ8Em5ggXCpVEFRCoxgMyyj0l/dIGGG/XwmbglVccEpKmuiXt7ZHWB+W9kNUH7oHSO1AbHisd66ziWAfFY53A"
    "sQ6AJkyagRqN1Rkx8otdETO/OEjA2orDE1OfZmi7FuE8TfxpfSpO/Qn8PfObs/ppo1uf1SfwNRXn6B5p5kEc0DytO4D+QGACgdYdjJ7Cbpz7kNQcUh7I0hhi"
    "BliFqbjwh/XT5rA+KYGmgEBTiqDprJ4dDgA4Nc7hY4wfc/hI8OMCPkYW3EqVKZvPX3v4/yfWx57Tkd9pDlaNmocKqCoPFZDK4hh//4tjdC+l8qXT7Gg7ZX8s"
    "qd8+0LlXuiTYPI/iDB4gxdjbIMJxA2lCNNs6i/mdCW4OB06F19dsBFVyfdarO042PsmYDO3ImLaGG5sws3zCiTCVmHI1s0KiugeQyMuNVKtn4b/lIhJhoMkr"
    "lSEDUmrukU6O3LvfSfQWWQ4fgw1poclhoVrLwFhyn0yVc1Y0Y0SDopCHyJCug8mU2L/C0feYOhc0+l5xKDnPCls99wQPuCKXJIE426JvOnlWHC60TiaWqoZL"
    "Hv0iO8YPjcNqfjcq8A5EvFhGgD+OMqBYmcqwnF7/ntTMctvrKNT1YzPZACNcVj++NgRpofbP36v2/UmS5KNC7X9+be2l3ZYzZznsqytXXqs/9yRCedNs2lfu"
    "s5sK8RTZRd7dVITH7eAHodn1m7NYWgGDY+zsFPROgQdbnzy1QxqbGhAUgYCCAA64YRBYZHUwkaiSK89qES6FEsaU12D5cS+Yvf4SpHE1fNtUhp4d+qPUbUmp"
    "L8rH8MrGuSqGS6u11o7oyR783L+5DG8IVeLZGiV4P6gS7xYLd5M7F5dkEkt3IuakKLeszI/ltSQT8QViijL3XYcOmkKPJZUUe0i7NvzcnvkFvSjLqjt/rep6"
    "oeo8jSbsisJxP8cwINJOZyXmj56Roib73+qnWPHWFlB0R7t532s0UumQusOPSPhudT/se81mhmeg0cgEzkt6fZ3dIs5Ieh+z1TJj6i0TZFmR/DkrTlcF8Om7"
    "V6rFCYzZUUjhArVZhRZYS9HHNfootycE0qIh8f+kM+Rb7b4z2YWuhE3Hb8QtYqFXWh8s4BDPY2pqgzqzgRbhNSbRYmEFGAeaY+1XrowzELNOxOdYswf0g8hR"
    "OJnm80KjkvUzG49dZj+uoOuTMUa3u8jmLD4dw+aIslfBqxqyydbqUMTIGtnY3+Dp4Upl5wAfjLiDDPYXxq4i7KTsfrReO69nOdrtT/AJGeF3ZlUvMqcB2O8L"
    "+2kRxlR8uY9cD8WxpldpsmL7bQTNDPJk8cTEa04MbbibZiZ2Oq439QIdDU8YnqkzVXEmaF8hy33pefEUC6iACUQ6tnALwFZ8J2zf83JTdSyL3p1+QM9HgXqU"
    "hgEoW4/5YUBzlSDIOwyA3oeNjcAluL7muPaRh6/MkZfiMXUY5EE9xp41Y2JVxK4RcbkugeUq4jBvcGv0tn4YQej6mr8z+AZohU3IucUtkRI4Duj59ioHWMf9"
    "VacC2wdIVo9LxFcsuWHICKOW8qMFAEeu9H5bV+ZYBk5pOFCdcAfAPVXdx2rUWyAhwjzze7WlADMX7dUAM8QM9ZgcFZpKcqeEwhbk0/SN2wz9R63YaJhMshs1"
    "ZU/EepitacSKnpGM4EYZPanGPgCpX/xh8/Vc6st/Ok6CfKdLZ1s4GQs014qcVfiX/0wyDD8aehro3cXHUhP66ZLfScXHVZ3g1j/e2PrPbp7VOJR8t1mvCCNR"
    "sogc4y/uGH9ZMUaS7NyU5X5zy/22opy8bWTBX/9dGDEW80U6gsrrgXWQx42gT3aA++OGfxufArMx9IMe/1I8a2LcDPCvMO/6mhXzgzM9X0U7/vq11N3ihxWT"
    "b71oih9u3vlunq/ZSPnY2RD5eHmfpCCBKOcpHRcny5Lj4uRZ77isVaRqlKE7ynDFKM/JYYPa9vG4ZM2f/L1HfvdOuy3fbx0+FEuySC7SUEqKqDdX8zgccQT6"
    "ifMHuZaxl/l3AclVJ8SWN1EcpWmADgwtMoUMYythDnSzig7CT6fZ9bXrmNNGZXNCNemulB7txjWU1vNYwhbxLiOlgrJ10KGaYjjprgoM6HGZI8VDY094Il3O"
    "Civ23OFrqUZ6VnucsxcKrLknG9Ad6FmdWZTsO1sjj+yRw+Uie/b/2XsX9raNXVH0ryi63d6iRcmS7aStFMVf4iRtuvJqHk1b1zuHlsYWW5pUScqx4vC/XwDz"
    "Hg5lOUnX3ufe/a3VWJwnBoPBYDAYgBCgRhFLuA9pl3ucZ+fPtJ32a/b3kqW0yTvcjWlrnpqFV87DD3BZScRykfE2OmU3h7PZf+DdIzm2+o9cxRah2No65AHs"
    "7q61ZicKxhFmGIaTnQiqJRh8w0nPeHr86RPZ1YGo9+lTRwb6UJ1mJALK8D5IBT9nnbZprP4kPU0IC8VRu4tXkDRz3fZxm2IBCVoa7oCEUxnskhMTBjcrFFWj"
    "DdKDFS5DfmIW1gNk6laL3KhP6OOUzOu479ytLf27H6XxORENrANmfLkh7BRdwMSghCpHQeorCfSRvp7AI5FNGWgAVjSQR+He/pHj+p3/6hz98aF3vH0QwI/Z"
    "cTf4Zmdsn3N0kNa7qR0uAS1lMTA9QUnPNHDSyPoIjmmi2r2hlhQLNEDgLZMjdgaTHVM4VKA+5WY8xYsCFZprXHPmLsNbIme6bkVg6zRsbT1nca/7cjZUoM5b"
    "6wPHqQqI6xH68lUTitDh6n2aoWl9axaVEUWKw8O1RFnd7lNpkpRJ6pUTl2wMa8Yw9EWv7kDUcEImo1JlK5nr1VGCPI9hJwCRMl4vZ2WfPrXFTWYbJR3Om/cG"
    "FIdAMSwVi4C3/OlTb6jWe9mfxyyP8ul89elTjakkJgkrNpQgIeP5mwxZtrZIPSQHi0mwuq2HJ5pgripOLtSPsEI0YwZTbebUts2lavlKYwb14fjlZh+lx8cT"
    "ofwSs0YaGw2UDH1sueXHMTX11O3GRhxtVA2UMngrNx5WsVyN2hN8ATAcDUCE3YDtAddLiduVwMmCSs/edod2X9MOr5i0+ydwhKuxynHaYUlYdNt97cg/D9Gr"
    "fxvJLe18k1GuKRZCPsgwIl/UJmZLWcU0wSw69JJljZx2g4WNXX6chxksjAgZtSV2XElCc/iTFnjWaGSMUjQBWhsJR0Vh45g61pTHtmWOEqO4kMWVqm78Gg80"
    "uBjM7tEaGKsObGlLh1lpUlA6wbM948aeJqWQ2tRo+7qxsU9f9Dkj0NUtsF3dgLFM1kNt8HQNt2gtcEnEfj2jNQUmOsPSEXaNsDf2TbVqSl1vooK00hHG4kRo"
    "5fUFD10k3uiKGivo++mwNY1S0vPRNmQElzGNn+X9CgD2NPvA8sOIAobTxUobV1mUt0f0McuWMFDxcYoaC/FbnJ/5Bx6Gz/BLIODnjG5P5GljZHzsWl971te+"
    "aoAlvAF+upepP4lmDT4hs74RWaTyGKmfLNJFfhdFxEFPJpdJtRm63xrXe6U4UI3ovr7qiHkbO3NJSlH7LFDfdkvaw2jbbtOZGvZ2URcPVlJLllZmgCIe75Xm"
    "+ED8xRdV4mmVPMWI07o6qtsX2NIuOU/Q2ovCtY5uDcPTOGHFCKS4aDYb2fHGbw3l3T1T8V75W02sQwIXf2ZlvaxC4cetaNlSytogOLLzzHpjEFzNGCxy5hYk"
    "X6ymQ16dj1a80jC5SOrKHNtOjV9NhfQmOQK+b1rzw1xRs1lKxmbqnM7TUByT5+IsfZlnZzn6Z2QyhWhJWWiW7Jw3Yo4t63ZDil6mHk3h0U32t7WlfgLkEQUi"
    "BXZd6RYfpTOrvQjas9qRUPGm5JdsLUIDN7pCg8Fb9XBsvA7+AtZgdkrjMru1qlI2r8tXUykrQ9dZcsHevno68bzEyA+w6KislC0KFAT+Gp+CYOitIPdLYS87"
    "m/0YpbPEKoxEK4qLB0QqnLQCCimuXlMv24THYXpxiuYpvDG6o2RbW0m/WAg1967VKiwBX5O2gJHYB6DJrpYqEjwBxfgHzRTx0NY/S7KTKCEzFCDtUhge4wOx"
    "khV4cxyoU715jpNHnogHKCwSsTKyxH0HLd7Bp9GZaXlfHpSjSHpOwRimL/L4DL2VRGmWrs6zZaEcmpRzODXNWFrG5ht/Cu3WbmsSWOZT9tJOhGNVUf7I8HiD"
    "qzchmruiv/eLVTqtXzGbOzeQ9XkM/E+/76HjD+AJ24GzWohBICsgYs4noWWgrkM9mJqRmTFOQWNQ4Z09QLeSO35dEQfrluYB71SRVwZe3KIWzswqBtbqdUyU"
    "ikrStDTB4w+ngqnWHNYJQkaoqQQehRZIb3DI/WEWA8071NgoS3118ZH7xCQwgxngsjL15hgkLsEVhBnm5al8MtrKVSuKreIzekAKXsZkS/MFbnAF65Se/di1"
    "gHXi0AK8LSqsbhI8V1rRACmFs48rzhFHLNTsdJSGgteN4iowHp9xzcXOf+GJfdTpoy5kfAIiyJ394CCE7+CbHa7lzxAA4zoxOhriWfDWreho91gcm6OjveNx"
    "NpmxKcicb189OczOFyDFpmUH9wV6KhaV2Ql8BeMyX5F0H6vgqB1JSlCnIH0wHNnbgS0ASsOrQkiCdL92QtpvKVMBE4KfQgjHxfcWRIrvhM5NHsIc0TyzBHLg"
    "bJOsP51H+SEM5H6JtkjxhLeMOyE9D38AX50jdLKBnR+HVyhejdIKn6vzNMNwC4TU6RL9VUjIuAuUhy+evcT1no9jKSqhPuc1iX+dDAQBs40/C5Qm4wmFqOR8"
    "IgscI+14klXriCxeQ2TVlIgB2FhDC2jugEFZnRbEFrqmYX4EJxqluNk3IVYRv/bXZ09/LMuFYClhBqcwlnbaPzx6gwf/WwNUmML++ugCkPw0LoBhAGNoI1to"
    "h8Zjd+dULKkNNzIEColcSHP0He4OlFEb6s6WBbemNRKCKydha8tnf/YY5D6uHRu1fnzz5mXrNRVuDVo5m7L4gmGYzRC4CgwCz+VuLMvYjGVphc/l6kQlEclf"
    "qEv0zgjNhdl4ajaemo2LoJGplpnUzw6rNb+WDmAibw39k7QQ894wUXximi4S7lrBhNFWB2exWa5ka0EhfedaOEzquDlQNcnzhlhshjw6gWPf/3jI7V3Y5PW4"
    "O1gJ9SJudUecoRZqIo6nICIQBOo8j2fsGbA53n09rWN3dy6SD6yvUbsE8WRnkURx2g5cjX1d0gmyfl0y8giZ5D4AywL6UEpW+rgGySITMprClkdGM3CrZDQ1"
    "WKe4HJ+WzEgam3+2NPZ/seglxSW5iQuTm0cJuaB6/rrTnsPWNNrZ+fDhQ//DXj/Lz3aG33///c7lvDxP2mE7Pj9D2pBKtQz6jcSxzrthZbRYGooIHpVQGbVh"
    "cMtGGg3/6R2QAiFB/vCVYPgcsUDMVdS4Z8vuG3k077tNgmubTMGKJT6h7AzC24Ghq3CPS0B7kX18cgq4ECsKivpFPsUHeHIxLD9jMZj3+SzGK0T4MefRxmW3"
    "6PrEPf7VgKQi/DAmF1IwlvZ/BqmlAIo4ZuKjYdvDRD8+hw7x4gfH1e2GdybcxUveTxmbFW8XqDvH6KREWXmAa0LomOKgcvTb+tKHBVmH1a17Zl+AMLH4EV1x"
    "jI8g4cfUxZsiK5cZWgcFmsoaE66zal7MwTFPdE/aeX2TieThvlNX2aSTXGlEx0iS2vUVn5KDgv+V3yOjAJI8qsR4Tp87z0z531CmzukNIiTzHyodK4s2kIY+"
    "5NHi9cRonBIOxN/RUhR54xZ5I4q8oSLQMsi3pamUSXXigfF7dIbF49RTXCYeGL+peJTGRVbm2WJlltepB+bHyFATpn2WwlEUzlKELvlhpAdm4dMkXvxGJemX"
    "TLHLoCeNkheinyrNKlZyiUY46+PfVoHzeHEeLehxgPytUy0UrYIQ1a8y8xADpItqsog9jDNkk7BqnxldOGn1UkgLvhVfoP0ArHl61q2kgMUXMr6TsonxKT8/"
    "n8X/wqK+5lLF5ibpWDstKzBQ2ryz80f/TxDpzjrffPrjINiJg3sDccJTJUg38seImvhjB0qf7QQYMZ5Pf3zwZvQorLFLfQPC+aZEoWKF5zbehG6T+3Y8XOYX"
    "TGgeAYanxFUfxhdxQa5J4EyKz3jIaaG+0fMdOqmhUauvS9O9W4wGnii68EMnKWFlkful57EMZL4tszdAKAtUTJTWxaduPUWvN/KrABHvtnmz5PhyQbMhOGkw"
    "4/ZTN7RT6ptPRh4gFtGUzb604fu+pjl6jYtbWYVnFMY175G29D+ujBKiTn2u+Ls3ol10iXFfFijEPbWTKlpH8usOt7Zu8ftdTVnWlZRTd+wW1g8NCEmEIOMp"
    "A8f0AF8jweYtLNxMpyjDcYG4LACXqVOr2EG5rov7iHBE+SZD1RJT1mtwchyvAXbCQlYtCUydKBeBs5Rq01G55OijVz150k5F+/syXnwV4wJfSJfb6VHeExZa"
    "Ed6whckkl+/AsruTZEzPMCbGU6is20l6WbCzi7dU6VF83CvC6O4gyNBNC3+qQU50onsDdA43icXbigQfmHE3LkmI9QxXg614p5Prt2FTalcgshN3O0VvGux0"
    "ILE7PIafAS+NCBERWG1kDFlvX4wf34nk9E4kvst9oOHU3xuirDdRHRbORMeBtABXSWiaiChD+aP4hV+YH3B/znS3+1Rzh4y7j40CFM3hkGX6oYIDq4b685iO"
    "HDKxnWl2vliiORzsa+Vjerdut8lhkzZhwpBJyJFPS6FlLJijGIYFYNn9sZ1yjC+cJg4QyFlCMfwQC5hjrdDMS3ZU6N/CfP45f+b07P6v73+5//TtIzX5RGzR"
    "SdFBbxz9S/R1YCeBbLB0kj7CVAK5wrQmk2mY4ibZGYYDOMhCdZkxFxncohUaoQyZhqFcQ3FG4hPM2wZRwJpBMl6rl0K50p+TW04AGXAYjWBhzoY2YrAOpflL"
    "iLYwk8JOq4GGuYR0EroFO78n0Pro5esnT188x9OuAbx6fMmxN82KzrwUTfVnGf2E5nrooRwIDyeUfHVC8Ty+3O9kffQ6+CoryUzh/mVcdKKwRFeKZLzjwIdD"
    "h+Zw0d8ST7I5BqzeCZ/YNyl/ZN9sB90Xyixn7DlHK1pM38NVzSY9ZnqdHOr9MIdD3rVjiPGIyPAxVIiHQncg6HTqKFW2YVelCPs8ikOO22KUhyex/E3yIjeK"
    "EizMZ8Ukw8qS5ZHjhLpJBCo9icIyTRo6KTtKDCtOWoKrC5ZjydF+/3ZItzVS0hLCcJaLFGEN1a4qxXO8UPihC0vha0FKdGFZoTs+aXj/2WMTouOplr4dKZJf"
    "FaBGf4hPEEL+nnl3m8js5RPgdrfQg8/A61H8UYIm1MwSPn+V1iPRb9Jq5PJVNIuX6oHFSnyK5xQRaUvup2cJm+Qi6VE64wnSU/thkk3/+hAXbBKJFEmEk0xL"
    "thb3ZjwI8QUycDUe6a3J6qXnwqG2NsUo8+CuzRzEc++7g3EAUk0sPu/F8NmDz9wpTntmcTAYwfZ4S+mY9LA+fSpgf0SH0/FBPunFI2zFuHV1AOyW27k0POZY"
    "75qY5qNFHhGh11oxHV0T+7xIAeekiJSiGiaJWL0eVGtOgVBwo0KetowsmBeB1l/xpkz8/m0Mm+Z22Yu3WVfmJpDCuvF2KVJ+0/5HcHPJwmQTR/NId9iaJD34"
    "+M2mvlL+sqmwlL/q1FiaXy5llvp3jUZL48Ml11L/bmA/fHQy2WAnv8rJDksaIh9qWOohGuMNSz1EY7xhaQ/RGXNYmkO0RhyW1hDtIYelOURrxH5WxgdpZPzf"
    "Po3VaWKFKTB5o47ZcKYZ8Wnie5JJvhqsZ5ky0WS8cALiTNewu71ItNG3YOl4UWdodvNOTo8hMjg1AatFXppOenvbeXdvu+jtbke9jHhlDr+LbtTN5KYdp3F5"
    "GJXncOp/lZ3X3XEDxDl6hQ8L9O6GNyfb9NgT3d1D1edZukxj1H6sbYRDxgXcDjDlYKfoUXM7HYAm6PJXpNqZV9SjTnY6ACnkxr002MnGyTYMa4r/cJCScIoR"
    "AKJkqrvM9auRHBipXGFdBqhIYejxdqfYzoOqqs7sSZV41w4wVomU0i/5r4skfK9+fZC/xPSfrNmH4WhBAbjSSXsKElIeL0AKwQcBk/5t7+6rkUkw7YkdeEEa"
    "D7kLgzRVsJnciadYkO7oxF5McaKAnGNrExVDMvZSqQXgjYfa/yAIC3Cg7Bh9wSYHQuh2KQ/GwOCNI3CBrv+KXjI2ayTdSXJvABU7Rkm19ybBTh50h8F2PkJJ"
    "eLq1laBDj96QDiZArNCiDArEG/z0CZtDR4xw3sYHecejzirBI6V1uEC3hQFdhuEnnltXSaC8UcdHCVSE0xK2Qs/6yNbD7qa7ezc/yHiRXX9HOT9qAJyqM0zC"
    "MzF0F1qTrbwhyXn69Kk9nWf5zJend4zmMgf926P+7u2xcXZYZB/gTKC1MK//XkY5m3Xm+MIZqE8VmnsKLalQrAstPYUyLATHCNQlkPdcmB0mv4C4oAH5FePX"
    "JRqj+tkEQHoZzuG/JfyXwX/CXej7tVVWUGUFVVZQZSWrfFhb5SNU+QhVPkKVj/I5PGpk2lNeDHaqOnphBAL4NSCbC01BvgZeu/wHT3kbWKu84fwBpSeADllf"
    "Z0ody58f1E+M2SGFK49UJXiJ/S6FXjnynObHjqIAvR7SLYm3Weqpiu1Li7OqUvxwOVapf9u8q5S/NpOnyuZRGYA2DkyXEU/TrIFJ9+jG0NS49BjD0hyXNcyw"
    "1OMyBukIUQ3S0z8zXR2xGxgO4tN/Zuq0MPMsMd98ys0avceSB8+IfuboBRbdfMpQNR30Ubu7HXdRXtjulNsoF4CAk4KAE6NQ04PkrFvAWSPVfb1O9JWTGJRt"
    "a690oz0dE2cbDi0VZXe9hvm72x0oDtsgL5Y2FCtlgdiwtHhjDf4mIDUB5Sm+R8WvAW9PjUKB2TwOUQTkphNbanIFFS0OP1ojD4nDu4z8JyIX0XGe//ALRRgA"
    "4gH7GLPcVEtcDJR7nqFypr8rpaCLPZ8AdOETgC7QHYNoSLo/udhVTk/2bPZLMxnDHpDDf3CGCSPUyYrUFaSuIHUFqSuLEzcfcy8GMpLehQwxezFUSVIKuthV"
    "SbsyaU8l7QU3PHgi8kTnir/h+6uhRISdvCuxYifvSRSZyTc4GkLvmgHZ47fSDSRY6QYmrHSJjuqRQ7UOJWmqvX8N1T7VInRo3CJIoXoDqt37ArJ9+m8lW5n6"
    "EVI/QupHSP34v8T8307M99cTs8GDn27OgxvDSlqsdqhodnfCrmGrcl7QXqU8SLVeH4c46jgJdCkoOsE3YjJgzGt63AvI5ZaNogASoWuj4DVCKIURgvdCVKmR"
    "Da9qjTBZl10bLYFN6P0GxH0TKr4Jud6MLKvqaVKL2ekw0VebM9H1ZCd5pdHLnjICuI4Yn/7PJMb//xIOJ47Dz5ALvVTy8zKaoYeDG4iEXyQI2vvoa2cfDcLX"
    "zh76D++U/95t8d+//1XVoc1ofPOtec7LzxDcNiarvRvT1dN/kK5kipLL/pfSvpDSXl5PaYZU9XCtpt9LVDxsocmipCp/Q5aklPMdFR1oGKA/b1MBjzbqk7xX"
    "UIAiill7UIwK1FEnFLEonOKfezq+0IFubFR0odzcLrBnF9g9tmn1WdKJSA2awH9TrhoNQpG6gtQVpJIC9H+SMvJ/nBbxv1n9d7FO/VdVD+21YVCyjM3+JLkm"
    "Nru8WhudJaGrMBqdQJpzghk9qqftje4noXnxOjpNQiUVjp4aH3ujV0noW8OjQ3/63uhlEhoDGz1M8Fk8LfbnjYu94bQP9dEOvW0oSsm2kF8yL8vsEPWok1vD"
    "imRIYeXKywkj5aAiXSuZszuGyLwc2n+ZRrvMyjQbVAbKuvwQKZz9vcSHKiwQAXPM/mn/SjBYTxAY7MmM9Gnb1XZMV+mEgbq9LRnRxNq3I9n4oT3rvYlx20WG"
    "s0zukmI8MVpn5lZ30YSzt8Fo2Ct3Cu3l1JDDo6CKu13Lt8g6+24Hbq+R901slMnGWRo46xCetZ6qWoptJe4xEa/bh9dmvG4bLg3DDYdk0rDVjqRRb2yc3o25"
    "FX3Xmpj02JqUUoaQHjd1jtKTx35/f/A1XgbYCwyfvnNw6P6XWS8RhrtWhzzgk+HF0CZAIFvlazHWvBVoEkkzolAMjlXIwe52OeKPbYyz66dP9lE2OBiORGWD"
    "+xyU284mM0KDZD1sNP1y/E1mHn+T8SSj5/poyiuWewzLXcZrRz+FE/RI2Ig+6bZ0a+sWO2IGK5HMAzFbw/OaXV4xw9omZpObZxMTFCc3MYtd+a8cNast9e9N"
    "rxGNylZb8navSTa4bhzG4iHZwB6HBOBmokHzQP8hhNMG8SQ54g/ZjgMNUk144HvoA72HPm94IGZtovb+mbOU073UCki3jgWa9mfnclVTgC4rQXiIyi7wQUpJ"
    "VvNhyS3lTbtzn2tfqolLkmoyXpNRTXOAsm1X5WOCTaKy9uIlG629RhCP36yq0sidj5wqOdzVJKAgvK7zv6X0Q5yGg1F/kHe4ISjyNzbQDFbeBJYKDludaFlM"
    "w8TN5aSZLDk7/Ay45G9srBnGqAlGZZ1XFcSh38zzpeUuvRGgY7Rqn0b0VpzP8MOktjdeN3ucfZoyiIAmyqd+PNXbuAwzT+rK3jZPCmqxG4Wsm9l2iZXMNLtz"
    "KjO+9/nMGis7j3K4OaBypewBeboJyKrlbgJgT63WFejXdo9Tc5rU8rVBmMkq7w30rpGYMriWqWuASxlbrnzkJcBIqhoJJPqlV2LJ642shehjKsPRXKsB8tCW"
    "DekN9T0Wd6738HlaHasJU1/jgVRsMB/1BvOg8QWyN8KDUM/Mo4VUzMyzhLZMLTD+iCl2jASfyEwVaxIzOVNQ2UJkVjuUfnoK8GO8b713CfvcAkEbWWIvZofU"
    "nJMuAa3WCWFyfDWRwILfIxHwAUiBgJf+MjULD0khZyUsG2GrodcvWWkIbfA+T64yA2b8Q5jrcNdx10tQjzWBH7JaCJRhwy3a2byUYhS6JeZRQ1I1QPStmxZx"
    "uZqwahYXi4y7mtyAjWBrin/gR63FUv+2eIqxiixyMJZCP+PKJA6z0SH6KGWXHXxYLIoY/VndO66hzvJsmc4OsRG0UpW1jeRaOdWZ62ZKmsSaDck0u4RbM8KH"
    "CGa1yHiJgT9rXbFptLL6wQQjr+bHkqGT6zwy68g0u4RbE7jMLPtg1uMpZq5eRnDKqx67t8DoX0SpyF9oen3s9WgsX0GkFsn+yM6BDucsZybxSh/8nOIOVQSj"
    "twu5Uknzwh8yyjVgTq6g+/rlqzUIzEPbWS5khmW9IUnyRpLch17Y6HAGot8U/Ci82rLwB/W64K38JZD3l9/T7TQ6ZziPwpt7HBUToWrilgkPjJScv5wRHm3P"
    "owXGuRLnp87t4W4I/wUq01BPwddLgMFKQbRKqNWJ09F22Z4GeNr703xZlMtzqhvFKhHG8QiJoywkRPi2VWRfxOwD+iwnzyZyACoV2C9WeS0eJkM12qd/MSs5"
    "bu3tFrH0Yw6WW05AWxnUNLX3fWMa5K7EsTP+IZGnUU6FLwXJdkpR4l2WJzNyiWAR8w9JANPfWLek2BZuE0mW/XW/7LxN8MOkfSoBC+DHRFksqGEwpE9c2dAy"
    "Lx0ys+EnGIK8kIHVJTLUGdup2/kxCcSr8f5tfE8e0h/9D/0Rv+gFuTai8IDiZHvAMmfZfTWrqQM9rZ+JB/+CxOqzrPOMXU8Sv1AtwK++yrRWhi6AX7qQ962y"
    "XLLihzqr6hVc0h9r0Zbih7V2JesRn04Ayk3fU9efQ6snzWorQHjQI5+ATybxIDsOu6FyBvdxsoNQedHnQ6LyufFykf8OkB1JNbYc8OXWlid19emTxoKdpW/W"
    "wlJhXs+C3L5uDQOxx4XSU78swekODitC6vpF72J/+e9+KNJg2rmNvAioHWPQidk9zabIgDfmJrvbSbktZIJt3YI8XMhBkluvHStJuPTCyEZS9MDwWqdRPk7R"
    "uwA0dLG1FdPPqFjAwLe2cp4R5aQPhgKTNJS5kzikLHxGKNhLbfmjRwAS32rD20B85KgpxQD5/vmLe8uYlbRzvuZiiBIs3q0XLOg0Jt6D7+yJF+9D49Glcysv"
    "evkSQYPzaK43k5FXpDwojHm5oCc2QCWQiQfxXKgTq11IXdjYL3Rd1Vqgt2uHiSk5Vz59R82nKCmlBUM03pGlHHYnxLpGFuaZOi3pqp/mEAX9OgPVQqg5YiHO"
    "WkhUO57FKKUoKgE2M6vqXQPlaJnrTyVz/a1krl8dmeu3TZf69/Wl7hVp9mWABVekudMo0uyiQEN7pRRx7O8953tYyx84+UJEEj1OlyfsYZzzhSy6fSr9wQh3"
    "NZ2e/Sl3b/3VMz+HVlGsa/X2dqG7cQtv/uWB4LjGV/FZtHvRwnlrbEpqDptMiU3m3JlflFOwCs760jWs7+8Npb3U4Sh/g9D0a8J/6zxKU/akziyR3xmExTBQ"
    "lagVeUIc/DUJFNC2OBiTZ5c3eZQWIlR572+02MF/V/TvxyD80yMzpnWZMf0CmfHPJKiq3+zlSoorL6f/fQNO38jZdbNtP0vmPCiucd3fvFx3f/uz+G5nXxYM"
    "/kne6+Gn61nm701zoLH/k8b+z6lj4wbLnN+aoVuZHvqV6aM54y7b8+qhXuTlPDvLo8U8nh7SkhRz8jHLzuUJD7mhcehM2GkpT7o5QSZfF2YLOYUnWVlCA7E8"
    "gNKy5QIG/CzMHbu+iAW+jXs7E+OKmjgc9McEpuR/NUio615YYJXihwFdSX80iCX+a6CipD8OPsj5JH4d4NdIGFrBFMVnaeeqCnlmoDx44yHpxempuPbTFzay"
    "Ldm4DIRFHeloWoPwFAq+QylzNKTfP5J4CR8ZtfrraCB+/Qa/PoiCc1GoCjT4MnKWNtjHRNW8sqSVqbwjZVaLyaJL9ShKp/0m553SuNPb3EgRDm+vpQKMymWg"
    "jOPpVgOe9IiGwTUNN2Wok1dHk1NPkVqw09ndVgSBdl0dSWI9g7jcYqnZWtdobRcWqGqha7WwK9wlpT3yvtotw2gSdzGie9xjFMCphogaHq4djWfadxTczaPz"
    "0YWuN8670gLNJBK0Eeua6dwDctSbsFrh32CYUc9M5/TCr+TcTY+2T5OLSV8rem2rRS3NNZ0mxE5pbOI1FQjsHFAGSPAzVOWcfWjEynTOvuR86HTBxdSc6Rxi"
    "ZmJSdKrkaXqSdB5nbQoRKv1UJiOT86wqVZA4UJ2vyaJc3fyTvWXV9xS9df28kSj/U9LpgRQP/+vdVvJ8Vf1s96MEsSjxCin/WiekNAknbqNf+fRpSDM/J19B"
    "7viCs9m/1mNT4/Gbz8Hj/fOTmEkxr6q+cfz5GLm6o3K6Xqocklg5aOryFcB/P2eROW/C47q4I+T7TrwBXnm9UvAps3YpdTqfeXMmWlad6EUpO9Cd4eIqpxbm"
    "rEFq1LGpz0f1NGOnp/E0Zo55OrdL/F5HzDULGnbG3GzLiTU4ZlCT+WqSP01uFAOnH8sW6yPLs1oM3wYAMOKUdG86sJsBEq85oIXDEnoJX9Hp8aN8umM2qOeA"
    "oEML7cB9Ntjf/W538P1t1NrDcQ8T2Yz7EupgPPiwv//dd3cGe9uxv8SuLpH7S+zpEqm/xP5xmG7H28P+4Pvd2/vf+QvdPg7j7fyaQnegr73h7b3vd7c7ANA2"
    "eh72l/wW+7yuue+gudv7d3a/3d/upNtpLwYsCKvdJzlqifG480/NC6Bsd/fbxnkByHf39na/WzMxqkjzzKgizVPT/+72d4Pv7mynTT3d1mXipq5wYr7d3xvu"
    "3qZZ6e/uf/vt4LvGqdF95s0zs7/7/WB/z54Z8Xph41VL5d1Ue+mpzl//yCfatgT/nkeqrzWeHtfALt0SoW15V1Fk9hvBX3sDbLWXsHxxQ5h5lesAlUZlfkjR"
    "hbAXWlmthnDxMODWUPRya+C9uSKebFUObnzbZNhu+bRz5io1bO/HMYwNzezxLUjfagNd/djYkXc+wsXe5/WhGnF7qDCEXzxtARN6EBUgUqznP2M0gZ8ILh8y"
    "YB0TxdLhc1d/gpgMLEF9pvC5fzyRHBLXP6TcNlJiqnIHqrg8F5K/tapiwe+goMtNq4pNHU35nOUxUPOPUX6epfG0MN44ps1yEg9WNDXtjkQMIKaEwkm5gQBU"
    "zOWtZjEXEmOTY+RmwyJsxTQSVE3dXGwq5tLcxbBcBPEondZNXeBkf2K8Po6nG4RAkf6xLqFbHvX72thG06STO3Gg6FVYriKc1CIX5W7YomLDCEX+aCkYmZV1"
    "ZHwiI9hoGgRGjND4AIN+jWTYESaCAjZEAatEAJRKRj2yL0UlirR7UzPisQrny0CK84fXhDMSkDacMkSIzbfpjJ3GKZu1RNNtNO/B+pUKbQjY/is74s66j61Y"
    "ctwGED0amjaBpgkVha+nAvhDfFsFptz4zAg8LlLED5wjtDdT9nRG3Txbns1TCl4JPagvM8cqjy7OE1VefZk5VvlizlhKZenXhHP2NNAwUbplCtdHA6civmDW"
    "mHSi/q1bkSl25ws2XSaRjRudqH8b0IgUZxBxGqtBqy8zx54Q1PxNM4reFOsvM8df/pU1G/Vkb1mrqdPsjOrCX/5l5yYRqRl4nKrY/LZzrVonCTAfWUV+GOkO"
    "McLZOOUEK37rVBur8YwXwx/i20E7Hvtfq2Lq08qzqmSLaArcm8qL3zrVKlnilRWwCJbyWTK+7VyrVpQs5tEbYIBUR32ZOVb5GVuUc1VefZk59fLvcuBnugJ9"
    "Wnn11a+r6E8rz0Ys7CbTONGVzAQnv7His6j4q1YZEz3lfI08Bv5r1sdvO9dX6xU7NSvBp5XX1JELrEyrl/K2EMWJVRu+7Vxfrd/dar/b9X5vrMhtwYyKZDZp"
    "51sVP8Q5I0MBqqW+zBx/eXwvSmocu6JK9pZtbmqKRm5uQ5DoKdfcyJ9ZnNZbwVRfSWc3467VxWamfK7nKr4AGY3Raxk97MQYbeId5Cwq5mgPxRel+DDSrcJn"
    "3HaKyp5JYzL5y55vPCjyicZfMsUqs8iS1VmW8jskKmuluCWa6z6OUGast8DT/aWbW3ubxmVRb4ySvWXtgf8Vp6ncUeSHkW6LHFm+mL8h/bCQOowEJ79e8TmP"
    "EKMrigQn355uEF9BxhPwqS8zx7MrZIcY1hnkUWNvUGn1UjZucyZ0ADGb3ceCHLe1ZG9Zqym86GGXwicdH7edVCtjV4+L+ERQpfitU+0dFE7sz6LFgnHxVX9a"
    "eVaVZcHyhzymaKw+jHTPQMgSHofRTikyShuyQIrOTlt2iYPY+nTq3xuM1ubbUHI/30WgQ2wzCrGtc/R5SqXhG5wCDj5wZonNRDiLwUml5I7Xo4K1y/bIKtC/"
    "iJIlm6RwEKJfAY/eNqbCU39hV4z2Vb3YXVPXdszirb63pvrT66vvr6n++trq5+t6X11ffV3vBVtTfcYvxry1RdGqsk9w/PxXCImNfutUD0mjjM1yY2XyBCff"
    "FtyBY2AYTaOqnVQrYx+oLoW/bA9NG3mx8UGR5+xv5+BaNBwkhhMj15H4P0qJ/yMT37UC90vocqm3cSetXsrm+UICoZcn3Jo7cAqUU1UGf4pi+DOoc/Vnoqz8"
    "oNLywy5/sjxfyOLiN5UWv+uFX6vtX32ZObYSgLYp2bz6og7UV+CvIiIOxHaKW8JTl4MhAshZieNbOMu0gCi8CS4kFlCYhiMWou1gbBb38RymHiNaO2+xSKIp"
    "hYyVg3XSaMhOWuBs3zpT47iW6ivZ2I54kRDXEj3l/AoWORwzgcZiJgR+ZYusayZwujUS/DoUWdX4pprGt7/iE6kKtao/MZSltTSv8kUCYHwTAMa3A0B6oYCm"
    "nxxe+ukr6QBqp9VL2bPDThM0G7iQ1c0EJ9+tiE+qgf28QiYk65pp9VJWCwmqeuVA5QcNVX4E3vL2cGupvpI2W8sUT8sUQ8tq3VGS3ZedVCtjH4LwXtVYxMY3"
    "9Wh8B35dmKxpJlBVMyG4Ro9Wa+SVu/q8OQ3NPreYcD3ZbvC5ny07+ZpB+TJ8jNPfRFDXbtHyFLuomeDkN1aUA3XSaJROGvQeo63FG6Fdd+9Y1MVEaflWyKdX"
    "4hpsxqbZjGF1rAqCRnsp9ertW1Lix+yHVDAPjItCM7nPG8IrGL5ttduuPwWP47HXJZ7pCMeH8yg/pBYw/OgYryjEQHjLb189OczOF3C+ScsOg/P6gnWYuq3Q"
    "wwYpUQxNeF94mycPIudCog9YKJ/AQC9fwIh32vLeqIe7Kzto93faoxIdRhdl3hmErDsMJOoKfSv0KL3ed+CTlFt1zx4sT09Z/gPLYOvIV9IHtsgVD2N3Bhvc"
    "sNl1Svv7Kz0htG/VsIz2D+L0XwMphCJFw7DJbBgmyL58ay4sbuIijfMi9UXhC65qp9SUR6sKU3y6bF9hiQslp1uQbPP4ZFmyUevNnLWMXlpp9gEAwHd0RUt5"
    "jJ+1ogLQyhC1OXzkZ0sURPpt+Y5OPIsXTz5ZMX9pdDKJP30abjDbtXplLemG/l3qLfr68c2irsKnMVo7jWZpMY/ZJjeq/3uBWrtAvaqAjK8qfXUaU2ifK/Nc"
    "yI7iY+kgEn+PtaYEL9dz2Lou5NQI95t2ZCCzsRStSqSTTTyByhtVjFwMe6FsBsNe05S8hT72dvk+GQdwlsL8sdHCJA9zijYk8ih6H6peKHpyEYjo7I8wyHbe"
    "B7rAGyYdZZ7uaHPutwVHN8nCrDJG2MRDKGx9MaWw9Y/SkE4NURkBSoD3W+fqQq8cgEtEdS74cAEQJBfaMLgmA3E2xKmWHrp4s5GkeW6ZYpz3I21hEiFCcV8s"
    "yCSfgHcmSK0dTQNxh/cRpvQH6J4/diwwaibREaq54Se3hocfmlPxyHjm+LibPygkxkeNCVj8q/ggmo6KNED3HmZvZieafOTdecHvzlP37jztLwtS2NIyfIu/"
    "OyLNLkXW0a9QX0qtGd9inHYhOXi75JR2qrSehiuxQNNQMcYOUqKY00zOKWmr7+uJhTnLXG1Opmc3w0VR8ygEq0T7EspNX0I5LTYefBPbzm9KD7mgh0jTQ65n"
    "KJcoydfSg1iHuaQH3hiRudWYd7pzOd0Rn+5cTLfyw1flLhaRI8CeZWBYXCS8YvicEa0NOrk3A3YT6VEtkVOEvlsWxadP4nOWRx/Q54tO4TgorAWfBIZVccgm"
    "iXK3itx03O0aDDhBp6s5WiT+gF11GHCoKMcnFkRH3OsGWacQj5A0NJUAnqBvGZB0yTyLWWBMDUfJ/PGyzupjrFFUO5o+02QqEq/dLjXwE9qzTaUTBh3VUExR"
    "bluzWFcDuf9qIK/EcJLp5OrtL3i9AF2OYnTnzV6JgzqdRig9l+niBC7Ti/DR30t6LAALEBUQ9aqRp4jTSkatv/2lXjnROU6daRVOAfRXbMGi8l0uUufhYRKd"
    "L95kj2ZnTKUuw2cx7sds5hSfVeEcGnkOJ0AQIx7HCUzBaBGK72fx4jxa2JnndiZenEa5yDsNrc8z8elr5sLKs6qtKiFYLTcyVTOOdkpMnuYwSPbkHDjvg7gk"
    "La7P+MooIO2vajU7AQjHJQqx+ASezUAS9nZ5ykA82rgbKu1rmsTWbMHfvV/py7kVXc2N2imcgdoVHo5f8ELu2VjWFYcmW+xUBml491vCabbmlyuiW+xSf3VR"
    "Zpfv0rl4B9OYJRdwdn2KB2NHlM3JHyRm2HKI8ntuCokUER0nEQ/78TnLlmVHi6TBFTp3xmDNjmiZoj03yJUDqDiWwgpIklF/qqXcSTsCZK3Os2WhQ+bmWVG8"
    "yOOzOD1oF8Arehl9tEftOJ0myxlrh1F/TtK09PtiStghn7cyjII+HJJSA1hjGvonSXbSAQDXFKrTWRnaL8tyOZXhFRnekCfyw4ye50GyJIXA1w8I87l4qY9n"
    "NUJj2oRGODSTaK+ro9y8tRU3C/VNLYX+6a2qpXOocleEPk3NvG93DOeXpstlw2eaMLFdnmC+DmMgPXFG+PAJHetf5weZCpKvQ7tFw6m7UTSo9dI3Oqi7UW7o"
    "sW+Uuc79cVMTjTWudVzc1GJzFZ+f4aZWrFLyVE8zWYhnCJbpbLPnUlfTppzHo25NnNg+JiDQCF/isfgRMunduVLaNOvMacbZk51w8/tbQwUEvj9G3wGDcXE3"
    "HeeTAiHg0LGjHJ3rM4xbk02i/mUvxUgz8GsFv1bICI0w8veek06n/+jl6ydPXzynI2pyd0AhykUTvYyayzEaTi9BeWZ1N0WHUvDjHsa7BHDLOF2SzAVpEzxV"
    "rHg0hv4lfV0G6tGEIRWDzLeNRRDAoJfhbwKRmDUZKqtqmMLuDoyuJrfiijfGe73FezVhgbHfxbcGKNvBL+gFHZdYaZEJmpyQWJ16FxnGiUyy6V/v4oLJd1Jy"
    "EUowndgRR3SZfMscABBTwZW2+Po5Ud6KedFhvQ04GuGTqzARdBQmko4ySUdT6fs4nFLL88mtvGNFOCk6QTCeT8qDW/PRXPS5RF40k8EbFuF5eIopZ0BJs6Oz"
    "44l4kXCKv2seW2u+WhFKOIotACqjUwym1FngH+g6AokzOujcgo0cO9jaOut2Q+rqqhiJwS1Gi4rS+kVtmHNeQwAUjPAHH/nVHCRVqApDrohkbs3wlZqJcUic"
    "qXAMfIGUuIw8y5ofUGbqgHKX0QvIJZxK3NWf62IpHDRT86B5CqtfFebrU8XAKHRBwBAsLpqDDHcbva4H4/yuaj7Hp0pAxwvATg7DxlvoFE+DKnrEFR5aAJFh"
    "mRVwLEAfuaMChJHsoJOhVnZJ9fgRMRiVdLLLtraW+JBKnhwrFbIC13356VPndLKEnfwa1Fwlkxm9BzVI8RwQULpIPXdrJqYv33NsY6644VReAiymXm+efO99"
    "DMtceqihs1R1xlIGGw7TvHwyHBjvmoDIY0sbh95UJc/ml/2IShInJmwn5XLlEkvDuu+k6hz4ILvsr55FyLXspDjtpn2UxXO+xcTTv8iEfTuHlXAkJ3sAa3rg"
    "vPaNPVFI2CRGTOJt0R9pm3gJVe5NirGlWTifojYC2QqMaJx10SZceHaIhJjAdbaKv0UUD11jL3DmK3bni6NHi0zYah/ObMkKQwXg/Kk9VMcKSI3Q9edTX+j6"
    "vH+WrBZzfFT86ZP6aB+0OQNVUjrxJFvB+5/8LINkAKekeYSnUZa32v/ZLbv/2W7NMlbQaYZdxkVZoProFIq2TqPzOFm1/rOb9/nP5yB1d9tw1Bmb8RJmnKMi"
    "SqfhPFyGsxA4JQIFJxpD+99/T0F6Zi+WJU45LBw3CcpkJHOUnXarHQTmQidNqVzod4sxnCK5iV15BOv+WFjUnbdH2YSnbLNuCtSjPuIwkvIdgGrZiCXrawkB"
    "z631d3s0NWvNzVpLM2dmt1eT9hBlgDqr9ZPPa31h5pzb/doyoZgn6riS1H4l3e8U/Xm0zUJcCqOoqhb2GQApCRkxzs3pVFDD2XRyhVGmIA+O/iPjKOg+qTpF"
    "c//TKWl4P8QpenS4v5zFmaj66ZNI/MBO/opLMwuOKafTCo+ctW7w7SpgrJIaiIuvdrVjHqDF7c4rViygLENzqU6bVJT8FqEtLnLo+kcdxL03QPUjasMlkHCx"
    "sPE9kNLe9YsknjIMG3E27eup6ci7ccIsqtSsVgA9jB8Jv/jeSGxOKz0V6dTnDsN54ZkG6nrHNG/lu42VgjpdoaVEs5sw7p/BfycoTYnUHFKBW6JqEq90SJgv"
    "/oajrfSZBqs929bJ/W9vB2P9cNR41jwQPiEKHpc5qr3wz8wXp0bFoa6IEZ/rFdFj3Wq6zrW38xD0ch1CPQ9mx3Jfd9DZOFAKnNNH/4Rn8N9JDeLd7ToiYRCX"
    "00bnKHoEBMr7qfSW+UH+EmM7cQaEQtiVcghOagnp1ItU3qh7kg7Js2RmfrIkWhTMSsqX/IXBrWFFtZXHGtXUs6m0l5Ct2QWub3oATWcL2TKsu0e6sGxbwxE6"
    "YwPI3CpunO+HLIE1G9QgqXTelZSXZHwZ1cHW1i0TAiuEn0BJqKvJUkrOQvSMy0mH9UwcBTtDtmdjrY4oELZ0aAwt7mCL0jDEp50FSkbra7wCPABmxUZGQtBP"
    "sw8debvxeiodr77hv+6X4SOVdn9qO2N9Om2MduE1nCFuKcToJMYXWGhsLbVZxFnpoR39El79o1j5CKDkPtcd/gDpKoQA/MbslE3RGQK6IkkXy7ITGIHXFkm0"
    "UqTCNxrDjeEMcJcqMkyybKHK4gcnKyP3UTpTC4ZfWYqv2TLnltWCcfD9C/o+iaZ/vUKn+6LZuHgJqSb9ziNKwoK4y+RZohzyFdkyR5eYCl6eQJbHbXa+KKXt"
    "0XsiPza7rwDCAK9nOT53lQkCUeQdT7hDo5sHGUsHJElCnr1kAMWodn8Oe95r6txVevnAH9ahjZAEsJW2NTRNBDSJ2k3iMzaLo0cJtyr+ko7PsSFPx3XS8na5"
    "Fr7XJVQ8/3LweDubA+n02wQjv292AROLoKzDIuQwe/EIt4L4UzZNv0t05iH1TzZpe89T5tUQ8YNRi/604qIVJTCe2aq14PXxmIQtD2XLdVxu2gVWF/1AI3BU"
    "a8lFibWwJezMXUM23rlWl9hwObadw1tTw9EtZsXwXiERrpHPY0QIR374SySILdpiPyILOY/JhtBlWAocmInkLH2EH/2TOOWOitFZEN+UnNE5/KFrcDOblwU1"
    "jmWzJenODndOZKMdg6UGKu+lwQQ7Nbbo0m61iJbkDI/m/9Z18+8hPukPVA+QS67n0WWnaV57LooGwXYN1FB3hnNQ62dif/5Hx0KmCDHH51+jODAx2ufCj5Uk"
    "J1nvAeYOwsuOv3QBSKlrM6S3/BvM547hKw1BkZCKziy2NyMwoAmUkhuMoqjWNXRUQxm3225JqausqmWPDiueRstjwWFUQr3NWmVjQ+YWPq1G4M2iViRCc8vn"
    "k4TeJm+KKKPKV8VVU7ubosutvx5j60o3Io3TJkWs4d070pEAiuLaigLGZouRUPAqwWAcuvGDjvQZbu/bSigrhQrC5ZDBaE2xymDGen6FlFu6McUEZjjDXsNQ"
    "66WRsXNbrvvEQM1+wiYuG/YHgmHx0xaBaSOUt6ARXj/AqYnA2xiFd493NDUhB0BRI5iGqrYZ3YzhWdJ8GV6LKbO8H1/W7nI91r6cR545GLCRZoJTCXlCHsVN"
    "ho3NPM1ow5C+6poFtYPOl0IdwjY3UtsuziJ1fsPZ4/JWuEZgENNmC2ZfbYPlUEvTkCsXNFM5Iwtz+5JaUZIE1TL6JUuWdUUHHozpH/7WulDlyqaCNQLdZBkL"
    "5cErpSg4VMqDlyrtoaM8eFJXu+0O9r+Tmqo0SlYFvcW2pev7IkMyRFmwf3paknsQ46JJmbF/J1572RVIc5xOVw/i9JCbDpfmfmDvrrJawAOXiaqk+nWiLMv2"
    "0RvhqmR2WX0JpsEkV7bckYUqbOifrKNGvfOx7chS2XCQP8vuBAPbqo1tR+ZKvfLzqS/oI5l/oKmLdMIgoosJZ1XC5zeSFA9XGEp3DH8v0WQ2BYm2PRLhXN4X"
    "6D8zzM2v+7NZjHa/0sDhPRpsi7QnpJ8vVz+rpmydDczo4ySLyjv7fE7vbMsDyvsPWf4X2epObpv3QAU9SGuP+KVQliUGbAwNTjVw/HMNVC/QZUkNIA7JbYDE"
    "dXwgWrJQsCkGnsNCzePp+uFjp9wb/vvz+JIfPV+xM1RByUCOTrrsRoZhwK5llzIOw3s0DOSoFDrJ99FsxhP2pdHX+ZJbb7/jvqoH/nTVnciHYx1/ZSZ1vgwg"
    "Y+rt2aCKpqIJ5jjxNM7NAq+KBBG323E35k8wJl5AlPWM43oaHdmRZXh6lHfL40mK1+DFhHFjHzg2srG8DmI7xdiL004KqwUGhDc2fuwUxrgkSmr+FMXoUnd0"
    "8STdtmdhrPavJnwLIcScXckymyiiAz3RKHSIe3/bqIbm1/HuCDTMtQnbZl0m14DbcFisH0wYTUwWZIR8vzER5neH+kZdRJbQ5N44vbCYwmEvB4KsCjRbWY9I"
    "Km9N2DYLh3iNpEkPkQEoIQLMiQDx5RAQH73VKrvsOLiKcD/+BXGKTQruUlVFdMG4HW+UgMRAEpwckomncGO6MhCAWwzvkmGXGmCgcVgpGlx0y4lP1Lrlf6TH"
    "40Zqu/kcVWgpn+WNQ9zbtscxNsesMWYOvQyqNSzWRZ+eNQdhgM+u07fefUvcfbUbad4zhpUfVOs3OClcr4XQ3AaO1kLZ3TueDKvG3as2WDX3ntGu68cRPJxM"
    "Fw+sC5iwxgAJldhyTWsdWAXxvUn/dmD5gRZUFwcltBQfo9Ur/Km4PKHND+7jORhSHgNlcaMj+o+u0m3Zw2cg5IgR2/kY2pP3tnq+CtF6IVsPwoZ+C+rXBNHs"
    "btiLFQ4jGGQkBhlpMwrWjcblUY7jzY+3iy4OOzrejqvqfdNgLGNA0WJhtljoFqm9gtoTMvyD6aT9xx9Hf/xx/Mcf/dEff+y0w49ccAcW9+hy0WkfWdnH7bB9"
    "Boezx1Dv6L/crBeU2u4+mMI+T35dOtB6H+q0g24bCvw4nex0Ogejd4fdoz92RsfBdrAjT2CqxrvDNrQfhD9g4XeHLx52gwNfqRcPsccgfIvlDkZ/9KFwN8Bf"
    "R51+N/jjODjwV+TN/wXVfHWaa/xiYwYG+uO0+8O0+3ba/WvabX8DeHk3nRy15UMyGLf8WcDvkyxlRftYnIf+9IriygBGnmvoIfGMG+p/+vTnlCe8yeGgicZl"
    "HalKT7MZm0D2KbDF5+SDwa1PRbBO8OmTjJqWZXTFJ2V8uRWIlSE/3y9TMkVUyn2zUOEUkq4b+PFNjkueSNAuOi7up/E5KcL5IxB6kEfPfGEA5CGiiFXVEU8X"
    "X7L1IkphJXxkz8WYrBcpcuJ2/ih2zsL2+3agkj5OkRZlKw4uDenml2mfXbKpeNlziwFvy7MPLXqDzB0QvMyzBctLPE3iPjRqHUYpmgRSm61SNjpqtbulNie5"
    "kpMwQsfzIfdwLr735DdxoxH6ng8XohdR5LZOkYXuHFe4savZ3drSv20fGeJ+Tftz3trqoR4lNp4v6qrccwaMrBN3h8GYSr6b8lfX0FpJtsKquL/mAB9kpH09"
    "yglszCKGGRnum8P79GlQS5NW69djfxqlLY1+s40W2t3WJkRZkQpSMNYNf0h/CwBqk2Es/O2LH9y5yKdPbCK9i8vf5HpcUiB/qVD8hXGss9QMTCDTSFsA/ODB"
    "Sqxia2ZSZWVeUUvTeZzM4NSkW7KM+KwACsrmMzbtxkt0J0BPk+ntKA2Cv8U3nxHkY7lX4Yss1StZpsoyhbRIoQdGSHomgFg0DtRDB7OoyU6iiyhOMFYdSCeV"
    "yULsDFVjRi9K+dyU5F5P8rwjGbVNzfexUS+SASScc6V41DeTZGQGoTAfettySBdPi3GtfWE74ELn9mFDSqvSbMqIdiG2AbcBMyBGYGDNQo6ODMZ5qwc/E4Sx"
    "Vh+/nzM2K97SE/sbNhfWSvZT3RhaePn6M+LhftWuz73t2lB8PdJAmphw+nDb9yL163U3vgHWFTzrkP6PgrbxrNgLatO15JK12VQzcW/ccnhTXBtdX0vnXwLF"
    "Jmi1wvU0sRe7kK/25mh0GrsJ8qzeboy56ztegy9X4DX6IQObwBaTXTytr1VYtXiOUPqr3cxW/mvZPcTHmlqMgl2XWZxQBI1lNsmQGELX0dbBwJT8Q2YdC6ws"
    "fa4og+vOB2rfvu6MoAveKtc8wBGXbTUx702Od3etMmtxTzAthK8FjIkLeCTcqcf87dbJErhV2foQFel/llAMZkdIwfLaQ+GVY0xcgKTydYw6vo0opJfyEfIZ"
    "oB8KCRVnHgcgm2rR9SGMQr0qmkcXrBWpAn1umsdPIhoE9aP4isDoRiVYKqURPihKD7gkmOXEB6T1VIdOwgKlSkz+8lFQu434VLK3CahKpLpjO5ibkqVFTDfk"
    "wFKARt9dEyZUtM41kBTl8WSArhC/fGR67bfQMR4OTxl+m+OB3iwP26aHMYImPv6iVWfDQ2sGATJYE19zRQ28sOQAgvxcKQWcOJUo+ArzcbpkTOOvwCYkX7RZ"
    "RYpv87pxE6NANTJX1hFP+YX7okCl83MglrqEY5namBvcgacFnTuyXWx7Nif0y1Rrwb81BuIdKJUW+EAz3z4P8msdM7n2tW24aHqSniZ418zIiwgnnlt4RcBd"
    "sX0NRuPrrHXCpmgA6l+3sneHD6rkuLAdxv1jUGZpi9eXPfVbbwFonmbD0EJ/miyaNcHs+NL6b0asA40C2h9J5GHMg0Xnq6P8GH1OrckOKg8p3jfE49ArxcmL"
    "aUucmeTcvk/72dHCnhGmrJDnZHyiXuv7x6h4DJXe8CJNvQcj2z97EXgbe5SWsKquaal2bMVbtbo09QMrYad8sDLaP8qOfeLU63rJ++lMcwaodhQdVyCNchlT"
    "y3HaEveraHorU107IbW2T6etQvn59dfiKpEzUlIFS3X0e3HHR+aOy5NiCvTJ3tM9ZWXJ4KawPbbOsGaz/fSQv50mZl28l2YFqhs0nTG1o7Ej6VsSvH1WVm0Y"
    "h+TrAWg8RPsPC/Y9Xq1HtkGPqX4TzlDRCD2ygMQa3oemmn+0M9lLVYV/mi8gDZKeXPEdazQIjWU2GoYm9xjthvaCHu05LeplMbnCXRuaM7feYejfRke7Tjue"
    "5Tk5skq42sqwIZeE5bWZkjU2lBHs7djOv44rTI4ccIu14K7XDl5f2I/W47AJCB9W1urSri37WRD4Ub+JTmnTKjcFS21x4XUFNgRoA/3Ksbyk/Nvrn4yuDbTp"
    "+nsu/RcT148Hmaj30Y1nR7rWLuR1pc0qlNljiZ7teKtIxyDCPFi9ffvk4aR0fAWpBn0cRmUSrwFojydCXH+/sBymvdcqFiNRbTu1lCcSJrqQBVBNfc1YvggG"
    "VFwJnIzQwQOIZ2WUaLtRplAmzTNDLBWnbwvXpJdq9piLr4rXkGC9ZDnPsfqQucoGlGLD18xBeJPK+sNCu7Qh4mhTG6aBNWXmKHtDr4imX7EozNQ2gH5CPLOv"
    "pnYKUzuvT+0UpnYOU6v8rBpljqbH4XIyp1mm/mYTdrS0T5Wz4Go2ybrdEHMmINxwvzVz1zFNYfiNzdHtDBUT98/zEI21yB2NfGiC7p/uJsEVnM2OZtLBdjbp"
    "9ZIQPZFkx2N2NBX0B71CmcmUw5CFmA2DagZAXUUjJIBUlAY5QuHnDJuG5opQDTKiQ6IP2BCrTqJKvI6BereouPec4buaH7UexqdktlmKk37RQu8a5DofHUu2"
    "kFZaM1bSG5d+6zBhUUq55K0GTk55CzYGsghorbJlDrR+mkecryxz1vowZymUQP8ceG4vpizFY4iwrnSZRVLlDB3TfCYtK0qNtfjF9adriRNfvEV14kSbm8i0"
    "uTHKSB955PM84VNo+S/e2kruTUyXRUCkSDqFSTogsh8lnHQynHLMhi4dp0o+0uFEwB3t5fAnwWahKQxjD23kwBK8CM5BDKSZ+7dg2HRUZNhHIeZcZKPJVGaa"
    "TBll4MzTjGjU3s1Q08dadKxK7uZ6vnq9HJkUNADI7/UKch00RVxl9hRkgLe5SIpCrDCZIzfBRzE3mo+Iz8fUmI8I/esz3lZlujHMCaSIDLjGOflQg9ISsIBD"
    "Ft0UDN4xuV1UnTYRg3H2WnPysfZG4VKSejJ9BBknCVtNKS0QdNiDwtp6Is/WkzlEmUwyudFM/ZQ2N+z3k2Acq62K+CXMg/R8GkjnZuSyx90uUliMsepsnAqb"
    "cuOsm+FZco7WoIIjo0GxDts2hwVmYtUxq/bjFC+BhM82w1JEOwY1sJV7sFW4izOa6HeVIbpajJAqkKoB6hBPw0DwBSeOkCxbcyJckRJjSowpsSKgvx13Np7N"
    "RPvk+dXn2YhUFHgrQxHSn2UzebpH42vtv+D9NIkX0kbufZKBkPkqy9BomasCZGWYF2UJ2Se9rymgCKdHUscDSLjiYdnpXdbor5B/PUpno7+qsfuMoKjv0+LR"
    "Ejn/X2RJlJZk6xSMIzKeDulFYEmyZVZJDqrKvlZ5YS2vwCUuH2CbmsDCHoLADpI9V5ppnc/7k9UhYO3Qm1fG54yCYRmwm/kfyGjbn0lv53Z3B0M1HdmCP+/o"
    "DU1HJeTlRlcrDVdAqnvpMOU9A4kDo8exN24WB6Ve8J2VnrMFK2PuHny4I12zoDJUe0JhKd5EqmfS/Sl6k38HssjjOI2LuVHyI8uz1wkg/n4pnA4Nahn4Qu/W"
    "gLvJsKV4Trv99+jT/gJfhAjbLfLXwN/+lcbDSrvSjDVWQ30jQ/cJ4q/9rnP9WE3sbzpjCOVjEtE6AX28i/IFfVVx8Yo7PnKgEN1KH0ocqK0t9RpZTfzWlrAK"
    "dLo33l4gNoCp0GSbuIDOXyObXyb0ctWHRG816uN+7WGmMXT7ZWbdq7V4XlojOKbqPbKp0+1L0HLZQMsmCg9QFVabBnJ95fThDMdptDqNZrDK66MWKBQtl+Eg"
    "HAZU+MWyvL70MBwEFbmchyT2mHydcv0vmTDKdqRJhgACPlNXoEDWrlxlkC9sJymfxDvoNTXdAfEGH+cuOsMwV9YelFDQ+5vKgFkDJ/17GybDDuB4XynMjTWq"
    "HVG4xhHHhrXjLRF14FreKQm0jP5iD/QjZnMHKYOgTk6KK9bCBSp2WTZzUmMtHgxGpSYrvaDPvJ01kZYqQa4IxEStAa0+0TteKIpVOkXvjr6mMOSici9ntK1/"
    "+1qcR0l9FSLFNI0In8gFFS9iX2gYE4g0yaEpHL42Nm6QvbvsWPA9foCP8cFoOvOQQbBup44CpX6I8IoFDuWQ+RJvZiiqAvpnh5P6ImGkBCxUWK6jAZ47sqMh"
    "/EFHrvhd7kAFTGE7hd6eFAqddeAdU9NSuE7MuMFqAAJ9hqW9HB9zUdhxM5HkMA+FRTdPSZHCEEoCgwZR1fulsjzjT4/oQtdgz9YVLq/PqyjG7xyE1C4zFhbx"
    "/EZe+fEse3mwnZJB9d2BMI4XnYy9WzT82Y4rtj0xezf5hHOw0gWQDUcTP9Do6P7ewJ10UzJVygBXMrWfk2t5Hm2AWn+PmiIeqNMUWrP0Gb4xQ8yjaytM8Dys"
    "jUxHwK0/R9IQx/aqbvWQb9QDHNvQgXnlYuVK2gjJl6mSCK4ELjjLd+4i6zuGnnnznlRDVKK/kzFMKRwf76WehQ2rVDmyMLaqEKklljlK9MNwd9U62QCPKJWH"
    "cszxmoKcGq9md871q485WKOGwbkD3mysih/RYNmBJfQORs6uwBpHrndFZ/D1Y7kthqTaX8jYuPEtyWOcfv4vRWu1+uCktIuLWb5Ql3aQ9LgkP4hHxdYW/OwM"
    "t/LggPXiUYwloRZnAVe8nGKpWnYf6Bf+j/gK7IC0j/+HiR+XI/HWkgVouCZjW9xC9hJIe/uYzBNoMbXKKp4MKv/ByIttTWfGESO2ODvGe0cXyY8u0MD7CvUE"
    "o/apaLQd8qBf1FrI7xLp8+7goDccDatAh+TQSLg3GRx08obBq3fzhnQeFtKbklnYWxJaKDA+8z18L4N4UqIqOZc7BdznnXgHn+/0JmwbRFPhdg6jn6Ta67vb"
    "cC8nvnp3Mgg+D79Aa/cGB2w0+MqYhkaHox5gWtEH4jkyFsLdwbiGOxbeorCL2v2WNQ1DIMCwkIYeml7zmwGPFW3AMYXc945SSRstk47x3ZJeSyq0ai/WoVhM"
    "QL3inUc/M04POnHf0BRNfgnlNyoBfglGTn554NUiHPwy+mv0zqrLDjxqBV4OTc3dU9e18qjpIsOzAQnZs8DQnBvInvWTTKGJvPAJnhkk+wTPSAqeA9SLRVz+"
    "LLkgmlqBzn/TXofzcq3XYZLU1LEHSL18xs6zfPWMe1xXxXB/5wqwwVpNVEW60vtGgF39kK8uLWLnxmlVKBtzFe6CrjlqMhLKXY44lWGcU36RYWtsH6ywv/vp"
    "TJn8TycJXnWoS0AMEzDFWMOYPJkGNZ1lbuosc4oBAt3wV4EizM8Ur4lMHfM8QO98k7liCfMJJoRGgStBR3NT9QiwdLtzx+OL6SeD9FpMYKIzxxgjAZEThViq"
    "JFljQLc64lDbKn1AaDX3mGv4n087f06FMrYDpxp8h5r3iQLRMIYMlXNl3YUODNB79M2hDTlmKlTvojIOhE5uBwqY4H4PQHZ01XfG2aGmlwooU2rCTGxqFtxp"
    "or6Ak02qaJA+JV/gjBOICM9FR8qFh0XiZPT2V5p9SHkSuXusI0GVR1dQldoaPCcA8zbKCDhWCmON0rD3xquNAY8rzJHf7WoZB5a/xHyKT4o9LlGkXyQsq9AJ"
    "6K/rT7WLRB/+v+5wej09IHox7Xd2Is/24tyrxypHJXPMkXk43JU119qAJTWHqZ2DWzSBfKPJAEbUV3cOg3ANa9INTWtbSb3NQ08ZbQ5UM6wRIHsNa0o1orWG"
    "NS46qiqU42hq1rWlWd+uRBM07MFAUx+eopt158FghYfWGnGbtK05i6OsQXZ710c1lZcFeCQQMQfqCGRRmRBIcoyYbT/9gMk1Wc8I9xhelVPX6ArwWXquseBw"
    "T8Gox1bwv8JiZGNfRbmAZdBEWFUWZvTeHat74aJvwoR3oGX1ntujuLhxj48SL8gdmHZpK4QGPR+pDUPI8FltKu/pw7J2v6c1SybjL3zox1tfvPRHPaGJH7w8"
    "P8oMoDDwdQ1j48SDxSlIbVO0kcg0gN6rxrGwwIhsFF6znx3TIU6CtrUlGqFRiLIW9uXie5zlmlluUOYfYPy2MCFvsrygAKsnJz/G3tVIPRYBqP3dXqzdLlmY"
    "xMfj0kdNGBPeoiLSfE3yqr7RbAxGr+cD5DPB8IhcXl7j3YBqRomC6cTIdDTHoUMP7FUxcY+AmzyUzvpStgK54hENM9hoxUG4kj7aUL8nn9jydYqvANBso2ah"
    "YY8JF26MC9dmHonDPDIL+iRkwnBJrE25BHH1IBLEo+i/2KrAOEu1RYbuLkzpa/0g19ClXEpfTJifA4lDmurw9XmQNB2SHR25Zz9XqvLmnbvbHQv/ckisbiC4"
    "lCLn4hHn76yjvIPu7XL7j10QjT1pw2Zp7JVxZkG5+r05YhaSNxMUuKv1VzKuntQzcpqUtbPiQYYxQU0z5J0i3AT9Egqzjut4OleWw8VE+o5VkZTKgzShl/PC"
    "T08clsGo1MEcucRUHHCb+FGpTNLsw1bELXcTvguaUo+YT92OuiYZ/WmEqzeiQmb25pnbZ3VysKWtnwxXQuUYzeOcs11o6X8SLj2YLprUy0RDzOBe0H6d8mty"
    "flFu+9I3zpRTPCU3HSCnYRQC2U0riqMJmPfpWpijYEkl17x2tlKaLbT3OsjlBEXeCSqclcYNpB05JT7+9Ik8GeHtw/0kuS+jNTrOJIXA6b6UsrfF3nDM7sE5"
    "p9cTT6K4BZCJx0peOcKpbtt/zeLZlH1bsHkzgatGBPWLz1KSaCeOWuy/JsOao8KUW92inaq8IySbRgyQXllqbs2HM/8OUFNKZSRBkQ6Fu5ktbDygk3JxH2Oo"
    "65wgtxYizIC3VgYFk+XVTYpVmG66G6Y7YGEUTffKa+WiZr0LchzUvTjPsCUF5faZ5dobUisqvFDo1PQdsb55jq19sbBFiXFcNxyMm6wGbSPDPERL4kmhzygb"
    "CeZxUAlBA5FSSQTTBFgCudRreY6TiCOBPozDa96jEkGZK5jZ7zs7jQjzwq9yg2oDGdQ635KJswNp7IRiZutVQ36ZE681Bda8emq6Nqoz17G8f21GQroWCaiY"
    "qn6zHphdI11MapLJMBCGwL/XQnHXWTsqtj3RGN6mMUY2HLVQrdtS9w/ogCLNWkmWnsEHemJAPxQwIuO5wFBqNi/4G2cQGjIzMgnC+/vU8J2hyvap5IFOGLl5"
    "qI0U1xc/6euLR8W6UK7aXdE5K+YvWa4e5OPL6WFFEVG14RJV6ovEhnplLYnfrHAYjRUmGhPJOoZZvUVfPyGryuyn1y+e15tU6brNuHgCJTAW5YxoJWHRBZsJ"
    "Irk1CG/Qa/WTbW++tmFBbD97rM65J11PmDyKYynMyDBer4h/IOJ08pCBmBTLsBnGteYFf3o7GeD+1UpN71XBFQ9IIjQComS3u2H0PtoQMY5yzR6V4C3r8Gl7"
    "1CdiFG4HanS6C1JauOX4EEWh6mcb/z885dDrmQO0q7Cl/5raAOMtL01Xj6mfOsjpN1PL1go2v2iFkaBKYJIdJr4CFHeBl0wjTETPzITY2NgYlW9KS4eTm1Ho"
    "c5IToEO6DsM+bw0C6aWm5DdKFzKOCZs70XZ5bgdtzOG/AM2weEoPk/AfaV1xjrFNxVKN4HhVUUxhdw7PKbqpuayjS57AtB0omqm+zGLgY0oqOocD2iMMDmo8"
    "AuJvZfUr/LspuTIS28IiQlGZmiHXRjWpC3s5ZLiWYFPjVGPvLhIuN9px/7Z9IjDGRBGdU3doJQ8RLdIdVsxjMIu4NBLzgahIltwOb3T6xJ+1DvGn6M3AXK2F"
    "y4n6udLvCLAZmQM/V5OezKriwtsUVrirWxVnGqqsk1dktkgIN0ajd6CGTfBBdrk7avV13aDF3TLwbfBDCyPIxDmb0Q7IqVNFeRTQHpQUwnoQDuDIhHPxC0Ms"
    "i1jqAJwauG+yEW6Hq9wAan7/+qUwA13ZMEeXoQQ+qBxqr080/GetOAocKQhEVubt+2ojUZu1kZjd2gJdntpQWmT2nEZUHbEuQMaK4rRwRnELyNmhrv7lPU14"
    "+G2SGX3r/FWgGoZJ8QF4eXdS8h/IcDkxa/K/FPsYtS1LrmTJlVGSCPyllNIs3qfohV1LL7qBNUTDFNHQo6sOoqinBxTsdDT8ZnoIBVc6YWUWtNKDiqw2CnxZ"
    "aKHtVqf0rXf+w5kVlwHIcs7soHmYmPPPQJlRfUOESZ5MNT0soJL79ZusvqCurS1Wi9iVOgYivQsTF6LDwmmxGhx8mXqeG6hV7VbmDarKZR6lRSIEM8/SbF7Y"
    "fy+jxIr9SFVksuQ9ciFY6YjEis0t8QmniwRVvr/OZSi1eG6HUst9IsjTUszgUykQFEZwObQwaxI4qJwtckBpW+TwbbBmPf7h1i7xpzbXv/nO9jRO2d7Ntran"
    "FNfN3cAIPAWddwuboSHh5wDHK24ImLtLMeVsixCoFtbrvzs+bOt19xqoCZpXreiq11Q0akQ35CdyxNEGjOSppAaOnrqMKEQ+c/Ag9hUg3hNLEZzF2SzSuYVB"
    "E3WwTK7DrjJOh6KzrOzE80BYbouEdB7spOqoCliIJ3MoRk/hoAMfgN5ryWsH4r3YWYdyX89r5iD1zUFam4O4PgekgeVugfb9K94pode8L8PHJHkzJjvkfSOj"
    "xHbMHKLTmx4HhOqlmCvVyyFbazkqvdTqt5zpDI7bKspCIC2aMNyotG1Vdu6Q+DzLzwFmM+kwS4AOzZS3F/pzIVohVZjxAiqlhtzUKbblJi4v6sV4TL2qsDeW"
    "J+fnbBbDGF/RuIS3UrXTRHp/4W/I1M6T8F8FC6fyl9iD5hq3q8jFrdQBLclihtPiI3zKeYR+L47QnQ3NYYqyOywt8Q6fvvEdqaPaZ6Y63za0oHciDG0r5C/a"
    "R1Nau/zCHruAY7n9JS7zYabDon8G/52otAjSIkiLIC2oKESq0md02nLe2nTTep4Cc9gLyCLRKkYTpsrkWGbMKS6lxGnUubpgeckuOZmMbg3CGVuU8zewyEdA"
    "JPTxLo9Lhl8ljOgZLC82oy8UWPhooV4lo8mT8qf9Wjg0/pEl0F1bHpnsZPUK3LSYJg/IUtvFHdHaHmnNnPvLMpOe04fipspwJlaPdYiNSxW+dDxKt3lnfvwC"
    "s57rNU2AmrAE/ThFBDoeT9JGgsk5weSKYPIawSRzxR9pKFNWdKZo+GH2G2bzvlCN8PFKfgDVkcog79fffodphnKX+M8K//kYhI2tCzA+p5PusNZN2p3sItk2"
    "odVxvB9yPa1vAiutjKOFrCb06HgsAjlxxCnrOTdEnNK9mUFqeWlx08dCbFqVQ2fZWlUsNHAzxY4Wc8EnwnP5S7Cj02Z2NBmiI5wh0sb+/rd73+/uA9l99/3+"
    "t99h9OFUch44IcTyt3FjxnZ28YJ0h2JSwu8M2VeCKDDJjjsu60VAexMKfB/Ou+jMOuOY6UXAdeYh/RuEInEOX5CBf9TrXVgs9AYsHcVjpuK9JOE0CKfdyV74"
    "BWmVeXn/CPB2HWPLiLFN1zK2xGBs00bGZnEvh139kMczwaqkxvVMzfeF+rVytqhLlfOe//o51YrmD3Ot1g8LwO7VJS2nDv/qL1PYHMldXqyfTddc62RmWKxG"
    "JmVrWrOaplUu1QyDd1ziWr3EtXqJa1UO+IQP4aU83z1rJmak5Dvffnvn+zvGWzQZ9np4h8/5Ee6iw3AX/rcH/xuE++Ft+N8d+N+38L99ShlCyi6k7IXfHkvS"
    "tw2H9oMwl9SS4zjorpekrwJjpKJInK+lokJuj5xCckUhREUj1kAZmfSmbtAJnI2tHc2zC4XmVX5gWE84BjINmhLePIjasgUZ2F3cImpnZG3DMscAeGvrRHFt"
    "4Y7RyA1CyFWqS/mQWy57yEMVSUo/ImU8pzx0R9o5t0Qx3WGQd9Fxjq+oUqCtnJ5PAX3lR7v44yP82DueMMral1m3ZdYdmfUt/sCs72TW96rBgcwbDmXmcFfl"
    "7slGh/tYjnJvy2aHd1Tutyr3O5X7vWx5dyBzd4ey5d1dlbvHc2N367JRNM3OF4CgB+iaGh/eLeYsRzpwZsQ+0dikJmfeq/NYRYZMjXncs6mIbWBTrvgh37HR"
    "JL9WLAz5xZt5+Ggu1vv9dZuXse2S7BqW4cD4n/N1rKTt8bWSK9tAcj0aUrvDsH9H/Bjgz6HoT34ebyTgrtsH7l+yQu4D+Pozo6cO9gyrdPvQptMlqp/OfdYG"
    "4Sue/CRViU/lMy25DxzOuW3UIvvQ2Q2/C8KX88lRf7h7O+zvDuGfPfhvf/9O2L+9i/98t3scPpxPbndfzqWZ8JP5ZHcQPp9Pro5+PR4NwqPfjkfD8Oin49Fu"
    "ePTz8WgvPPrX8Wg/PPrmeHQ7PPr9eHSnCh9wyFjauSriGbqltuV/82hQBeFHXvwdPwr/nUL9IHzME39Kwqv3STZ7mUQgcY9ezMP3RfyRPc1mxehH+jg7j4rR"
    "D/Nq8tscUPlWSVJIln/NzQcFv8wnnWGXG4v9DbL27SAAEejdfDLc+WUe/gmo4RTdwe1myI1Pn+LFp/WJXz0n1/gehNDWu7n93TMS3qGI9Iv+7rkJVB9PdbKA"
    "TDjWMsHflgRLQ+J64Dws4bBX4gFQvOpGnbEqQD+mLE74ryQ72+0wIN3ecBeoY7j7bTAuXa2KQUG9VMmzKK1DnZ3d27e1YP3rvKZxlFactC4mE3xGiWqRaYZc"
    "DRJ+xTsC4/s35/v3QDePM6ykGBBcUXaHYR4dC7XXd0p+ydEi7u7D+TjXR6bCXAsgKgk5vwiET5vhTjHO7+0fRJOX86O89113vzcEkufv8qMJTog4Umsxa7jT"
    "KXDyk0kvA1qaToZd/AuklIQgq9J/+K/8S+5974Qz+G8Bgu35ZDc8nQzDs/oCX2zPtpewM9dzznnOqp5zSjmOumFpHhtBIv+Pve3dnT0gWzi93ts9GIx6+P7+"
    "CAVMOFl0Ic/81UWeyHwp+Pd4TBHLQSxCcGHbuKDPeXhOn+rYcQR7h/zf8XglpNdTKiSY3KVk9JfXyWBn4QIo1i22vFAFLsJzT4HTaMrN+1S5VXiKXJtP6iX6"
    "Rdjf2op7PfGc3uQ7pcF2mOI6qXGm/N1akdjDm7Kzt304D+kf02QIdhF0EwwLcoEuP/BVg0qjcEztl89ePXrWny5P2FsYFusX07goshxZJtkU6W5/mls2P/2L"
    "mH1YZDm/SpbpqKjnLVjJupWfzYUlLUCA540N47EnWecKoRu1H5FedlpG6dkSGMSb7BAB/aUdLrn5WjG6YukFbI2jK25PxmOrwhBZgpfjMrmswjgFIeeRWO0y"
    "/fn8aI/tHVdhtizXZFch35Ffz6MZSLrfIPs/zaMzMhPiaf/njxT+V+L/FgBxjHZJLVQeLs8XrVNcOuPmfDiCjHX9i4gHYbpg073WxQuC7KH0NGGUEzhocS8B"
    "+e7DFsfF2C0ADe22FE6MFr65Kpew+euE/ydOp8lyxlp3QSg8z9J7BlTIY88jYPFB60on/1GeJe8fAy5IWmlNsLP9TmvQH4TGP8P+oBWMzVo0uMweG9TmWlw0"
    "a6iNnDegq++2lhdQgwkSeXvRqbVX63K3dQpVYOpAoMXqOxotrR6AetvqBAr0oHRr20aeOYIyQRAI7W+wzJsM7x1gK2uJZQaz0hEFQuwxaAX9/OzE6qV/2epO"
    "dB/9y1ov+dfoZWX3snJ7Ocm/zlh6a8dy8gUYc7ByDi2dx5dQMwkBR2HrFHoPah2qYidQ7MRTzKRg7Eo1ew7lz7H8yiElh+YTGsGbjFMsDavTssqY1Sv+4/+E"
    "9NwE2c06FXZlMM9/zW0rW80okS8Ck9+UQf5/nhkiGgx22MTvvg5TayajdXSOIGpCx/F1gAe5Q4QjfT1t9RFXRW31/TdR5TeaKtuyvfXTvn7SlcqGT7qUysae"
    "XNwDaunUU0sJYardzUhpZ6f16sfWNMtyQAKGXShWRcnOxy2SlKjZXoztYlMgyKCjJagvusCmKZSbaLYjoQwNsFoWsdEutgu0tI3crodkZdI/tjgzdkhOKNig"
    "u6nGp8A4qf3JBCm05dC02Yr63f+4uhy3YNCd1hAoDTguVAOUty4N0miJqB1G+8NN27/8aG40Znpre9Lq4Wh5971liCD0LgQAq/UA7K4BAPh7vW3EmGj64/qm"
    "926Cu83GxnHbQ+Sm7Ow65O5/MXJXHgQQEAKAa5B7ex1y3cFx3PaGoumPNk8RP8WuNautN1WmgRfXxMCJs8BobcnFbjBk4HTy8k9xcKUCdhYPMb+2Zmm4Tbgs"
    "Te4wwKpa1v45rufbO6jqRovXUu1QUKSB93L/VPI2gst7cTYPyqAN2Z4gmj4LLOIC7ixKhQk2MPbTQK2RYUMjxasfHhiAcZg2bXS3oVFo89FnN7rX3Ogzp9Gw"
    "9a3DPte2vH+Tlod3btL07eamH9aa3r3d0LavhR+i8/Oo3kR/t7bxOysRacwrPaylPpvoP5f86q000d9TASHS4XWkUm9195pWkRBv3ure9a0+24gC603v36Tp"
    "tSRYb/v29W0//DwalC0QLd6UBr3iM+VMufxqdik9U3qY5tSRdgW7P0+Epzttr2ZadcrrziQ7k7edh8v8gkn3ePye84Qx2NYYirSQSFar+rpLegkRNaBDlWWV"
    "Mw3yyJGKzkGzfTRaRPuiKnxgtEBBfPlbJmMANfDN21r09DdqGRXNQeQMz5IUE7lvv5QK6AmoiGxrvaCqwlPTYo4ev5H1nmEB6A2Xps0XRi1Vy4/SsLXAy75y"
    "znBz5dHRNLp0wGgAZm4Cwy8lr4XENvAatdQttv0OtqAo7hiYLUkY7y0zJwMOdahWepsnD6KCbTohTzM6MLec6n5EmFGzecW3ZZwU9boaJWE+dbNLDvmPUTqD"
    "gzI6hpvNRtchScIpa5H1q0U7+NZZQPiUR6MTvvWwqKhmgUa+2W7e8RnGa9msYyjq7bgKrfcDUzKQ33TG5NMQaVXvXT6G2b1aOqYdvwMBQ9sGk1Cv759xcwh/"
    "98paQnWuUpye8emveEKCr/2zyxuioVa/ESTryY8BmPMUyIEPr0JuCFLBH8U1zQt/M2fOinh/V4Uvyy+hir2vQBUWBDekir0vogqr5y+hir1/hirWwMcNVr4I"
    "RGnzch2UspwHUJFVg/WGFLz3RRT80+fSD4f+iygoim1BQNvgbjr8x/myKJfno5Zd/ToZhZskQvOyeE1acUsgqvLP3wLU05IvXO0rc7ZOk6gsWSrCsb84PQXg"
    "18ulJkh8YAiUr5114oQyeLXEBhFDmedQAEYbWmm4wR8N7W2KOQ2m00Djln5B+eYLmb1Oi9swtmyI7TL8PctaoPkjkOsEkEaQJdZqcLfdfgmwN5nr4uHGOPO3"
    "04g6pRH/OtgTvX8m8jy1N8LdGVrPop+Pz8CXrtuII44LKMVfTCjc6EcStVUhH+nKElVYsPo5QKoebwb1vj4JqIcLDewFwdBlLOhUcukC9zWYzP4/yWQsaOno"
    "66JxQwiNuo2TL4ix4cGId510xOtAuw6cXpfnPDhiuOeOokDHaCVFvsEaP2O4njz9LNJobKqJSNBBR1MNNQNrCtUISPJAMfM+BuAH3anYtPa9nX3WJrP/RZvM"
    "/gZscl+yyTVA738p0Pv/DUDfiLnv33BntPrNkezY/cvY1mCtR5CudO3Kpgd/eCtj3Bf5UeQp6UMUhbvkrd98Zo3K/5ZZVf4ZbjCV2qfD5vP36w3aFzVu0Ppv"
    "N279txu0/vuNW//9Bq0/WNm0vWEXvNpG/XwN2XL/K8mWn0+pny9b7t9MtrS5Hmx84sxpCUHyAdymDFs304gvKvQSRNEFj53XaSXstAxbOYaFCmGvXoStk6ws"
    "s/OwhXcWeKWdt+rSktsMQhsDvORNtiY9fZbkvP9vlJyfs0atDh6yN77WQNPpmlIHW7hepcNLeRQ6lIEC0P2vcPDVMtUNxBLhncLcEgzZrPW3+u1db6YYx5ec"
    "NZDYpY3N4Y99pKFwK2fXQKmc7px9DSXjq2j1z+gY14BHFPYlAFID14MoinmA5DnXgHkzVagPzq+rCf3TYvI5izYktzd5HKVnCa5prLVGxXafsk0FG09x+j6J"
    "8hXZ9Knby82PvQYw9WbWgPZAFrbh08n8yGvBeR7PFg50m8Imq66B6JkqYgKkUt0Z48b4nwEKr7gGkOeigAmGSKuRzeImS88AYrFuxaF6wFlqKklAcA3JyIcw"
    "/zzh/MlqZCOf1WBebZak89d/YK44KHKmREdV+NFzEX4/SW5mm/B6Hi2Y1n+p+s0yldKUiWKuvGHnlz5AlzP2OfBBtUawHvH8H8TjWxsuVB+dZfLtswMQCney"
    "2k2hMus2gkaFmwF7ZgB2UTqGJr6zBXfgtQY6LuHsCpOThmOFpjYsVDuBqOn05XIQHHC1y7ZnUTqHo0q0sa5NA+xppFnLJgo81L7iTPVaPbeGYf4G2gPuhtA6"
    "9a+F9KnwIFmHUuZU4dPS1mI+WiYsR/4llYXu7V7DQU2IuVx/2dBI82rC8lLbWqtjWpM0guvTuN4I1CY9axOYVvlGEA0N9c2vSjWg3mauuTJ1NN21C1Mnv/SA"
    "TsFUvhBup42NgKY6TRDzTB+4XEPvhfca6c+C2G1mI5DF7UADzCKXhTWw6RSn765vjmKngeaN1NHhePQOjodCB9Iv3B/2/vn94elX3R/2/h37w9OvtT/s/UP7"
    "w+uvSAL7/zwJvP5a+Nz/h/B56GjvDtG3G4/ItSl5cnPjvYeCW+kG1sj3vI4sZh6HrJzShZC7+nyIrwKvuwY0wDJqNXEjXuRFDv+E1vZ5eM3lxnWs3ABj7TWH"
    "uLKWRV6kpI1vRfBv2JLr3sMk3Rqcp7vTSl4CpSRzrURiz6hVtxF0u6aqZ+oOhf/boBm/Bsf/DOozajeRnrWlBM07jTChn7HTOGUveXzZmBUdE9rwiqEARjRD"
    "UaJHmy1oA2LdgNS78ssYfPGWYbKC0E5GrbvRXXntCvjcjiZlVYXLghmq2KsGC+qGnq3KHtp5M2etJD7Jo3xFcEH5wtAyQ/FVC6YhWialuEz5b+y9qoLwXWqL"
    "/A/z6AOFb7z27ogVcy7kyyrexYQgcW6EjwAApij5EK2KVlQUy3OAzlamFLItqChvk1urbJm3pCOu1kUctfj+JE/h3IK/zGqN4NHhtJWyKSsKQAkhvGkpmHgI"
    "r2aihesWgg8jsu5XRAZQ86CRVv5NMPxDE4I0+E1hXarE5QPyBnz9q5M4TdmMD1lVa7q2/JdD6CA0FM6+570X0zeFh8B380jWRc5Ts9HCYwlGE+cSSStKZyBu"
    "xcn5D9HyjLUAea2otZgDFGd5tJjHU3y3s1xYTh2ZDIun6k2YcejRra/l7I8Tk5yzNFmhX4nsw+hqA47zFG9UgYp0NS9KgSIpk+PlcXbhNF6ub92prDg5T+9P"
    "ObZPswt96HNziJ+b7Txlp+XnQ4G1G8DAq+YGODCrBsgrav6zIaHqDaDQfff/W923/zVuK4/+fv+KNLdfGoMJeQBLk7p8WGC327ILB9hHl8PlGEcQdx07tR12"
    "Kc3/fmdGD0uy84Ddc869Z0+JNRpJI1mWZkajmRm0UF6JmPNk/HRSoPAMQvJkPIMMyCkR8ZyO5p9OBy8/gxR+7j+DGp5ZIugN89Onk4OlZxCDxgczSMGsEiEv"
    "voaOFzPJuJlJxU0FEe/CLLyO2KPWh4ryM9hp8rDGr38SNr8o2ABmBeeyQV3NumfGM5+HfvaUQcJy9vhcA8weGYRpY3Lgp59gL8meMByy6Nwl87U/fh8O8uFT"
    "uiTL2t0a+WNyF/UZM+3+GZlaR6GyX9hT1ypVeBYpQ1a1Vpm5SAzwALM2sszgy7jO4DHiiaXKUHqHOYbWFMmirL5DmyhRGIZvcA9iWBh8HSmikjm0TDIfGAeb"
    "FoJi4JZ8KQni2zUMjb3FjEacc9Yts8UHXvHSt7VKlBV1zDWGV4Q4s2nkcU5347zHZEAYk140pnoVDzAawnL2cmVy7SpmGs1Z4/TU9mTRGeusyYfXDKeaqDFI"
    "2TjyA7aRMrwgV5gg1rgn4gxJPdRJ9QcDHmV70cpgtozqC1Fyjn5f5OtvjoPyCjKqNaNLzC6TpoXaUcMPKdFGUSbCzHoZf/8t4KV4vwplt05eruo0DRuLKTTo"
    "66mJLofxhm7qo4eewnS0NHRCVnB6EmxocZ1e4ykDhWs7+4JiUFZr1HCsXG3uGC9Qb054btXDXrtFonPpOBXvGWXFfT+KSgpwJSHFM9zoV3RCVma4IsgmY3S6"
    "mvHB5DdvsAtPqFTuewB9mSaTsa6EkyBuomR0M4DpkspqMm+p5VujwSyukUEZ1KrucMKA2pRwx/V8EJ5AiVF81uJntMgzHm1mXWraqmfW5zxAzxSs6qDDzikv"
    "Oo9X35any79fiXtY0lyhf43sUaqrEt2qmhlnHjSbioMOSgJXlPCJ8HiuSGta1PGohoETOcy+ihMpreL/fl7Epng5nmAGof9uruA2MY9gqL3FOjrLkqtXK8rO"
    "Wi2MljAMMVpcHYVZ/oTG9OKPau8r2prVznlmnp9j2IjFOs6AkVm+CjNRXffHQKs7id+W3PbMmflvudM90ve91cPPlCcSjywCeM8ZFBER/crTfvZa9cE4cPqc"
    "+uO9FGOmLHUM81rE3ABKi5Kz7qh8w9pgWbtjKa6IjyVTlvsWRM6ta+piH05fPn/MymuNAJSuqtvlETqExoCcGS/jdKi+pgKsrNWFxgA9pVVPLHRIdZ1AJX7M"
    "b0afCXT9bHUZBcU3b1YYRRR5XhtXchyOnMVBGL32s0+PGfUShUU1s47RBcoLqJ/QSruMmf/EkfrmdCzQ7Pxq8i0wFe98PGZ5FPfAvWVrM1mrp+gGdYL4B4p1"
    "hG6TMx1TN3euyF5uQGv/EVrEoH7ObHHinMwRKmzYq2h9z65fHp2KU0ApUfAaFrE6vJTCpSMuKl7N/hjouS6RNKTpu9EXPw5HviUWLNMDUU4ThpDLISDeY0uS"
    "saloMLNymww0/ZmkGBtT78CSO6pNW3VlGqm39rDqJkRmTpnO1/4X6EyW5PBp3T+dQqMaXab0x/51GIX4zVbgFVLmfLwy3SfSg/nTaVZVzKJXeUmvJlNlW9QB"
    "I8zyl0dn+fJclE2dXoU+KTHNM/X5qEMtWqS2gmIHnXO/99kTiaqsS6NOW3JwstV+OD48uxK+9q/I+foPmrbHwq6XkOuz+vKLH918y/6U6lu2T0Mo+LiOFSVm"
    "9g7eeTzw08GBtmp/Xf8qalzUw0wUudL2jkV9rCozs5f7yQg+nyxjAzHoZ93z/a/sZ2Wd83r6/vD5y6OrQBVTrynr5sG87s4tuHyfT96dfvtOU6VP6vX4Ln1i"
    "t6nkzH4/x3gWr8MY1vSv7KxW07weHn44v6IgGlejMB75X+b1ycad2Yt3FHPlGy05ZmWz9p87A6t6EzJxZtD+KuYms4Oy3uUJxFu1zXsTe29eHh1ehbLAFR2D"
    "zl1JqkuU3gqLfXhrZ1qstMeyfKUaTObPyNBZPz2jxPihHZlk4Repw22C9LKzNDUmszsYAAtzEk1uw/ixjelll24syfKnt6YKL9Uc9wF9Jg0FHtugVXypJtE+"
    "zg/Y/iSKUFB/ZItm6eWGNIqSQHy5b+HlP3pUrfLL9lIUeUIPRclHNtQ5eHpTnYNHNoZBnZ7eHIWEWqZB1EOTWefrcAwT7IjdsejpkkipKlO6U9kjlW34Wyhn"
    "z9ah6r1wH5QlziGthoNZyhPNAAeQxdo5WE6zURptq03L/Eer3zL90XJMG6Rz6MuSlGO3v5JsbK2CZpoWZYIpCIFBLS4RuFQspQ6eSYWspULbdsbymtxIRIGz"
    "cMBma0P/o01P8aQcGcfHKOpKGidexSxVI10UJYwKtZKeibH00q8h5C6tPmn8ksrmvqTQyC0G33iFgTG+prGilhlDjxppGdiIDD4Eg56Rmb1Y41TEZX1kvms/"
    "aU78+yiSg8aj33z1qPFqZhBplGhawVnmDVL+iFH6JhTQjDKzPXE4/HvvA4yZDFQPGe+HYc7I9cbXjF5lhZUnPU+bQUvXP/c0YGLsasE3WWGbwX9vif3KttFF"
    "FVZ6yshPGGYPvnIwyhX+N4blW1EhB+gsxBtG3258tPr+e8Pz1UTM+9LOc/sM/uyJyws/kxDHx2eKo5LR0Qla3OzVoY9ehZ/eFLFu+HT+LXp5Xtn0eWXT51/X"
    "y0c1Rb0c+bcvwihn6df2VFVUIkHllMhQOU/v9Vc1y0cgjL/RCMiKyqTInDIpMucrRuBrmqUR8NWZ29cOQVFTiZgiq0RNkfX0UfjalmkguBnh1w4Cr6VERqKM"
    "hg0SOPjpHX9Ka9TZlI2Z/9Wd5bWUmufgUvMc/PTOPqU16ixaq319Z3ktpeY5uNQ8Bz+9s09pjTqbz9GMLNt4rus7ZBuGtkMHPr2Tj2uHunfLYpYCL8N1X9nX"
    "9tSqrkSMlV+iy8p/+lB8Y0KE1c+RbuYZJf5g4ZnJ3mQQJnjpDZAXxUMkXB4sUOMl+xQosgavC+jr8/dBITzuAoeqbeRuQyeDdLnc5BiPVxxpE/oqMLWvB36+"
    "rC9bomwv9qP7TCpcsbSpY32Rsj8nLA7ueZauLbKy4KsKSycVqDg2zikW3S7CAvw2qzquEHVodMlImMXVPQ7gl0IMMsgo6nE+SA0SlBGW9IiizUiZWfhFKWcW"
    "vkivch5n4TgN8aiIX7xBKL7u0tnDXEOzV8CfMe4PQiu8aCYKND4XDQWgmI2phzNwHPXJYeB+QSx3naBRL0tkXipnayycLQDNYj7X0OkCOpQiLQb6W8imVn/N"
    "U4kn9BkrWNRvxFmy75P/RN9LpgmLbyeYfS9VMNfd0sHBWWn5mU3LrHOiR9FjvpVlaZJr4sR74GGJX6M7eimTc8G7t5RpP5FWYRy6wb74o3HEso0/stHGBLE2"
    "igLNP+g8fcByPxj+BxpCL2v/iYam/fok5gqLQf07D5fF5KZ2dUXVXx0cvjs/Pj46u7paWSnD6IqEnwfDwzsW57RD7U+yPBnxdD1ltyG8wLSOtsa5H0Y9WFnv"
    "yPqvx6a4TVW1/TmMB8ln+EL4Q1O2e7Vrfvnv907fvHrzsld7zQMTcG0MGkBkNajmfJgyBn2EMSF98AjNMEgp2rMrJm8vzb39wzOQ6kZhcF6oV71NzBkMDtGl"
    "ExpOxjx9PEZeAQEdDgjxOFFZuHKbG/Tp9KeWTVCslMpE46H/gnOd7VanTSB0C8R90xXps8LE29tq/4jw0XUIA0yeAbzvIwtykibwlX4JECzp2Y/CsRdHOoh/"
    "Yl4QYZRqMigZwLhFD1qQamQtssmYMxViNdMWYVwXyfNyRstjEDWkj0CMY+z0yfIDY3dzOMyUIax03EIZeIMs/4XRgit2Rw0k0N6H+XA/ZQPoGXzjGUf8bAIR"
    "tcQSxc5Dnt4/sEYKrWLMh1/Pjt+IxxivwAY4axvAbIS7Ifz0zG8LpkMqu9EMczY6JDDev3f4fs2ryuU4MO/isg8SRCMCUTL2Wv34p1z4LejHa2sSLYRXIKjI"
    "L+JLp8+a40k2bECFcmNA8aN4R6/DL/CKfjdeJV/q6G6p96eRQ1+19y5BWBpQ0Hfvlt45mgNxpsW7yTgg+cwdglgTYJ+ZE8DjkeMaLbflth2X6ekWDAbM09Br"
    "bz97tv3jdstNvWZnNYb5AD+pnDuSzUb5o661XBf7oud550P43M+HVPlh7J4PzWvN9bHw6VunmzGjuHFB7SNNbuvS7cJKcshLH/hAWnML4FtI7+FQ85nYctch"
    "pyW446aslAd4kQ6fIgzggjXd+EARtRf4DRiUKEl7oToqYoPed23JaFOhJg8lszfJE3EX7Lu2upbcUGgqqIxo5n3cOOTNsHhRM1hoiWYQrTC90iKDFUDhzkpE"
    "4LGwHsKbRt68/7n5I/7PoTKFRzss3hCDDzM4wogvhP/T+pwCbf6+eIGHM3rBUOgvAK7nzS+OcOhPMYKV+AO7+7DpB0mG1Tv9inrRITE6rdxD5X3jbIgs/rTo"
    "HyxXOBWBB8Mf+NyLt5Wh02VBGrUz8r802mwd1vt1GL22/ma5ACGvFGvvoqgkpgvzWpaaXfderoHNqpDUfXzlOOrGTOInDTQd+FBZU6CcPxVxiqTwQh+fCF7k"
    "fmf0hwPt6SjBxfThixEKgt5RIB+lTOi9UiBhGeDdKghe/kRxev7qUr1C6OXrijzRgGGIIIRNHwQmHa3JGcSXPprNFTiYHcMkbxi4A9h0wpjWT3nxjGt240kU"
    "CbrCETtgEYjOLQ64CqIk+ESf73UwpahT40lumc9gk1PO3XJdcZGPVX/neVpr0gEd0QlclUEqRzGoK+HM7s4Tuy5mAEryFvUaJs5fkZ3P6d3uv6FzwttGZZ0L"
    "O5jPGpoqaparekr3b5DbfYeu0VnFbOB/7vxownDgDOx8Fjqa2ZHKYC8/DxHRNac6v+SEWW6zJb7yqb7KkBNdxck1q7IMfZOqOhLfoCu+rsm4jwu9+UVo3wOp"
    "iBDYUM6iVSMwXmgBj9epzwL3PHAPA8fdC9Rest52SuG/zgGlWEc/SCYq92YNwJpJW18rTGudn56COHSevMM3IIbzLGh+cXO9od/n4d6buB/n4f7FcYEtBHll"
    "MJuEPUWCQP19Huq9gfpxHqogYDKe3XYsm56Mf5+DdC+RPs5BwuamxA/QlqjiMNAQ09jhoGA9kHucorzCHULTAFDXiGikCZuEGp1iBxLiyh1tMV9YJrjXvSGl"
    "Q5le0vVEUaLSEk0iyWbknew97gE0bz73g09oAEBi2nM/CwOS2k4AjPJdt9NqyYzCHJtAsGik94bCaWlHJ+Wic4hHtauBK/swiGQfkAXNfXqi+9i/sfubFGo5"
    "T6Eb3seEciYksT5PvpQGeJEnKLtotSZuH/dqFND9miwn8e04P0OpwkWcjseG/KnrneTiqVKkufdNkQZ4QimwyDUlpr38bRjn7W3uU+wCRYu224F/XfjXcjdB"
    "oNhyt+HfM/i3SRAUMjoA6brPLh2Qg7j40g8L11DCHVMM/CQgLJRq2tTouvpZNwAFaF37v/jlUhDnp0JLbGEz5Inr5IvcCjnvVYxhXSryhZss4XVIvlQZR696"
    "pzG2E2ikjz67Dkfj/L7h/P13g1/1ZXEuZX85ELg2YIAVlAK4pIL8tZRaiNeWQSgxZoqfNppbos+V+xpfPqBTpjcW78+Yg0sAMX9e08yyYtNmsQKqcodlmPim"
    "E1ynnt/nDA2fvXarjR//vh8MmZdG9EjS+D9i9bzU3C3GVojKfOYWrxpn9rNOe8vlV5xIuMh637XsGQCz9eIS5HX4k3kPU9f3uM+JBlSwtQXLl+MmGqjd3t4B"
    "UCRBm91nz2DoAw2FmnXcoQR1u1tbm5tdpy9XjNpEuhUjrbjjDhoU4kLl89CCDyFXiggNQ8pTrJni5ID/rp1Cb5Bd5JeoSocf6AlMBXgQKhWhftnoAmsxnTTq"
    "cbvu1uNO3fUdF5MdTG6q5CYmuyrZxWRbJm+w7I0qi0/1G1UWn+o3qiw+UQlRVdtMdoyqqKGiLJFR1DwWVCRFqmOkukZqU6Ym2OYEUCOexDYnXZXENhGFJwNI"
    "5XU3UPUAYMhziPYA25SALgE2FYBGJrgpMKj3wQ3HYIsWvJAWLRuNZrLCSbWFrZCx+XqlfzlSVOQfVq6nxOJQxBTiy8TKymIck5f1cp2p1XMq9DEUnRL3/cw4"
    "hpRrZsNiaOVS64p1UxbvXw3xxNKgSkQ5FhK71okZiI77eUhvGnYC92roCpaboLkBFUB88wIqNxqR0VEZJryrF9Dgmxp+Ab6x65dwq3oJtmqXYLNyAZ0UdTef"
    "ue1m0eqkqH3dzupqw9BR4MAkVI1PYBCqgfVqCvoDjdJWQWkQW7WrZmOz+gJu1V9kmA0gPKdrThVfn9OMGRtkcrq2psrT1YO56Su4mulc7aTgtL/u+/Gdn0nT"
    "2oXHgyZ66SCw7hR1yhPQU59AOWz+0WkyIs1217umfRTEkSTWD2y6BE2DiFmb9YmvcixYRNLM4eCWvU9FLROCo5bnGndy2ke9PJaPJqv8a0Jw++D2yK8Cf9Nz"
    "F9i++c4NbSXVpzBJ1SlMwk9h4DXGGbEojTrdGebe5+oKYdExTVJ1TJOWzmj6eC4Sea2CFwgagfOQiJObi+DStGeRXLk8vsnd71pOPwM874Hcafdi7lbb5Q6t"
    "IckfpKFaLOy53JEwcoqbI2lfFK2B1LYN3EMErAN6lpJ5++iEGmC+Zt9567h+M8QTbVjGfWklpqr3rU+JHNc2fOA8xUFReNMggQIYYC5YAFcqD4oYSISxpw6K"
    "2E9xf22NOUGDCb16UmnsUz040FAMjeDhOnSl2FtUxyVDZPS2r9ECBOR9hqdV2QWDgZZDd3E5VWg5oOU/mVXkUARLqIaIAVPtXrBVA38tJ0YN8O3BJCD3li5f"
    "L4GE03L5jqfqhXBtg0haJSXUKuz6kq5iZPpLTYJlX71489LUw+drJCxQ1nL0yhdwC6Lsi7w8FOlTdhPJXZ0vT2mRg7GCtZxM5MhFiIUmQCw/k0iA374r127k"
    "WfUHPC8MnrO/QpbyU8bDqALa9fYUmDxojpPIj3PvD1otxVUdVKSQEkUCXqQJ4HRsCOF1Negb1F6QUEUUjCL5iEud94Yn0RZBHbpvKZC+XRD0PkK/xKn1hg58"
    "La8KmoYghfKV3Do1bbv8426Zat3UH4STTMncQ5b7HuOJe7QsQN0tneYY8QeXLmqdABnlxINePOe/spIc1XyCAjxXE2EY7QpF7r4P8mUW+vE+RvfGE7ovLtWA"
    "WrtZODO7RSdw2Z8pdH01X4tXFZtPhPJzwNyPO1xwFAQz0esoiQt1O6lcpIZYvhGn4JQ5z6Jpx7wwNAGdA+4Vds+Cdw+8IwskPqZBJMD8BJ5PCHQwFQa1PFGe"
    "knAkj4YXrUsvVxr3U0z3udkA+/nn9vZKt/NsewfEc0x1VjqtzWfKCA1Ane5KZ2tLLi7pT+1WdzfupT+3Nzu7jfhvr9t+trnpwkMD0Dwv3W312s7Kyk53Z2e7"
    "9WwFx78Hpdrd3Ub4twe17xByCI21N9dTZ40/dtfTFWAhARkrTdfb7c5P0NbfmOnGsIGuhCi601CyAFg32OmFncrx2IMOdPUc5KqK3O5ma2ubsikqGo4hfIiv"
    "Ob/i+TSWqEwVtjF7Mi3tYDj4SILla9zH71IdY/sRt475LaqALmX5wOeqcUAZ8eAZWqJC7fR0gZHkXi8u1BtoJsi8NmefTPXiXDXiOgPiYUMSf9fxZ734ZcJm"
    "wjBuRBXSTXLbQ2JMVVG/6O4JLODKOiJ0U6dk4CCRCnXoEmpPacwBkgtXY/Jvn469jsK5LRZIjiVlW0JNQdpM+UZDKYs6Nk0za9FQKgQmJfvfDmeFrS4mlzZ1"
    "HPduiQK8bbPcPZSbXPPFPGvcDd3bYbm/UZJ82sshV4uKVszJ3cbM8dEsEBT6nIGoRIdFZkH9hZ6D93O5ZqpLlYtpnbezuNL5Lw8GMZIRb2nRCzNY23Km8zUf"
    "kK85SAYs8IdsAEyLxU489618O2dyzS+gEvdzkOXcQtDHFwcr9TYHkvipgDsI5P7SbX310gc11eUXRpUqhzNWntdrRsUYNYHgZuxZcTSC1Jkxboou7cMLJLjX"
    "3WpttYoMVaUX5wX0FGpW6JtoKIkifaaG+aUvQY89y9IKzbVcNtqzO3vvE2/y0m/kavFwrJOaIiThLlO6/KlDg3IYReE4Y5zfvcF9DY1SI24uuimTprWoANP+"
    "58e3k8hPy+y+X4lk8f04uQ8xBqx3w/ARLXwPhPkvAFN8D8JBuzXvb5MiywSS7yc6oSARAKStQjuCad0DNKWBiWpv2xN9HMu8bucJ098uuXDiiwKl+W+/71Gs"
    "ziuri3gjRfn25lMp10ouR/n25kLKbyzKS0W8G0U5nW9dYyq59c4y/nD4ZdzxXvMErIzjQDz9f2hznC5lcAxEEw/fR7Pj2NPMjXPN2LjyJaph6dXwUOMGvgm0"
    "Whd+A/FqETdgt6+vaMhJbLxBo3nc/GFs0b3aNsq6fBtb76ARtDBKTpVldJ90GKFTMnOWpl14AyggQ2h8oyiY04bVohTI1pOR56OM9PLInjH/CAh8dtT26u1W"
    "qy5SXS/EleMlOntRzlk+IoRs+ESg9u0CoK15WxbUXPp2ilwzg1cWDuQmMKQ0mlPf4owtHNvizH6NEDaChQ4PvLlMcRyVgf9vixR9ZX8OL/AqQQVjyON0s98b"
    "za1VErBPXll6fk9aJX+GrYHU3XiQXC0kFEJL0a40alRsmX4m7aHSUn7R4cyjCvGdy0WaVKjd1ZQHDHRKspB2dAi8RVYIEygwCHPr0Oyls0BwCIZhNEhZDHL6"
    "TJ5fx5nH9JsWbEUh1N5W8N2OOXyVzDM3qFYWFwWJtxUD4/THw1nM9Kicc0sBOvZ5O5oKGJYYxt8AKoNJAaxuOABgo7M7HvZGwz4d7H74/SPMfJhu8N8t/Hft"
    "TJl9AJVLHnzwSKkIKrqloeUrkvVR8psw97jyvAqSrJotv0/MbCsDlcjPwxy9BArNaSDBAjCMJICrf65ySo/YIATS+EEXv7ThZbjWvIqrFCYdPaesMNmibOHf"
    "1V5b/aCcq7qRGZmlUEber0Y+Rvb2Ip9AaAz1eK7ELLeQJyH0hRxJUHAkVQW8IOY5T+IAzXLLULwE9zcxKC7zfhNB8c7TCN55FL07C8n1DXJ3SjNMUCutmbo8"
    "aU2lw6wKXNRypPKFyPw+0QE5k2K194sJPyKjU++lCT0bJQmwAW8JintL8clsdXdwm0dGyDaynHEsXaBWOzLNm78xpn2UeU4g/Rj4d+zN0fGB9wH7eQR7qi0K"
    "fUlkhg26Z7Al3qJkhYHuQfJZ4uKvwlx49ZMHkco2sMQNlhCXTLG1TFw47IiUxmB1dZDJRG2JLBNKBYijeRHJR74IxwGlY+iVL566XjoUj2Qoq1xABxKDi7pH"
    "kZ7seqcyfeBnsGSrYm8TAcfwIt4XWclJyALGBUxMnrFbioXp3UuEM5ie3DiXTzPFhn5QIHkOqAD6JP4zUWDhjFYvA7zujpWPccxB7CgQnukII6uCeyvTLH2n"
    "cvVDLeqsMESM1DPfntJAADSle8aRbo8md6r33VaLXicM5jHsDF6H2zRj+gTyTxK8x9ppdQTslDuLAQg1PhmFMW4o1i3XLT2vAKOa4vXx27ND7+Ho8MV5r+W+"
    "fnVwcHTYa7unr17+ct7ruKfH53vnh5BzcHx09DtknOy96XVwCqsJ8IFpKdH9MOCwoXebiyc+DDIJrETXuy8Sm17Gq/mirvy2W0Qf7ozvY/FkzlgmwfTdFBM5"
    "k+CQb68qZyhzUAVS9n0+3xDcLjdjB/DjGo9QqLvCnIp2j/zRNayYqulfEpEBkMAfK/hLCX9D9/MU/FjCT4YJziMB/qsA32d4OqpynsscGVhD5byROecJma1x"
    "6AuChrH2FroEwRWQDfh0U+Y6A8r6UtzKxkloeAjQBhdNReeOr17uMYPbDDOzTbS7Vfw7Hd426bSywgcLCOpRGDDylDKVxI+je3V8vakBi35iA+aKMC4g1cvR"
    "lo1QWo82DQx7Qbqxc83yI8q+k6J7SybNnQKXjTeJ6hxhJUIy5Al9PeMQnIAVN+z/UJmqujaBcm0re6ZBTEooZ4Jfg7mb/wMnIOfbuwfePlOpb6rEqtelyIyq"
    "qt00wNgOqHp9m0awxqDypaey+0JrlSWTNOCGDiXI339nQqz2qzVkfpWGzF9OQ+Yvfyvfr7yVjwJkRvcLSVOWecbF/EJTJj4KJROHKyt4T59rGaqYoZ7+bnq1"
    "fT/+Ia9RvbX6Wr5Wb9Zd1hwBwwKDIDVfvgfDwnIfpFC/r5ryV1aKZ7IzXlmpS6m6roDNPDlKPrN0H1+Ssyt1aJlb8iYgVpUqAslFFNBnK9xcVph5ibcEcDX3"
    "0drCVwk8ai2QKB4qImT04Liplim4zpAQblUCWoapWKCRMIsoZEKVaa9QCwlOqsKInE35WlkVniWXjrVgmuNlggJFLo+IoxZHNHmPNCQ+XIDBeVc3dRM0/Qj0"
    "rn7CcNJ8ODL57EaOuqZ4HcaDAilCG/eCfD4Zc++7Nqk2hAYF1vNa5kAFaHkm/YkkN7Vfzl8f0bAcRiShOw9QstW/Tpn/afodxacU4yGmbjQ1XkdxceNhamp7"
    "HM1vBL5J5Tfip9D0HUH6q78ip3mTJiP8bqQTiYuwOZmEg0svLNxI2AOkz6iHKbTDyWDoHoFMxBuGMhu3MrwftrLSCC9yUT33LFZBOSqAU92QMeV2hHKtI8pD"
    "36CcXaKRXHyRispTSXs8LU1Vk3huDWEN4oNuy5AESluFu4lfkJb95PczJA3zfLFQwty8yC77GaxhsPok9HlDhfBh1+k0Wspq9V4BMwW7eo8vtifJBS9+CdVw"
    "O8REGh4KgJQ/VIYEOHwy9akN7XKSaLV0i2m5Nge4+y1sWyJW0mKaTAtyqiysKykSZmdoKKzaIqsu4APTXCa4vwazWcvwTjZcaas3p+nzZOzK5+dJnicjfXwQ"
    "DhzBnIFJxiw+hBYHSxOelKZM2fxz/lj9O8njV/csAjnwMSQunFDjYShJgUdOyLI0VllJCEpnG1gIhApVbjmnsuRxkFcXLDIqy53Dc3VBLecxY8udZRkjcgps"
    "rVU5gpaoNYxjlp7KqpNJrqX4yzDf2SM/0/MknWR2txH2mA7nk2tWNdsBjtYIGsRPg3Lzv8VJXkUCwp9ARrnVEmG4pPxpEgJlbRoAtLB52htfRZBGTpwDtY0y"
    "URc2qqmaVJIHQmbGBgZ9htZREFihoqwcILqKZq3hVZ+30eBJEt1XfhRFxhJNo1oXxGJsEZZ98WR9KOZ2RRyXvbwhzG4t1J128VOtRHDNkl2wTrdij11IlIv8"
    "8rIvrqqC5FIiPcTXgArLyj3VMocRZFbaz3xjQuURJO4cJD6QwAmtokhYSEF4BaISxZMTNi/N1ZxEgYpxENWYHE7l2l19hoUzQwpYiVGLVt6QuX6okrmE0rNW"
    "lOKOW3NpbMEGNT++HyUpa/4gmxlwK+WeoSiqrP5tXFQj5UWy0qjVf1jjw7H2Q/0HOvXOw3jCpj7xvvAq8KdgrJMm+nKgOyj4INKOSzFrfOTMjdHR5FWoKWMp"
    "WqJTaZnQ4CBnXiSC5fanFs+tCWYz5IVZHHcYcNeoSvpjjn61CKXLQjrIdOkAckAWwLrrhu4M9SKZZMXFnNU+A35fqTgcV18C00UmLR8bUTYKsRSZYJjUs5pi"
    "IZoEiA9GZjrT8CKTchC/d6TVlmm1ZVZtGXp+UFCVr8Y+nJaE++WExZlOBlEsBHlM+RoMUUYsyVlCRpSyvq1y4K8cm0qLW3uZ8A5X5xZF8I6Ez07d6YO8nFCh"
    "zY4Nr4q0f6Bns7SsK3IeTFxgccmztfSlOwdT+Gp0qytwpo2N/9P458Y/N5y/Gxf++l+Xaz2e3HU2wmaOEUqZs8t6saFWW2NKss+bqDDafcC/vfu4wRdCl4Md"
    "l99PzM37ibm8u9ZD3dfUeKcrK/I1/mw5JMki/I74FzaMYPV2l/A/POd2IdN3CBLBoe7mJI36pWuKqQMv+iIWk8bafxg6N5YbD9M3HuZljRR2HJDsuRcwusJQ"
    "m6NI2VWtCEcPTs+C0PoSNhiNryvu+bnyZh9685xqNjA5UEB9UiTkpMSICyWG/ekVKiv8CNRkN87h6zGpiLUZv5v3GlWnB/qmYKjEerX9hMYBFr1hMokGtWtS"
    "N0HNsEQFGOht1KyTpyUcQWWeN3MNSM2FNYWFNTUX1pRfLfKLlS/jOr2VlcX72ZukVifkei0bsyC8CWFjg7Z/cPkSptl8sQtR76VVcb16oxQ+f2u8elcUdpQ+"
    "pqivb07LxNltcBaDhZCga7yJGAC+YRqGRXi1AHa/lZWEf7NiKvGUa6tLnB559Mux4qSyNlfs2Jm1Y2fWjp2JHVtDEM6++c1ScawRF07A3ShwDHweaWRlxZeh"
    "SJDN4oMgMw18HqwD8UU0Dx2fgwz8gPzvID5/MvA5yKw/4d66qAMyocENZIyPRIg8HBT2E58uWpduEOAo8gBKCt4muFEFv2VLlYibtxJmoHH9fIPr5CXDoCPI"
    "KImEJBMa3HxJ8sqvdQGYXpRMuUOLVBUPSbzcW72UTJVKFbFzqJgWSkfPM8ckCse/8yHBJwkxcMYpkz6S6Oyd0G1gBZ5RySQe+8GnvSi8jXGN5sykCStj6WyO"
    "by2xUpsvT8Fo+XIz1y94i0TjE7SFJX/kkqJOauRCWvgXigT3ojOxVffzq5w6h6aKO/wp7Ye6iju/CA3+Mn0k2ZJPreMVNyaZz/RS8wxtj0386LEpGkH2CMdm"
    "KpTeua70Jhfx9R7nO84zTePfvIYXzq1CV1b4oSkMHBrD3OKJoZbtwJxJtbT0CmXg6IcJ8HHjN1p/kdzSkSimiajdFB+p/Bme9CCcTFZd/own0+Lxxk+dXl3c"
    "PzBrIXJkPa+r6hmwOAvze+AnDC0GS3HzQ+/t/Fa+HJffYqoD7W98woAHSYqfWj0LJpkgIKC72PRr4PyVJCNCwQeRNisJo9FLf3IreiJTek4J/1jsIaLAsQhe"
    "pSWMInch+0zI+OCJWIZ+lsH33XhAmwSEm8NznObD5BaW8GEYmOPza4QehtkNjktK2myYYwkaJ1xzJfzM4Vo4FE+gU/ePLyn8PmrkagaEMIf5+zf00NbFYfXy"
    "FxelkKxGoY+VhVBuENZB+MgC36znFAjYg6RRVR7MqIozNFLasPT+iUnP+yXo8dGvNgrDDPnf1K8k0bK6lvUf6/VrpuQzh6wwF8QqiO+PA0O/ZPTnE2q2yWC5"
    "3su8xLhV53uRdriMh+D0xrMGnTPrC1rIrUj4ciUSGryMLFyHpVpK4510JKOsPKCmhmRCgxt9e0yn3sfUKb24Yc29sJ7C+y+36SepUUpsvCd0Li1B9BKFVBr5"
    "2LhLARPMIp50fWnf3oib5GnIcdvb2gCFNC5GEzNrCGUNIVcjhH8x82s/Oj6Qk/BDZs6wYmO78RvmsBiD4tjF0MZRFv3yuKJSAyyL3y9fnFaR/SiZDJTyPNRq"
    "Ol++prNxCoMlC77LGrNR6TKUxLzNzHPpYgBz31KRcvA+Q/UGVzd5uSUo5VJQSrmglNuCkrzesdtIxZPxcY3KH5Z994kqL12IKuPxGB92WdVq4bA6Vf5JoUQR"
    "fgB1MeSeFGRMjRyJi1WpGAB6Fwpnp1qpQrpK1bNRqkLMyjVisJwWGUEvWYDN9Qhpx2L8irteQnRLx4aXn3NHxjS8RdLIM7vEAgZck1bKgNgYqHLNJK5OJ8Fg"
    "VfU5C6WlzVyzdxzOg0o8t8oW0CrMqnr4gZNeh3KrY6SryoKAjysUDbUBMcfcyKqqh3t41GmQri7LNlMGguNY3FMWXkdMMFD0XEBNFpLf6ESPS2zAuUgdYmNY"
    "b5+uPaUDkopTPW3mGqW004u0OL3ItdMLDTmiSxOEyh9h9LJPKsOcv+KunaYcUzDLuVlxpgbyXVp4GtFHGA/VhCjrmMpc3S5Pu5VXQK3WmN6adnqXX/ZTrZRw"
    "aov2XlNskLY4EnC43KYTYK6Evr4GFgl914/YHYuWJ8zFG8O3LOdj8fxexDm/b9Rxra+7sTDZc/r67k4jeYQtoXtqxWM6ShmbTk1zPVK+YmHB4H9i93ispJR9"
    "+QzTNRgdOihDEzaNR9ROyTSmS++YguqC/O5i+fpNUpMlazfI5dbQGLb29u2rgx5ec1aN9Thz2MDLAhqXiIHQpsXZGTfNnWugbrbftMo6lV7hhcNHdaChBaa0"
    "c8T1bt7K2ZicrXHj/zFZVs8yh/GuEiPXgsdMuS1p8/TrMJ5kJUcnz8xM0+HJj1rmWRqYJbfMTLMk3h4oi6ver3j15WT/ReGmvy0AyU1eANFe/eT16eHrlzxK"
    "bJJW+X0TV4avUukMVVzsvhqLGzN6yFotwszVdTRJK6+B6O6+DSYYNThFMK42hlbSLum/ShoP+Op73LILb2K89CcZ+mJ7Dk3VXa4SynoPcS+fupM4RN0qJFl8"
    "B93tPVCUEn5Y5Wb8JpkEtqfuZxIvFYRN3QjWqXwyCEFSltDvAHFwjmZDEtLCaPHjV3GuAcYwuTEag2py6oYYTUfeQ5LwN8OLLuteTt1kks/JngrH5/jm4OP4"
    "Hua5e5P6xIAL2L/+GcO/HP+hRwUKR1jD27qT0bh2g2Pcn50P3He/KH/np/doUXDHgm7t7pgoUzoDDU+Mb42PZNo5qPFx7tsIIR4O8eEu5RFpNTH0F7W4dllC"
    "uU6SqKa9ihl18LcyI5O/oVImdVG+LK1r3z/kk4YzLQD/m0+t2uGbd6/3Tq7Ofz85vNp/+/zw6u07hRLGQTQZsNpPweSaXU3u4IORTmmu5Nv6WRtmbBsDA9DI"
    "NASh3HkhJ8wHompO7aEo88+cYwVJRr2tefjY4KVqTr/A29ionSaDNLydsOwHqmmd1B41yXMXqNQWfz/qNUPF9puvrapmi7L/zNdqdFDboDbccikHymVhrGg0"
    "y1IXV2uDJJ9bQaPWbrZq60W/nb4+KGKNuA55HBfu5bMhpqNr980Vs8GoRHvXtE+OKM6WOfbFS/H0+VjbVVOo1pOjISFVHTJoD2+gd34UAbl4jUcNA7TVqLWg"
    "1474Z5Dyz1zQwfHsNpp/uVjWhSErZX2xCJjqCVFrEbdOTEKjxG109QLmM+1EnITNBm+u+IPvS5+PeplmentdW/OKr75Vu4SXrH0LVEVFy/DRwmjhehJCw+0+"
    "/PxUi+Fnba08QjSyYe1nTy4+ZZR/5lzYN6FTM6l9mNAoX2aAXAJjA3o3F3Q0tDu6juO0Kr/6osdPrK5UT/Vbtl5gJK7a8pkCXBewszUDx/5S4Odf7rW4mNZr"
    "uQPuszLL0YkKJd6jYoacqEwbr6QPuismHHLxTUvnFQLu4rkiIxmNw0hZcTXKrIUzRZmTzncoDg5aiDQpgmer5Tx8EtfJFPuCXL7OsTQKjy0cET7GBF0S8eys"
    "oXgQnp1hQ+eJWGQwEjZ5cGQ/t0ToBSKukZLrR+nA8YrigBGv1ZA+/q4wXHo8GSMgpT5Y/spsJ7VXiCOYYuRnMSlcYy9CFaO4r48yGhnBOMureuYrkG5urBfz"
    "27DhzH0xBrqjGrZ6Vk2AOTsUBdak+cciEkx8p+T7xpg8mtMbPcSfPRgVwFkl7V5UQYuyloB6PNQl1OMhhn4pXO6oGaN48ioW3PbjU8z8zJr5n4akoAJuMEtS"
    "/HzRxdKvGJEK527p4GVqzapHfFtsxreVWx+XuF2nfV5VXxCzvyAAsGlF5YXJoTJZ6I1dZfQAz7dc8mGvhV94dAJ1P2a9L9LpfrvV6brSrqL3AQXJ3VwZWvR+"
    "5csdlxlJOoi9j0OycJPWgRqC991c2YkKuvG0PLFNmel93DhGF8tuLvzYFi9AlGww98UQXpi9WFn3dcWZ748oa1EYmsy70CJztTEQw4UepQtDciXWS3cjLyEl"
    "zH6EZ5+BRxcyNHHGHQIoL24+9xOKlYXYPE7sW34XwLgbbVfifXC1VvCCI341E3zgvRkYJ/loXTYgvckAAwiQt6eH50PdieoAHWhzFyunL5+fJ/w+OH41xhE/"
    "XuuVg//nsCHrcPrwlIz9IMzRafcEPT5xkz2zFehbdTN9ado3q9KcV2oaB2+b9o7sf7r9Fmp0ULU4FpEn0SyYR5kSDqh8AmCQ6F7bRuboGnLL5eh4BDCvUiyK"
    "mI6DS0boxqv7Q9gFO7v7Q+AI4Hm/CMmhrQih407QdI3PnsZfQ5cibohkDqmpORWG5akQGFMhmlYuG+ZdYzVb+7kIKCHWsV2+BT1uC3R6eqnl9y1H+ayzqaho"
    "qVdVTXE+ylcAuuQcNqWuo482YCj18Iis5JzBaOjvvzO8wswiUtYTEo8XvZHrESfcIi0Wf7Rn1FUYooE3sEWp1fASkMxXVWAxeXNaw4Z5w2ir6eLc6azijIlL"
    "MwaXRDk/Ur6saXuBHZmvWJbQn5uaJX1mLh7FV9WGr+pgaH5Whef+l0D55Sr9Xce/sACuil+8ov7H8KIBz87//CE37su+xgTCPFxvi0V3qlMAa7xEEAxkocOf"
    "uT+Imod+dPOcF85E8bomjtaVP/ECM3ORT+WYUKmOOtXrE8SgDZskqGK1L3NRfYMAtLZbWTFb4jDLhwBWURsowXg0gfauWY2FIMekhogNMohe3Xd1+SUF2ucQ"
    "XmL0t6EXqc8BVtBfhhfxJbyEgRdmL8IYpBLgt3eFb8iNRmd14vQ60lfkRvdHd+ylGwN3ZKC31wgBRD4gu7s6dnqvhv3Rz6+Glqnav7LwduTjTUQ/BiH++4d0"
    "6uJ1nDxJoDvwDms+adKjqBZE4RgktawWwkbMHVKwAZQYTZW4+nnIkC9itREIdKPJCGvKkFVMap3Wv+QI3KBFH07lWwxFZHCTQOTamv6BbIxdMbnZl3Fjna2y"
    "jQ6woPIelQtbSb57u+bFvfwnNFeCx84qRjM0qr3RedQbDFOIfzZu+0Nr8ZGfvAsfE++TyBkBREixAnIDEO2FC6gxrdDQ28VTtGFTalcEmg9vvckFcwEZAIBr"
    "eQRgZz0G0MwVrLw2DWeuZBXY4lXc4WwLL/u0rnX5pIJ318DtcL2zeue4DTx/CXdbPVrs1gC22gh/3twN13fWNnuwvXZX71zCTKrWQbVPBrQO4oHGiY/eXfJU"
    "+jpX5xOfEyO3BM9D4J6X9c4k8efE5uW2GNL13rtMBOSVJU3fUo9s1ig8hwZuSKJkO0HLlW/RcnYPH9toedeEZrmF7ct2z33lftCs4anDYJb+mnHIh95zOiOy"
    "LUG93/Agj4fEeMPkozW1glBmVIAqnQbfVIQrhqXIDlkcerFgcA5jfvduQWTh9crQwvPjDbd5UA4enIOH5khnhQIWITpTy6N9OCP2MA/gqQ2DiN85pvEUsl8G"
    "nJfH+nK/F73NnhpHudxLETV5Zp9s38GZK30iy94JEaTX7Lh5ClvZ2Mdw9Ogo2VDulUZhRuhkzmitF0PBwxj5cd7nUZCu8RD8pzZb3+HBYeDBCJCMHGpza1UN"
    "n2sk2DyHxRm68mY/tXbbIJVwh7tcftGo4erueUGX+VKr7NYesXSoMk9aNlTpRy8ZpZJfsVwoq1/vY2SkK7/19/FsB+Fk8J7A7rjpdhz3Ea64n+BZXEYe0r9M"
    "i/K66YO80h3zAmfkszxsPyLeqPKkXRWlZglv2b0qnGqH2MU8zrxzXz2r2XXFYcCras7k50fv9totCj4PyzlG+NrehBV8c/NZ98fOJoimOz9uPtvZ3IG5Ki9L"
    "pNL9OTxnjnLrdoGKJe1uCr8LDOukcQs43mCrjU7h2j0SoloYN0JnNXcDnsZDUUz3fSPstkhFkApky0OvvRLvpr2sn/DcYTMF5u8W/rtGhqsEsxQynnEDNYLa"
    "GNaGPr3W8414lVk8echtj3gF+UZodCcuusOc1QBGUnUH07I7MXQgVcRFQFwExEVIMPMa+VrbeVy97sJ6p1PZP7FdRYu2K582oWhuFOyEAnbxlSFSW+yCIO/m"
    "VmtO1bqY39W+OLw3vpFrw3kP/GgPBI7EmvVHwSxXiGKH99EjjfxwY5QQmgEFajihnIaBRz/oqjYZsMir/3J6/qJu5EMVMVk90lLihyCDoXEaSSQN86iHl5iS"
    "5v1G+mStxAHhQSFMMxM/X1RALLrUThJFyc0NN0Ka0ZKOQm0ZZfLFhbT2lKNZHK3q9gY6CrZnlskXF9Lae+1/mT+OowIB29Lx80UFZDtEo7qIg760qiJJFtOB"
    "vcKnPbThkPuflnmMHpB4JqvOfAmTSIbUrGTS0D3gTP6Hu+mgioGNAAb3HlXkQHUOL21lhfsKxNyQMvEKKu9GX9s9qTLNBv00cPcD9yRw3INAaaPbTpO0e/9Q"
    "Ft6N/cA6RuJ9467+5MLzwQ5EIT/CSYq863k4YmtiOwQhCocmB9ABi3K/r9XS5IfhpxQ8+x2K/HtUFojFWKB0jVHg/j4P997E/TgP9y+Om6R4m8ufT8aBIkND"
    "/30e+n0J/eM8dCSGuwIhtlvFqqD+U8eQYqwRco+LSjll1B7WIjgNYUX7POQubP8INODr8AusmG8QBC97gL5/jfi7+9GMnK53IrLEDPH2ciNt+rz9PjEyy77O"
    "f0OE08N3r85eHb/B76d5+vL5HrnYPYEKyG84N+9EuHDwfSiS4p6ocvvd7YqMq72z8/2rdutLu3UlcrvPdto/2tlbeu62nbut5z6zc3f03B0zt2M23GmVsjt6"
    "dtvI3vyyqWW2zKq3zMwfrUyjQ2ar22Zm28o0OtsxMnfMkl0r0yi5aWUao7QlM5+fQKbM2N78UbV3eL7fuTrcU5nHMuPk3SkU6Tw/OXnXlpl/mZmbRuZzmXnW"
    "hbyDD+cqZ9/O6cqcEztH9fuA5xyoY7DvOeBQAX4VgMIHvRgnATjnqdKUFV1/3d5WVf0mQM8U5B8cgsPTLsat/eN2AVcT6oWAVY3YGyPPHLBXIq80XqeUUVDM"
    "p1WpIzSh/M/8qEwJNa/o+/bvvZTxB7zcwyptnXEHJo+17Y2WEOQwajKypinTrBEococQLW/gUYRfEpdXCvMlfmuD3zlTGzMsTZn3gPb7vYepi+tRD5pKIZ1E"
    "A7REPjo+wBwum1l5XM0K2VMZVLugsykhRaxqpIcjoWtj3J5tjZ+Ig4rFE/K2MyNYETMjsKpC6vyGty4iZje3HHSlkCZ0tQSjYU6u7YacZmFn6BgDyJyeILds"
    "3F5Jr2q3wejdrDG6gb0hkus8aZI0qwvIhawDG0KaLzxGUnxa5RgUJFd6yFbTzfSlxkuRJ7VerQ7k0pWXKd4hTjNorfAtgZqGWPfu/33Q4AygG7t0NppBlY3f"
    "ArQcscpnZgUL/D5gxXiIpuoubFi0JnBfN+5ve3lAoEGx7OxwgP19dmirYGE89NOB7gqfFiAz+MGQgyI/YEZIKfrEGV2ROZtck1d5LZoCVVRyIepdJwJuQtAg"
    "Y+dq7+jkl72d6u16qz0LTd+ONndmYen7EgVvrcTSNyiKCFuBZW3nW52ZaPq2vtWtRDO3983qJs1tfrM9A8kYiGqqzG1/szsDyRiszUokkw3Y3JqBZNS0PQNJ"
    "H/VNZK3IhNQ7z+Qjj6nCJpimHWV/OIk/ecNQAY7Ca29QJNWO82ssYHjPRsAOEg4bM++vSD5a8/S1wqkA0anNIJBJTt2YZ8PnKaNY0XuSV/FDX0sJdd5wSDB1"
    "icz7nnpMcadkZFp6j3Re4P2aq2eL2rNE5VTA8HpO1RbrtYWNbssUdeXlU7FLDkMpyZJRsxRdxQ5ny8nLFeY60eqS0v+qqiDHv3olOf/lVY38T+zMv9HjH7bZ"
    "+rZhzYi1qDPg3OWPYdwQarn1glqh1JLbNo9r3rVJVZs67CB46Wk/SdJBpra+HKWuWTgzh6wwdclX8zW2ytbiVW4J4GloYtPlA9EqBqklw5zzHH6ak/txBw2j"
    "HdccBh/1jUPYeje0munEypEDQIFklNKFjgukNC9nkFOotrmQqWbbL346SuIwyLoeo89kHKl4Wwf0yUlvId57I7lknNP/RpRTeU4YC8UrsKYXLVf+a2u/8nld"
    "T7TUeafIattOD5E36Hakr13XCHGA+msxWVc7LiyFelq631Wa5NRxlZYZntuuysm0HHhuO9N4kfo4JNWwbqeKquHqs6G+nCJMeqKIlQ2UuumM2YuCogLK7Gio"
    "mDnv8Gb+NOgbyimOJvWPuxWwXpt14e3kfLxz3wgWSje5tG5rp6QMPvTccf0lQo7m3Fq8gqPFKqXp6FC/cm4dS5ljMudsqgpxzgEVl3G8d5l6VrvoKcHsO7Kb"
    "GvDMh6UD0GUmqSBKF2dpg0PNFcDH928xgo7X3WpxZoHgB6n/mcNZroCnzB8UyMR+AGEsEcYT5e1OzAt+WqHjCn0/d7PliW+e3bMzNvaare1NXbg4EobaBkz6"
    "B2Cxfx2xRtsxc2evK4RwWlHlqVVlxzFzZ1Z5FfjBkHkP5AGMTGPdm+SOP/AO8mcUxUS2fEA/WPyJ953fzpVfVcm4kxriil9qrHBF9vffCLoTgDtM8qa5/wB6"
    "XNWGHBGQHu6/BR6oApG+4UkkTnnvQgCnUX4JPAWStekCzSVCOBmupKKSBpeTIAhwqX3eust0R2GubFlvV20MwKhwmTakSHPo50Vso7BZyJIbHZD4wlXe4Iag"
    "GJd0AhSrTJKvEvmrIMBvcCK4t1U36X8Omoz7t80u2p1Lbz10ryxQ6PreerYqe72WuolnJOMCv3XpdSRBjWQdvXgXmTuXXiNZ8x2RY0ztcndxKYkds+11s+31"
    "b9H26cy2pxUfHz/v4OucGeBZOqlsfA4qv7ElSl4FYq3MQQ4fmYtYR8GLRQzhLQU317E2hwPjb2rOc+KSylJ2WwOHd0xFqEPB7vz47f4v3oOKbIkRLds8vuUV"
    "RbcUzwKhi304hw0Nr7KbXhmQ2pmxRrzzxMy2M75YQTK8QwkvQyYg5lznRUIEwxtj9yvCfnh7icyoAJXjc3hHRqYFhm2Nzu3+YFrqhR/jqyO/bR0NTmFdVU5b"
    "y8kUlMatFKXDO00E3IS8fSeVMCjHvg2fGCDbKrgwfjPHXxjBeVhEcK4s4Q1jkfWkINlWwaWoXiJM9sCkuhwneyCp3nki0TuPo3lxqOzEJLkUKztRFO9H/gg4"
    "76cSbhdfjn5RamE3Iqsb1eW8iHrDLwl4H4MikaFGZxJqAK5m+cgLoDdOqO0+Z6Rt+aIBX8V5Z3OHwGcmmGCfNRipazbhf1Jl88zO3NraasvMnVLm9pbM+9HO"
    "o4wrgL47e134ecF1mesUOt5drhJd76hIbHpnRcLcBlhEGWiVY7xsTaNQ9bp5iRnvV6IQHfZbPJI6nqlqWERRxbXwXThgiVy291DB9Z5dvzxqn0r/NNeZhOHF"
    "J9M7TSizKC4Hv5VgYBzmEsMAn1eCsf7HDEllBY0av3JVEwHSaiKyTM3hkVw+1yr70qiR8WuBLUcvDkn5MjUJhnH5rMaFT+oRpaXtpdoW9hMdLHQjS3/kVsG5"
    "r7/Utj0RUJRH74umP0jTFrsIfrkLUvP2s2fPOu0t9D6FvUj98R7d4zzkHMp7AH745VRs8Eva0qoCc0xoX4SRYBsk7Rijloj4yNIElVAolzJBxycBV/JoSwDO"
    "omQskd5JmDr/QKwMT3jloezvrnAkxv3TKK9luVu/umIZ8AOTiNVd6byoNcVgIf3/9X8BKyGdXDU1CQA="
)

def _three_js_source():
    """@brief Gömülü Three.js (r128 UMD) kaynak metnini döndürür.
    
    @return Üretilen sonuç.
    """
    import gzip, base64
    return gzip.decompress(base64.b64decode(_THREE_MIN_GZ_B64)).decode("utf-8")
